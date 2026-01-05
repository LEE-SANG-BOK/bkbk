#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook


_JIMOK_TOKENS = {
    "답": "답",
    "전": "전",
    "임": "임야",
    "대": "대",
    "도": "도로",
    "구": "구거",
    "잡": "잡종지",
}

# Extract patterns like:
# - "151-3 답"
# - "151-3답"
# - "638구"
_OCR_PAIR_RE = re.compile(r"(?<!\d)(\d{2,4}(?:-\d{1,4})?)(?:\s*)(답|전|임|대|도|구|잡)\b")


def _discover_ocr_texts(case_dir: Path) -> list[Path]:
    cands: list[Path] = []
    for rel in ["_analysis/drawings/text", "attachments/derived", "_analysis/tmp_ocr"]:
        base = case_dir / rel
        if not base.exists():
            continue
        cands.extend(sorted([p for p in base.rglob("*.txt") if p.is_file()]))
    return cands


def _parse_parcel_tokens(parcel_no: str) -> list[str]:
    s = str(parcel_no or "").strip()
    if not s:
        return []
    tail = s.split()[-1]
    if "," not in tail:
        return [tail]

    parts = [p.strip() for p in tail.split(",") if p.strip()]
    if not parts:
        return []

    first = parts[0]
    base = first.split("-", 1)[0].strip()
    tokens: list[str] = []
    for p in parts:
        if "-" in p:
            tokens.append(p)
            continue
        if base:
            tokens.append(f"{base}-{p}")
    return tokens


def _choose_single(counter: Counter[str]) -> str | None:
    if not counter:
        return None
    best = counter.most_common()
    if not best:
        return None
    if len(best) >= 2 and best[0][1] == best[1][1]:
        return None  # ambiguous
    return best[0][0]


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Auto-fill PARCELS.jimok by parsing OCR text extracted from official drawings (best-effort, conservative)."
    )
    ap.add_argument("--xlsx", required=True, type=Path, help="Path to case.xlsx(v2)")
    ap.add_argument("--out", type=Path, default=None, help="Write to this path (default: in-place)")
    ap.add_argument(
        "--ocr-text",
        type=Path,
        action="append",
        default=[],
        help="OCR text file to parse (repeatable). If omitted, auto-discovers under case_dir.",
    )
    ap.add_argument(
        "--append-src-id",
        type=str,
        default="",
        help="When filling a row, append this src_id token to PARCELS.src_id (e.g., S-CLIENT-DRAWINGS-2025-12-29).",
    )
    ap.add_argument("--dry-run", action="store_true", help="Print planned changes without writing")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    out_path = args.out.expanduser().resolve() if args.out else xlsx
    case_dir = xlsx.parent.resolve()

    ocr_paths = [p.expanduser().resolve() for p in (args.ocr_text or []) if p]
    if not ocr_paths:
        ocr_paths = _discover_ocr_texts(case_dir)
    if not ocr_paths:
        raise SystemExit(f"No OCR text files found under {case_dir} (use --ocr-text to specify)")

    counts_by_parcel: dict[str, Counter[str]] = defaultdict(Counter)
    parsed_pairs = 0
    for p in ocr_paths:
        try:
            text = p.read_text(encoding="utf-8", errors="ignore")
        except Exception:
            continue
        for num, tok in _OCR_PAIR_RE.findall(text):
            parsed_pairs += 1
            mapped = _JIMOK_TOKENS.get(tok)
            if not mapped:
                continue
            counts_by_parcel[str(num)] += Counter({mapped: 1})

    if parsed_pairs == 0:
        raise SystemExit("No (parcel_no, jimok) patterns found in OCR texts (check OCR quality / file selection).")

    wb = load_workbook(xlsx)
    if "PARCELS" not in wb.sheetnames:
        raise SystemExit("Missing PARCELS sheet in case.xlsx(v2)")
    ws = wb["PARCELS"]

    header = [c.value for c in ws[1]]
    hm: dict[str, int] = {}
    for idx, h in enumerate(header, start=1):
        if not h:
            continue
        hm[str(h).strip()] = idx

    def _col(name: str) -> int:
        c = hm.get(name)
        if not c:
            raise SystemExit(f"PARCELS missing required column: {name}")
        return c

    col_parcel_no = _col("parcel_no")
    col_jimok = _col("jimok")
    col_src = hm.get("src_id") or 0

    append_src = str(args.append_src_id or "").strip()

    changes: list[tuple[int, str, str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        parcel_no = ws.cell(row=row_idx, column=col_parcel_no).value
        if parcel_no is None or (isinstance(parcel_no, str) and not parcel_no.strip()):
            continue

        cur_jimok = ws.cell(row=row_idx, column=col_jimok).value
        if cur_jimok is not None and str(cur_jimok).strip():
            continue  # don't overwrite

        tokens = _parse_parcel_tokens(str(parcel_no))
        if not tokens:
            continue

        chosen: str | None = None
        if len(tokens) == 1:
            chosen = _choose_single(counts_by_parcel.get(tokens[0], Counter()))
        else:
            per = []
            for t in tokens:
                per.append(_choose_single(counts_by_parcel.get(t, Counter())))
            per = [x for x in per if x]
            if per and len(set(per)) == 1 and len(per) == len(tokens):
                chosen = per[0]

        if not chosen:
            continue

        # src_id update (optional)
        src_before = ""
        src_after = ""
        if col_src and append_src:
            src_val = ws.cell(row=row_idx, column=col_src).value
            src_before = "" if src_val is None else str(src_val).strip()
            parts = [s.strip() for s in re.split(r"[;,]", src_before) if s.strip()] if src_before else []
            if append_src not in parts:
                parts.append(append_src)
            src_after = ";".join(parts)

        changes.append((row_idx, str(parcel_no).strip(), chosen, src_after))

        if not args.dry_run:
            ws.cell(row=row_idx, column=col_jimok).value = chosen
            if col_src and append_src and src_after:
                ws.cell(row=row_idx, column=col_src).value = src_after

    print(f"OCR files: {len(ocr_paths)} (parsed_pairs={parsed_pairs})")
    print(f"PARCELS rows updated: {len(changes)}")
    for row_idx, parcel_no, jimok, src_after in changes[:30]:
        extra = f", src_id+={append_src}" if append_src and src_after else ""
        print(f"- row {row_idx}: {parcel_no} -> jimok='{jimok}'{extra}")

    if args.dry_run:
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK wrote {out_path}")


if __name__ == "__main__":
    main()

