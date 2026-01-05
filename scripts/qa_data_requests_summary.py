#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any


try:
    import openpyxl
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
        f"Import error: {e}"
    )


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _is_empty_row(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def _as_bool(v: Any) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if not s:
        return None
    if s in {"true", "t", "1", "y", "yes"}:
        return True
    if s in {"false", "f", "0", "n", "no"}:
        return False
    return None


def _as_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


_MISSING_ENV_RE = re.compile(r"missing env\s+([A-Z0-9_]+)", re.IGNORECASE)


def _extract_env_vars(note: str) -> list[str]:
    """
    Best-effort parse of planner notes like:
    - "(disabled: missing env SAFEMAP_API_KEY)"
    - "(disabled: missing env AIRKOREA_API_KEY or DATA_GO_KR_SERVICE_KEY)"
    """
    if not note:
        return []

    # Primary: "missing env X"
    envs = [m.group(1).upper() for m in _MISSING_ENV_RE.finditer(note)]

    # Secondary: "...missing env A or B" patterns may only capture A.
    # Grab trailing tokens around "missing env" and split by common separators.
    lowered = note.lower()
    if "missing env" in lowered:
        tail = note[lowered.rfind("missing env") :].replace("missing env", "")
        tail = tail.replace("(", " ").replace(")", " ")
        for token in re.split(r"[\s,;/]+", tail):
            t = token.strip()
            if not t:
                continue
            t = t.upper()
            if t in {"OR", "AND"}:
                continue
            if re.fullmatch(r"[A-Z][A-Z0-9_]{2,}", t):
                envs.append(t)

    # de-dupe preserving order
    seen = set()
    out: list[str] = []
    for e in envs:
        if e not in seen:
            seen.add(e)
            out.append(e)
    return out


@dataclass(frozen=True)
class DataRequestRow:
    req_id: str
    enabled: bool | None
    connector: str
    purpose: str
    output_sheet: str
    run_mode: str
    last_run_at: str
    last_evidence_ids: str
    note: str


def _load_data_requests(case_xlsx: Path) -> list[DataRequestRow]:
    wb = openpyxl.load_workbook(case_xlsx, data_only=True)
    if "DATA_REQUESTS" not in wb.sheetnames:
        return []

    ws = wb["DATA_REQUESTS"]
    headers = [c.value for c in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    required = [
        "req_id",
        "enabled",
        "connector",
        "purpose",
        "output_sheet",
        "run_mode",
        "last_run_at",
        "last_evidence_ids",
        "note",
    ]
    missing = [c for c in required if c not in header_map]
    if missing:
        raise SystemExit(f"DATA_REQUESTS headers missing columns: {missing}")

    out: list[DataRequestRow] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue

        req_id = _as_str(row[header_map["req_id"]])
        if not req_id:
            continue

        out.append(
            DataRequestRow(
                req_id=req_id,
                enabled=_as_bool(row[header_map["enabled"]]),
                connector=_as_str(row[header_map["connector"]]).upper(),
                purpose=_as_str(row[header_map["purpose"]]).upper(),
                output_sheet=_as_str(row[header_map["output_sheet"]]).upper(),
                run_mode=_as_str(row[header_map["run_mode"]]).upper(),
                last_run_at=_as_str(row[header_map["last_run_at"]]),
                last_evidence_ids=_as_str(row[header_map["last_evidence_ids"]]),
                note=_as_str(row[header_map["note"]]),
            )
        )

    out.sort(key=lambda x: x.req_id)
    return out


def _load_json(path: Path) -> dict[str, Any] | None:
    if not path.exists():
        return None
    try:
        obj = json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None
    return obj if isinstance(obj, dict) else None


def _load_validation_report(path: Path) -> dict[str, Any]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(obj, dict) or "results" not in obj or not isinstance(obj["results"], list):
        raise SystemExit("validation_report JSON must be an object with a top-level 'results' list")
    if "stats" not in obj or not isinstance(obj["stats"], dict):
        obj["stats"] = {}
    return obj


def _append_rule(
    obj: dict[str, Any],
    *,
    rule_id: str,
    severity: str,
    message: str,
    fix_hint: str | None,
    path: str | None,
) -> None:
    obj["results"].append(
        {
            "rule_id": rule_id,
            "severity": severity,
            "message": message,
            "fix_hint": fix_hint,
            "path": path,
        }
    )


def _recompute_stats(obj: dict[str, Any]) -> None:
    results = obj.get("results") or []
    err = sum(1 for r in results if (r.get("severity") == "ERROR"))
    warn = sum(1 for r in results if (r.get("severity") == "WARN"))
    info = sum(1 for r in results if (r.get("severity") == "INFO"))

    stats = obj.setdefault("stats", {})
    stats["error_count"] = int(err)
    stats["warn_count"] = int(warn)
    stats["info_count"] = int(info)


def _build_summary(rows: list[DataRequestRow], run_json: dict[str, Any] | None) -> dict[str, Any]:
    enabled_rows = [r for r in rows if (r.enabled is True)]
    disabled_rows = [r for r in rows if (r.enabled is False)]
    unknown_rows = [r for r in rows if (r.enabled is None)]

    executed_est = [r for r in enabled_rows if (r.last_run_at or r.last_evidence_ids)]

    missing_env_map: dict[str, list[str]] = {}
    disabled_other: list[dict[str, Any]] = []
    for r in disabled_rows:
        envs = _extract_env_vars(r.note)
        if envs:
            for env in envs:
                missing_env_map.setdefault(env, []).append(r.req_id)
        else:
            disabled_other.append({"req_id": r.req_id, "note": r.note})

    for k in list(missing_env_map.keys()):
        missing_env_map[k] = sorted(set(missing_env_map[k]))

    summary: dict[str, Any] = {
        "generated_at": _utc_now_iso(),
        "counts": {
            "total": len(rows),
            "enabled": len(enabled_rows),
            "disabled": len(disabled_rows),
            "enabled_unknown": len(unknown_rows),
            "executed_est": len(executed_est),
        },
        "disabled_missing_env": missing_env_map,
        "disabled_other": disabled_other,
    }
    if run_json:
        summary["run_json"] = {
            "executed": run_json.get("executed"),
            "skipped": run_json.get("skipped"),
            "warnings_count": len(run_json.get("warnings") or []),
            "evidences_count": len(run_json.get("evidences") or []),
        }
    return summary


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Summarize DATA_REQUESTS status (enabled/disabled/needs keys) and (optionally) append to validation_report. "
            "This is a post-process helper to improve QA usability without touching core QA code."
        )
    )
    ap.add_argument("--case-xlsx", required=True, type=Path, help="case.xlsx(v2) path")
    ap.add_argument(
        "--data-requests-run",
        type=Path,
        default=None,
        help="Optional _data_requests_run.json path (default: <case_dir>/_data_requests_run.json)",
    )
    ap.add_argument("--out-summary", type=Path, default=None, help="Write summary JSON to this path")

    ap.add_argument("--validation-report", type=Path, default=None, help="validation_report_*.json to augment")
    ap.add_argument("--out-report", type=Path, default=None, help="Output path for augmented report JSON")
    ap.add_argument("--in-place", action="store_true", help="Overwrite validation report in place")

    args = ap.parse_args()

    case_xlsx = args.case_xlsx.expanduser().resolve()
    if not case_xlsx.exists():
        raise SystemExit(f"case.xlsx not found: {case_xlsx}")

    case_dir = case_xlsx.parent
    run_json_path = (args.data_requests_run.expanduser().resolve() if args.data_requests_run else (case_dir / "_data_requests_run.json"))
    run_json = _load_json(run_json_path)

    rows = _load_data_requests(case_xlsx)
    summary = _build_summary(rows, run_json)

    if args.out_summary:
        out_summary = args.out_summary.expanduser().resolve()
        out_summary.parent.mkdir(parents=True, exist_ok=True)
        out_summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"OK wrote summary: {out_summary}")

    if args.validation_report:
        report_path = args.validation_report.expanduser().resolve()
        if not report_path.exists():
            raise SystemExit(f"validation_report not found: {report_path}")

        out_report = args.out_report.expanduser().resolve() if args.out_report else report_path
        if out_report == report_path and not args.in_place:
            raise SystemExit("Refusing to overwrite validation report without --in-place (or pass --out-report)")

        rep = _load_validation_report(report_path)

        counts = summary["counts"]
        rj = summary.get("run_json") or {}
        msg = (
            f"DATA_REQUESTS summary: total={counts['total']} enabled={counts['enabled']} disabled={counts['disabled']} "
            f"executed_est={counts['executed_est']} (run_json executed={rj.get('executed')} skipped={rj.get('skipped')})"
        )
        _append_rule(
            rep,
            rule_id="DATA_REQUESTS_SUMMARY",
            severity="INFO",
            message=msg,
            fix_hint="If you expected more auto-fill, run `eia-gen run-data-requests` (or generate-xlsx-both --enrich).",
            path="DATA_REQUESTS",
        )

        missing_env = summary.get("disabled_missing_env") or {}
        if missing_env:
            env_parts = [f"{env}({len(reqs)} reqs)" for env, reqs in sorted(missing_env.items())]
            _append_rule(
                rep,
                rule_id="DATA_REQUESTS_DISABLED_MISSING_ENV",
                severity="INFO",
                message="DATA_REQUESTS has disabled rows due to missing env vars: " + ", ".join(env_parts),
                fix_hint="Set env vars in `eia-gen/.env.local` (or shell env) then re-run planner/runner, or keep them disabled.",
                path="DATA_REQUESTS",
            )

        disabled_other = summary.get("disabled_other") or []
        if disabled_other:
            # Keep message short; details remain in DATA_REQUESTS.note.
            reqs = [str(x.get("req_id") or "") for x in disabled_other if str(x.get("req_id") or "").strip()]
            msg2 = "DATA_REQUESTS has disabled guidance rows: " + ", ".join(reqs[:10])
            _append_rule(
                rep,
                rule_id="DATA_REQUESTS_DISABLED_GUIDANCE",
                severity="INFO",
                message=msg2,
                fix_hint="Either keep them disabled, or fill required params_json fields and set enabled=TRUE.",
                path="DATA_REQUESTS",
            )

        _recompute_stats(rep)
        out_report.parent.mkdir(parents=True, exist_ok=True)
        out_report.write_text(json.dumps(rep, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"OK wrote augmented report: {out_report}")


if __name__ == "__main__":
    main()
