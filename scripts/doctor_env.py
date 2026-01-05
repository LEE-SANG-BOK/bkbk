#!/usr/bin/env python3
from __future__ import annotations

import argparse
import os
import platform
import shutil
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional


@dataclass(frozen=True)
class CheckResult:
    name: str
    ok: bool
    detail: str
    fix_hint: str = ""


def _load_dotenv_files(repo_root: Path) -> list[Path]:
    """Best-effort dotenv loader for doctor checks (never prints values).

    We intentionally keep this dependency-free so `doctor_env.py` can be run even
    outside the project venv.
    """
    loaded: list[Path] = []
    for fname in [".env.local", ".env"]:
        p = repo_root / fname
        if not p.exists():
            continue
        try:
            for raw in p.read_text(encoding="utf-8").splitlines():
                line = raw.strip()
                if not line or line.startswith("#"):
                    continue
                if line.startswith("export "):
                    line = line[len("export ") :].lstrip()
                if "=" not in line:
                    continue
                k, v = line.split("=", 1)
                k = k.strip()
                if not k or k[0].isdigit() or not all(ch.isalnum() or ch == "_" for ch in k):
                    continue
                v = v.strip()
                if len(v) >= 2 and v[0] == v[-1] and v[0] in {"'", '"'}:
                    v = v[1:-1]
                if not v:
                    continue
                if not os.environ.get(k, "").strip():
                    os.environ[k] = v
            loaded.append(p)
        except Exception:
            # Fail-open: dotenv parsing is best-effort.
            continue
    return loaded


def _check_import(module_name: str) -> CheckResult:
    pip_name_map = {
        "PIL": "Pillow",
        "docx": "python-docx",
        "yaml": "PyYAML",
        "fitz": "PyMuPDF",
        "pytesseract": "pytesseract",
        "skimage": "scikit-image",
    }
    try:
        __import__(module_name)
        return CheckResult(name=f"python:{module_name}", ok=True, detail="import OK")
    except Exception as e:
        return CheckResult(
            name=f"python:{module_name}",
            ok=False,
            detail=str(e),
            fix_hint=f"pip install {pip_name_map.get(module_name, module_name)}",
        )


def _check_cmd(cmd: str) -> CheckResult:
    p = shutil.which(cmd)
    if p:
        return CheckResult(name=f"cmd:{cmd}", ok=True, detail=p)
    return CheckResult(
        name=f"cmd:{cmd}",
        ok=False,
        detail="not found in PATH",
        fix_hint=f"Install '{cmd}' (e.g., via Homebrew/apt) or disable related features.",
    )


def _check_app_bundle(app_basename: str) -> CheckResult:
    if platform.system() != "Darwin":
        return CheckResult(name=f"app:{app_basename}", ok=True, detail="n/a (non-macOS)")

    candidates = [
        Path("/Applications") / f"{app_basename}.app",
        Path.home() / "Applications" / f"{app_basename}.app",
    ]
    for p in candidates:
        if p.exists():
            return CheckResult(name=f"app:{app_basename}", ok=True, detail=str(p))

    return CheckResult(
        name=f"app:{app_basename}",
        ok=False,
        detail="not found in /Applications",
        fix_hint=f"Install {app_basename}.app (or adjust your post-processing mode).",
    )


def _check_soffice_any() -> CheckResult:
    """
    Align with scripts/postprocess_docx_to_pdf.py:
    - Accept 'soffice' on PATH
    - Or LibreOffice.app bundled soffice
    """
    p = shutil.which("soffice")
    if p:
        return CheckResult(name="soffice:any", ok=True, detail=p)

    if platform.system() == "Darwin":
        candidates = [
            Path("/Applications/LibreOffice.app/Contents/MacOS/soffice"),
            Path.home() / "Applications/LibreOffice.app/Contents/MacOS/soffice",
        ]
        for c in candidates:
            if c.exists():
                return CheckResult(name="soffice:any", ok=True, detail=f"{c} (from LibreOffice.app)")

    return CheckResult(
        name="soffice:any",
        ok=False,
        detail="not found",
        fix_hint="Install LibreOffice (soffice) or use Pages/Word backend on macOS.",
    )


def _check_font_file_env(env_var: str) -> CheckResult:
    """
    Optional: deterministic watermark/caption fonts for reproducible outputs.
    - If env var is not set: OK (optional)
    - If set: must exist on disk
    """
    raw = os.getenv(env_var, "").strip()
    if not raw:
        return CheckResult(name=f"env:{env_var}", ok=True, detail="not set (optional)")

    p = Path(raw).expanduser()
    if not p.is_absolute():
        p = p.resolve()
    if p.exists() and p.is_file():
        return CheckResult(name=f"env:{env_var}", ok=True, detail=str(p))
    return CheckResult(
        name=f"env:{env_var}",
        ok=False,
        detail=f"not found: {p}",
        fix_hint="Set the env var to an existing .ttf/.otf font file path (or unset it to use system fallback).",
    )


def _check_bundled_fonts_dir(repo_root: Path) -> CheckResult:
    """
    Optional: when you bundle a font pack in the repo (e.g., assets/fonts/*.ttf),
    ensure at least one font file exists.
    """
    d = repo_root / "assets" / "fonts"
    if not d.exists():
        return CheckResult(name="repo:assets/fonts", ok=True, detail="not present (optional)")
    if not d.is_dir():
        return CheckResult(name="repo:assets/fonts", ok=False, detail=f"not a directory: {d}")

    fonts = sorted([p for p in d.iterdir() if p.is_file() and p.suffix.lower() in {'.ttf', '.otf'}])
    if fonts:
        return CheckResult(name="repo:assets/fonts", ok=True, detail=f"{len(fonts)} font(s) found")
    return CheckResult(
        name="repo:assets/fonts",
        ok=False,
        detail=f"no .ttf/.otf files in {d}",
        fix_hint="Place a font file under assets/fonts/ (or remove the folder to rely on system fonts).",
    )


def _fmt(res: CheckResult) -> str:
    tag = "OK" if res.ok else "MISSING"
    s = f"[{tag:7}] {res.name}: {res.detail}"
    if (not res.ok) and res.fix_hint:
        s += f"\n          fix: {res.fix_hint}"
    return s


def _split_ids(raw: object) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []

    # Support "A;B", "A,B", "A / B" and line breaks.
    # NOTE: Do NOT treat "/" as a separator globally because paths contain "/".
    s = s.replace("\n", ";")
    s = s.replace(",", ";")
    s = s.replace(" / ", ";")
    s = s.replace(" /", ";")
    s = s.replace("/ ", ";")
    parts = []
    for p in s.split(";"):
        t = p.strip()
        if t:
            parts.append(t)
    return parts


def _load_sources_yaml_source_ids(sources_yaml: Path) -> tuple[set[str], Optional[str]]:
    """
    Supports:
    - v2: sources: [ {source_id: ...}, ... ]
    - legacy: sources: { SRC_ID: {...}, ... }
    Returns (source_ids, error_message)
    """
    try:
        import yaml  # type: ignore
    except Exception as e:
        return set(), f"PyYAML not available: {e}"

    try:
        obj = yaml.safe_load(sources_yaml.read_text(encoding="utf-8")) or {}
    except Exception as e:
        return set(), f"failed to parse YAML: {e}"

    raw = obj.get("sources")
    if raw is None:
        return set(), "missing top-level key: sources"

    ids: set[str] = set()
    if isinstance(raw, list):
        for item in raw:
            if not isinstance(item, dict):
                continue
            sid = item.get("source_id") or item.get("src_id") or item.get("id")
            if sid:
                ids.add(str(sid).strip())
        return ids, None

    if isinstance(raw, dict):
        for k in raw.keys():
            ids.add(str(k).strip())
        return ids, None

    return set(), f"unsupported sources type: {type(raw).__name__}"


def _extract_column_indexes(headers: Iterable[object], *, targets: set[str]) -> list[int]:
    idxs: list[int] = []
    for i, h in enumerate(headers):
        if h is None:
            continue
        name = str(h).strip()
        if name in targets:
            idxs.append(i)
    return idxs


def _check_case_dir(case_dir: Path) -> list[CheckResult]:
    req_subdirs = [
        "attachments",
        "attachments/inbox",
        "attachments/normalized",
        "attachments/derived",
        "attachments/evidence",
    ]
    out: list[CheckResult] = []
    for rel in req_subdirs:
        p = case_dir / rel
        out.append(
            CheckResult(
                name=f"case-dir:{rel}",
                ok=p.exists(),
                detail=str(p) if p.exists() else f"missing: {p}",
                fix_hint="Run `eia-gen init-case ...` to generate skeleton, or create the folders manually.",
            )
        )

    # case.xlsx / sources.yaml existence (best-effort)
    case_xlsx = case_dir / "case.xlsx"
    out.append(
        CheckResult(
            name="case-file:case.xlsx",
            ok=case_xlsx.exists(),
            detail=str(case_xlsx) if case_xlsx.exists() else f"missing: {case_xlsx}",
            fix_hint="Run `eia-gen init-case ...` or place your case.xlsx in the case folder root.",
        )
    )
    sources_yaml = case_dir / "sources.yaml"
    out.append(
        CheckResult(
            name="case-file:sources.yaml",
            ok=sources_yaml.exists(),
            detail=str(sources_yaml) if sources_yaml.exists() else f"missing: {sources_yaml}",
            fix_hint="Copy a template sources.yaml into the case folder root (or generate it via init-case).",
        )
    )

    # sheet sanity (only when openpyxl is available and case.xlsx exists)
    if case_xlsx.exists():
        try:
            import openpyxl  # type: ignore

            wb = openpyxl.load_workbook(case_xlsx, read_only=True)
            sheets = set(wb.sheetnames)
            required_sheets = {
                "META",
                "PROJECT",
                "LOCATION",
                "FIGURES",
                "ATTACHMENTS",
                "SSOT_PAGE_OVERRIDES",
            }
            missing = sorted(required_sheets - sheets)
            out.append(
                CheckResult(
                    name="case-xlsx:required-sheets",
                    ok=not missing,
                    detail="OK" if not missing else f"missing sheets: {missing}",
                    fix_hint="Re-generate case.xlsx from v2 template (`eia-gen init-case`) if structure is broken.",
                )
            )

            # (best-effort) sources.yaml: ensure all referenced src_id values exist.
            if sources_yaml.exists():
                src_ids, err = _load_sources_yaml_source_ids(sources_yaml)
                if err:
                    out.append(
                        CheckResult(
                            name="case-sources:parse",
                            ok=False,
                            detail=err,
                            fix_hint="Fix sources.yaml format (or ensure PyYAML is installed).",
                        )
                    )
                else:
                    referenced: set[str] = set()
                    for ws in wb.worksheets:
                        headers = next(
                            ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple()
                        )
                        src_idx = _extract_column_indexes(headers, targets={"src_id", "src_ids"})
                        if not src_idx:
                            continue
                        for row in ws.iter_rows(min_row=2, values_only=True):
                            for i in src_idx:
                                if i < len(row):
                                    referenced.update(_split_ids(row[i]))

                    sentinels = {"S-TBD", "SRC-TBD", "S-UNKNOWN", "SRC-UNKNOWN"}
                    missing_src = sorted(x for x in referenced if x and x not in src_ids and x not in sentinels)
                    out.append(
                        CheckResult(
                            name="case-sources:missing-src_id",
                            ok=not missing_src,
                            detail="OK" if not missing_src else f"missing: {missing_src[:30]}",
                            fix_hint=(
                                "Add the missing source_id entries into sources.yaml (sources: [...]). "
                                "Tip: copy an existing entry and adjust title/publisher/citation, "
                                "or run scripts/fix_missing_sources.py to append stubs."
                            ),
                        )
                    )

            # (best-effort) file path existence for key columns.
            # - Absolute paths are checked as-is.
            # - Relative paths are resolved from:
            #   1) case-dir root
            #   2) case-dir/attachments
            #   3) case-dir/attachments/{normalized,derived}
            missing_paths: list[str] = []
            checked = 0
            max_checks = 200
            max_missing = 30
            file_cols = {"file_path", "override_file_path", "boundary_file", "photo_folder"}
            for ws in wb.worksheets:
                headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
                path_idx = _extract_column_indexes(headers, targets=file_cols)
                if not path_idx:
                    continue
                for row in ws.iter_rows(min_row=2, values_only=True):
                    for i in path_idx:
                        if i >= len(row):
                            continue
                        for raw_path in _split_ids(row[i]):
                            if not raw_path:
                                continue
                            checked += 1
                            if checked > max_checks:
                                break
                            p0 = Path(raw_path)
                            candidates: list[Path] = []
                            if p0.is_absolute():
                                candidates = [p0]
                            else:
                                candidates = [
                                    (case_dir / p0).resolve(),
                                    (case_dir / "attachments" / p0).resolve(),
                                    (case_dir / "attachments" / "normalized" / p0).resolve(),
                                    (case_dir / "attachments" / "derived" / p0).resolve(),
                                ]

                            if not any(c.exists() for c in candidates):
                                missing_paths.append(f"{ws.title}:{headers[i]}={raw_path}")
                                if len(missing_paths) >= max_missing:
                                    break
                        if checked > max_checks or len(missing_paths) >= max_missing:
                            break
                    if checked > max_checks or len(missing_paths) >= max_missing:
                        break
                if checked > max_checks or len(missing_paths) >= max_missing:
                    break

            out.append(
                CheckResult(
                    name="case-files:paths-exist",
                    ok=not missing_paths,
                    detail="OK" if not missing_paths else f"missing: {missing_paths[:max_missing]}",
                    fix_hint=(
                        "Fix file_path/override_file_path/boundary_file/photo_folder values in case.xlsx "
                        "(relative paths are resolved from case-dir, case-dir/attachments, and attachments/{normalized,derived})."
                    ),
                )
            )
        except Exception as e:
            out.append(
                CheckResult(
                    name="case-xlsx:required-sheets",
                    ok=False,
                    detail=f"failed to read workbook: {e}",
                    fix_hint="Ensure openpyxl works and case.xlsx is not corrupted.",
                )
            )

    return out


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Environment doctor for eia-gen (local only). "
            "Checks python modules and external tools used by PDFs/SSOT/Page overrides and optional post-processing."
        )
    )
    ap.add_argument(
        "--no-dotenv",
        action="store_true",
        help="Do not load repo-root `.env.local`/`.env` before env checks.",
    )
    ap.add_argument(
        "--case-dir",
        type=Path,
        default=None,
        help="Optional case folder to validate attachments/* skeleton (expects case.xlsx parent style).",
    )
    ap.add_argument(
        "--strict-case",
        action="store_true",
        help="Exit non-zero if any case structure/sources/paths checks fail (only when --case-dir is set).",
    )
    args = ap.parse_args()

    print("== eia-gen doctor ==")
    print(f"python: {sys.executable}")
    print(f"version: {sys.version.splitlines()[0]}")
    print(f"os: {platform.system()} {platform.release()} ({platform.machine()})")
    repo_root = Path(__file__).resolve().parents[1]
    if not args.no_dotenv:
        loaded = _load_dotenv_files(repo_root)
        if loaded:
            names = ", ".join(p.name for p in loaded)
            print(f"dotenv: loaded {names} (best-effort, no override)")
    venv_dir = repo_root / ".venv"
    venv_python = venv_dir / "bin" / "python"
    # NOTE: Don't compare resolved executables; venv python is often a symlink to the base python.
    # Use sys.prefix to detect whether we're *running inside* the venv.
    if venv_python.exists() and Path(sys.prefix).resolve() != venv_dir.resolve():
        print("")
        print(f"hint: This repo has a venv: {venv_dir}")
        print(f"      Consider running: {venv_python} {Path(__file__).name} ...")
        print(f"      (or) source {venv_dir}/bin/activate")
    print("")

    # Core python deps (from pyproject.toml)
    checks: list[CheckResult] = []
    for mod in [
        "openpyxl",
        "PIL",
        "pydantic",
        "yaml",
        "docx",
        "shapely",
        "pyproj",
        "typer",
        "rich",
        "fitz",
        "pytesseract",
        # Optional: overlay polygon extraction helpers (scripts/extract_overlay_polygons.py).
        "skimage",
    ]:
        checks.append(_check_import(mod))

    # Optional PDF/OCR deps: prefer external binaries (tesseract/pdftoppm) for portability.
    for cmd in ["tesseract", "pdftoppm", "pdfinfo"]:
        checks.append(_check_cmd(cmd))

    # Optional post-processing tools
    if platform.system() == "Darwin":
        checks.append(_check_cmd("osascript"))
        checks.append(_check_app_bundle("Pages"))
        checks.append(_check_app_bundle("Microsoft Word"))
    checks.append(_check_soffice_any())  # LibreOffice (PATH or app bundle)

    # Optional fonts for deterministic watermark/caption rendering (golden tests stability).
    checks.append(_check_font_file_env("EIA_GEN_WATERMARK_FONT_PATH"))
    checks.append(_check_font_file_env("EIA_GEN_FONT_PATH"))
    checks.append(_check_bundled_fonts_dir(repo_root))

    # API keys (best-effort). Some keys have safe fallbacks and are treated as optional.
    optional_env_fix_hints = {
        "VWORLD_API_KEY": (
            "Optional: set VWORLD_API_KEY for higher-quality/stable geocoding; "
            "a best-effort fallback is available when it's not set."
        ),
    }
    keys = ["VWORLD_API_KEY", "SAFEMAP_API_KEY", "DATA_GO_KR_SERVICE_KEY", "OPENAI_API_KEY"]
    for k in keys:
        v = os.getenv(k, "")
        is_set = bool(v.strip())
        is_optional = k in optional_env_fix_hints
        checks.append(
            CheckResult(
                name=f"env:{k}",
                ok=is_set or is_optional,
                detail=("set" if is_set else "not set (optional)" if is_optional else "not set"),
                fix_hint=(
                    optional_env_fix_hints[k]
                    if (not is_set and is_optional)
                    else f"Export {k} in your shell or put it in `.env.local`/`.env` (if you use dotenv)."
                ),
            )
        )

    if args.case_dir:
        case_dir = args.case_dir.expanduser().resolve()
        checks.extend(_check_case_dir(case_dir))

    # Print grouped report
    ok = sum(1 for c in checks if c.ok)
    total = len(checks)
    print(f"checks: {ok}/{total} OK")
    print("")
    for c in checks:
        print(_fmt(c))

    # Exit code: if a core python dep is missing, treat as failure.
    core_missing = [
        c
        for c in checks
        if c.name in {"python:openpyxl", "python:PIL", "python:docx", "python:pydantic", "python:yaml"} and not c.ok
    ]
    if core_missing:
        raise SystemExit(2)

    if args.strict_case and args.case_dir:
        case_prefixes = (
            "case-dir:",
            "case-file:",
            "case-xlsx:",
            "case-sources:",
            "case-files:",
        )
        case_missing = [c for c in checks if (not c.ok) and c.name.startswith(case_prefixes)]
        if case_missing:
            raise SystemExit(1)


if __name__ == "__main__":
    main()
