#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
from dataclasses import dataclass
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any, Iterable

import openpyxl
import yaml


_WS_RE = re.compile(r"\s+")
_KEEP_RE = re.compile(r"[^0-9A-Za-z가-힣]+")


def _compact(text: str) -> str:
    s = _WS_RE.sub(" ", (text or "")).strip()
    s = _KEEP_RE.sub("", s).strip().lower()
    return s


@dataclass(frozen=True)
class SpecCandidate:
    kind: str  # "FIGURE" or "TABLE"
    scope: str  # "EIA" or "DIA"
    spec_id: str
    caption: str
    compact: str


def _load_yaml(path: Path) -> dict[str, Any]:
    obj = yaml.safe_load(path.read_text(encoding="utf-8"))
    return obj if isinstance(obj, dict) else {}


def _iter_specs_from_yaml(path: Path, *, scope: str) -> Iterable[SpecCandidate]:
    obj = _load_yaml(path)
    figures = obj.get("figures") if isinstance(obj.get("figures"), list) else []
    tables = obj.get("tables") if isinstance(obj.get("tables"), list) else []

    for f in figures:
        if not isinstance(f, dict):
            continue
        sid = str(f.get("id") or "").strip()
        cap = str(f.get("caption") or "").strip()
        if sid and cap:
            yield SpecCandidate(kind="FIGURE", scope=scope, spec_id=sid, caption=cap, compact=_compact(cap))

    for t in tables:
        if not isinstance(t, dict):
            continue
        sid = str(t.get("id") or "").strip()
        cap = str(t.get("caption") or "").strip()
        if sid and cap:
            yield SpecCandidate(kind="TABLE", scope=scope, spec_id=sid, caption=cap, compact=_compact(cap))


def _score(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    seq = SequenceMatcher(a=a, b=b).ratio()

    # Char n-gram Jaccard is often more stable for Korean captions (spacing varies).
    def _ngrams(s: str, n: int) -> set[str]:
        if len(s) < n:
            return {s} if s else set()
        return {s[i : i + n] for i in range(len(s) - n + 1)}

    def _jaccard(x: set[str], y: set[str]) -> float:
        if not x or not y:
            return 0.0
        inter = len(x & y)
        union = len(x | y)
        return inter / union if union else 0.0

    g2 = _jaccard(_ngrams(a, 2), _ngrams(b, 2))
    g3 = _jaccard(_ngrams(a, 3), _ngrams(b, 3))

    # Weighted blend; keep within [0,1].
    blended = 0.60 * seq + 0.25 * g2 + 0.15 * g3
    return max(seq, blended)


def main() -> None:
    ap = argparse.ArgumentParser(
        description="Suggest our spec_id candidates for coverage_matrix_seed.xlsx via fuzzy matching (caption ↔ spec captions)."
    )
    ap.add_argument("--in-xlsx", type=Path, required=True, help="Input coverage_matrix_seed.xlsx path")
    ap.add_argument("--out-xlsx", type=Path, default=None, help="Output xlsx path (default: <in>.suggested.xlsx)")
    ap.add_argument("--top-n", type=int, default=3, help="Top N suggestions per row (default 3)")
    ap.add_argument(
        "--min-score",
        type=float,
        default=0.35,
        help="Do not write suggestions below this score (default 0.35)",
    )
    ap.add_argument(
        "--spec-figure",
        type=Path,
        default=Path("spec/figure_specs.yaml"),
        help="EIA figure_specs.yaml path (default: spec/figure_specs.yaml)",
    )
    ap.add_argument(
        "--spec-table",
        type=Path,
        default=Path("spec/table_specs.yaml"),
        help="EIA table_specs.yaml path (default: spec/table_specs.yaml)",
    )
    ap.add_argument(
        "--spec-dia-figure",
        type=Path,
        default=Path("spec_dia/figure_specs.yaml"),
        help="DIA figure_specs.yaml path (default: spec_dia/figure_specs.yaml)",
    )
    ap.add_argument(
        "--spec-dia-table",
        type=Path,
        default=Path("spec_dia/table_specs.yaml"),
        help="DIA table_specs.yaml path (default: spec_dia/table_specs.yaml)",
    )
    args = ap.parse_args()

    in_xlsx = args.in_xlsx
    if not in_xlsx.exists():
        raise SystemExit(f"Input xlsx not found: {in_xlsx}")

    out_xlsx = args.out_xlsx
    if not out_xlsx:
        out_xlsx = in_xlsx.with_suffix("").with_name(in_xlsx.stem + ".suggested.xlsx")

    repo_root = Path(__file__).resolve().parents[1]  # eia-gen/

    def _resolve_path(p: Path) -> Path:
        if p.is_absolute():
            return p
        return (repo_root / p).resolve()

    candidates: list[SpecCandidate] = []
    for p, scope in [
        (_resolve_path(args.spec_figure), "EIA"),
        (_resolve_path(args.spec_table), "EIA"),
        (_resolve_path(args.spec_dia_figure), "DIA"),
        (_resolve_path(args.spec_dia_table), "DIA"),
    ]:
        if p.exists():
            candidates.extend(list(_iter_specs_from_yaml(p, scope=scope)))

    by_kind: dict[str, list[SpecCandidate]] = {"FIGURE": [], "TABLE": []}
    for c in candidates:
        by_kind.setdefault(c.kind, []).append(c)

    wb = openpyxl.load_workbook(in_xlsx)
    ws = wb.active

    # Header mapping
    header_row = [c.value for c in ws[1]]
    header_map: dict[str, int] = {}
    for idx, h in enumerate(header_row, start=1):
        if not h:
            continue
        header_map[str(h).strip()] = idx

    def _col(name: str) -> int:
        if name not in header_map:
            raise SystemExit(f"Missing required column '{name}' in {in_xlsx}")
        return header_map[name]

    col_kind = _col("kind")
    col_caption = _col("caption")
    col_our = _col("our_spec_id")

    # Ensure suggestion columns exist (idempotent: do not keep appending on repeated runs).
    # If some suggestion columns already exist, we reuse/overwrite them.
    def _ensure_suggest_columns(top_n: int) -> dict[str, int]:
        nonlocal header_map
        start = ws.max_column + 1

        for i in range(1, int(top_n) + 1):
            for suffix in ("id", "score", "scope", "caption"):
                name = f"suggest_{i}_{suffix}"
                if name in header_map:
                    continue
                ws.cell(row=1, column=start).value = name
                header_map[name] = start
                start += 1
        return header_map

    _ensure_suggest_columns(int(args.top_n))

    # Fill rows
    for r in range(2, ws.max_row + 1):
        kind = str(ws.cell(row=r, column=col_kind).value or "").strip().upper()
        if kind not in ("FIGURE", "TABLE"):
            continue

        our_id = str(ws.cell(row=r, column=col_our).value or "").strip()
        if our_id:
            continue  # already filled by human

        # Clear previous suggestion values for this row (overwrite-safe).
        for i in range(1, int(args.top_n) + 1):
            for suffix in ("id", "score", "scope", "caption"):
                cidx = header_map.get(f"suggest_{i}_{suffix}")
                if cidx:
                    ws.cell(row=r, column=cidx).value = None

        cap = str(ws.cell(row=r, column=col_caption).value or "").strip()
        cap_compact = _compact(cap)
        if not cap_compact:
            continue

        scored: list[tuple[float, SpecCandidate]] = []
        for c in by_kind.get(kind, []):
            s = _score(cap_compact, c.compact)
            scored.append((s, c))
        scored.sort(key=lambda x: x[0], reverse=True)

        out_items = []
        for s, c in scored[: int(args.top_n)]:
            if s < float(args.min_score):
                continue
            out_items.append((c.spec_id, f"{s:.3f}", c.scope, c.caption))

        for i in range(int(args.top_n)):
            if i < len(out_items):
                sid, ss, scope, sc = out_items[i]
                ws.cell(row=r, column=header_map[f"suggest_{i+1}_id"]).value = sid
                ws.cell(row=r, column=header_map[f"suggest_{i+1}_score"]).value = ss
                ws.cell(row=r, column=header_map[f"suggest_{i+1}_scope"]).value = scope
                ws.cell(row=r, column=header_map[f"suggest_{i+1}_caption"]).value = sc

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"OK wrote: {out_xlsx}")


if __name__ == "__main__":
    main()
