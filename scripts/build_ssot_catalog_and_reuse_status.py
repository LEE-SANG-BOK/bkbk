#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


@dataclass(frozen=True)
class ChapterRange:
    name: str  # e.g. "CH2_지역개황"
    start: int
    end: int


_CH_FILE_RE = re.compile(r"^CH(?P<num>\d+)_(?P<title>.+)_p(?P<start>\d+)-(?P<end>\d+)\.pdf$", re.IGNORECASE)


def _load_chapter_ranges(split_dir: Path) -> list[ChapterRange]:
    ranges: list[ChapterRange] = []
    for p in sorted(split_dir.glob("CH*_p*.pdf")):
        m = _CH_FILE_RE.match(p.name)
        if not m:
            continue
        try:
            start = int(m.group("start"))
            end = int(m.group("end"))
        except Exception:
            continue
        if start <= 0 or end <= 0 or end < start:
            continue
        title = m.group("title").strip()
        num = m.group("num").strip()
        ranges.append(ChapterRange(name=f"CH{num}_{title}", start=start, end=end))

    ranges.sort(key=lambda r: r.start)
    if not ranges:
        raise SystemExit(f"No chapter split PDFs found under: {split_dir}")
    return ranges


def _chapter_for_page(ranges: list[ChapterRange], page: int) -> ChapterRange | None:
    for r in ranges:
        if r.start <= page <= r.end:
            return r
    return None


def _read_hits(index_json: Path) -> list[dict[str, Any]]:
    data: dict[str, Any] = json.loads(index_json.read_text(encoding="utf-8"))
    hits = data.get("pass2", {}).get("hits", [])
    return [h for h in hits if isinstance(h, dict)]


def _read_sheet_rows(case_xlsx: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(case_xlsx, data_only=True)
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(c or "").strip() for c in rows[0]]
    out: list[dict[str, Any]] = []
    for raw in rows[1:]:
        if raw is None:
            continue
        # Stop when first 3 columns are empty (conventional "end" marker).
        if all((raw[i] is None or str(raw[i]).strip() == "") for i in range(min(3, len(raw)))):
            continue
        obj: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if i >= len(raw):
                continue
            obj[h] = raw[i]
        if obj:
            out.append(obj)
    return out


def _to_int(v: Any, default: int | None = None) -> int | None:
    if v is None:
        return default
    try:
        return int(float(v))
    except Exception:
        return default


def _to_float(v: Any, default: float | None = None) -> float | None:
    if v is None:
        return default
    try:
        return float(v)
    except Exception:
        try:
            return float(str(v).strip())
        except Exception:
            return default


def _sort_key_num(ch_name: str) -> int:
    m = re.match(r"^CH(\d+)\b", ch_name)
    if not m:
        return 999
    try:
        return int(m.group(1))
    except Exception:
        return 999


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--index-json",
        type=Path,
        default=Path("output/pdf_extract/changwon_gingerfarm_2025_twopass/combined_index.json"),
        help="combined_index.json (extract_pdf_index_twopass.py 결과)",
    )
    ap.add_argument(
        "--split-dir",
        type=Path,
        default=Path("output/pdf_split/changwon_sample_gingerfarm_2025"),
        help="장(章)별 분할 PDF 디렉터리",
    )
    ap.add_argument(
        "--case-dir",
        type=Path,
        default=Path("output/case_new_max_reuse"),
        help="신규 케이스 디렉터리(내부에 case.xlsx 존재)",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path("output/pdf_split/changwon_sample_gingerfarm_2025/CATALOG_AND_REUSE_STATUS.md"),
        help="생성할 Markdown 파일",
    )
    ap.add_argument(
        "--keywords",
        type=str,
        default="위치도; 용지도; 현황도; 현황평면; 계획평면; 배수; 종단; 횡단; 단면; 구조물; 부대공; 토지이용계획도; 조사지점도",
        help="치환 우선 후보(그림 캡션) 키워드(세미콜론 구분)",
    )
    args = ap.parse_args()

    chapter_ranges = _load_chapter_ranges(args.split_dir)
    hits = _read_hits(args.index_json)

    case_xlsx = (args.case_dir / "case.xlsx").resolve()
    if not case_xlsx.exists():
        raise SystemExit(f"Missing case.xlsx: {case_xlsx}")

    overrides_raw = _read_sheet_rows(case_xlsx, "SSOT_PAGE_OVERRIDES")
    overrides: list[dict[str, Any]] = []
    override_pages: set[int] = set()
    for r in overrides_raw:
        sample_page = _to_int(r.get("sample_page"))
        override_page = _to_int(r.get("override_page"), 1) or 1
        file_path = str(r.get("override_file_path") or "").strip()
        if not sample_page or not file_path:
            continue
        overrides.append(
            {
                "sample_page": int(sample_page),
                "override_file_path": file_path,
                "override_page": int(override_page),
                "width_mm": _to_float(r.get("width_mm")),
                "dpi": _to_int(r.get("dpi")),
                "crop": str(r.get("crop") or "").strip() or None,
                "src_id": str(r.get("src_id") or "").strip() or None,
                "note": str(r.get("note") or "").strip() or None,
            }
        )
        override_pages.add(int(sample_page))
    overrides.sort(key=lambda o: int(o["sample_page"]))

    inserts_raw = _read_sheet_rows(case_xlsx, "APPENDIX_INSERTS")
    inserts: list[dict[str, Any]] = []
    for r in inserts_raw:
        ins_id = str(r.get("ins_id") or "").strip()
        file_path = str(r.get("file_path") or "").strip()
        if not ins_id or not file_path:
            continue
        page_raw = r.get("page")
        if page_raw is None:
            page_raw = r.get("pdf_page")
        inserts.append(
            {
                "ins_id": ins_id,
                "order": _to_int(r.get("order"), 0) or 0,
                "file_path": file_path,
                "page": _to_int(page_raw, 1) or 1,
                "caption": str(r.get("caption") or "").strip() or None,
                "note": str(r.get("note") or "").strip() or None,
                "src_id": str(r.get("src_id") or "").strip() or None,
            }
        )
    inserts.sort(key=lambda x: (int(x.get("order") or 0), str(x.get("ins_id") or "")))

    # Catalog: group table/figure captions by chapter.
    by_ch: dict[str, dict[str, list[dict[str, Any]]]] = {}
    for r in chapter_ranges:
        by_ch[r.name] = {"tables": [], "figures": []}

    for h in hits:
        kind = str(h.get("kind") or "").strip().lower()
        if kind not in {"table", "figure"}:
            continue
        page = _to_int(h.get("page"))
        if not page:
            continue
        ch = _chapter_for_page(chapter_ranges, int(page))
        if not ch:
            continue
        label = str(h.get("label") or "").strip()
        text = str(h.get("text") or "").strip().replace("\n", " ")
        by_ch[ch.name]["tables" if kind == "table" else "figures"].append(
            {"page": int(page), "label": label, "text": text}
        )

    for ch_name, obj in by_ch.items():
        obj["tables"].sort(key=lambda x: (int(x["page"]), str(x["label"])))
        obj["figures"].sort(key=lambda x: (int(x["page"]), str(x["label"])))

    # Overrides grouped by chapter
    overrides_by_ch: dict[str, list[dict[str, Any]]] = {r.name: [] for r in chapter_ranges}
    for o in overrides:
        ch = _chapter_for_page(chapter_ranges, int(o["sample_page"]))
        if ch:
            overrides_by_ch[ch.name].append(o)

    # Candidate pages (keyword in FIGURE captions)
    keywords = [k.strip() for k in str(args.keywords or "").split(";") if k.strip()]
    kw_lower = [k.lower() for k in keywords]
    candidates: list[dict[str, Any]] = []
    for ch_name, obj in by_ch.items():
        for f in obj["figures"]:
            t = str(f.get("text") or "")
            if not t:
                continue
            t_lower = t.lower()
            if not any(k in t_lower for k in kw_lower):
                continue
            candidates.append(
                {
                    "chapter": ch_name,
                    "page": int(f["page"]),
                    "label": str(f.get("label") or ""),
                    "text": t,
                    "status": "OVERRIDDEN" if int(f["page"]) in override_pages else "NEED",
                }
            )
    candidates.sort(key=lambda c: (int(c["page"]), str(c["label"])))

    # Write
    total_pages = sum(r.end - r.start + 1 for r in chapter_ranges)
    overridden_pages = len(override_pages)
    remaining_pages = max(0, total_pages - overridden_pages)
    pct = (overridden_pages / total_pages * 100.0) if total_pages else 0.0

    lines: list[str] = []
    lines.append("# 샘플 PDF(기허가) 장별 카탈로그 + 신규 케이스 치환 현황")
    lines.append("")
    lines.append(f"- 샘플 PDF: `{args.index_json}`")
    lines.append(f"- 장별 분할 PDF: `{args.split_dir}`")
    lines.append(f"- 신규 케이스: `{case_xlsx}`")
    lines.append("")
    lines.append("## 1) 샘플 PDF 장(章) 구성(커버 OCR)")
    toc_path = args.split_dir / "TOC_sections.md"
    if toc_path.exists():
        lines.append(f"- `{toc_path}`")
    else:
        lines.append(f"- (missing) `{toc_path}`")
    lines.append("")

    lines.append("## 2) 장별(章) 캡션 카탈로그(표/그림)")
    for r in sorted(chapter_ranges, key=lambda x: _sort_key_num(x.name)):
        obj = by_ch.get(r.name) or {"tables": [], "figures": []}
        tables = obj.get("tables") or []
        figures = obj.get("figures") or []
        lines.append(f"### {r.name} (p{r.start:03d}-{r.end:03d})")
        lines.append(f"- tables: `{len(tables)}` / figures: `{len(figures)}`")
        if figures:
            lines.append("**Figures**")
            for f in figures:
                lines.append(f'- p{int(f["page"]):03d} `<그림 {f["label"]}>` {f["text"]}')
        if tables:
            lines.append("**Tables**")
            for t in tables:
                lines.append(f'- p{int(t["page"]):03d} `<표 {t["label"]}>` {t["text"]}')
        lines.append("")

    lines.append("## 3) 신규 케이스: 샘플 페이지 치환(SSOT_PAGE_OVERRIDES) 현황")
    lines.append(f"- overrides: `{len(overrides)}`")
    lines.append(f"- total ssot pages: `{total_pages}`, overridden: `{overridden_pages}`, remaining(sample): `{remaining_pages}`")
    lines.append(f"- overall progress: `{pct:.1f}%`")
    lines.append("")
    for r in sorted(chapter_ranges, key=lambda x: _sort_key_num(x.name)):
        ovs = overrides_by_ch.get(r.name) or []
        total = r.end - r.start + 1
        lines.append(f"### {r.name} (p{r.start:03d}-{r.end:03d})")
        lines.append(f"- total pages: `{total}`, overridden: `{len(ovs)}`, remaining(sample): `{max(0, total - len(ovs))}`")
        for o in ovs:
            src = f" src=`{o['src_id']}`" if o.get("src_id") else ""
            note = f" note={o['note']}" if o.get("note") else ""
            lines.append(
                f"- p{int(o['sample_page']):03d} -> `{o['override_file_path']}`#p{int(o['override_page'])}{src}{note}"
            )
        lines.append("")

    lines.append("## 4) 신규 케이스: 부록 삽입(APPENDIX_INSERTS) 현황")
    lines.append(f"- inserts: `{len(inserts)}`")
    for ins in inserts:
        cap = f" cap={ins['caption']}" if ins.get("caption") else ""
        src = f" src=`{ins['src_id']}`" if ins.get("src_id") else ""
        note = f" note={ins['note']}" if ins.get("note") else ""
        lines.append(
            f"- {ins['ins_id']} order={int(ins.get('order') or 0)} `{ins['file_path']}`#p{int(ins.get('page') or 1)}{cap}{src}{note}"
        )
    lines.append("")

    lines.append("## 5) (초안) ‘치환 우선’ 후보 페이지(샘플 내 캡션 키워드 기반)")
    lines.append(f"- keywords: `{'; '.join(keywords)}`")
    lines.append("")
    if not candidates:
        lines.append("- (no candidates)")
    else:
        for c in candidates:
            lines.append(
                f"- [{c['status']}] {c['chapter']} p{int(c['page']):03d} FIGURE `{c['label']}` {c['text']}"
            )
    lines.append("")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    args.out.write_text("\n".join(lines), encoding="utf-8")
    print(f"OK wrote {args.out}")


if __name__ == "__main__":
    main()
