#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
import re
import shutil
import subprocess
from typing import Any


try:
    import openpyxl
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
        f"Import error: {e}"
    )


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _is_empty_row(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if str(v).strip() != "":
            return False
    return True


def _as_int(v: Any) -> int | None:
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        return int(float(s))
    except Exception:
        return None


def _as_float(v: Any) -> float | None:
    if v is None:
        return None
    try:
        s = str(v).strip()
        if not s:
            return None
        return float(s)
    except Exception:
        return None


@dataclass(frozen=True)
class SsotOverride:
    sample_page: int
    override_file_path: str
    override_page: int
    width_mm: float | None
    dpi: int | None
    crop: str
    src_id: str
    note: str

    resolved_file_path: str
    file_exists: bool
    file_size_bytes: int | None


_PDFINFO_PAGES_RE = re.compile(r"^Pages:\s*(\d+)\s*$", re.MULTILINE)


def _pdf_page_count(pdf_path: Path) -> int | None:
    """
    Best-effort PDF page count via poppler `pdfinfo`.
    - Returns None if pdfinfo is missing or parsing fails.
    """
    if shutil.which("pdfinfo") is None:
        return None
    try:
        out = subprocess.check_output(
            ["pdfinfo", str(pdf_path)],
            stderr=subprocess.STDOUT,
            text=True,
        )
    except Exception:
        return None

    m = _PDFINFO_PAGES_RE.search(out)
    if not m:
        return None
    try:
        return int(m.group(1))
    except Exception:
        return None


def _load_ssot_overrides(case_xlsx: Path) -> list[SsotOverride]:
    wb = openpyxl.load_workbook(case_xlsx, data_only=True)
    if "SSOT_PAGE_OVERRIDES" not in wb.sheetnames:
        return []

    ws = wb["SSOT_PAGE_OVERRIDES"]
    headers = [c.value for c in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers) if h}

    required = [
        "sample_page",
        "override_file_path",
        "override_page",
        "width_mm",
        "dpi",
        "crop",
        "src_id",
        "note",
    ]
    missing = [c for c in required if c not in header_map]
    if missing:
        raise SystemExit(f"SSOT_PAGE_OVERRIDES headers missing columns: {missing}")

    case_dir = case_xlsx.resolve().parent

    out: list[SsotOverride] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue

        sample_page = _as_int(row[header_map["sample_page"]])
        override_file_path = str(row[header_map["override_file_path"]] or "").strip()
        override_page = _as_int(row[header_map["override_page"]])

        if not sample_page or not override_file_path or not override_page:
            # ignore malformed rows (keep script safe)
            continue

        width_mm = _as_float(row[header_map["width_mm"]])
        dpi = _as_int(row[header_map["dpi"]])
        crop = str(row[header_map["crop"]] or "").strip()
        src_id = str(row[header_map["src_id"]] or "").strip()
        note = str(row[header_map["note"]] or "").strip()

        p = Path(override_file_path)
        if not p.is_absolute():
            p = (case_dir / p).resolve()

        exists = p.exists()
        size = int(p.stat().st_size) if exists else None

        out.append(
            SsotOverride(
                sample_page=int(sample_page),
                override_file_path=override_file_path,
                override_page=int(override_page),
                width_mm=width_mm,
                dpi=dpi,
                crop=crop,
                src_id=src_id,
                note=note,
                resolved_file_path=str(p),
                file_exists=bool(exists),
                file_size_bytes=size,
            )
        )

    out.sort(key=lambda x: x.sample_page)
    return out


def _load_validation_report(path: Path) -> dict[str, Any]:
    obj = json.loads(path.read_text(encoding="utf-8"))
    # report can be ValidationReport (dict with results/stats)
    if not isinstance(obj, dict):
        raise SystemExit(f"Unexpected validation report JSON type: {type(obj)}")
    if "results" not in obj or not isinstance(obj["results"], list):
        raise SystemExit("validation_report JSON must have a top-level 'results' list")
    if "stats" not in obj or not isinstance(obj["stats"], dict):
        obj["stats"] = {}
    return obj


def _append_rule(obj: dict[str, Any], *, rule_id: str, severity: str, message: str, fix_hint: str | None, path: str | None) -> None:
    obj["results"].append(
        {
            "rule_id": rule_id,
            "severity": severity,
            "message": message,
            "fix_hint": fix_hint,
            "path": path,
        }
    )


def _recompute_stats(obj: dict[str, Any]) -> None:
    results = obj.get("results") or []
    err = sum(1 for r in results if (r.get("severity") == "ERROR"))
    warn = sum(1 for r in results if (r.get("severity") == "WARN"))
    info = sum(1 for r in results if (r.get("severity") == "INFO"))

    stats = obj.setdefault("stats", {})
    stats["error_count"] = int(err)
    stats["warn_count"] = int(warn)
    stats["info_count"] = int(info)


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Generate SSOT_PAGE_OVERRIDES summary from case.xlsx and (optionally) append it to validation_report.json. "
            "This is a post-process helper to improve QA usability without touching core QA code."
        )
    )
    ap.add_argument("--case-xlsx", required=True, type=Path, help="case.xlsx(v2) path")
    ap.add_argument("--out-summary", type=Path, default=None, help="Write summary JSON to this path")

    ap.add_argument("--validation-report", type=Path, default=None, help="validation_report_*.json to augment")
    ap.add_argument("--out-report", type=Path, default=None, help="Output path for augmented report JSON")
    ap.add_argument("--in-place", action="store_true", help="Overwrite validation report in place")

    args = ap.parse_args()

    case_xlsx = args.case_xlsx.expanduser().resolve()
    if not case_xlsx.exists():
        raise SystemExit(f"case.xlsx not found: {case_xlsx}")

    overrides = _load_ssot_overrides(case_xlsx)

    by_sample_page: dict[int, list[SsotOverride]] = defaultdict(list)
    for o in overrides:
        by_sample_page[int(o.sample_page)].append(o)
    duplicate_sample_pages = sorted([p for p, xs in by_sample_page.items() if len(xs) > 1])

    # PDF page range checks (best-effort)
    pdfinfo_available = shutil.which("pdfinfo") is not None
    pdf_pages_by_override_path: dict[str, int] = {}
    out_of_range: list[dict[str, Any]] = []
    if pdfinfo_available:
        for o in overrides:
            if not o.file_exists:
                continue
            p = Path(o.resolved_file_path)
            if p.suffix.lower() != ".pdf":
                continue

            if o.override_file_path not in pdf_pages_by_override_path:
                pages = _pdf_page_count(p)
                if pages:
                    pdf_pages_by_override_path[o.override_file_path] = pages

            pages = pdf_pages_by_override_path.get(o.override_file_path)
            if pages and (o.override_page < 1 or o.override_page > pages):
                out_of_range.append(
                    {
                        "sample_page": o.sample_page,
                        "override_file_path": o.override_file_path,
                        "override_page": o.override_page,
                        "pdf_pages": pages,
                    }
                )

    summary = {
        "generated_at": _utc_now_iso(),
        "case_xlsx": str(case_xlsx),
        "case_dir": str(case_xlsx.parent),
        "count": len(overrides),
        "missing_files": [o.sample_page for o in overrides if not o.file_exists],
        "duplicate_sample_pages": duplicate_sample_pages,
        "pdfinfo_available": bool(pdfinfo_available),
        "pdf_pages_by_override_file_path": pdf_pages_by_override_path,
        "override_page_out_of_range": out_of_range,
        "overrides": [o.__dict__ for o in overrides],
    }

    # default summary output path
    out_summary = args.out_summary
    if out_summary is None:
        out_summary = case_xlsx.parent / "_out_check" / "ssot_page_overrides_summary.json"

    out_summary = out_summary.expanduser().resolve()
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK wrote SSOT summary: {out_summary}")

    if not args.validation_report:
        return

    report_in = args.validation_report.expanduser().resolve()
    if not report_in.exists():
        raise SystemExit(f"validation report not found: {report_in}")

    report_obj = _load_validation_report(report_in)

    # Summary entry
    if overrides:
        head = ", ".join([f"p{m.sample_page}->{Path(m.override_file_path).name}:p{m.override_page}" for m in overrides[:6]])
        if len(overrides) > 6:
            head += f" …(+{len(overrides)-6})"
        msg = f"SSOT_PAGE_OVERRIDES: {len(overrides)}건 적용됨. {head} (details: {out_summary.name})"
    else:
        msg = "SSOT_PAGE_OVERRIDES: 적용된 치환이 없습니다."

    _append_rule(
        report_obj,
        rule_id="I-SSOT-OVERRIDE-001",
        severity="INFO",
        message=msg,
        fix_hint="SSOT_PAGE_OVERRIDES 시트에서 sample_page/override_file_path/override_page를 확인",
        path="SSOT_PAGE_OVERRIDES",
    )

    # Missing file warnings
    for o in overrides:
        if o.file_exists:
            continue
        _append_rule(
            report_obj,
            rule_id="W-SSOT-OVERRIDE-001",
            severity="WARN",
            message=(
                f"SSOT 치환 파일이 존재하지 않습니다: sample_page={o.sample_page} "
                f"override_file_path='{o.override_file_path}' (resolved='{o.resolved_file_path}')"
            ),
            fix_hint="case_dir 기준 상대경로가 맞는지 확인하거나 attachments/normalized에 파일을 배치",
            path="SSOT_PAGE_OVERRIDES",
        )

    # Out-of-range override_page warnings (PDF only)
    if summary.get("override_page_out_of_range"):
        _append_rule(
            report_obj,
            rule_id="W-SSOT-OVERRIDE-003",
            severity="WARN",
            message=(
                "SSOT_PAGE_OVERRIDES의 override_page가 PDF 페이지 범위를 벗어났습니다. "
                f"(count={len(summary['override_page_out_of_range'])})"
            ),
            fix_hint="override_page(1부터) 값을 PDF 총 페이지 수 이내로 수정하세요(pdfinfo 기준).",
            path="SSOT_PAGE_OVERRIDES",
        )

    # Duplicate sample_page warnings (order-dependent behavior)
    if duplicate_sample_pages:
        head = ", ".join([f"p{p}(x{len(by_sample_page[p])})" for p in duplicate_sample_pages[:8]])
        if len(duplicate_sample_pages) > 8:
            head += f" …(+{len(duplicate_sample_pages)-8})"
        _append_rule(
            report_obj,
            rule_id="W-SSOT-OVERRIDE-002",
            severity="WARN",
            message=(
                "SSOT_PAGE_OVERRIDES에 sample_page 중복이 있습니다(치환 결과가 행 순서에 의존). "
                f"{head}"
            ),
            fix_hint=(
                "sample_page 당 1행만 남기고 정리하세요(중복 삭제). "
                "향후에는 UPSERT 정책(키=sample_page)을 강제하는 쪽이 안전합니다."
            ),
            path="SSOT_PAGE_OVERRIDES",
        )

    # Attach machine-readable summary
    report_obj.setdefault("_meta", {})["ssot_page_overrides_summary"] = {
        "summary_path": str(out_summary),
        "count": len(overrides),
        "missing_files": summary["missing_files"],
        "duplicate_sample_pages": duplicate_sample_pages,
        "pdfinfo_available": summary["pdfinfo_available"],
        "override_page_out_of_range": summary["override_page_out_of_range"],
    }

    _recompute_stats(report_obj)

    out_report = args.out_report
    if out_report is None:
        if args.in_place:
            out_report = report_in
        else:
            out_report = report_in.with_name(report_in.stem + ".with_ssot_overrides.json")

    out_report = out_report.expanduser().resolve()
    out_report.parent.mkdir(parents=True, exist_ok=True)
    out_report.write_text(json.dumps(report_obj, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK wrote augmented report: {out_report}")


if __name__ == "__main__":
    main()
