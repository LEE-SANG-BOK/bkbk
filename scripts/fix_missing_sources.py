#!/usr/bin/env python3
from __future__ import annotations

import argparse
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any


SENTINELS = {"S-TBD", "SRC-TBD", "S-UNKNOWN", "SRC-UNKNOWN"}


def _split_ids(raw: object) -> list[str]:
    if raw is None:
        return []
    s = str(raw).strip()
    if not s:
        return []
    s = s.replace("\n", ";")
    s = s.replace(",", ";")
    s = s.replace(" / ", ";")
    parts = []
    for p in s.split(";"):
        t = p.strip()
        if t:
            parts.append(t)
    return parts


def _extract_referenced_src_ids(xlsx: Path) -> set[str]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    referenced: set[str] = set()
    for ws in wb.worksheets:
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        idxs = [i for i, h in enumerate(headers) if h is not None and str(h).strip() in {"src_id", "src_ids"}]
        if not idxs:
            continue
        for row in ws.iter_rows(min_row=2, values_only=True):
            for i in idxs:
                if i < len(row):
                    referenced.update(_split_ids(row[i]))
    referenced = {x for x in referenced if x and x not in SENTINELS}
    return referenced


@dataclass(frozen=True)
class SourcesDoc:
    raw: dict[str, Any]
    source_ids: set[str]
    sources_kind: str  # "list" | "dict"


def _load_sources_yaml(path: Path) -> SourcesDoc:
    import yaml  # type: ignore

    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    sources = raw.get("sources")
    if sources is None:
        raise ValueError("sources.yaml missing top-level key: sources")

    if isinstance(sources, list):
        ids: set[str] = set()
        for item in sources:
            if not isinstance(item, dict):
                continue
            sid = item.get("source_id") or item.get("src_id") or item.get("id")
            if sid:
                ids.add(str(sid).strip())
        return SourcesDoc(raw=raw, source_ids=ids, sources_kind="list")

    if isinstance(sources, dict):
        ids = {str(k).strip() for k in sources.keys()}
        return SourcesDoc(raw=raw, source_ids=ids, sources_kind="dict")

    raise ValueError(f"unsupported sources type: {type(sources).__name__}")


def _make_stub(source_id: str) -> dict[str, Any]:
    today = date.today().isoformat()
    return {
        "source_id": source_id,
        "kind": "TBD",
        "title": "",
        "publisher": "",
        "issued_date": "",
        "accessed_date": today,
        "url": "",
        "local_file": "",
        "license": "",
        "retrieval": {"method": "manual_upload", "endpoint": "", "params": {}},
        "citation": {"short": f"{source_id} (TBD)", "full": f"{source_id} (TBD)."},
        "notes": "AUTO-STUB: case.xlsx references this src_id; fill metadata before submission.",
    }


def _apply_stubs(doc: SourcesDoc, missing_ids: list[str]) -> None:
    if doc.sources_kind != "list":
        raise ValueError("sources.yaml uses dict-form sources; auto-append is only supported for list-form sources")
    sources = doc.raw.get("sources")
    assert isinstance(sources, list)
    for sid in missing_ids:
        sources.append(_make_stub(sid))


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Find src_id references in case.xlsx and append stub entries into sources.yaml (optional). "
            "This reduces iteration friction when QA/doctor reports missing source_id."
        )
    )
    ap.add_argument("--xlsx", required=True, type=Path, help="case.xlsx path")
    ap.add_argument("--sources-yaml", required=True, type=Path, help="sources.yaml path")
    ap.add_argument("--apply", action="store_true", help="Write back into sources.yaml (default: dry-run)")
    ap.add_argument("--out", type=Path, default=None, help="Write to a different output path (implies --apply)")
    ap.add_argument("--backup", action="store_true", help="Create a .bak copy when writing in-place")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    sy = args.sources_yaml.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"missing xlsx: {xlsx}")
    if not sy.exists():
        raise SystemExit(f"missing sources.yaml: {sy}")

    referenced = _extract_referenced_src_ids(xlsx)
    doc = _load_sources_yaml(sy)
    missing = sorted(x for x in referenced if x not in doc.source_ids)

    if not missing:
        print("OK: no missing source_id (case.xlsx references are all present in sources.yaml)")
        return

    print(f"missing source_id count: {len(missing)}")
    for sid in missing:
        print(f" - {sid}")

    if not (args.apply or args.out):
        print("")
        print("dry-run only. To apply:")
        print(f"  python scripts/fix_missing_sources.py --xlsx {xlsx} --sources-yaml {sy} --apply --backup")
        print("")
        print("stub preview (YAML):")
        import yaml  # type: ignore

        preview = [_make_stub(sid) for sid in missing[:5]]
        print(yaml.safe_dump(preview, sort_keys=False, allow_unicode=True))
        return

    out_path = args.out.expanduser().resolve() if args.out else sy
    if out_path == sy and args.backup:
        backup = sy.with_suffix(sy.suffix + ".bak")
        backup.write_text(sy.read_text(encoding="utf-8"), encoding="utf-8")
        print(f"backup written: {backup}")

    _apply_stubs(doc, missing)

    import yaml  # type: ignore

    out_path.write_text(
        yaml.safe_dump(doc.raw, sort_keys=False, allow_unicode=True),
        encoding="utf-8",
    )
    print(f"OK wrote: {out_path}")


if __name__ == "__main__":
    main()

