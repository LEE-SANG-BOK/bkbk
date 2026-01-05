#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import subprocess
import sys
from datetime import datetime
from pathlib import Path


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _bin_python(repo_root: Path) -> Path:
    cand = repo_root / ".venv" / "bin" / "python"
    if cand.exists():
        return cand
    return Path(sys.executable)


def _bin_eia_gen(repo_root: Path) -> list[str]:
    cand = repo_root / ".venv" / "bin" / "eia-gen"
    if cand.exists():
        return [str(cand)]
    return [str(_bin_python(repo_root)), "-m", "eia_gen.cli"]


def _tee_run(*, cmd: list[str], cwd: Path, log_path: Path) -> None:
    log_path.parent.mkdir(parents=True, exist_ok=True)
    with log_path.open("w", encoding="utf-8") as f:
        f.write(f"$ {' '.join(cmd)}\n\n")
        p = subprocess.Popen(cmd, cwd=str(cwd), stdout=subprocess.PIPE, stderr=subprocess.STDOUT, text=True)
        assert p.stdout is not None
        for line in p.stdout:
            sys.stdout.write(line)
            f.write(line)
        rc = p.wait()
        if rc != 0:
            raise SystemExit(rc)


def _load_yaml(path: Path) -> dict:
    try:
        import yaml
    except Exception as e:  # pragma: no cover
        raise SystemExit(
            "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
            f"Import error: {e}"
        )
    obj = yaml.safe_load(path.read_text(encoding="utf-8"))
    if not isinstance(obj, dict):
        raise SystemExit(f"YAML must be an object: {path}")
    return obj


def _required_anchors_for_eia_ssot(spec_dir: Path) -> list[str]:
    sections = _load_yaml(spec_dir / "sections.yaml").get("sections") or []
    tables = _load_yaml(spec_dir / "table_specs.yaml").get("tables") or []
    figures = _load_yaml(spec_dir / "figure_specs.yaml").get("figures") or []

    required: list[str] = []

    def _anchor(x) -> str:
        return str((x or {}).get("anchor") or "").strip()

    for s in sections:
        sid = str((s or {}).get("id") or "").strip()
        if sid.startswith("CH0_") or sid.startswith("CH1_") or sid.startswith("SSOT_"):
            a = _anchor(s)
            if a:
                required.append(a)

    for t in tables:
        try:
            ch = int((t or {}).get("chapter") or 0)
        except Exception:
            ch = 0
        if ch == 1:
            a = _anchor(t)
            if a:
                required.append(a)

    for f in figures:
        try:
            ch = int((f or {}).get("chapter") or 0)
        except Exception:
            ch = 0
        if ch == 1:
            a = _anchor(f)
            if a:
                required.append(a)

    # Appendix inserts are rendered outside SSOT page ranges.
    required.append("[[BLOCK:APPENDIX_INSERTS]]")

    # Stable order: keep duplicates out while preserving order.
    out: list[str] = []
    seen: set[str] = set()
    for a in required:
        if a in seen:
            continue
        seen.add(a)
        out.append(a)
    return out


def _assert_docx_has_anchors(*, template_path: Path, required_anchors: list[str]) -> list[str]:
    try:
        from docx import Document
    except Exception as e:  # pragma: no cover
        raise SystemExit(
            "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
            f"Import error: {e}"
        )

    doc = Document(str(template_path))
    found = {p.text.strip() for p in doc.paragraphs}
    missing = [a for a in required_anchors if a.strip() and a.strip() not in found]
    return missing


def _read_case_center_lon_lat(xlsx: Path) -> tuple[float | None, float | None]:
    try:
        import openpyxl  # type: ignore
    except Exception:
        return (None, None)

    try:
        wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
        if "LOCATION" not in wb.sheetnames:
            return (None, None)
        ws = wb["LOCATION"]
        headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), tuple())
        hm = {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}

        idx_lon = hm.get("center_lon")
        idx_lat = hm.get("center_lat")
        if idx_lon is None or idx_lat is None:
            return (None, None)

        def _to_float(v: object) -> float | None:
            if v is None:
                return None
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).strip()
            if not s:
                return None
            try:
                return float(s)
            except Exception:
                return None

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row is None:
                continue
            lon_raw = row[idx_lon] if idx_lon < len(row) else None
            lat_raw = row[idx_lat] if idx_lat < len(row) else None
            if lon_raw is None and lat_raw is None:
                continue
            return (_to_float(lon_raw), _to_float(lat_raw))
    except Exception:
        return (None, None)

    return (None, None)


def main() -> None:
    ap = argparse.ArgumentParser(description="Run eia-gen quality gates end-to-end (doctor/template/QA/optional generate).")
    ap.add_argument(
        "--case-dir",
        required=True,
        help="Case folder containing case.xlsx and sources.yaml (e.g., output/case_new).",
    )
    ap.add_argument(
        "--out-dir",
        default="",
        help="Output dir for gate logs/reports (default: <case-dir>/_quality_gates/<timestamp>).",
    )
    ap.add_argument(
        "--mode",
        default="check",
        choices=["check", "generate", "both"],
        help="check=doctor+template-check+QA, generate=generate-xlsx-both, both=check then generate",
    )
    ap.add_argument(
        "--skip-unit-tests",
        action="store_true",
        help="Skip running unit tests.",
    )
    ap.add_argument(
        "--template-eia",
        type=Path,
        default=None,
        help="Override EIA template for scoped QA/generate (disables case-local template preference).",
    )
    ap.add_argument(
        "--template-dia",
        type=Path,
        default=None,
        help="Override DIA template for scoped QA/generate (disables case-local template preference).",
    )
    ap.add_argument(
        "--use-ssot-eia-template",
        action="store_true",
        help="Shortcut for --template-eia templates/report_template.sample_changwon_2025.ssot_full.scaffolded.docx.",
    )
    ap.add_argument(
        "--skip-template-check",
        action="store_true",
        help="Skip template-check gates.",
    )
    ap.add_argument(
        "--skip-doctor",
        action="store_true",
        help="Skip doctor_env gate.",
    )
    ap.add_argument(
        "--doctor-relaxed",
        action="store_true",
        help="Do not fail the doctor gate on case structure/sources/paths checks (prints warnings only).",
    )
    ap.add_argument(
        "--verify-keys",
        action="store_true",
        help="Run verify-keys gate (default: presence mode; no network calls).",
    )
    ap.add_argument(
        "--verify-keys-strict",
        action="store_true",
        help="Fail the gate when verify-keys reports any required failures.",
    )
    ap.add_argument(
        "--verify-keys-strict-ignore-category",
        action="append",
        default=[],
        help="When --verify-keys-strict, ignore verify-keys failures with these categories (repeatable).",
    )
    ap.add_argument(
        "--verify-keys-center-lon",
        type=float,
        default=None,
        help="Override WGS84 lon for verify-keys (defaults to LOCATION.center_lon when available).",
    )
    ap.add_argument(
        "--verify-keys-center-lat",
        type=float,
        default=None,
        help="Override WGS84 lat for verify-keys (defaults to LOCATION.center_lat when available).",
    )
    ap.add_argument(
        "--verify-keys-wms-radius-m",
        type=int,
        default=1000,
        help="Radius used to build a WMS bbox for verify-keys.",
    )
    ap.add_argument(
        "--verify-keys-timeout-sec",
        type=int,
        default=12,
        help="HTTP timeout used by verify-keys.",
    )
    ap.add_argument(
        "--verify-keys-mode",
        default="presence",
        choices=["presence", "network"],
        help="Verify-keys mode: presence=env-only, network=call endpoints (can be flaky due to provider outages).",
    )
    ap.add_argument(
        "--verify-keys-wms-use-cache",
        action="store_true",
        help="Pass --wms-use-cache to verify-keys (allows cache hits; default forces network for WMS checks).",
    )
    ap.add_argument(
        "--enrich",
        action="store_true",
        help="Run DATA_REQUESTS planner+runner before QA/generate (updates the XLSX in-place and may write evidences).",
    )
    ap.add_argument(
        "--enrich-overwrite-plan",
        action="store_true",
        help="Overwrite existing DATA_REQUESTS rows before running (use with --enrich).",
    )
    ap.add_argument(
        "--enrich-fail-on-warn",
        action="store_true",
        help="Fail the gate when DATA_REQUESTS run emits warnings (use with --enrich).",
    )
    ap.add_argument(
        "--write-next-actions",
        action="store_true",
        help="Write QA next-actions markdown summaries (best-effort).",
    )
    ap.add_argument(
        "--append-data-requests-summary",
        action="store_true",
        help="Append DATA_REQUESTS summary into check_full EIA validation report (best-effort).",
    )
    ap.add_argument(
        "--append-ssot-overrides-summary",
        action="store_true",
        help="Append SSOT_PAGE_OVERRIDES summary into check_scoped_default EIA validation report (best-effort).",
    )
    ap.add_argument(
        "--submission",
        action="store_true",
        help="Run QA/generate in strict 'submission mode' (treat missing core sheets/rows as ERROR).",
    )
    ap.add_argument(
        "--upgrade-xlsx",
        action="store_true",
        help="Run `eia-gen xlsx-upgrade-v2` in-place before running gates (creates .bak by default).",
    )

    args = ap.parse_args()

    repo_root = _repo_root()

    if args.use_ssot_eia_template:
        if args.template_eia is not None:
            raise SystemExit("Use either --template-eia or --use-ssot-eia-template (not both).")
        args.template_eia = repo_root / "templates" / "report_template.sample_changwon_2025.ssot_full.scaffolded.docx"

    if args.template_eia is not None and not args.template_eia.is_absolute():
        args.template_eia = (repo_root / args.template_eia).resolve()
    if args.template_dia is not None and not args.template_dia.is_absolute():
        args.template_dia = (repo_root / args.template_dia).resolve()

    if args.template_eia is not None and not args.template_eia.exists():
        raise SystemExit(f"EIA template not found: {args.template_eia}")
    if args.template_dia is not None and not args.template_dia.exists():
        raise SystemExit(f"DIA template not found: {args.template_dia}")

    case_dir = (repo_root / args.case_dir).resolve() if not Path(args.case_dir).is_absolute() else Path(args.case_dir)
    xlsx = case_dir / "case.xlsx"
    sources = case_dir / "sources.yaml"

    if not xlsx.exists():
        raise SystemExit(f"missing: {xlsx}")
    if not sources.exists():
        raise SystemExit(f"missing: {sources}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    out_dir = Path(args.out_dir).resolve() if args.out_dir else (case_dir / "_quality_gates" / ts)
    out_dir.mkdir(parents=True, exist_ok=True)

    py = _bin_python(repo_root)
    eia_gen = _bin_eia_gen(repo_root)

    summary: dict[str, object] = {
        "case_dir": str(case_dir),
        "out_dir": str(out_dir),
        "started_at": ts,
        "mode": args.mode,
        "steps": [],
    }

    def _step(name: str, cmd: list[str], log_name: str) -> None:
        steps = summary.get("steps")
        if not isinstance(steps, list):
            steps = []
            summary["steps"] = steps
        steps.append({"name": name, "cmd": cmd, "log": log_name})
        _tee_run(cmd=cmd, cwd=repo_root, log_path=out_dir / log_name)

    def _capture_qa_stats(tag: str, qa_dir: Path) -> None:
        def _read_stats(report_path: Path) -> dict[str, int]:
            obj = json.loads(report_path.read_text(encoding="utf-8"))
            stats = obj.get("stats") or {}
            return {
                "error_count": int(stats.get("error_count", 0) or 0),
                "warn_count": int(stats.get("warn_count", 0) or 0),
                "info_count": int(stats.get("info_count", 0) or 0),
                "placeholder_count": int(stats.get("placeholder_count", 0) or 0),
                "s_tbd_citation_count": int(stats.get("s_tbd_citation_count", 0) or 0),
            }

        out: dict[str, object] = {}
        for kind in ["eia", "dia"]:
            rp = qa_dir / f"validation_report_{kind}.json"
            if not rp.exists():
                continue
            out[kind] = {"path": str(rp), "stats": _read_stats(rp)}

        qa = summary.get("qa")
        if not isinstance(qa, dict):
            qa = {}
            summary["qa"] = qa
        qa[tag] = out

    if not args.skip_unit_tests and args.mode in ("check", "both"):
        _step(
            "unit-tests",
            [str(py), "-m", "unittest", "discover", "-s", "tests", "-p", "test_*.py"],
            "gate_ut_unittest.log",
        )

    if not args.skip_doctor and args.mode in ("check", "both"):
        doctor_cmd = [str(py), "scripts/doctor_env.py", "--case-dir", str(case_dir)]
        if not args.doctor_relaxed:
            doctor_cmd.append("--strict-case")
        _step(
            "doctor",
            doctor_cmd,
            "gate_00_doctor.log",
        )

    if args.upgrade_xlsx and args.mode in ("check", "both", "generate"):
        _step(
            "xlsx-upgrade-v2",
            [*eia_gen, "xlsx-upgrade-v2", "--xlsx", str(xlsx)],
            "gate_00_xlsx_upgrade.log",
        )

    if not args.skip_template_check and args.mode in ("check", "both"):
        # Gate-1A: EIA "normal" template (ignore SSOT-only blocks).
        eia_normal = repo_root / "templates" / "report_template.sample_changwon_2025.scaffolded.docx"
        if eia_normal.exists():
            _step(
                "template-check:eia-normal",
                [
                    *eia_gen,
                    "template-check",
                    "--template",
                    str(eia_normal),
                    "--spec-dir",
                    "spec",
                    "--ignore-missing-anchor-prefix",
                    "[[BLOCK:SSOT_",
                    "--out",
                    str(out_dir / "template_check_eia_normal.json"),
                ],
                "gate_01_template_eia_normal.log",
            )

        # Gate-1B: EIA SSOT template (ensure CH0/CH1/SSOT + chapter-1 table/fig anchors exist).
        eia_ssot = repo_root / "templates" / "report_template.sample_changwon_2025.ssot_full.scaffolded.docx"
        if eia_ssot.exists():
            required = _required_anchors_for_eia_ssot(repo_root / "spec")
            missing = _assert_docx_has_anchors(template_path=eia_ssot, required_anchors=required)
            (out_dir / "template_check_eia_ssot_required_anchors.json").write_text(
                json.dumps(
                    {"template": str(eia_ssot), "required_count": len(required), "missing": missing},
                    ensure_ascii=False,
                    indent=2,
                ),
                encoding="utf-8",
            )
            if missing:
                raise SystemExit(
                    f"EIA SSOT template is missing required anchors ({len(missing)}): {missing[:10]}"
                )
            _step(
                "template-check:eia-ssot",
                [
                    *eia_gen,
                    "template-check",
                    "--template",
                    str(eia_ssot),
                    "--spec-dir",
                    "spec",
                    "--allow-missing-anchors",
                    "--out",
                    str(out_dir / "template_check_eia_ssot.json"),
                ],
                "gate_01_template_eia_ssot.log",
            )

        # Gate-1C: DIA template (scaffolded is the default expectation).
        dia_tpl = repo_root / "templates" / "dia_template.scaffolded.docx"
        if not dia_tpl.exists():
            dia_tpl = repo_root / "templates" / "dia_template.docx"
        if dia_tpl.exists():
            _step(
                "template-check:dia",
                [
                    *eia_gen,
                    "template-check",
                    "--template",
                    str(dia_tpl),
                    "--spec-dir",
                    "spec_dia",
                    "--out",
                    str(out_dir / "template_check_dia.json"),
                ],
                "gate_01_template_dia.log",
            )

        # (Optional) Case-local templates when present (what users typically edit).
        case_tpl_eia = case_dir / "templates" / "report_template_eia.docx"
        if case_tpl_eia.exists():
            _step(
                "template-check:case-eia",
                [
                    *eia_gen,
                    "template-check",
                    "--template",
                    str(case_tpl_eia),
                    "--spec-dir",
                    "spec",
                    "--ignore-missing-anchor-prefix",
                    "[[BLOCK:SSOT_",
                    "--out",
                    str(out_dir / "template_check_case_eia.json"),
                ],
                "gate_01_template_case_eia.log",
            )

        case_tpl_dia = case_dir / "templates" / "report_template_dia.docx"
        if case_tpl_dia.exists():
            _step(
                "template-check:case-dia",
                [
                    *eia_gen,
                    "template-check",
                    "--template",
                    str(case_tpl_dia),
                    "--spec-dir",
                    "spec_dia",
                    "--out",
                    str(out_dir / "template_check_case_dia.json"),
                ],
                "gate_01_template_case_dia.log",
            )

    if args.verify_keys:
        lon = args.verify_keys_center_lon
        lat = args.verify_keys_center_lat
        if lon is None or lat is None:
            xlon, xlat = _read_case_center_lon_lat(xlsx)
            if lon is None:
                lon = xlon
            if lat is None:
                lat = xlat

        cmd = [
            *eia_gen,
            "verify-keys",
            "--mode",
            str(args.verify_keys_mode),
            "--wms-radius-m",
            str(int(args.verify_keys_wms_radius_m)),
            "--timeout-sec",
            str(int(args.verify_keys_timeout_sec)),
            "--out",
            str(out_dir / "verify_keys.json"),
        ]
        if lon is not None:
            cmd.extend(["--center-lon", str(float(lon))])
        if lat is not None:
            cmd.extend(["--center-lat", str(float(lat))])
        if args.verify_keys_strict:
            cmd.append("--strict")
        if args.verify_keys_wms_use_cache:
            cmd.append("--wms-use-cache")
        for cat in (args.verify_keys_strict_ignore_category or []):
            s = str(cat or "").strip()
            if not s:
                continue
            cmd.extend(["--strict-ignore-category", s])

        summary["verify_keys"] = {
            "enabled": True,
            "strict": bool(args.verify_keys_strict),
            "mode": str(args.verify_keys_mode),
            "wms_use_cache": bool(args.verify_keys_wms_use_cache),
            "strict_ignore_category": [str(c).strip() for c in (args.verify_keys_strict_ignore_category or []) if str(c).strip()],
            "center_lon": lon,
            "center_lat": lat,
            "wms_radius_m": int(args.verify_keys_wms_radius_m),
            "timeout_sec": int(args.verify_keys_timeout_sec),
            "report": str(out_dir / "verify_keys.json"),
        }

        _step(
            "verify-keys",
            cmd,
            "gate_03_verify_keys.log",
        )

    if args.enrich:
        summary["enrich"] = {
            "enabled": True,
            "overwrite_plan": bool(args.enrich_overwrite_plan),
            "fail_on_warn": bool(args.enrich_fail_on_warn),
        }

        plan_cmd = [*eia_gen, "plan-data-requests", "--xlsx", str(xlsx)]
        if args.enrich_overwrite_plan:
            plan_cmd.append("--overwrite")
        _step(
            "plan-data-requests",
            plan_cmd,
            "gate_04_enrich_plan.log",
        )

        run_cmd = [*eia_gen, "run-data-requests", "--xlsx", str(xlsx)]
        _step(
            "run-data-requests",
            run_cmd,
            "gate_04_enrich_run.log",
        )

        # Capture the run report into the gate folder for auditability.
        run_report = case_dir / "_data_requests_run.json"
        if run_report.exists():
            captured = out_dir / "data_requests_run.json"
            captured.write_text(run_report.read_text(encoding="utf-8"), encoding="utf-8")
            summary["enrich_report"] = str(captured)

            if args.enrich_fail_on_warn:
                try:
                    obj = json.loads(captured.read_text(encoding="utf-8"))
                    warnings = obj.get("warnings") or []
                    if isinstance(warnings, list) and warnings:
                        raise SystemExit(2)
                except SystemExit:
                    raise
                except Exception:
                    # If the report is unreadable, don't silently pass in strict mode.
                    raise SystemExit(2)

    if args.mode in ("check", "both"):
        # Gate-2A: full-spec QA (case completeness).
        _step(
            "check-xlsx-both:full",
            [
                *eia_gen,
                "check-xlsx-both",
                "--xlsx",
                str(xlsx),
                "--sources",
                str(sources),
                "--out-dir",
                str(out_dir / "check_full"),
                *(["--submission"] if args.submission else []),
            ],
            "gate_02_check_full.log",
        )
        if args.append_data_requests_summary:
            rep = out_dir / "check_full" / "validation_report_eia.json"
            if rep.exists():
                _step(
                    "qa-data-requests-summary",
                    [
                        str(py),
                        "scripts/qa_data_requests_summary.py",
                        "--case-xlsx",
                        str(xlsx),
                        "--validation-report",
                        str(rep),
                        "--out-summary",
                        str(out_dir / "check_full" / "data_requests_summary.json"),
                        "--in-place",
                    ],
                    "gate_02_data_requests_summary.log",
                )
        _capture_qa_stats("check_full", out_dir / "check_full")

        # Gate-2B: template-scoped QA (final output quality).
        case_tpl_eia = case_dir / "templates" / "report_template_eia.docx"
        case_tpl_dia = case_dir / "templates" / "report_template_dia.docx"
        scoped_cmd = [*eia_gen, "check-xlsx-both"]
        if args.template_eia is not None:
            scoped_cmd.extend(["--template-eia", str(args.template_eia)])
        elif case_tpl_eia.exists():
            # Prefer case-local templates when present to avoid accidental SSOT sample reuse.
            scoped_cmd.extend(["--template-eia", str(case_tpl_eia)])

        if args.template_dia is not None:
            scoped_cmd.extend(["--template-dia", str(args.template_dia)])
        elif case_tpl_dia.exists():
            # Prefer case-local templates when present to avoid accidental SSOT sample reuse.
            scoped_cmd.extend(["--template-dia", str(case_tpl_dia)])

        if args.template_eia is None and args.template_dia is None and not case_tpl_eia.exists() and not case_tpl_dia.exists():
            scoped_cmd.append("--scope-to-default-templates")
        scoped_cmd.extend(
            [
                "--xlsx",
                str(xlsx),
                "--sources",
                str(sources),
                "--out-dir",
                str(out_dir / "check_scoped"),
                *(["--submission"] if args.submission else []),
            ]
        )
        _step(
            "check-xlsx-both:scoped",
            scoped_cmd,
            "gate_02_check_scoped.log",
        )
        if args.append_ssot_overrides_summary:
            rep = out_dir / "check_scoped" / "validation_report_eia.json"
            if rep.exists():
                _step(
                    "qa-ssot-overrides-summary",
                    [
                        str(py),
                        "scripts/qa_ssot_overrides_summary.py",
                        "--case-xlsx",
                        str(xlsx),
                        "--validation-report",
                        str(rep),
                        "--out-summary",
                        str(out_dir / "check_scoped" / "ssot_overrides_summary.json"),
                        "--in-place",
                    ],
                    "gate_02_ssot_overrides_summary.log",
                )
        _capture_qa_stats("check_scoped_default", out_dir / "check_scoped")

        # Additional scoped QA: "normal template" view (shows input backlog when SSOT reuse is the default).
        eia_normal = repo_root / "templates" / "report_template.sample_changwon_2025.scaffolded.docx"
        if not eia_normal.exists():
            eia_normal = repo_root / "templates" / "report_template.sample_changwon_2025.docx"
        dia_tpl = repo_root / "templates" / "dia_template.scaffolded.docx"
        if not dia_tpl.exists():
            dia_tpl = repo_root / "templates" / "dia_template.docx"
        if eia_normal.exists() and dia_tpl.exists():
            _step(
                "check-xlsx-both:scoped-normal",
                [
                    *eia_gen,
                    "check-xlsx-both",
                    "--template-eia",
                    str(eia_normal),
                    "--template-dia",
                    str(dia_tpl),
                    "--xlsx",
                    str(xlsx),
                    "--sources",
                    str(sources),
                    "--out-dir",
                    str(out_dir / "check_scoped_normal"),
                    *(["--submission"] if args.submission else []),
                ],
                "gate_02_check_scoped_normal.log",
            )
            _capture_qa_stats("check_scoped_normal", out_dir / "check_scoped_normal")

        if args.write_next_actions:
            # Best-effort helper: turn validation_report*.json into “sheet/column” hints.
            for kind in ["eia", "dia"]:
                rep = out_dir / "check_full" / f"validation_report_{kind}.json"
                if not rep.exists():
                    continue
                _step(
                    f"qa-next-actions:{kind}",
                    [
                        str(py),
                        "scripts/qa_next_actions.py",
                        "--validation-report",
                        str(rep),
                        "--out-md",
                        str(out_dir / "check_full" / f"next_actions_{kind}.md"),
                    ],
                    f"gate_02_next_actions_{kind}.log",
                )

    if args.mode in ("generate", "both"):
        case_tpl_eia = case_dir / "templates" / "report_template_eia.docx"
        case_tpl_dia = case_dir / "templates" / "report_template_dia.docx"
        gen_cmd = [*eia_gen, "generate-xlsx-both"]
        if args.template_eia is not None:
            gen_cmd.extend(["--template-eia", str(args.template_eia)])
        elif case_tpl_eia.exists():
            # Prefer case-local templates when present to avoid accidental SSOT sample reuse.
            gen_cmd.extend(["--template-eia", str(case_tpl_eia)])

        if args.template_dia is not None:
            gen_cmd.extend(["--template-dia", str(args.template_dia)])
        elif case_tpl_dia.exists():
            # Prefer case-local templates when present to avoid accidental SSOT sample reuse.
            gen_cmd.extend(["--template-dia", str(case_tpl_dia)])
        gen_cmd.extend(
            [
                "--xlsx",
                str(xlsx),
                "--sources",
                str(sources),
                "--out-dir",
                str(out_dir / "generate"),
                "--no-use-llm",
                *(["--submission"] if args.submission else []),
            ]
        )
        _step(
            "generate-xlsx-both",
            gen_cmd,
            "gate_03_generate.log",
        )

    (out_dir / "summary.json").write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"OK quality gates complete: {out_dir}")


if __name__ == "__main__":
    main()
