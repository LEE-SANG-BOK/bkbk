#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import zipfile
from dataclasses import dataclass
from datetime import datetime, timezone
from hashlib import sha256
from pathlib import Path
from typing import Any


try:
    import openpyxl
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
        f"Import error: {e}"
    )


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _as_str(v: Any) -> str:
    return "" if v is None else str(v).strip()


def _as_bool(v: Any) -> bool | None:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if not s:
        return None
    if s in {"true", "t", "1", "y", "yes"}:
        return True
    if s in {"false", "f", "0", "n", "no"}:
        return False
    return None


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _is_empty_row(values: list[Any]) -> bool:
    return all(_is_empty(v) for v in values)


_TS_RE = re.compile(r"^[0-9]{8}_[0-9]{6}$")


def _find_latest_quality_gates_generate_dir(case_dir: Path) -> Path | None:
    qdir = case_dir / "_quality_gates"
    if not qdir.exists():
        return None
    candidates = [p for p in qdir.iterdir() if p.is_dir() and _TS_RE.match(p.name)]
    if not candidates:
        return None
    latest = max(candidates, key=lambda p: p.name)
    gen = latest / "generate"
    return gen if gen.exists() else None


def _load_sheet_header_map(ws: Any) -> dict[str, int]:
    headers = [c.value for c in ws[1]]
    return {str(h).strip(): idx for idx, h in enumerate(headers) if h is not None and str(h).strip()}


def _load_single_row(wb: Any, sheet: str) -> dict[str, Any]:
    if sheet not in wb.sheetnames:
        return {}
    ws = wb[sheet]
    if ws.max_row < 2:
        return {}
    headers = [c.value for c in ws[1]]
    values = [c.value for c in ws[2]]
    out: dict[str, Any] = {}
    for h, v in zip(headers, values):
        key = str(h).strip() if h is not None else ""
        if key:
            out[key] = v
    return out


def _resolve_case_relative_path(case_dir: Path, raw: str) -> Path:
    p = Path(str(raw or "").strip()).expanduser()
    if not p.is_absolute():
        p = (case_dir / p).expanduser()
    return p.resolve()


def _load_figures(case_dir: Path, wb: Any) -> list[dict[str, Any]]:
    if "FIGURES" not in wb.sheetnames:
        return []
    ws = wb["FIGURES"]
    if ws.max_row < 2:
        return []
    header_map = _load_sheet_header_map(ws)

    def _get(row: list[Any], key: str) -> Any:
        idx = header_map.get(key)
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    out: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue
        fig_id = _as_str(_get(row, "fig_id"))
        if not fig_id:
            continue
        file_path = _as_str(_get(row, "file_path"))
        src_path = _resolve_case_relative_path(case_dir, file_path) if file_path else None
        out.append(
            {
                "fig_id": fig_id,
                "doc_scope": _as_str(_get(row, "doc_scope")),
                "figure_type": _as_str(_get(row, "figure_type")),
                "title": _as_str(_get(row, "title")),
                "caption": _as_str(_get(row, "caption")),
                "file_path": file_path,
                "src_path": str(src_path) if src_path else "",
                "src_exists": bool(src_path.exists()) if src_path else False,
                "gen_method": _as_str(_get(row, "gen_method")),
                "fallback_mode": _as_str(_get(row, "fallback_mode")),
                "src_id": _as_str(_get(row, "src_id")),
            }
        )
    out.sort(key=lambda x: x.get("fig_id") or "")
    return out


def _load_source_register_figure_usage(gen_dir: Path | None) -> dict[str, Any]:
    """Best-effort: map figure_id -> evidence info from the latest source_register.xlsx."""
    if not gen_dir:
        return {"available": False}
    sr_path = (gen_dir / "source_register.xlsx").resolve()
    if not sr_path.exists():
        return {"available": False, "path": str(sr_path)}

    wb = openpyxl.load_workbook(sr_path, data_only=True)
    out: dict[str, Any] = {"available": True, "path": str(sr_path), "figures": {}}

    # Evidence Register: evidence file path for related_fig_id
    if "Evidence Register" in wb.sheetnames:
        ws = wb["Evidence Register"]
        headers = [c.value for c in ws[1]]
        hm = {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}
        idx_rel = hm.get("related_fig_id")
        idx_fp = hm.get("file_path")
        idx_eid = hm.get("evidence_or_fig_id")
        if idx_rel is not None and idx_fp is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                rel = _as_str(row[idx_rel] if idx_rel < len(row) else "")
                if not rel:
                    continue
                cur = out["figures"].setdefault(rel, {})
                if idx_eid is not None:
                    cur["evidence_id"] = _as_str(row[idx_eid] if idx_eid < len(row) else "")
                cur["evidence_file_path"] = _as_str(row[idx_fp] if idx_fp < len(row) else "")

    # Usage Register: whether the figure was actually used in EIA/DIA
    if "USAGE_REGISTER" in wb.sheetnames:
        ws = wb["USAGE_REGISTER"]
        headers = [c.value for c in ws[1]]
        hm = {str(h).strip(): i for i, h in enumerate(headers) if h is not None and str(h).strip()}
        idx_kind = hm.get("content_kind")
        idx_id = hm.get("content_id")
        idx_report = hm.get("report")
        idx_section = hm.get("section_path")
        idx_eids = hm.get("evidence_ids")
        idx_flag = hm.get("qa_flag")
        if idx_kind is not None and idx_id is not None:
            for row in ws.iter_rows(min_row=2, values_only=True):
                kind = _as_str(row[idx_kind] if idx_kind < len(row) else "")
                if kind.upper() != "FIGURE":
                    continue
                fid = _as_str(row[idx_id] if idx_id < len(row) else "")
                if not fid:
                    continue
                cur = out["figures"].setdefault(fid, {})
                cur.setdefault("used_in", [])
                cur["used_in"].append(
                    {
                        "report": _as_str(row[idx_report] if idx_report is not None and idx_report < len(row) else ""),
                        "section_path": _as_str(row[idx_section] if idx_section is not None and idx_section < len(row) else ""),
                        "evidence_ids": _as_str(row[idx_eids] if idx_eids is not None and idx_eids < len(row) else ""),
                        "qa_flag": _as_str(row[idx_flag] if idx_flag is not None and idx_flag < len(row) else ""),
                    }
                )

    return out


def _scan_xlsx_terms(xlsx: Path, terms: list[str]) -> dict[str, int]:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    counts = {t: 0 for t in terms}
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in ws.iter_rows(values_only=True):
            for v in row:
                if v is None:
                    continue
                s = str(v)
                for t in terms:
                    if t in s:
                        counts[t] += 1
    return counts


def _scan_text_terms(path: Path, terms: list[str]) -> dict[str, int]:
    if not path.exists():
        return {t: 0 for t in terms}
    txt = path.read_text(encoding="utf-8", errors="ignore")
    return {t: txt.count(t) for t in terms}


def _scan_docx_terms(path: Path, terms: list[str]) -> dict[str, int]:
    if not path.exists():
        return {t: 0 for t in terms}
    try:
        with zipfile.ZipFile(path, "r") as zf:
            xml = zf.read("word/document.xml").decode("utf-8", errors="ignore")
    except Exception:
        return {t: 0 for t in terms}
    return {t: xml.count(t) for t in terms}


def _docx_zip_content_hashes(path: Path) -> dict[str, str]:
    hashes: dict[str, str] = {}
    with zipfile.ZipFile(path, "r") as zf:
        for name in sorted(zf.namelist()):
            data = zf.read(name)
            hashes[name] = sha256(data).hexdigest()
    return hashes


def _docx_same_contents(a: Path, b: Path) -> bool:
    if not a.exists() or not b.exists():
        return False
    try:
        return _docx_zip_content_hashes(a) == _docx_zip_content_hashes(b)
    except Exception:
        return False


def _md_escape(value: str) -> str:
    return (value or "").replace("|", "\\|")


def _md_table(headers: list[str], rows: list[list[str]]) -> str:
    h = "| " + " | ".join(_md_escape(x) for x in headers) + " |"
    sep = "| " + " | ".join(["---"] * len(headers)) + " |"
    body = "\n".join("| " + " | ".join(_md_escape(x) for x in r) + " |" for r in rows)
    return "\n".join([h, sep, body]) if body else "\n".join([h, sep])


@dataclass(frozen=True)
class TableGuide:
    sheet: str
    purpose: str
    key_columns: list[str]
    notes: str


_CORE_FIELDS: dict[str, list[tuple[str, str]]] = {
    "META": [
        ("case_id", "케이스 ID(내부)"),
        ("created_at", "생성일"),
        ("author", "작성자"),
        ("output_targets", "산출물 대상(ENV/DRR)"),
    ],
    "PROJECT": [
        ("project_name", "사업명"),
        ("project_type", "사업유형"),
        ("doc_env_required", "ENV(소규모환경영향평가) 필요여부 (Y/N/판정필요)"),
        ("doc_drr_required", "DRR(재해) 필요여부 (Y/N/판정필요)"),
        ("total_area_m2", "총면적(m²)"),
        ("main_facilities_summary", "주요시설 요약"),
        ("max_occupancy", "최대 수용인원(명)"),
        ("expected_visitors_day", "예상 이용객(일/명)"),
        ("water_supply", "용수공급"),
        ("sewage_treatment", "오수처리"),
        ("stormwater_plan", "우수처리/배수계통"),
        ("client_name", "의뢰자"),
        ("proponent_name", "사업자(시행자)"),
        ("author_org", "작성기관"),
        ("submit_date", "제출일"),
        ("approving_authority", "승인기관"),
        ("consultation_agency", "협의기관"),
        ("src_id", "근거 출처(src_id)"),
    ],
    "LOCATION": [
        ("address_road", "도로명주소"),
        ("address_jibeon", "지번주소"),
        ("admin_si", "시/군"),
        ("admin_eupmyeon", "읍/면/동"),
        ("center_lat", "중심점 위도"),
        ("center_lon", "중심점 경도"),
        ("crs", "좌표계(EPSG)"),
        ("boundary_file", "경계파일(geojson 등)"),
        ("bbox_wkt", "bbox(WKT, 선택)"),
        ("src_id", "근거 출처(src_id)"),
    ],
}


_TABLE_GUIDES: list[TableGuide] = [
    TableGuide(
        sheet="PARCELS",
        purpose="지번별 현황(지목/면적/용도지역 등)",
        key_columns=["parcel_no", "jimok", "area_m2", "zoning", "ownership", "src_id"],
        notes="지번·지목·면적은 공적자료(지적/토지이용계획확인) 근거(src_id) 필수",
    ),
    TableGuide(
        sheet="FACILITIES",
        purpose="시설/규모(야영장, 편의, 체육, 판매 등)",
        key_columns=["type", "name", "qty", "qty_unit", "area_m2", "capacity_person", "note", "src_id"],
        notes="수량/단위/면적/수용인원 중 해당되는 값은 채워야 표/본문이 비지 않음",
    ),
    TableGuide(
        sheet="UTILITIES",
        purpose="용수/오수/우수 등 기반시설(배수시설 포함)",
        key_columns=["util_type", "description", "capacity", "discharge_point", "drawing_ref", "src_id"],
        notes="배수계통은 도면 참조(drawing_ref)와 연결 권장",
    ),
    TableGuide(
        sheet="FIELD_SURVEY_LOG",
        purpose="현장조사 로그(일시/조사자/범위/경로/사진폴더)",
        key_columns=["survey_date", "surveyors", "weather", "scope", "route_desc", "photo_folder", "src_id"],
        notes="생태/경관 현황의 ‘조사 근거’ 역할(최소 1행 권장)",
    ),
    TableGuide(
        sheet="ENV_SCOPING",
        purpose="스코핑 매트릭스(중점/현황/제외)",
        key_columns=["item_id", "category", "item_name", "status", "reason", "src_id"],
        notes="status=FOCUS/BASELINE/EXCLUDE + 사유가 보고서 핵심",
    ),
    TableGuide(
        sheet="ENV_BASE_AIR",
        purpose="대기질 현황(자동: AIRKOREA)",
        key_columns=["station_name", "period_start", "period_end", "pollutant", "value_avg", "unit", "src_id", "evidence_id"],
        notes="DATA_REQUESTS: REQ-AIRKOREA-ENV_BASE_AIR → ENV_BASE_AIR",
    ),
    TableGuide(
        sheet="ENV_BASE_SOCIO",
        purpose="인구/주거/사회통계(자동: KOSIS)",
        key_columns=["admin_code", "admin_name", "year", "population_total", "households", "housing_total", "src_id", "evidence_id"],
        notes="DATA_REQUESTS: REQ-KOSIS-ENV_BASE_SOCIO → ENV_BASE_SOCIO",
    ),
    TableGuide(
        sheet="ENV_BASE_WATER",
        purpose="수환경 현황(자동 가능: NIER_WATER / 현재는 params 필요)",
        key_columns=["waterbody_name", "relation", "distance_m", "parameter", "value", "unit", "sampling_date", "src_id", "evidence_id"],
        notes="DATA_REQUESTS: REQ-NIER-WATER-ENV_BASE_WATER (params_json.mgt_no 필요)",
    ),
    TableGuide(
        sheet="ENV_BASE_NOISE",
        purpose="소음·진동 현황(대상지/주변 대표 지점)",
        key_columns=["point_name", "receptor_type", "day_leq", "night_leq", "unit", "src_id"],
        notes="측정/조사 근거(src_id)와 지점명(point_name) 일치 필요",
    ),
    TableGuide(
        sheet="ENV_BASE_GEO",
        purpose="지형·지질/토양 요약(문헌/지도 근거)",
        key_columns=["topic", "summary", "source_map", "src_id", "evidence_id"],
        notes="지도(예: 지질도/토양도) 또는 문헌 근거를 source_map/evidence_id로 연결",
    ),
    TableGuide(
        sheet="ENV_ECO_EVENTS",
        purpose="생태 조사 이벤트(날짜/계절/팀/구역)",
        key_columns=["date", "season", "survey_team", "weather", "area_desc", "evidence_id"],
        notes="ENV_ECO_OBS는 eco_event_id로 이벤트와 연결",
    ),
    TableGuide(
        sheet="ENV_ECO_OBS",
        purpose="생태 관찰 기록(식물/동물)",
        key_columns=["eco_event_id", "taxon_group", "scientific_name", "korean_name", "count", "protected_status", "src_id"],
        notes="보호종은 protected_status에 명시 + 근거(src_id) 필수",
    ),
    TableGuide(
        sheet="ENV_LANDSCAPE",
        purpose="경관/조망점(사진/시뮬레이션 연결)",
        key_columns=["viewpoint_name", "lat", "lon", "description", "photo_fig_id", "src_id"],
        notes="photo_fig_id는 FIGURES의 fig_id와 연결(사진 시트 생성 가능)",
    ),
    TableGuide(
        sheet="ENV_MITIGATION",
        purpose="저감대책(공사/운영 단계)",
        key_columns=["stage", "target_item", "measure", "location", "responsible", "src_id"],
        notes="요약/본문/부록 표가 여기서 파생됨",
    ),
    TableGuide(
        sheet="ENV_MANAGEMENT",
        purpose="협의의견(조건) 이행관리",
        key_columns=["condition_text", "compliance_plan", "status", "evidence_id"],
        notes="조건별 1행 권장(증빙은 evidence_id로 연결)",
    ),
    # DRR
    TableGuide(
        sheet="DRR_HYDRO_RAIN",
        purpose="강우 입력(자동: KMA_ASOS)",
        key_columns=["source_basis", "return_period_yr", "duration_hr", "rainfall_mm", "data_origin", "src_id", "evidence_id"],
        notes="DATA_REQUESTS: REQ-KMA-ASOS-DRR_HYDRO_RAIN → DRR_HYDRO_RAIN",
    ),
    TableGuide(
        sheet="DRR_HYDRO_MODEL",
        purpose="유출 해석(사업 전/후 비교)",
        key_columns=[
            "model",
            "cn_or_c",
            "tc_min",
            "k_storage",
            "peak_cms_before",
            "peak_cms_after",
            "vol_m3_before",
            "vol_m3_after",
            "critical_duration_hr",
            "src_id",
        ],
        notes="계산 근거(산정서/모델결과)를 evidence_id 또는 src_id로 연결 권장",
    ),
]


def _summarize_table_fill(wb: Any, guide: TableGuide) -> dict[str, Any]:
    if guide.sheet not in wb.sheetnames:
        return {"sheet": guide.sheet, "exists": False}

    ws = wb[guide.sheet]
    if ws.max_row < 2:
        return {"sheet": guide.sheet, "exists": True, "row_count": 0, "key_columns": guide.key_columns, "missing": {}}

    header_map = _load_sheet_header_map(ws)
    rows: list[list[Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue
        rows.append(row)

    missing: dict[str, dict[str, Any]] = {}
    for col in guide.key_columns:
        idx = header_map.get(col)
        if idx is None:
            missing[col] = {"present": False, "filled": 0, "total": len(rows), "ratio": 0.0}
            continue
        filled = sum(1 for row in rows if idx < len(row) and not _is_empty(row[idx]))
        total = len(rows)
        missing[col] = {"present": True, "filled": int(filled), "total": int(total), "ratio": (filled / total) if total else 0.0}

    return {
        "sheet": guide.sheet,
        "exists": True,
        "row_count": len(rows),
        "key_columns": guide.key_columns,
        "missing": missing,
    }


_MISSING_ENV_RE = re.compile(r"missing env\\s+([A-Z0-9_]+)", re.IGNORECASE)


def _extract_env_vars(note: str) -> list[str]:
    if not note:
        return []
    envs = [m.group(1).upper() for m in _MISSING_ENV_RE.finditer(note)]

    lowered = note.lower()
    if "missing env" in lowered:
        tail = note[lowered.rfind("missing env") :].replace("missing env", "")
        tail = tail.replace("(", " ").replace(")", " ")
        for token in re.split(r"[\\s,;/]+", tail):
            t = token.strip()
            if not t:
                continue
            t = t.upper()
            if t in {"OR", "AND"}:
                continue
            if re.fullmatch(r"[A-Z][A-Z0-9_]{2,}", t):
                envs.append(t)

    seen: set[str] = set()
    out: list[str] = []
    for e in envs:
        if e not in seen:
            seen.add(e)
            out.append(e)
    return out


def _load_data_requests(case_xlsx: Path) -> list[dict[str, Any]]:
    wb = openpyxl.load_workbook(case_xlsx, data_only=True)
    if "DATA_REQUESTS" not in wb.sheetnames:
        return []
    ws = wb["DATA_REQUESTS"]
    header_map = _load_sheet_header_map(ws)

    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = list(row)
        if _is_empty_row(values):
            continue
        req_id = _as_str(values[header_map["req_id"]]) if "req_id" in header_map else ""
        if not req_id:
            continue
        params_json = _as_str(values[header_map["params_json"]]) if "params_json" in header_map else ""
        params_obj: dict[str, Any] | None = None
        if params_json:
            try:
                parsed = json.loads(params_json)
                if isinstance(parsed, dict):
                    params_obj = parsed
            except Exception:
                params_obj = None
        out.append(
            {
                "req_id": req_id,
                "enabled": _as_bool(values[header_map["enabled"]]) if "enabled" in header_map else None,
                "priority": values[header_map["priority"]] if "priority" in header_map else None,
                "connector": _as_str(values[header_map["connector"]]).upper() if "connector" in header_map else "",
                "purpose": _as_str(values[header_map["purpose"]]).upper() if "purpose" in header_map else "",
                "output_sheet": _as_str(values[header_map["output_sheet"]]).upper() if "output_sheet" in header_map else "",
                "params_json": params_json,
                "params_obj": params_obj,
                "run_mode": _as_str(values[header_map["run_mode"]]).upper() if "run_mode" in header_map else "",
                "last_run_at": _as_str(values[header_map["last_run_at"]]) if "last_run_at" in header_map else "",
                "last_evidence_ids": _as_str(values[header_map["last_evidence_ids"]]) if "last_evidence_ids" in header_map else "",
                "note": _as_str(values[header_map["note"]]) if "note" in header_map else "",
            }
        )

    def _prio_key(x: dict[str, Any]) -> tuple[int, str]:
        pr = x.get("priority")
        pr_i = int(pr) if isinstance(pr, (int, float)) and pr is not None else 9999
        return (pr_i, str(x.get("req_id") or ""))

    out.sort(key=_prio_key)
    return out


def _summarize_data_requests(rows: list[dict[str, Any]]) -> dict[str, Any]:
    enabled = [r for r in rows if r.get("enabled") is True]
    disabled = [r for r in rows if r.get("enabled") is False]
    unknown = [r for r in rows if r.get("enabled") is None]

    executed_est = [r for r in enabled if _as_str(r.get("last_run_at")) or _as_str(r.get("last_evidence_ids"))]

    disabled_missing_env: dict[str, list[str]] = {}
    disabled_missing_params: list[dict[str, Any]] = []
    disabled_other: list[dict[str, Any]] = []

    for r in disabled:
        note = _as_str(r.get("note"))
        envs = _extract_env_vars(note)
        if envs:
            for env in envs:
                disabled_missing_env.setdefault(env, []).append(str(r.get("req_id") or ""))
            continue

        params = r.get("params_obj") or {}
        if isinstance(params, dict) and any(isinstance(v, str) and not v.strip() for v in params.values()):
            blanks = sorted([k for k, v in params.items() if isinstance(v, str) and not v.strip()])
            if blanks:
                disabled_missing_params.append({"req_id": r.get("req_id"), "blank_params": blanks, "note": note})
                continue

        disabled_other.append({"req_id": r.get("req_id"), "note": note})

    for k in list(disabled_missing_env.keys()):
        disabled_missing_env[k] = sorted(set(disabled_missing_env[k]))

    return {
        "counts": {
            "total": len(rows),
            "enabled": len(enabled),
            "disabled": len(disabled),
            "enabled_unknown": len(unknown),
            "executed_est": len(executed_est),
        },
        "disabled_missing_env": disabled_missing_env,
        "disabled_missing_params": disabled_missing_params,
        "disabled_other": disabled_other,
    }


def build_portal(
    *,
    case_dir: Path,
    out_md: Path,
    out_json: Path | None,
) -> dict[str, Any]:
    case_xlsx = (case_dir / "case.xlsx").resolve()
    sources_yaml = (case_dir / "sources.yaml").resolve()

    if not case_xlsx.exists():
        raise SystemExit(f"case.xlsx not found: {case_xlsx}")
    if not sources_yaml.exists():
        raise SystemExit(f"sources.yaml not found: {sources_yaml}")

    gen_dir = _find_latest_quality_gates_generate_dir(case_dir)
    report_eia = (gen_dir / "report_eia.docx") if gen_dir else None
    report_dia = (gen_dir / "report_dia.docx") if gen_dir else None
    validation_eia = (gen_dir / "validation_report_eia.json") if gen_dir else None
    validation_dia = (gen_dir / "validation_report_dia.json") if gen_dir else None

    wb = openpyxl.load_workbook(case_xlsx, data_only=True)
    meta = _load_single_row(wb, "META")
    project = _load_single_row(wb, "PROJECT")
    location = _load_single_row(wb, "LOCATION")
    figures = _load_figures(case_dir, wb)
    source_register_fig_usage = _load_source_register_figure_usage(gen_dir)

    try:
        from eia_gen.services.xlsx.status import compute_xlsx_status

        xlsx_status = compute_xlsx_status(case_xlsx).to_dict()
    except Exception as e:
        xlsx_status = {"error": str(e)}

    data_requests = _load_data_requests(case_xlsx)
    data_requests_summary = _summarize_data_requests(data_requests)

    banned_terms = ["S-CHANGWON-SAMPLE", "진전", "진저"]
    scan = {
        "case_xlsx": _scan_xlsx_terms(case_xlsx, banned_terms),
        "sources_yaml": _scan_text_terms(sources_yaml, banned_terms),
        "report_eia": _scan_docx_terms(report_eia, banned_terms) if report_eia else {t: 0 for t in banned_terms},
        "report_dia": _scan_docx_terms(report_dia, banned_terms) if report_dia else {t: 0 for t in banned_terms},
    }

    template_dir = case_dir / "templates"
    template_eia = (template_dir / "report_template_eia.docx").resolve()
    sample_template = (Path(__file__).resolve().parents[1] / "templates" / "report_template.sample_changwon_2025.scaffolded.docx").resolve()

    template_check = {
        "template_eia": str(template_eia) if template_eia.exists() else "",
        "sample_template": str(sample_template) if sample_template.exists() else "",
        "same_contents_as_sample_scaffolded": _docx_same_contents(template_eia, sample_template) if template_eia.exists() else False,
        "template_terms": _scan_docx_terms(template_eia, ["샘플", "소규모환경영향평가서", "관광농원"]) if template_eia.exists() else {},
    }

    core_missing: list[dict[str, Any]] = []
    for sheet, fields in _CORE_FIELDS.items():
        cur = meta if sheet == "META" else project if sheet == "PROJECT" else location if sheet == "LOCATION" else {}
        for col, label in fields:
            v = cur.get(col)
            if _is_empty(v):
                core_missing.append({"sheet": sheet, "column": col, "label": label})

    table_summaries = [_summarize_table_fill(wb, g) for g in _TABLE_GUIDES]

    portal: dict[str, Any] = {
        "generated_at": _utc_now_iso(),
        "case_dir": str(case_dir.resolve()),
        "paths": {
            "case_xlsx": str(case_xlsx),
            "sources_yaml": str(sources_yaml),
            "latest_generate_dir": str(gen_dir) if gen_dir else "",
            "report_eia": str(report_eia) if report_eia and report_eia.exists() else "",
            "report_dia": str(report_dia) if report_dia and report_dia.exists() else "",
            "validation_eia": str(validation_eia) if validation_eia and validation_eia.exists() else "",
            "validation_dia": str(validation_dia) if validation_dia and validation_dia.exists() else "",
        },
        "meta": meta,
        "project": project,
        "location": location,
        "xlsx_status": xlsx_status,
        "core_missing": core_missing,
        "table_summaries": table_summaries,
        "figures": figures,
        "source_register_fig_usage": source_register_fig_usage,
        "data_requests": data_requests,
        "data_requests_summary": data_requests_summary,
        "contamination_scan": scan,
        "template_check": template_check,
    }

    # Write JSON (optional)
    if out_json:
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(json.dumps(portal, ensure_ascii=False, indent=2), encoding="utf-8")

    # Markdown
    lines: list[str] = []
    case_id = _as_str(meta.get("case_id")) or "CASE"
    proj_name = _as_str(project.get("project_name")) or "(미입력)"
    lines.append(f"# Case Input Portal — {case_id}")
    lines.append("")
    lines.append(f"- project: {proj_name}")
    lines.append(f"- generated_at(UTC): `{portal['generated_at']}`")
    lines.append(f"- case_dir: `{portal['case_dir']}`")
    lines.append(f"- case.xlsx: `{portal['paths']['case_xlsx']}`")
    lines.append(f"- sources.yaml: `{portal['paths']['sources_yaml']}`")
    if portal["paths"]["latest_generate_dir"]:
        lines.append(f"- latest_generate_dir: `{portal['paths']['latest_generate_dir']}`")
    if portal["paths"]["report_eia"]:
        lines.append(f"- report_eia.docx: `{portal['paths']['report_eia']}`")
    if portal["paths"]["report_dia"]:
        lines.append(f"- report_dia.docx: `{portal['paths']['report_dia']}`")
    lines.append("")

    lines.append("## 0) 한 줄 요약(현 상태)")
    xs = portal.get("xlsx_status") or {}
    if isinstance(xs, dict) and "sheet_stats" in xs:
        total_rows = xs.get("total_rows")
        fill = xs.get("total_data_fill_ratio")
        lines.append(f"- 입력행수: {total_rows}, 채움률(데이터): {fill:.2%}" if isinstance(fill, float) else f"- 입력행수: {total_rows}")
        zero_sheets = [s["sheet"] for s in xs.get("sheet_stats") or [] if s.get("row_count") == 0]
        if zero_sheets:
            lines.append(f"- 0행(비어있음) 시트: {', '.join(zero_sheets)}")
    else:
        lines.append(f"- xlsx_status: {xs.get('error') if isinstance(xs, dict) else 'N/A'}")

    counts = data_requests_summary.get("counts") or {}
    lines.append(
        f"- DATA_REQUESTS: total={counts.get('total')} enabled={counts.get('enabled')} disabled={counts.get('disabled')} executed_est={counts.get('executed_est')}"
    )
    lines.append("")

    lines.append("## 0-1) 지금 당장 채워야 하는 것(사람이 입력)")
    if core_missing:
        by_sheet: dict[str, list[str]] = {}
        for m in core_missing:
            sheet = str(m.get("sheet") or "").strip()
            label = str(m.get("label") or "").strip()
            if not (sheet and label):
                continue
            by_sheet.setdefault(sheet, []).append(label)
        for sheet in sorted(by_sheet.keys()):
            labels = [x for x in by_sheet.get(sheet, []) if x]
            if labels:
                lines.append(f"- `{sheet}` 시트: " + ", ".join(labels))
        lines.append("- 빈칸이어도 문서는 생성되지만, 해당 문단/표는 `【작성자 기입 필요】`로 남습니다.")
    else:
        lines.append("- (없음) 핵심 필드 빈칸이 탐지되지 않았습니다.")
    lines.append("")

    lines.append("## 1) 필수(핵심) 입력 — 빈칸 목록")
    if not core_missing:
        lines.append("- (없음) 핵심 필드 빈칸이 탐지되지 않았습니다.")
    else:
        rows = [[m["sheet"], m["column"], m["label"]] for m in core_missing]
        lines.append(_md_table(["sheet", "column", "meaning"], rows))
    lines.append("")

    lines.append("## 2) 시트별 입력 가이드(핵심 컬럼 + 현재 채움률)")
    for g in _TABLE_GUIDES:
        s = next((x for x in table_summaries if x.get("sheet") == g.sheet), None) or {}
        row_count = s.get("row_count")
        lines.append(f"### {g.sheet} — {g.purpose}")
        if not s.get("exists"):
            lines.append("- (없음) 이 시트가 case.xlsx에 없습니다. (템플릿 업그레이드 필요)")
            lines.append("")
            continue
        lines.append(f"- 현재 입력행수: {row_count}")
        lines.append(f"- notes: {g.notes}")
        miss = s.get("missing") or {}
        rows2: list[list[str]] = []
        for col in g.key_columns:
            info = miss.get(col) or {}
            if not info.get("present"):
                rows2.append([col, "MISSING_COLUMN", "", ""])
                continue
            filled = info.get("filled", 0)
            total = info.get("total", 0)
            ratio = info.get("ratio", 0.0)
            rows2.append([col, f"{filled}/{total}", f"{ratio:.0%}" if isinstance(ratio, float) else "", ""])
        lines.append(_md_table(["column", "filled/total", "ratio", "note"], rows2))
        if (row_count or 0) == 0:
            lines.append("")
            lines.append("- 빈 행 템플릿(복붙용):")
            lines.append(_md_table(g.key_columns, [["" for _ in g.key_columns]]))
        lines.append("")

    lines.append("## 2-1) 그림(위치도 등) — 실제로 들어갔는지 빠른 확인")
    if not figures:
        lines.append("- (없음) FIGURES 시트가 비어있습니다.")
        lines.append("")
    else:
        sr_figs = (source_register_fig_usage.get("figures") or {}) if isinstance(source_register_fig_usage, dict) else {}
        rows_fig: list[list[str]] = []
        for f in figures:
            fid = str(f.get("fig_id") or "")
            used = sr_figs.get(fid) if isinstance(sr_figs, dict) else {}
            used_in = used.get("used_in") if isinstance(used, dict) else None
            used_in_str = ""
            if isinstance(used_in, list) and used_in:
                parts = []
                for u in used_in[:3]:
                    if not isinstance(u, dict):
                        continue
                    rep = _as_str(u.get("report"))
                    sec = _as_str(u.get("section_path"))
                    flag = _as_str(u.get("qa_flag"))
                    parts.append(":".join([x for x in [rep, sec, flag] if x]))
                used_in_str = ", ".join(parts)
            ev_fp = _as_str(used.get("evidence_file_path")) if isinstance(used, dict) else ""
            ev_id = _as_str(used.get("evidence_id")) if isinstance(used, dict) else ""
            rows_fig.append(
                [
                    fid,
                    _as_str(f.get("title")) or _as_str(f.get("caption")),
                    "OK" if bool(f.get("src_exists")) else "MISSING",
                    (_as_str(f.get("file_path")) or "")[:60],
                    (ev_id or "")[:24],
                    (ev_fp or "")[:60],
                    used_in_str[:60],
                ]
            )
        lines.append(
            _md_table(
                ["fig_id", "title/caption", "src_exists", "source(file_path)", "evidence_id", "evidence(file)", "used_in"],
                rows_fig,
            )
        )
        if isinstance(source_register_fig_usage, dict) and source_register_fig_usage.get("available"):
            lines.append(f"- source_register.xlsx: `{source_register_fig_usage.get('path')}`")
        lines.append("")

    lines.append("## 3) API/자동 채움(DATA_REQUESTS) — 무엇이 어디를 채우는가")
    if not data_requests:
        lines.append("- (없음) DATA_REQUESTS 시트가 비어있습니다.")
        lines.append("")
    else:
        dr_rows: list[list[str]] = []
        for r in data_requests:
            dr_rows.append(
                [
                    str(r.get("req_id") or ""),
                    str(r.get("enabled")),
                    str(r.get("connector") or ""),
                    str(r.get("output_sheet") or ""),
                    _as_str(r.get("params_json"))[:80],
                    _as_str(r.get("last_run_at")),
                    _as_str(r.get("last_evidence_ids")),
                ]
            )
        lines.append(
            _md_table(
                ["req_id", "enabled", "connector", "output_sheet", "params_json(80)", "last_run_at", "last_evidence_ids"],
                dr_rows,
            )
        )
        lines.append("")

        miss_env = data_requests_summary.get("disabled_missing_env") or {}
        miss_params = data_requests_summary.get("disabled_missing_params") or []
        if miss_env:
            lines.append("### 3-1) (disabled) 환경변수(API KEY) 필요")
            rows3 = [[env, ", ".join(reqs)] for env, reqs in sorted(miss_env.items())]
            lines.append(_md_table(["env_var", "req_ids"], rows3))
            lines.append("")
        if miss_params:
            lines.append("### 3-2) (disabled) params_json 빈칸")
            rows4 = [[str(x.get("req_id") or ""), ", ".join(x.get("blank_params") or []), _as_str(x.get("note"))] for x in miss_params]
            lines.append(_md_table(["req_id", "blank_params", "note"], rows4))
            lines.append("")

    lines.append("## 4) 오염(샘플 혼입) 스캔")
    scan_rows: list[list[str]] = []
    for area, counts_map in scan.items():
        for term, cnt in (counts_map or {}).items():
            scan_rows.append([area, term, str(cnt)])
    lines.append(_md_table(["area", "term", "count"], scan_rows))
    lines.append("")

    lines.append("## 5) 템플릿(서식) 체크")
    lines.append(f"- case template(EIA): `{template_check.get('template_eia')}`")
    lines.append(f"- sample scaffolded: `{template_check.get('sample_template')}`")
    lines.append(f"- same_contents_as_sample_scaffolded: `{template_check.get('same_contents_as_sample_scaffolded')}`")
    tt = template_check.get("template_terms") or {}
    if tt:
        lines.append(
            f"- template terms: 샘플={tt.get('샘플',0)}, 소규모환경영향평가서={tt.get('소규모환경영향평가서',0)}, 관광농원={tt.get('관광농원',0)}"
        )
    lines.append("")

    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_md.write_text("\n".join(lines), encoding="utf-8")

    return portal


def _copy_portal_deliverable(*, portal_md: Path, portal: dict[str, Any], case_dir: Path) -> None:
    try:
        from eia_gen.config import settings
    except Exception:
        return

    deliverables_dir = str(getattr(settings, "deliverables_dir", "") or "").strip()
    if not deliverables_dir:
        return

    tag = str(getattr(settings, "deliverables_tag", "") or "").strip()
    if not tag:
        meta = portal.get("meta") if isinstance(portal.get("meta"), dict) else {}
        tag = str((meta or {}).get("case_id") or "").strip() or case_dir.name
    tag = re.sub(r"\s+", "_", tag)
    tag = re.sub(r"[^\w.\-]+", "_", tag, flags=re.UNICODE).strip("._-") or "case"

    ts = ""
    paths = portal.get("paths") if isinstance(portal.get("paths"), dict) else {}
    gen_dir = str((paths or {}).get("latest_generate_dir") or "").strip()
    if gen_dir:
        p = Path(gen_dir)
        if p.name == "generate" and _TS_RE.match(p.parent.name):
            ts = p.parent.name
    if not ts:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")

    try:
        dst_dir = Path(deliverables_dir).expanduser().resolve()
        dst_dir.mkdir(parents=True, exist_ok=True)
        dst = dst_dir / f"input_portal.{tag}_{ts}.md"
        shutil.copy2(portal_md.resolve(), dst)
        print(f"OK copied deliverable: {dst}")
    except Exception as e:
        print(f"WARN deliverables: copy failed: {e}")


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Create an integrated 'input portal' markdown for a case directory: "
            "what to fill in case.xlsx, which APIs feed which sheets, and what's still blank."
        )
    )
    ap.add_argument("--case-dir", required=True, type=Path, help="Case directory containing case.xlsx + sources.yaml")
    ap.add_argument("--out-md", type=Path, default=None, help="Output markdown path (default: <case_dir>/INPUT_PORTAL.md)")
    ap.add_argument("--out-json", type=Path, default=None, help="Optional output JSON path")
    args = ap.parse_args()

    case_dir = args.case_dir.expanduser().resolve()
    out_md = (args.out_md.expanduser().resolve() if args.out_md else (case_dir / "INPUT_PORTAL.md"))
    out_json = args.out_json.expanduser().resolve() if args.out_json else None

    portal = build_portal(case_dir=case_dir, out_md=out_md, out_json=out_json)
    print(f"OK wrote: {out_md}")
    if out_json:
        print(f"OK wrote: {out_json}")
    _copy_portal_deliverable(portal_md=out_md, portal=portal, case_dir=case_dir)


if __name__ == "__main__":
    main()
