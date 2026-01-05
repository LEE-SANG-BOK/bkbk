#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return not any(v is not None and _as_str(v) for v in values)


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def _sheet_rows(ws) -> list[dict[str, Any]]:
    headers = [_as_str(c.value) for c in ws[1]]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or _is_blank_row(row):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if i < len(row):
                d[h] = row[i]
        out.append(d)
    return out


def _find_row(ws, *, key_col: str, key_value: str) -> int | None:
    hm = _header_map(ws)
    if key_col not in hm:
        return None
    col = hm[key_col]
    for r in range(2, ws.max_row + 1):
        v = _as_str(ws.cell(row=r, column=col).value)
        if not v:
            continue
        if v == key_value:
            return r
    return None


def _write_cell(ws, *, row: int, col_name: str, value: Any) -> None:
    hm = _header_map(ws)
    if col_name not in hm:
        raise SystemExit(f"Missing column '{col_name}' in sheet '{ws.title}'")
    ws.cell(row=row, column=hm[col_name]).value = value


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _resolve_input_path(*, repo_root: Path, case_dir: Path, user_path: str) -> Path:
    p = Path(user_path).expanduser()
    if p.is_absolute():
        return p.resolve()

    # Try common roots for better UX:
    # - current working directory (workspace root or eia-gen/)
    # - case directory (case local assets)
    # - repo root (eia-gen/)
    candidates = [
        (Path.cwd() / p).resolve(),
        (case_dir / p).resolve(),
        (repo_root / p).resolve(),
    ]
    for c in candidates:
        if c.exists():
            return c
    # Fall back to case-dir resolution for consistent error messages.
    return (case_dir / p).resolve()


def _relpath(path: Path, base_dir: Path) -> str:
    try:
        return os.path.relpath(path.resolve(), start=base_dir.resolve()).replace("\\", "/")
    except Exception:
        return str(path)


def _resolve_path_from_attachments(ws_att, *, repo_root: Path, case_dir: Path, evidence_id: str) -> Path:
    r = _find_row(ws_att, key_col="evidence_id", key_value=evidence_id)
    if r is None:
        raise SystemExit(f"ATTACHMENTS row not found for evidence_id={evidence_id}")
    hm = _header_map(ws_att)
    fp = _as_str(ws_att.cell(row=r, column=hm.get("file_path", 0)).value)
    if not fp:
        raise SystemExit(f"ATTACHMENTS.file_path is empty for evidence_id={evidence_id}")
    p = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=fp)
    if not p.exists():
        raise SystemExit(f"Attachment file not found: {p} (from evidence_id={evidence_id})")
    return p


def _ensure_figures_row(
    wb,
    *,
    fig_id: str,
    doc_scope: str,
    figure_type: str,
    title: str,
    caption: str,
    source_origin: str,
    width_mm: int,
    src_id: str,
    sensitive: str,
) -> int:
    ws = wb["FIGURES"]
    row = _find_row(ws, key_col="fig_id", key_value=fig_id)
    if row is not None:
        return row

    hm = _header_map(ws)
    values: dict[str, Any] = {}
    values["fig_id"] = fig_id
    values["doc_scope"] = doc_scope
    values["figure_type"] = figure_type
    values["title"] = title or caption or fig_id
    values["caption"] = caption or title or fig_id
    values["source_origin"] = source_origin
    values["file_path"] = ""
    values["gen_method"] = "MASK_ANNOTATED_IMAGE"
    values["geom_ref"] = ""
    values["width_mm"] = width_mm
    values["crop"] = ""
    values["src_id"] = src_id
    values["sensitive"] = sensitive
    values["insert_anchor"] = ""

    ws.append([values.get(h, "") for h in hm.keys()])
    return ws.max_row


def _run_mask_to_polygons(
    *,
    repo_root: Path,
    py_exe: str,
    mask_path: Path,
    base_image_path: Path,
    out_annotations: Path,
    labels: str,
    max_polygons: int,
    min_area_ratio: float,
    closing_radius: int,
    opening_radius: int,
    hole_area_ratio: float,
    approx_tol: float,
    emit_number_badges: bool,
    badge_prefix: str,
    emit_legend: bool,
    legend_box: list[int] | None,
    roi: list[int] | None,
    exclude_rects: list[list[int]],
    debug_dir: Path | None,
) -> None:
    script = repo_root / "scripts" / "mask_to_polygons.py"
    if not script.exists():
        raise SystemExit(f"mask_to_polygons.py not found: {script}")

    cmd = [
        py_exe,
        str(script),
        "--mask",
        str(mask_path),
        "--base-image",
        str(base_image_path),
        "--out-annotations",
        str(out_annotations),
        "--labels",
        labels,
        "--max-polygons",
        str(int(max_polygons)),
        "--min-area-ratio",
        str(float(min_area_ratio)),
        "--closing-radius",
        str(int(closing_radius)),
        "--opening-radius",
        str(int(opening_radius)),
        "--hole-area-ratio",
        str(float(hole_area_ratio)),
        "--approx-tol",
        str(float(approx_tol)),
    ]
    if emit_number_badges:
        cmd.append("--emit-number-badges")
        cmd.extend(["--badge-prefix", badge_prefix])
    if emit_legend:
        cmd.append("--emit-legend")
        if legend_box and len(legend_box) == 2:
            cmd.extend(["--legend-box", str(int(legend_box[0])), str(int(legend_box[1]))])
    if roi and len(roi) == 4:
        cmd.extend(["--roi", str(int(roi[0])), str(int(roi[1])), str(int(roi[2])), str(int(roi[3]))])
    for r in exclude_rects or []:
        if len(r) != 4:
            continue
        cmd.extend(["--exclude-rect", str(int(r[0])), str(int(r[1])), str(int(r[2])), str(int(r[3]))])
    if debug_dir:
        cmd.extend(["--debug-dir", str(debug_dir)])

    subprocess.run(cmd, check=True, cwd=str(repo_root))


def _run_annotate_image(
    *,
    repo_root: Path,
    py_exe: str,
    base_image_path: Path,
    annotations_path: Path,
    style_path: Path,
    out_png: Path,
    font_path: str | None,
) -> None:
    script = repo_root / "scripts" / "annotate_image.py"
    if not script.exists():
        raise SystemExit(f"annotate_image.py not found: {script}")

    cmd = [
        py_exe,
        str(script),
        "--image",
        str(base_image_path),
        "--annotations",
        str(annotations_path),
        "--style",
        str(style_path),
        "--out",
        str(out_png),
    ]

    env = os.environ.copy()
    if font_path:
        env["EIA_GEN_FONT_PATH"] = font_path

    subprocess.run(cmd, check=True, cwd=str(repo_root), env=env)


def _preferred_python(repo_root: Path) -> str:
    venv_py = repo_root / ".venv" / "bin" / "python"
    if venv_py.exists():
        return str(venv_py)
    return sys.executable


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "One-shot: (mask PNG + base image) -> polygons -> annotations.yaml -> annotated PNG -> register into case.xlsx.\n"
            "Guardrail: Uses a user-provided mask only (no inference)."
        )
    )
    ap.add_argument("--xlsx", type=Path, required=True, help="case.xlsx path")
    ap.add_argument("--fig-id", type=str, required=True, help="FIGURES.fig_id to update/create")

    base = ap.add_mutually_exclusive_group(required=True)
    base.add_argument("--base-image", type=str, default="", help="Base image path (PNG/JPG).")
    base.add_argument("--base-evidence-id", type=str, default="", help="ATTACHMENTS.evidence_id for base image")

    mask = ap.add_mutually_exclusive_group(required=True)
    mask.add_argument("--mask-image", type=str, default="", help="Mask image path (PNG/JPG). White=foreground.")
    mask.add_argument("--mask-evidence-id", type=str, default="", help="ATTACHMENTS.evidence_id for mask image")

    ap.add_argument("--labels", type=str, default="", help="Polygon labels (comma/; separated)")
    ap.add_argument("--max-polygons", type=int, default=5, help="Keep up to N polygons")
    ap.add_argument("--min-area-ratio", type=float, default=0.002, help="Drop components smaller than ratio*(W*H)")
    ap.add_argument("--closing-radius", type=int, default=6, help="Binary closing radius (px)")
    ap.add_argument("--opening-radius", type=int, default=2, help="Binary opening radius (px)")
    ap.add_argument("--hole-area-ratio", type=float, default=0.001, help="Fill holes smaller than ratio*(W*H)")
    ap.add_argument("--approx-tol", type=float, default=6.0, help="Point decimation tolerance(px)")
    ap.add_argument("--roi", nargs=4, type=int, default=None, metavar=("X0", "Y0", "X1", "Y1"), help="ROI rectangle")
    ap.add_argument(
        "--exclude-rect",
        nargs=4,
        type=int,
        action="append",
        default=[],
        metavar=("X0", "Y0", "X1", "Y1"),
        help="Exclusion rectangle(s). Can be repeated.",
    )
    ap.add_argument("--emit-number-badges", action="store_true", help="Emit number_badge layers")
    ap.add_argument("--badge-prefix", type=str, default="#", help="Number badge prefix")
    ap.add_argument("--emit-legend", action="store_true", help="Emit legend layer")
    ap.add_argument("--legend-box", nargs=2, type=int, default=None, metavar=("X", "Y"), help="Legend box top-left")
    ap.add_argument("--debug-dir", type=str, default="", help="Optional debug dir (mask/contours)")

    ap.add_argument("--annotations-out", type=str, default="", help="Output annotations YAML path (default under case dir)")
    ap.add_argument("--style", type=Path, default=Path("config/figure_style.yaml"), help="Style tokens YAML (repo-relative)")
    ap.add_argument("--out", type=str, default="", help="Output PNG path (default under case dir)")

    ap.add_argument("--update-xlsx", action="store_true", help="Update FIGURES.file_path to the annotated PNG")
    ap.add_argument("--create-figure-row", action="store_true", help="Create FIGURES row if missing")
    ap.add_argument("--register-output-attachment", action="store_true", help="UPSERT output PNG row into ATTACHMENTS")
    ap.add_argument("--output-evidence-id", type=str, default="", help="Override evidence_id (default: DER-ANNO-MASK-{fig_id})")

    ap.add_argument("--doc-scope", type=str, default="BOTH", help="FIGURES.doc_scope when creating a row")
    ap.add_argument("--figure-type", type=str, default="AERIAL_PHOTO", help="FIGURES.figure_type when creating a row")
    ap.add_argument("--title", type=str, default="", help="FIGURES.title when creating a row")
    ap.add_argument("--caption", type=str, default="", help="FIGURES.caption when creating a row")
    ap.add_argument("--source-origin", type=str, default="REFERENCE", help="FIGURES.source_origin (OFFICIAL/REFERENCE/REF/UNKNOWN)")
    ap.add_argument("--width-mm", type=int, default=180, help="FIGURES.width_mm when creating a row")
    ap.add_argument("--src-id", type=str, default="S-CLIENT-001", help="FIGURES/ATTACHMENTS src_id")
    ap.add_argument("--sensitive", type=str, default="N", help="FIGURES/ATTACHMENTS sensitive flag (Y/N)")
    ap.add_argument("--font-path", type=str, default="", help="Optional font path (sets EIA_GEN_FONT_PATH)")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")
    case_dir = xlsx.parent.resolve()

    repo_root = Path(__file__).resolve().parents[1]  # eia-gen/
    py_exe = _preferred_python(repo_root)

    wb = load_workbook(xlsx)
    if "FIGURES" not in wb.sheetnames or "ATTACHMENTS" not in wb.sheetnames:
        raise SystemExit("case.xlsx must include FIGURES and ATTACHMENTS sheets (v2 template).")
    ws_att = wb["ATTACHMENTS"]

    fig_id = _as_str(args.fig_id)
    if not fig_id:
        raise SystemExit("--fig-id is required")

    # Resolve base/mask images
    if _as_str(args.base_evidence_id):
        base_img = _resolve_path_from_attachments(ws_att, repo_root=repo_root, case_dir=case_dir, evidence_id=_as_str(args.base_evidence_id))
    else:
        base_img = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=_as_str(args.base_image))
        if not base_img.exists():
            raise SystemExit(f"base image not found: {base_img}")

    if _as_str(args.mask_evidence_id):
        mask_img = _resolve_path_from_attachments(ws_att, repo_root=repo_root, case_dir=case_dir, evidence_id=_as_str(args.mask_evidence_id))
    else:
        mask_img = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=_as_str(args.mask_image))
        if not mask_img.exists():
            raise SystemExit(f"mask image not found: {mask_img}")

    # Paths (annotations + output)
    if _as_str(args.annotations_out):
        ann_path = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=_as_str(args.annotations_out))
    else:
        ann_path = (case_dir / "attachments" / "derived" / "annotations" / f"{fig_id}_mask.yaml").resolve()
    ann_path.parent.mkdir(parents=True, exist_ok=True)

    if _as_str(args.out):
        out_png = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=_as_str(args.out))
    else:
        out_png = (case_dir / "attachments" / "derived" / "figures" / "annotated" / f"{fig_id}.png").resolve()
    out_png.parent.mkdir(parents=True, exist_ok=True)

    debug_dir = None
    if _as_str(args.debug_dir):
        debug_dir = _resolve_input_path(repo_root=repo_root, case_dir=case_dir, user_path=_as_str(args.debug_dir))
        debug_dir.mkdir(parents=True, exist_ok=True)

    # style path (repo-relative default)
    style_path = args.style
    if not style_path.is_absolute():
        style_path = (repo_root / style_path).resolve()
    if not style_path.exists():
        raise SystemExit(f"style yaml not found: {style_path}")

    # 1) mask -> polygons -> annotations.yaml
    _run_mask_to_polygons(
        repo_root=repo_root,
        py_exe=py_exe,
        mask_path=mask_img,
        base_image_path=base_img,
        out_annotations=ann_path,
        labels=_as_str(args.labels),
        max_polygons=int(args.max_polygons),
        min_area_ratio=float(args.min_area_ratio),
        closing_radius=int(args.closing_radius),
        opening_radius=int(args.opening_radius),
        hole_area_ratio=float(args.hole_area_ratio),
        approx_tol=float(args.approx_tol),
        emit_number_badges=bool(args.emit_number_badges),
        badge_prefix=_as_str(args.badge_prefix) or "#",
        emit_legend=bool(args.emit_legend),
        legend_box=list(args.legend_box) if args.legend_box else None,
        roi=list(args.roi) if args.roi else None,
        exclude_rects=[list(x) for x in (args.exclude_rect or [])],
        debug_dir=debug_dir,
    )

    # 2) annotations.yaml + base image -> annotated PNG
    _run_annotate_image(
        repo_root=repo_root,
        py_exe=py_exe,
        base_image_path=base_img,
        annotations_path=ann_path,
        style_path=style_path,
        out_png=out_png,
        font_path=_as_str(args.font_path) or None,
    )

    # 3) sha256 + optional registration
    base_sha = _sha256_file(base_img)
    mask_sha = _sha256_file(mask_img)
    ann_sha = _sha256_file(ann_path)
    out_sha = _sha256_file(out_png)

    if args.register_output_attachment:
        out_eid = _as_str(args.output_evidence_id) or f"DER-ANNO-MASK-{fig_id}"
        out_row = _find_row(ws_att, key_col="evidence_id", key_value=out_eid)
        if out_row is None:
            out_row = ws_att.max_row + 1

        _write_cell(ws_att, row=out_row, col_name="evidence_id", value=out_eid)
        _write_cell(ws_att, row=out_row, col_name="evidence_type", value="파생이미지")
        _write_cell(ws_att, row=out_row, col_name="title", value=_as_str(args.caption) or _as_str(args.title) or fig_id)
        _write_cell(ws_att, row=out_row, col_name="file_path", value=_relpath(out_png, case_dir))
        _write_cell(ws_att, row=out_row, col_name="related_fig_id", value=fig_id)
        _write_cell(ws_att, row=out_row, col_name="used_in", value="FIGURE_MASK_ANNOTATE")
        _write_cell(ws_att, row=out_row, col_name="data_origin", value="DERIVED")
        _write_cell(ws_att, row=out_row, col_name="src_id", value=_as_str(args.src_id))
        _write_cell(ws_att, row=out_row, col_name="sensitive", value=_as_str(args.sensitive) or "N")
        _write_cell(
            ws_att,
            row=out_row,
            col_name="note",
            value=(
                "DERIVED: MASK_TO_POLYGONS+ANNOTATE "
                f"base_sha256={base_sha} mask_sha256={mask_sha} ann_sha256={ann_sha} out_sha256={out_sha} "
                f"base={_relpath(base_img, case_dir)} mask={_relpath(mask_img, case_dir)} ann={_relpath(ann_path, case_dir)} "
                f"params=max_polygons:{int(args.max_polygons)},min_area_ratio:{float(args.min_area_ratio)},"
                f"closing:{int(args.closing_radius)},opening:{int(args.opening_radius)},hole_area_ratio:{float(args.hole_area_ratio)},"
                f"approx_tol:{float(args.approx_tol)}"
            ),
        )

    if args.update_xlsx or args.create_figure_row:
        ws_fig = wb["FIGURES"]
        fig_row = _find_row(ws_fig, key_col="fig_id", key_value=fig_id)
        if fig_row is None:
            if not args.create_figure_row:
                raise SystemExit(f"FIGURES row not found for fig_id={fig_id} (use --create-figure-row).")
            fig_row = _ensure_figures_row(
                wb,
                fig_id=fig_id,
                doc_scope=_as_str(args.doc_scope) or "BOTH",
                figure_type=_as_str(args.figure_type) or "AERIAL_PHOTO",
                title=_as_str(args.title),
                caption=_as_str(args.caption) or "마스크 기반 주석 합성 이미지",
                source_origin=_as_str(args.source_origin) or "REFERENCE",
                width_mm=int(args.width_mm) if args.width_mm else 180,
                src_id=_as_str(args.src_id) or "S-CLIENT-001",
                sensitive=_as_str(args.sensitive) or "N",
            )
        _write_cell(ws_fig, row=fig_row, col_name="file_path", value=_relpath(out_png, case_dir))
        _write_cell(ws_fig, row=fig_row, col_name="source_origin", value=_as_str(args.source_origin) or "REFERENCE")
        _write_cell(ws_fig, row=fig_row, col_name="gen_method", value="MASK_ANNOTATED_IMAGE")
        if _as_str(args.caption):
            _write_cell(ws_fig, row=fig_row, col_name="caption", value=_as_str(args.caption))
        if _as_str(args.title):
            _write_cell(ws_fig, row=fig_row, col_name="title", value=_as_str(args.title))

    wb.save(xlsx)

    print(f"OK: wrote {out_png}")
    print(f"- fig_id={fig_id}")
    print(f"- annotations={ann_path}")
    print(f"- out_png_sha256={out_sha}")
    if args.update_xlsx:
        print("- updated FIGURES.file_path")
    if args.register_output_attachment:
        print("- upserted ATTACHMENTS row")


if __name__ == "__main__":
    main()
