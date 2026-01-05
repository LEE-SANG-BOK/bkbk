#!/usr/bin/env python3
from __future__ import annotations

"""
Verify markdown refs like:
  - `path/to/file.py:123`
  - `case.xlsx:SHEET#row14`

Goal: reduce "ref drift" in docs by catching:
  - missing files
  - line numbers beyond EOF (text files)
  - missing xlsx sheets / row indices (best-effort)

This is a local-only hygiene tool (no network).
"""

import argparse
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


@dataclass(frozen=True)
class RefIssue:
    severity: str  # ERROR | WARN
    md_path: Path
    md_line: int
    ref: str
    message: str


_FENCE_RE = re.compile(r"^\s*```")
_INLINE_CODE_RE = re.compile(r"`([^`]+)`")
_XLSX_REF_RE = re.compile(r"^(?P<path>.+?\.xlsx):(?P<sheet>[A-Za-z0-9_]+)#row(?P<row>\d+)(?:c(?P<col>\d+))?$")
_FILE_LINE_RE = re.compile(r"^(?P<path>.+?):(?P<line>\d+)(?::(?P<col>\d+))?$")


_TEXT_EXTS = {
    ".py",
    ".md",
    ".txt",
    ".yaml",
    ".yml",
    ".json",
    ".toml",
    ".ini",
    ".cfg",
    ".csv",
    ".tsv",
}
_BINARY_EXTS = {
    ".png",
    ".jpg",
    ".jpeg",
    ".gif",
    ".webp",
    ".pdf",
    ".docx",
    ".pptx",
    ".xlsx",
    ".zip",
    ".gz",
}


def _iter_inline_code_spans(md_text: str) -> Iterable[str]:
    for m in _INLINE_CODE_RE.finditer(md_text):
        yield m.group(1)


def _split_maybe_csv(s: str) -> list[str]:
    # Many refs are written as `a.py:1`, `b.py:2` (separate backticks),
    # but we also defensively split on commas/semicolons.
    tokens: list[str] = []
    for chunk in s.replace(";", ",").split(","):
        t = chunk.strip()
        if t:
            tokens.append(t)
    return tokens


def _repo_root() -> Path:
    # scripts/verify_md_refs.py -> scripts/ -> repo_root
    return Path(__file__).resolve().parents[1]


def _resolve_ref_path(raw_path: str, *, repo_root: Path, workspace_root: Path) -> Path:
    p = Path(raw_path)
    if p.is_absolute():
        return p

    # Many docs include `eia-gen/...` prefix even though docs live inside repo root.
    s = raw_path.replace("\\", "/")
    if s.startswith("eia-gen/"):
        p2 = (repo_root / s.removeprefix("eia-gen/")).resolve()
        return p2

    # Try repo-relative first.
    p2 = (repo_root / p).resolve()
    if p2.exists():
        return p2

    # Then try workspace root (for AGENTS.md, root-level docs, etc.).
    return (workspace_root / p).resolve()


def _is_text_file(path: Path) -> bool:
    ext = path.suffix.lower()
    if ext in _TEXT_EXTS:
        return True
    if ext in _BINARY_EXTS:
        return False
    # fallback: treat unknown as text if small and decodable (best-effort)
    return False


def _count_lines_text(path: Path) -> int:
    # Best-effort; do not crash on encoding issues.
    try:
        return sum(1 for _ in path.open("r", encoding="utf-8", errors="replace"))
    except Exception:
        return 0


def _check_file_line_ref(
    *,
    md_path: Path,
    md_line: int,
    raw: str,
    repo_root: Path,
    workspace_root: Path,
) -> list[RefIssue]:
    m = _FILE_LINE_RE.match(raw)
    if not m:
        return []
    raw_path = m.group("path")
    line = int(m.group("line"))

    resolved = _resolve_ref_path(raw_path, repo_root=repo_root, workspace_root=workspace_root)
    if not resolved.exists():
        return [RefIssue("ERROR", md_path, md_line, raw, f"file not found: {resolved}")]

    # For binary refs, existence is enough.
    if not _is_text_file(resolved):
        return []

    total = _count_lines_text(resolved)
    if total <= 0:
        return [RefIssue("WARN", md_path, md_line, raw, f"could not count lines (encoding?) for: {resolved}")]
    if line < 1 or line > total:
        return [RefIssue("ERROR", md_path, md_line, raw, f"line out of range: {line} (file has {total} lines): {resolved}")]

    return []


def _check_xlsx_ref(
    *,
    md_path: Path,
    md_line: int,
    raw: str,
    repo_root: Path,
    workspace_root: Path,
    check_rows: bool,
) -> list[RefIssue]:
    m = _XLSX_REF_RE.match(raw)
    if not m:
        return []
    raw_path = m.group("path")
    sheet = m.group("sheet")
    row = int(m.group("row"))
    col = int(m.group("col")) if m.group("col") else None

    resolved = _resolve_ref_path(raw_path, repo_root=repo_root, workspace_root=workspace_root)
    if not resolved.exists():
        return [RefIssue("ERROR", md_path, md_line, raw, f"xlsx not found: {resolved}")]

    if not check_rows:
        return []

    try:
        wb = load_workbook(resolved, read_only=True, data_only=True)
    except Exception as e:
        return [RefIssue("ERROR", md_path, md_line, raw, f"failed to open xlsx: {resolved} ({e})")]

    try:
        if sheet not in wb.sheetnames:
            return [RefIssue("ERROR", md_path, md_line, raw, f"sheet not found: {sheet} in {resolved}")]
        ws = wb[sheet]
        if row < 1 or row > ws.max_row:
            return [
                RefIssue(
                    "ERROR",
                    md_path,
                    md_line,
                    raw,
                    f"row out of range: {row} (sheet max_row={ws.max_row}) in {resolved}:{sheet}",
                )
            ]
        if col is not None and (col < 1 or col > ws.max_column):
            return [
                RefIssue(
                    "ERROR",
                    md_path,
                    md_line,
                    raw,
                    f"col out of range: {col} (sheet max_column={ws.max_column}) in {resolved}:{sheet}",
                )
            ]
        return []
    finally:
        try:
            wb.close()
        except Exception:
            pass


def _check_markdown_file(md_path: Path, *, check_xlsx_rows: bool) -> tuple[int, list[RefIssue]]:
    repo_root = _repo_root()
    workspace_root = repo_root.parent

    if not md_path.exists():
        return 0, [RefIssue("ERROR", md_path, 0, "", f"markdown not found: {md_path}")]

    issues: list[RefIssue] = []
    total_refs = 0
    in_fence = False

    for i, line in enumerate(md_path.read_text(encoding="utf-8", errors="replace").splitlines(), start=1):
        if _FENCE_RE.match(line):
            in_fence = not in_fence
            continue
        if in_fence:
            continue

        for span in _iter_inline_code_spans(line):
            for token in _split_maybe_csv(span):
                if token.startswith("http://") or token.startswith("https://"):
                    continue

                total_refs += 1
                issues.extend(
                    _check_xlsx_ref(
                        md_path=md_path,
                        md_line=i,
                        raw=token,
                        repo_root=repo_root,
                        workspace_root=workspace_root,
                        check_rows=check_xlsx_rows,
                    )
                )
                issues.extend(
                    _check_file_line_ref(
                        md_path=md_path,
                        md_line=i,
                        raw=token,
                        repo_root=repo_root,
                        workspace_root=workspace_root,
                    )
                )

    return total_refs, issues


def main() -> None:
    ap = argparse.ArgumentParser(description="Verify `path:line` refs in markdown files (SSOT hygiene).")
    ap.add_argument(
        "--md",
        action="append",
        default=[],
        help="Markdown file path (repeatable). Default: docs/11_execution_plan.md",
    )
    ap.add_argument(
        "--check-xlsx-rows",
        action="store_true",
        help="Also verify xlsx sheet/row refs like `case.xlsx:SHEET#row14` (slower).",
    )
    args = ap.parse_args()

    repo_root = _repo_root()
    md_paths = [Path(p) for p in args.md] if args.md else [repo_root / "docs" / "11_execution_plan.md"]
    md_paths = [p if p.is_absolute() else (repo_root / p).resolve() for p in md_paths]

    grand_total = 0
    all_issues: list[RefIssue] = []
    for md_path in md_paths:
        total, issues = _check_markdown_file(md_path, check_xlsx_rows=bool(args.check_xlsx_rows))
        grand_total += total
        all_issues.extend(issues)

    errors = [x for x in all_issues if x.severity == "ERROR"]
    warns = [x for x in all_issues if x.severity == "WARN"]

    for issue in all_issues:
        loc = f"{issue.md_path}:{issue.md_line}" if issue.md_line else str(issue.md_path)
        print(f"{issue.severity}: {loc}: `{issue.ref}`: {issue.message}")

    print(f"checked_refs: {grand_total}  errors: {len(errors)}  warns: {len(warns)}")

    if errors:
        raise SystemExit(1)


if __name__ == "__main__":
    main()

