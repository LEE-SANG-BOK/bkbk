#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any


try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except Exception:
    openpyxl = None  # type: ignore


@dataclass(frozen=True)
class Row:
    sample_page: int
    kind: str  # CHAPTER|FIGURE|TABLE
    label: str
    caption: str


_LINE_RE = re.compile(r"^-\s+p(\d+):\s+`([^`]+)`\s*(.*)$")


def _parse_block(lines: list[str], *, kind: str) -> list[Row]:
    out: list[Row] = []
    for ln in lines:
        ln = ln.rstrip("\n")
        m = _LINE_RE.match(ln)
        if not m:
            continue
        page = int(m.group(1))
        label = (m.group(2) or "").strip()
        caption = (m.group(3) or "").strip()
        out.append(Row(sample_page=page, kind=kind, label=label, caption=caption))
    return out


def _split_sections(text: str) -> dict[str, list[str]]:
    sections: dict[str, list[str]] = {}
    cur = None
    for ln in text.splitlines():
        if ln.startswith("## "):
            cur = ln.strip()
            sections[cur] = []
            continue
        if cur is None:
            continue
        sections[cur].append(ln)
    return sections


_HEADERS = [
    "sample_page",
    "kind",
    "label",
    "caption",
    "our_spec_id",
    "status(EXIST/NEED/IGNORE)",
    "note",
]


def _write_csv(path: Path, rows: list[Row]) -> None:
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(_HEADERS)
        for r in rows:
            w.writerow([r.sample_page, r.kind, r.label, r.caption, "", "", ""])


def _write_xlsx(path: Path, rows: list[Row]) -> None:
    if openpyxl is None:
        raise SystemExit("openpyxl is required for xlsx output. Install project deps first.")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "COVERAGE_MATRIX"

    ws.append(_HEADERS)
    header_fill = PatternFill("solid", fgColor="D9E1F2")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for j in range(1, len(_HEADERS) + 1):
        c = ws.cell(row=1, column=j)
        c.fill = header_fill
        c.font = header_font
        c.alignment = header_align

    for r in rows:
        ws.append([r.sample_page, r.kind, r.label, r.caption, "", "", ""])

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(_HEADERS))}1"

    # column widths
    widths = {
        "A": 11,
        "B": 10,
        "C": 10,
        "D": 60,
        "E": 22,
        "F": 22,
        "G": 40,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Create a coverage matrix seed from output/changwon_pdf_index/summary.md. "
            "Output can be .csv or .xlsx."
        )
    )
    ap.add_argument(
        "--summary",
        type=Path,
        default=Path("output/changwon_pdf_index/summary.md"),
        help="Path to PDF index summary.md",
    )
    ap.add_argument(
        "--out",
        type=Path,
        default=Path("output/changwon_pdf_index/coverage_matrix_seed.csv"),
        help="Output path (.csv or .xlsx)",
    )
    args = ap.parse_args()

    summary = args.summary.resolve()
    if not summary.exists():
        raise SystemExit(f"summary.md not found: {summary}")

    text = summary.read_text(encoding="utf-8")
    sec = _split_sections(text)

    rows: list[Row] = []
    if "## Chapters" in sec:
        rows.extend(_parse_block(sec["## Chapters"], kind="CHAPTER"))
    if "## Figures (captions)" in sec:
        rows.extend(_parse_block(sec["## Figures (captions)"], kind="FIGURE"))
    if "## Tables (captions)" in sec:
        rows.extend(_parse_block(sec["## Tables (captions)"], kind="TABLE"))

    rows.sort(key=lambda r: (r.kind, r.sample_page, r.label, r.caption))

    out_path = args.out.resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    if out_path.suffix.lower() == ".xlsx":
        _write_xlsx(out_path, rows)
    else:
        _write_csv(out_path, rows)

    print(f"OK wrote: {out_path}")
    print(f"Rows: {len(rows)}")


if __name__ == "__main__":
    main()
