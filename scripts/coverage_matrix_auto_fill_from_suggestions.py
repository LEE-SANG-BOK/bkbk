#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import openpyxl


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _as_float(v: Any) -> float | None:
    s = _as_str(v)
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i for i, h in enumerate(headers, start=1) if h}


def _col(hmap: dict[str, int], name: str) -> int:
    if name not in hmap:
        raise SystemExit(f"Missing required column '{name}' in coverage xlsx")
    return hmap[name]


def _append_note(existing: str, msg: str) -> str:
    existing = (existing or "").strip()
    if not existing:
        return msg
    if msg in existing:
        return existing
    return f"{existing} | {msg}"


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Auto-fill our_spec_id from suggest_1_* columns when score is high enough.\n"
            "This reduces repetitive manual copy/paste while keeping risk low.\n"
            "Recommended workflow:\n"
            "  1) suggest_coverage_spec_ids.py\n"
            "  2) this script (high-confidence autofill)\n"
            "  3) coverage_matrix_auto_status.py (optional)\n"
        )
    )
    ap.add_argument("--in-xlsx", type=Path, required=True, help="Input coverage_matrix_seed.suggested.xlsx path")
    ap.add_argument(
        "--out-xlsx",
        type=Path,
        default=None,
        help="Output xlsx path (default: <in>.autofilled.xlsx)",
    )
    ap.add_argument(
        "--in-place",
        action="store_true",
        help="Write back to --in-xlsx (overwrites file).",
    )
    ap.add_argument(
        "--min-score",
        type=float,
        default=0.92,
        help="Minimum suggest_1_score to auto-fill our_spec_id (default: 0.92).",
    )
    ap.add_argument(
        "--force",
        action="store_true",
        help="Overwrite non-empty our_spec_id (default: only fill empty).",
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not write any file; print summary only.",
    )
    args = ap.parse_args()

    in_xlsx = args.in_xlsx.expanduser().resolve()
    if not in_xlsx.exists():
        raise SystemExit(f"Input xlsx not found: {in_xlsx}")

    wb = openpyxl.load_workbook(in_xlsx)
    ws = wb.active
    hmap = _header_map(ws)

    col_kind = _col(hmap, "kind")
    col_our = _col(hmap, "our_spec_id")
    col_note = _col(hmap, "note")

    # suggestion columns are optional; if missing, we fail fast (this script is for suggested xlsx).
    col_sug_id = _col(hmap, "suggest_1_id")
    col_sug_score = _col(hmap, "suggest_1_score")
    col_sug_scope = hmap.get("suggest_1_scope")  # optional

    filled = 0
    skipped = 0
    unchanged = 0

    for r in range(2, ws.max_row + 1):
        kind = _as_str(ws.cell(row=r, column=col_kind).value).upper()
        our = _as_str(ws.cell(row=r, column=col_our).value)
        note = _as_str(ws.cell(row=r, column=col_note).value)

        # Avoid auto-fill for CHAPTER by default (risk of mis-mapping)
        if kind == "CHAPTER":
            skipped += 1
            continue

        if our and not args.force:
            skipped += 1
            continue

        sid = _as_str(ws.cell(row=r, column=col_sug_id).value)
        score = _as_float(ws.cell(row=r, column=col_sug_score).value)
        if not sid or score is None:
            unchanged += 1
            continue

        if score < args.min_score:
            unchanged += 1
            continue

        scope = ""
        if col_sug_scope is not None:
            scope = _as_str(ws.cell(row=r, column=col_sug_scope).value)

        ws.cell(row=r, column=col_our).value = sid
        msg = f"auto-filled our_spec_id from suggest_1 (score={score:.2f}{', scope='+scope if scope else ''})"
        ws.cell(row=r, column=col_note).value = _append_note(note, msg)
        filled += 1

    print(f"SUMMARY: filled={filled} skipped={skipped} unchanged={unchanged} min_score={args.min_score}")

    if args.dry_run:
        print("DRY-RUN: not writing output")
        return

    if args.in_place and args.out_xlsx is not None:
        raise SystemExit("Use either --in-place or --out-xlsx (not both).")

    out_xlsx = args.out_xlsx
    if args.in_place:
        out_xlsx = in_xlsx
    elif out_xlsx is None:
        out_xlsx = in_xlsx.with_suffix("").with_name(in_xlsx.stem + ".autofilled.xlsx")

    out_xlsx = out_xlsx.expanduser().resolve()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"WROTE: {out_xlsx}")


if __name__ == "__main__":
    main()

