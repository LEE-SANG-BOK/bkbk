#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return not any(v is not None and _as_str(v) for v in values)


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def _find_row(ws, *, key_col: str, key_value: str) -> int | None:
    hm = _header_map(ws)
    if key_col not in hm:
        return None
    col = hm[key_col]
    for r in range(2, ws.max_row + 1):
        v = _as_str(ws.cell(row=r, column=col).value)
        if not v:
            continue
        if v == key_value:
            return r
    return None


def _write_cell(ws, *, row: int, col_name: str, value: Any) -> None:
    hm = _header_map(ws)
    if col_name not in hm:
        raise SystemExit(f"Missing column '{col_name}' in sheet '{ws.title}'")
    ws.cell(row=row, column=hm[col_name]).value = value


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _resolve_path(case_dir: Path, p: str) -> Path:
    path = Path(p).expanduser()
    if not path.is_absolute():
        path = (case_dir / path).resolve()
    return path


def _relpath_or_str(path: Path, base_dir: Path) -> str:
    try:
        return path.resolve().relative_to(base_dir.resolve()).as_posix()
    except Exception:
        return str(path)


def _resolve_base_image_from_attachments(ws_att, *, case_dir: Path, evidence_id: str) -> Path:
    r = _find_row(ws_att, key_col="evidence_id", key_value=evidence_id)
    if r is None:
        raise SystemExit(f"ATTACHMENTS row not found for evidence_id={evidence_id}")
    hm = _header_map(ws_att)
    fp = _as_str(ws_att.cell(row=r, column=hm.get("file_path", 0)).value)
    if not fp:
        raise SystemExit(f"ATTACHMENTS.file_path is empty for evidence_id={evidence_id}")
    p = _resolve_path(case_dir, fp)
    if not p.exists():
        raise SystemExit(f"Base image not found on disk: {p} (from evidence_id={evidence_id})")
    return p


def _ensure_figures_row(
    wb,
    *,
    fig_id: str,
    doc_scope: str,
    figure_type: str,
    title: str,
    caption: str,
    source_origin: str,
    width_mm: int,
    src_id: str,
    sensitive: str,
) -> int:
    ws = wb["FIGURES"]
    row = _find_row(ws, key_col="fig_id", key_value=fig_id)
    if row is not None:
        return row

    hm = _header_map(ws)
    values: dict[str, Any] = {}
    values["fig_id"] = fig_id
    values["doc_scope"] = doc_scope
    values["figure_type"] = figure_type
    values["title"] = title or caption or fig_id
    values["caption"] = caption or title or fig_id
    values["source_origin"] = source_origin
    values["file_path"] = ""
    values["gen_method"] = "ANNOTATED_IMAGE"
    values["geom_ref"] = ""
    values["width_mm"] = width_mm
    values["crop"] = ""
    values["src_id"] = src_id
    values["sensitive"] = sensitive
    values["insert_anchor"] = ""

    ws.append([values.get(h, "") for h in hm.keys()])
    return ws.max_row


def _run_annotate(
    *,
    image_path: Path,
    annotations_path: Path,
    style_path: Path,
    out_png: Path,
    font_path: str | None,
) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    script = repo_root / "scripts" / "annotate_image.py"
    if not script.exists():
        raise SystemExit(f"annotate_image.py not found: {script}")

    cmd = [
        sys.executable,
        str(script),
        "--image",
        str(image_path),
        "--annotations",
        str(annotations_path),
        "--style",
        str(style_path),
        "--out",
        str(out_png),
    ]

    env = os.environ.copy()
    if font_path:
        env["EIA_GEN_FONT_PATH"] = font_path

    subprocess.run(cmd, check=True, cwd=str(repo_root), env=env)


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Render an annotated PNG from (base image + annotations.yaml) and optionally register it into case.xlsx.\n"
            "Guardrail: This script only overlays polygons/labels/badges/legend onto an existing image. No inference."
        )
    )
    ap.add_argument("--xlsx", type=Path, required=True, help="case.xlsx path")
    ap.add_argument("--fig-id", type=str, required=True, help="FIGURES.fig_id to update/create")

    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--base-image", type=str, default="", help="Base image path (PNG/JPG). Relative paths are case-dir based.")
    src.add_argument("--base-evidence-id", type=str, default="", help="ATTACHMENTS.evidence_id for base image")

    ap.add_argument(
        "--annotations",
        type=str,
        required=True,
        help="Annotations YAML/JSON path. Relative paths are case-dir based.",
    )
    ap.add_argument("--style", type=Path, default=Path("config/figure_style.yaml"), help="Style tokens YAML (repo-relative)")

    ap.add_argument(
        "--out",
        type=str,
        default="",
        help="Output PNG path (default: attachments/derived/figures/annotated/{fig_id}.png under case dir)",
    )

    ap.add_argument("--update-xlsx", action="store_true", help="Update FIGURES.file_path to the annotated PNG")
    ap.add_argument("--create-figure-row", action="store_true", help="Create FIGURES row if missing")

    ap.add_argument(
        "--register-output-attachment",
        action="store_true",
        help="Register the annotated PNG into ATTACHMENTS sheet (UPSERT by evidence_id).",
    )
    ap.add_argument(
        "--output-evidence-id",
        type=str,
        default="",
        help="Override evidence_id for ATTACHMENTS row (default: DER-ANNO-{fig_id})",
    )

    ap.add_argument("--doc-scope", type=str, default="BOTH", help="FIGURES.doc_scope when creating a row")
    ap.add_argument("--figure-type", type=str, default="AERIAL_PHOTO", help="FIGURES.figure_type when creating a row")
    ap.add_argument("--title", type=str, default="", help="FIGURES.title when creating a row")
    ap.add_argument("--caption", type=str, default="", help="FIGURES.caption when creating a row")
    ap.add_argument("--source-origin", type=str, default="REFERENCE", help="FIGURES.source_origin (OFFICIAL/REFERENCE/REF/UNKNOWN)")
    ap.add_argument("--width-mm", type=int, default=180, help="FIGURES.width_mm when creating a row")
    ap.add_argument("--src-id", type=str, default="S-CLIENT-001", help="FIGURES/ATTACHMENTS src_id")
    ap.add_argument("--sensitive", type=str, default="N", help="FIGURES/ATTACHMENTS sensitive flag (Y/N)")
    ap.add_argument("--font-path", type=str, default="", help="Optional font path (sets EIA_GEN_FONT_PATH for annotation)")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")
    case_dir = xlsx.parent.resolve()

    fig_id = _as_str(args.fig_id)
    if not fig_id:
        raise SystemExit("--fig-id is required")

    wb = load_workbook(xlsx)
    if "FIGURES" not in wb.sheetnames or "ATTACHMENTS" not in wb.sheetnames:
        raise SystemExit("case.xlsx must include FIGURES and ATTACHMENTS sheets (v2 template).")

    ws_att = wb["ATTACHMENTS"]

    # resolve base image
    if _as_str(args.base_evidence_id):
        base_img = _resolve_base_image_from_attachments(ws_att, case_dir=case_dir, evidence_id=_as_str(args.base_evidence_id))
    else:
        base_img = _resolve_path(case_dir, _as_str(args.base_image))
        if not base_img.exists():
            raise SystemExit(f"base image not found: {base_img}")

    ann_path = _resolve_path(case_dir, _as_str(args.annotations))
    if not ann_path.exists():
        raise SystemExit(f"annotations not found: {ann_path}")

    repo_root = Path(__file__).resolve().parents[1]
    style_path = args.style
    if not style_path.is_absolute():
        style_path = (repo_root / style_path).resolve()
    if not style_path.exists():
        raise SystemExit(f"style yaml not found: {style_path}")

    out_png = _as_str(args.out)
    if out_png:
        out_path = _resolve_path(case_dir, out_png)
    else:
        out_path = (case_dir / "attachments" / "derived" / "figures" / "annotated" / f"{fig_id}.png").resolve()
    out_path.parent.mkdir(parents=True, exist_ok=True)

    # render annotated image
    _run_annotate(
        image_path=base_img,
        annotations_path=ann_path,
        style_path=style_path,
        out_png=out_path,
        font_path=_as_str(args.font_path) or None,
    )

    out_sha256 = _sha256_file(out_path)
    base_sha256 = _sha256_file(base_img)
    ann_sha256 = _sha256_file(ann_path)

    # stopgap: register into ATTACHMENTS (for traceability/UX)
    if args.register_output_attachment:
        out_eid = _as_str(args.output_evidence_id) or f"DER-ANNO-{fig_id}"
        out_row = _find_row(ws_att, key_col="evidence_id", key_value=out_eid)
        if out_row is None:
            out_row = ws_att.max_row + 1

        _write_cell(ws_att, row=out_row, col_name="evidence_id", value=out_eid)
        _write_cell(ws_att, row=out_row, col_name="evidence_type", value="파생이미지")
        _write_cell(ws_att, row=out_row, col_name="title", value=_as_str(args.caption) or _as_str(args.title) or fig_id)
        _write_cell(ws_att, row=out_row, col_name="file_path", value=_relpath_or_str(out_path, case_dir))
        _write_cell(ws_att, row=out_row, col_name="related_fig_id", value=fig_id)
        _write_cell(ws_att, row=out_row, col_name="used_in", value="FIGURE_ANNOTATE")
        _write_cell(ws_att, row=out_row, col_name="data_origin", value="DERIVED")
        _write_cell(ws_att, row=out_row, col_name="src_id", value=_as_str(args.src_id))
        _write_cell(ws_att, row=out_row, col_name="sensitive", value=_as_str(args.sensitive) or "N")
        _write_cell(
            ws_att,
            row=out_row,
            col_name="note",
            value=(
                "DERIVED: ANNOTATED_IMAGE "
                f"base_sha256={base_sha256} ann_sha256={ann_sha256} out_sha256={out_sha256} "
                f"base={_relpath_or_str(base_img, case_dir)} ann={_relpath_or_str(ann_path, case_dir)} "
                f"style={_relpath_or_str(style_path, repo_root)}"
            ),
        )

    # optional: update FIGURES row
    if args.update_xlsx or args.create_figure_row:
        ws_fig = wb["FIGURES"]
        fig_row = _find_row(ws_fig, key_col="fig_id", key_value=fig_id)
        if fig_row is None:
            if not args.create_figure_row:
                raise SystemExit(f"FIGURES row not found for fig_id={fig_id} (use --create-figure-row).")
            fig_row = _ensure_figures_row(
                wb,
                fig_id=fig_id,
                doc_scope=_as_str(args.doc_scope) or "BOTH",
                figure_type=_as_str(args.figure_type) or "AERIAL_PHOTO",
                title=_as_str(args.title),
                caption=_as_str(args.caption) or "주석 합성 이미지",
                source_origin=_as_str(args.source_origin) or "REFERENCE",
                width_mm=int(args.width_mm) if args.width_mm else 180,
                src_id=_as_str(args.src_id) or "S-CLIENT-001",
                sensitive=_as_str(args.sensitive) or "N",
            )

        _write_cell(ws_fig, row=fig_row, col_name="file_path", value=_relpath_or_str(out_path, case_dir))
        _write_cell(ws_fig, row=fig_row, col_name="source_origin", value=_as_str(args.source_origin) or "REFERENCE")
        # gen_method is informational; materialize may still run depending on core.
        _write_cell(ws_fig, row=fig_row, col_name="gen_method", value="ANNOTATED_IMAGE")
        if _as_str(args.caption):
            _write_cell(ws_fig, row=fig_row, col_name="caption", value=_as_str(args.caption))
        if _as_str(args.title):
            _write_cell(ws_fig, row=fig_row, col_name="title", value=_as_str(args.title))

    wb.save(xlsx)

    print(f"OK: wrote {out_path}")
    print(f"- fig_id={fig_id}")
    print(f"- out_png_sha256={out_sha256}")
    if args.update_xlsx:
        print("- updated FIGURES.file_path")
    if args.register_output_attachment:
        print("- upserted ATTACHMENTS row")


if __name__ == "__main__":
    main()

