#!/usr/bin/env python3
from __future__ import annotations

import argparse
import hashlib
import os
import shutil
import subprocess
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _is_blank_row(values: tuple[Any, ...]) -> bool:
    return not any(v is not None and _as_str(v) for v in values)


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def _sheet_rows(ws) -> list[dict[str, Any]]:
    headers = [_as_str(c.value) for c in ws[1]]
    out: list[dict[str, Any]] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or _is_blank_row(row):
            continue
        d: dict[str, Any] = {}
        for i, h in enumerate(headers):
            if not h:
                continue
            if i < len(row):
                d[h] = row[i]
        out.append(d)
    return out


def _find_row(ws, *, key_col: str, key_value: str) -> int | None:
    hm = _header_map(ws)
    if key_col not in hm:
        return None
    col = hm[key_col]
    for r in range(2, ws.max_row + 1):
        v = _as_str(ws.cell(row=r, column=col).value)
        if not v:
            continue
        if v == key_value:
            return r
    return None


def _write_cell(ws, *, row: int, col_name: str, value: Any) -> None:
    hm = _header_map(ws)
    if col_name not in hm:
        raise SystemExit(f"Missing column '{col_name}' in sheet '{ws.title}'")
    ws.cell(row=row, column=hm[col_name]).value = value


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _sha256_file(path: Path) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _copy_into_inbox(*, src_dir: Path, inbox_dir: Path) -> list[Path]:
    """
    Copy images from src_dir into inbox_dir (non-destructive).
    Returns paths of copied files in inbox_dir.
    """
    exts = {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}
    copied: list[Path] = []

    for p in sorted(src_dir.glob("*")):
        if not p.is_file():
            continue
        if p.suffix.lower() not in exts:
            continue
        dst = inbox_dir / p.name
        if dst.exists():
            stem = p.stem
            for i in range(1, 1000):
                alt = inbox_dir / f"{stem}__{i}{p.suffix}"
                if not alt.exists():
                    dst = alt
                    break
        shutil.copy2(p, dst)
        copied.append(dst)
    return copied


def _run_ingest_attachments(*, xlsx: Path, src_id: str, data_origin: str) -> list[str]:
    """
    Runs core ingest-attachments so ATTACHMENTS rows get created with ATT-xxxx ids.
    Returns ingested evidence_ids (ATT-xxxx).
    """
    # Use installed console script if available; fallback to module invocation.
    # Prefer venv executable when running inside repo.
    repo_root = Path(__file__).resolve().parents[1]
    exe = repo_root / ".venv" / "bin" / "eia-gen"
    if not exe.exists():
        exe = shutil.which("eia-gen")  # type: ignore[assignment]
    if not exe:
        raise SystemExit("eia-gen executable not found (expected .venv/bin/eia-gen or PATH).")

    cmd = [
        str(exe),
        "ingest-attachments",
        "--xlsx",
        str(xlsx),
        "--src-id",
        src_id,
        "--data-origin",
        data_origin,
        "--copy-only",
    ]
    # NOTE: ingest-attachments prints summary only; we re-read workbook to get ids.
    subprocess.run(cmd, check=True)

    wb = load_workbook(xlsx)
    ws = wb["ATTACHMENTS"]
    rows = _sheet_rows(ws)
    out: list[str] = []
    for r in rows:
        eid = _as_str(r.get("evidence_id"))
        used_in = _as_str(r.get("used_in"))
        if eid.startswith("ATT-") and used_in == "INGEST":
            out.append(eid)
    return out


def _grid_from_count(n: int) -> int:
    if n <= 2:
        return 2
    if n <= 4:
        return 4
    if n <= 6:
        return 6
    return 6


@dataclass(frozen=True)
class AttachmentRef:
    evidence_id: str
    file_path: str
    title: str


def _attachments_for_fig(wb, *, fig_id: str) -> list[AttachmentRef]:
    ws = wb["ATTACHMENTS"]
    rows = _sheet_rows(ws)
    out: list[AttachmentRef] = []
    for r in rows:
        if _as_str(r.get("related_fig_id")) != fig_id:
            continue
        if _as_str(r.get("evidence_type")) != "사진":
            continue
        fp = _as_str(r.get("file_path"))
        if not fp:
            continue
        out.append(
            AttachmentRef(
                evidence_id=_as_str(r.get("evidence_id")),
                file_path=fp,
                title=_as_str(r.get("title")) or Path(fp).name,
            )
        )
    # Stable ordering: evidence_id (ATT-0001..), then file_path
    out.sort(key=lambda x: (x.evidence_id, x.file_path))
    return out


def _ensure_figures_row(wb, *, fig_id: str, doc_scope: str, caption: str) -> int:
    ws = wb["FIGURES"]
    row = _find_row(ws, key_col="fig_id", key_value=fig_id)
    if row is not None:
        return row
    # Append new row at the end (minimal required columns)
    hm = _header_map(ws)
    # Build row values aligned with headers
    values: dict[str, Any] = {}
    values["fig_id"] = fig_id
    values["doc_scope"] = doc_scope
    values["figure_type"] = "PHOTO_SHEET"
    values["title"] = caption or "현장사진대지"
    values["caption"] = caption or "현장사진대지"
    values["source_origin"] = "CLIENT_PROVIDED"
    values["file_path"] = ""
    values["gen_method"] = ""
    values["geom_ref"] = ""
    values["width_mm"] = 180
    values["crop"] = ""
    values["src_id"] = "S-CLIENT-001"
    values["sensitive"] = "N"
    values["insert_anchor"] = ""

    row_values = [values.get(h, "") for h in hm.keys()]
    ws.append(row_values)
    return ws.max_row


def _run_compose_script(
    *,
    images: list[Path],
    captions: list[str],
    out_png: Path,
    grid: int | None,
    title: str,
    font_path: str | None,
    style_path: Path,
) -> None:
    repo_root = Path(__file__).resolve().parents[1]
    script = repo_root / "scripts" / "compose_callout_composite.py"
    if not script.exists():
        raise SystemExit(f"compose_callout_composite.py not found: {script}")

    cmd = [
        sys.executable,
        str(script),
        "--out",
        str(out_png),
        "--style",
        str(style_path),
        "--title",
        title,
    ]
    if grid is not None:
        cmd.extend(["--grid", str(grid)])
    if font_path:
        cmd.extend(["--font-path", font_path])

    # Use absolute paths to avoid ambiguity
    cmd.append("--images")
    cmd.extend([str(p) for p in images])

    # captions are best-effort; avoid breaking delimiter parsing
    caps = [c.replace(";", ",") for c in captions]
    cmd.extend(["--captions", ";".join(caps)])

    env = os.environ.copy()
    subprocess.run(cmd, check=True, cwd=str(repo_root), env=env)


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Build a deterministic CALLOUT_COMPOSITE panel from case.xlsx.\n"
            "Input grouping rule (SSOT): ATTACHMENTS.related_fig_id == FIGURES.fig_id.\n"
            "If no related photos exist, optionally ingest from FIELD_SURVEY_LOG.photo_folder first."
        )
    )
    ap.add_argument("--xlsx", type=Path, required=True, help="case.xlsx path")
    ap.add_argument("--fig-id", type=str, required=True, help="FIGURES.fig_id to build")
    ap.add_argument("--doc-scope", type=str, default="BOTH", help="FIGURES.doc_scope when creating a row")
    ap.add_argument("--caption", type=str, default="현장사진대지", help="Caption/title for the composed panel")
    ap.add_argument("--grid", type=int, default=None, help="Force grid 2/4/6 (auto if omitted)")
    ap.add_argument(
        "--out",
        type=Path,
        default=None,
        help="Output PNG path (default: attachments/derived/figures/callout/{fig_id}.png under case dir)",
    )
    ap.add_argument("--update-xlsx", action="store_true", help="Update FIGURES.file_path to the generated PNG")
    ap.add_argument(
        "--register-output-attachment",
        action="store_true",
        help="Register the composed panel PNG into ATTACHMENTS sheet (evidence_id=DER-{fig_id} by default)",
    )
    ap.add_argument(
        "--output-evidence-id",
        type=str,
        default="",
        help="Override evidence_id for the composed panel ATTACHMENTS row (default: DER-{fig_id})",
    )
    ap.add_argument("--create-figure-row", action="store_true", help="Create FIGURES row if missing")
    ap.add_argument("--ingest-from-photo-folder", action="store_true", help="If no related photos, ingest from FIELD_SURVEY_LOG.photo_folder")
    ap.add_argument("--survey-id", type=str, default="", help="FIELD_SURVEY_LOG.survey_id to select (default: latest with photo_folder)")
    ap.add_argument("--src-id", type=str, default="S-CLIENT-001", help="src_id used when ingesting photos")
    ap.add_argument("--data-origin", type=str, default="FIELD_SURVEY", help="data_origin used when ingesting photos")
    ap.add_argument("--font-path", type=str, default="", help="Optional font path (overrides EIA_GEN_FONT_PATH)")
    ap.add_argument("--style", type=Path, default=Path("config/figure_style.yaml"), help="Style tokens YAML path")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")
    case_dir = xlsx.parent.resolve()

    wb = load_workbook(xlsx)
    if "ATTACHMENTS" not in wb.sheetnames or "FIGURES" not in wb.sheetnames:
        raise SystemExit("case.xlsx must include ATTACHMENTS and FIGURES sheets (v2 template).")

    fig_id = _as_str(args.fig_id)
    if not fig_id:
        raise SystemExit("--fig-id is required")

    # 1) gather related photos
    refs = _attachments_for_fig(wb, fig_id=fig_id)

    # 2) if empty, optionally ingest from FIELD_SURVEY_LOG.photo_folder
    if not refs and args.ingest_from_photo_folder:
        if "FIELD_SURVEY_LOG" not in wb.sheetnames:
            raise SystemExit("FIELD_SURVEY_LOG sheet is missing; cannot ingest from photo_folder.")
        ws_fs = wb["FIELD_SURVEY_LOG"]
        fs_rows = _sheet_rows(ws_fs)
        photo_folder = ""
        if args.survey_id:
            for r in fs_rows:
                if _as_str(r.get("survey_id")) == _as_str(args.survey_id):
                    photo_folder = _as_str(r.get("photo_folder"))
                    break
        else:
            for r in reversed(fs_rows):
                if _as_str(r.get("photo_folder")):
                    photo_folder = _as_str(r.get("photo_folder"))
                    break

        if not photo_folder:
            raise SystemExit("No FIELD_SURVEY_LOG.photo_folder found (set it or pass --survey-id).")

        src_dir = Path(photo_folder).expanduser()
        if not src_dir.is_absolute():
            src_dir = (case_dir / src_dir).resolve()
        if not src_dir.exists():
            raise SystemExit(f"photo_folder not found: {src_dir}")

        inbox_dir = (case_dir / "attachments" / "inbox").resolve()
        normalized_dir = (case_dir / "attachments" / "normalized").resolve()
        _ensure_dir(inbox_dir)
        _ensure_dir(normalized_dir)

        copied = _copy_into_inbox(src_dir=src_dir, inbox_dir=inbox_dir)
        if not copied:
            raise SystemExit(f"No images found in photo_folder: {src_dir}")

        # Close workbook before running CLI that writes xlsx
        wb.close()

        ingested_ids = _run_ingest_attachments(xlsx=xlsx, src_id=_as_str(args.src_id), data_origin=_as_str(args.data_origin))
        if not ingested_ids:
            raise SystemExit("No attachments ingested (inbox was empty or all files skipped).")

        # Re-open and assign related_fig_id
        wb = load_workbook(xlsx)
        ws_att = wb["ATTACHMENTS"]
        for eid in ingested_ids:
            rnum = _find_row(ws_att, key_col="evidence_id", key_value=eid)
            if rnum is None:
                continue
            _write_cell(ws_att, row=rnum, col_name="related_fig_id", value=fig_id)
            _write_cell(ws_att, row=rnum, col_name="used_in", value="FIGURE_CALL_OUT")

        wb.save(xlsx)
        wb.close()

        wb = load_workbook(xlsx)
        refs = _attachments_for_fig(wb, fig_id=fig_id)

    if len(refs) < 2:
        raise SystemExit(
            f"Need at least 2 photos linked to fig_id={fig_id}. "
            "Set ATTACHMENTS.related_fig_id and evidence_type=사진, or use --ingest-from-photo-folder."
        )

    # 3) compute grid and output path
    img_paths: list[Path] = []
    caps: list[str] = []
    for r in refs[:6]:
        p = Path(r.file_path)
        if not p.is_absolute():
            p = (case_dir / p).resolve()
        if not p.exists():
            continue
        img_paths.append(p)
        caps.append(r.title or p.name)

    if not img_paths:
        raise SystemExit("Resolved 0 image paths (check ATTACHMENTS.file_path).")

    grid = int(args.grid) if args.grid is not None else _grid_from_count(len(img_paths))

    out_png = args.out
    if out_png is None:
        out_png = case_dir / "attachments" / "derived" / "figures" / "callout" / f"{fig_id}.png"
    out_png = out_png.expanduser().resolve()
    _ensure_dir(out_png.parent)

    style_path = args.style
    if not style_path.is_absolute():
        style_path = (Path(__file__).resolve().parents[1] / style_path).resolve()

    # 4) run composer
    _run_compose_script(
        images=img_paths,
        captions=caps,
        out_png=out_png,
        grid=grid,
        title=_as_str(args.caption),
        font_path=_as_str(args.font_path),
        style_path=style_path,
    )

    out_sha256 = _sha256_file(out_png)

    # 4.1) optional: register composed output as an ATTACHMENTS row (stopgap for traceability)
    if args.register_output_attachment:
        ws_att = wb["ATTACHMENTS"]
        out_eid = _as_str(args.output_evidence_id) or f"DER-{fig_id}"
        out_row = _find_row(ws_att, key_col="evidence_id", key_value=out_eid)
        if out_row is None:
            out_row = ws_att.max_row + 1

        rel_out = out_png.relative_to(case_dir).as_posix()
        _write_cell(ws_att, row=out_row, col_name="evidence_id", value=out_eid)
        _write_cell(ws_att, row=out_row, col_name="evidence_type", value="파생이미지")
        _write_cell(ws_att, row=out_row, col_name="title", value=_as_str(args.caption) or fig_id)
        _write_cell(ws_att, row=out_row, col_name="file_path", value=rel_out)
        _write_cell(ws_att, row=out_row, col_name="related_fig_id", value=fig_id)
        _write_cell(ws_att, row=out_row, col_name="used_in", value="FIGURE_CALL_OUT")
        _write_cell(ws_att, row=out_row, col_name="data_origin", value="DERIVED")
        _write_cell(ws_att, row=out_row, col_name="src_id", value=_as_str(args.src_id))
        _write_cell(ws_att, row=out_row, col_name="sensitive", value="N")
        _write_cell(
            ws_att,
            row=out_row,
            col_name="note",
            value=f"DERIVED: CALLOUT_COMPOSITE grid={grid} inputs={len(img_paths)} sha256={out_sha256}",
        )
        wb.save(xlsx)

    # 5) optional: update FIGURES.file_path
    if args.update_xlsx or args.create_figure_row:
        ws_fig = wb["FIGURES"]
        fig_row = _find_row(ws_fig, key_col="fig_id", key_value=fig_id)
        if fig_row is None:
            if not args.create_figure_row:
                raise SystemExit(f"FIGURES row not found for fig_id={fig_id} (use --create-figure-row).")
            fig_row = _ensure_figures_row(wb, fig_id=fig_id, doc_scope=_as_str(args.doc_scope), caption=_as_str(args.caption))
        rel = out_png.relative_to(case_dir).as_posix()
        _write_cell(ws_fig, row=fig_row, col_name="file_path", value=rel)
        if _as_str(ws_fig.cell(row=fig_row, column=_header_map(ws_fig).get("figure_type", 0)).value) != "PHOTO_SHEET":
            _write_cell(ws_fig, row=fig_row, col_name="figure_type", value="PHOTO_SHEET")
        wb.save(xlsx)

    print(f"OK: wrote {out_png}")
    print(f"- fig_id={fig_id} images={len(img_paths)} grid={grid}")
    print(f"- out_png_sha256={out_sha256}")
    if args.update_xlsx:
        print("- updated FIGURES.file_path")


if __name__ == "__main__":
    main()
