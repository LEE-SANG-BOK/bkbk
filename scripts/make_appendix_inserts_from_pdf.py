#!/usr/bin/env python3
from __future__ import annotations

import argparse
import re
import shutil
import subprocess
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook


_RANGE_RE = re.compile(r"^\s*(\d+)\s*-\s*(\d+)\s*$")


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i + 1 for i, h in enumerate(headers) if h}


def _require_cols(ws, cols: list[str]) -> dict[str, int]:
    hm = _header_map(ws)
    missing = [c for c in cols if c not in hm]
    if missing:
        raise SystemExit(f"Missing columns in sheet '{ws.title}': {missing}")
    return hm


def _parse_pages(spec: str) -> list[int]:
    """
    Accepts: "1,2,5-7" or "1;2;5-7" or "1 2 5-7".
    Returns sorted unique pages.
    """
    tokens: list[str] = []
    for chunk in re.split(r"[,\s;]+", (spec or "").strip()):
        if chunk:
            tokens.append(chunk)
    pages: set[int] = set()
    for t in tokens:
        m = _RANGE_RE.match(t)
        if m:
            a, b = int(m.group(1)), int(m.group(2))
            if a > b:
                a, b = b, a
            for p in range(a, b + 1):
                pages.add(p)
            continue
        try:
            pages.add(int(t))
        except Exception:
            raise SystemExit(f"Invalid page token: {t!r} (expected int or range a-b)")
    out = sorted(p for p in pages if p >= 1)
    if not out:
        raise SystemExit("No valid pages parsed from --pages")
    return out


def _split_captions(raw: str | None) -> list[str]:
    """
    Captions are provided as semicolon-separated string for convenience:
      --captions "cap1;cap2;cap3"
    """
    if not raw:
        return []
    return [c.strip() for c in str(raw).split(";") if c.strip()]


def _resolve_pdf_for_render(*, pdf_path: Path, case_dir: Path) -> tuple[Path, str]:
    """
    Returns (resolved_pdf_path, file_path_value_to_store_in_xlsx).
    Prefer storing a case-relative path when possible.
    """
    resolved = pdf_path.expanduser().resolve()
    try:
        rel = resolved.relative_to(case_dir.resolve())
        stored = str(rel).replace("\\", "/")
        return resolved, stored
    except Exception:
        return resolved, str(resolved)


def _require_cmd(cmd: str) -> str:
    p = shutil.which(cmd)
    if not p:
        raise SystemExit(f"Missing command in PATH: {cmd}")
    return p


def _render_pdf_page_png(*, pdf: Path, page_1based: int, dpi: int, out_png: Path) -> None:
    pdftoppm = _require_cmd("pdftoppm")
    out_png.parent.mkdir(parents=True, exist_ok=True)
    prefix = out_png.with_suffix("")  # pdftoppm writes <prefix>.png
    cmd = [
        pdftoppm,
        "-f",
        str(int(page_1based)),
        "-l",
        str(int(page_1based)),
        "-r",
        str(int(dpi)),
        "-png",
        "-singlefile",
        str(pdf),
        str(prefix),
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    if not out_png.exists():
        raise RuntimeError(f"pdftoppm did not produce expected PNG: {out_png}")


def _max_ins_number(existing: Iterable[str]) -> int:
    best = 0
    for s in existing:
        m = re.search(r"(\d+)", s)
        if not m:
            continue
        best = max(best, int(m.group(1)))
    return best


@dataclass(frozen=True)
class UpsertResult:
    action: str  # insert | update | noop
    row: int
    ins_id: str
    order: int
    page: int
    caption: str


def _upsert_appendix_rows(
    *,
    xlsx: Path,
    file_path_value: str,
    pages: list[int],
    captions: list[str],
    width_mm: float,
    dpi: int,
    crop: str,
    src_id: str,
    note: str,
    order_start: int | None,
    order_step: int,
    dry_run: bool,
) -> tuple[list[UpsertResult], dict[str, Any]]:
    wb = load_workbook(xlsx)
    if "APPENDIX_INSERTS" not in wb.sheetnames:
        raise SystemExit(f"Missing sheet APPENDIX_INSERTS in: {xlsx}")
    ws = wb["APPENDIX_INSERTS"]

    hm = _require_cols(
        ws,
        ["ins_id", "order", "file_path", "pdf_page", "caption", "width_mm", "dpi", "crop", "src_id", "note"],
    )

    # Build index: (file_path, pdf_page) -> row
    index: dict[tuple[str, int], int] = {}
    existing_ins: list[str] = []
    max_order = 0
    for r in range(2, ws.max_row + 1):
        fp = _as_str(ws.cell(row=r, column=hm["file_path"]).value)
        pg = ws.cell(row=r, column=hm["pdf_page"]).value
        try:
            pg_i = int(pg) if pg is not None else 0
        except Exception:
            pg_i = 0
        if fp and pg_i >= 1:
            index[(fp, pg_i)] = r
        existing_ins.append(_as_str(ws.cell(row=r, column=hm["ins_id"]).value))
        try:
            max_order = max(max_order, int(ws.cell(row=r, column=hm["order"]).value or 0))
        except Exception:
            pass

    next_ins = _max_ins_number(existing_ins) + 1
    cur_order = int(order_start) if order_start is not None else (max_order + int(order_step))

    upserts: list[UpsertResult] = []
    for i, page in enumerate(pages):
        cap = captions[i] if i < len(captions) else ""
        if not cap:
            cap = f"추가 삽입(부록): {Path(file_path_value).name} p{page}"

        key = (file_path_value, int(page))
        target_row = index.get(key)
        if target_row is None:
            ins_id = f"INS-{next_ins:04d}"
            next_ins += 1
            target_row = ws.max_row + 1
            action = "insert"
            index[key] = target_row
        else:
            ins_id = _as_str(ws.cell(row=target_row, column=hm["ins_id"]).value) or f"INS-{next_ins:04d}"
            action = "update"

        new_values = {
            "ins_id": ins_id,
            "order": cur_order,
            "file_path": file_path_value,
            "pdf_page": int(page),
            "caption": cap,
            "width_mm": float(width_mm),
            "dpi": int(dpi),
            "crop": crop,
            "src_id": src_id,
            "note": note,
        }

        # Determine noop (best-effort)
        if action == "update":
            same = True
            for k, v in new_values.items():
                cur = ws.cell(row=target_row, column=hm[k]).value
                if _as_str(cur) != _as_str(v):
                    same = False
                    break
            if same:
                action = "noop"

        upserts.append(
            UpsertResult(action=action, row=target_row, ins_id=ins_id, order=cur_order, page=int(page), caption=cap)
        )

        if not dry_run and action != "noop":
            if action == "insert":
                ws.append([None] * len(hm))  # allocate row
                # ws.append increases max_row by 1; ensure we write to the last row
                target_row = ws.max_row
                upserts[-1] = UpsertResult(
                    action=action,
                    row=target_row,
                    ins_id=ins_id,
                    order=cur_order,
                    page=int(page),
                    caption=cap,
                )
            for k, v in new_values.items():
                ws.cell(row=target_row, column=hm[k]).value = v

        cur_order += int(order_step)

    meta = {
        "max_order_before": max_order,
        "next_ins_start": _max_ins_number(existing_ins) + 1,
        "file_path": file_path_value,
        "pages": pages,
        "dry_run": dry_run,
    }

    if dry_run:
        wb.close()
        return upserts, meta

    wb.save(xlsx)
    wb.close()
    return upserts, meta


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Populate case.xlsx:APPENDIX_INSERTS from a PDF with page list.\n"
            "Optionally renders low-DPI previews for human verification.\n"
            "This is a local convenience wrapper; docx insertion is handled by core renderer."
        )
    )
    ap.add_argument("--xlsx", type=Path, required=True, help="Target case.xlsx path")
    ap.add_argument("--pdf", type=Path, required=True, help="PDF path (prefer inside case attachments/normalized)")
    ap.add_argument("--pages", required=True, help="Pages to insert (e.g., '1,2,5-7')")
    ap.add_argument(
        "--captions",
        default="",
        help="Optional semicolon-separated captions aligned to pages (e.g., 'cap1;cap2;cap3')",
    )
    ap.add_argument("--width-mm", type=float, default=170.0, help="Insert width (mm)")
    ap.add_argument("--dpi", type=int, default=200, help="Rasterize DPI for insert (core materialize)")
    ap.add_argument("--crop", default="AUTO", help="Crop mode (e.g., AUTO)")
    ap.add_argument("--src-id", default="S-TBD", help="SRC id for citation/traceability")
    ap.add_argument("--note", default="", help="Note to store in APPENDIX_INSERTS.note")
    ap.add_argument("--order-start", type=int, default=None, help="Starting order (default: max+step)")
    ap.add_argument("--order-step", type=int, default=10, help="Order increment per page")
    ap.add_argument("--dry-run", action="store_true", help="Do not write xlsx; just print planned changes")
    ap.add_argument("--render-previews", action="store_true", help="Render preview PNGs for selected pages")
    ap.add_argument("--preview-dpi", type=int, default=140, help="Preview DPI (lower is faster)")
    ap.add_argument(
        "--preview-dir",
        type=Path,
        default=None,
        help="Preview output dir (default: <case>/attachments/derived/previews/appendix/<pdf_stem>/)",
    )
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"xlsx not found: {xlsx}")
    case_dir = xlsx.parent

    pages = _parse_pages(args.pages)
    captions = _split_captions(args.captions)

    pdf_resolved, file_path_value = _resolve_pdf_for_render(pdf_path=args.pdf, case_dir=case_dir)
    if not pdf_resolved.exists():
        raise SystemExit(f"pdf not found: {pdf_resolved}")

    upserts, meta = _upsert_appendix_rows(
        xlsx=xlsx,
        file_path_value=file_path_value,
        pages=pages,
        captions=captions,
        width_mm=float(args.width_mm),
        dpi=int(args.dpi),
        crop=str(args.crop),
        src_id=str(args.src_id),
        note=str(args.note),
        order_start=args.order_start,
        order_step=int(args.order_step),
        dry_run=bool(args.dry_run),
    )

    print(f"XLSX: {xlsx}")
    print(f"PDF : {pdf_resolved}")
    print(f"file_path(stored): {file_path_value}")
    for u in upserts:
        print(f"- {u.action:6s} row={u.row:4d} ins_id={u.ins_id} order={u.order} page={u.page} :: {u.caption}")

    if args.dry_run:
        print("DRY-RUN: no changes written.")
        return

    if not args.render_previews:
        return

    preview_dir = args.preview_dir
    if preview_dir is None:
        preview_dir = case_dir / "attachments" / "derived" / "previews" / "appendix" / Path(file_path_value).stem
    preview_dir = preview_dir.resolve()

    for u in upserts:
        # stable file name
        name = f"{u.ins_id}__p{u.page:04d}.png"
        out_png = preview_dir / name
        _render_pdf_page_png(pdf=pdf_resolved, page_1based=u.page, dpi=int(args.preview_dpi), out_png=out_png)

    print(f"PREVIEWS: {preview_dir}")


if __name__ == "__main__":
    main()
