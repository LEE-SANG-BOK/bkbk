#!/usr/bin/env python3
from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable, Optional


try:
    import openpyxl
    from PIL import Image, ImageEnhance, ImageOps
except Exception as e:  # pragma: no cover
    raise SystemExit(
        "Missing python deps. Install project deps first (see eia-gen/pyproject.toml).\n"
        f"Import error: {e}"
    )


_WS_RE = re.compile(r"\s+")
_KEEP_RE = re.compile(r"[^0-9A-Za-z가-힣]+")


@dataclass(frozen=True)
class Match:
    page: int  # 1-based
    score: int
    text: str
    compact: str
    preview_full_png: str
    preview_top_png: str


def _compact(text: str) -> str:
    s = _WS_RE.sub(" ", (text or "")).strip()
    s = _KEEP_RE.sub("", s).strip().lower()
    return s


def _normalize_keywords(keywords: Iterable[str]) -> list[str]:
    out: list[str] = []
    for kw in keywords:
        kw = str(kw).strip()
        if not kw:
            continue
        out.append(_KEEP_RE.sub("", _WS_RE.sub(" ", kw)).strip().lower())
    return [x for x in out if x]


def _require_cmd(cmd: str) -> str:
    p = shutil.which(cmd)
    if not p:
        raise SystemExit(f"Missing command in PATH: {cmd}")
    return p


def _pdf_page_count(pdf: Path) -> Optional[int]:
    pdfinfo = shutil.which("pdfinfo")
    if not pdfinfo:
        return None
    try:
        out = subprocess.check_output([pdfinfo, str(pdf)], text=True, stderr=subprocess.STDOUT)
    except Exception:
        return None
    for line in out.splitlines():
        if line.lower().startswith("pages:"):
            try:
                return int(line.split(":", 1)[1].strip())
            except Exception:
                return None
    return None


def _render_pdf_page_png(pdf: Path, page_1based: int, *, dpi: int, out_png: Path) -> None:
    pdftoppm = _require_cmd("pdftoppm")
    out_png.parent.mkdir(parents=True, exist_ok=True)
    # Use -singlefile so output path is deterministic: <prefix>.png
    prefix = out_png.with_suffix("")  # without .png
    cmd = [
        pdftoppm,
        "-f",
        str(int(page_1based)),
        "-l",
        str(int(page_1based)),
        "-r",
        str(int(dpi)),
        "-png",
        "-singlefile",
        str(pdf),
        str(prefix),
    ]
    subprocess.run(cmd, check=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    # pdftoppm writes prefix.png
    if not out_png.exists():
        raise RuntimeError(f"pdftoppm did not produce expected PNG: {out_png}")


def _preprocess_for_ocr(img: Image.Image, *, crop_top_ratio: float, contrast: float, threshold: int) -> Image.Image:
    # crop top (title area)
    w, h = img.size
    r = max(0.0, min(float(crop_top_ratio), 1.0))
    y1 = int(round(h * r)) if r > 0 else h
    y1 = max(1, min(y1, h))
    out = img.crop((0, 0, w, y1))

    # grayscale + contrast
    out = ImageOps.grayscale(out)
    if contrast and abs(float(contrast) - 1.0) > 1e-6:
        out = ImageEnhance.Contrast(out).enhance(float(contrast))

    # binarize
    t = int(max(0, min(int(threshold), 255)))
    out = out.point(lambda x: 0 if x < t else 255, mode="L")
    return out


def _tesseract_ocr(img: Image.Image, *, lang: str, psm: int) -> str:
    tesseract = _require_cmd("tesseract")
    with tempfile.TemporaryDirectory() as td:
        td_path = Path(td)
        img_path = td_path / "ocr.png"
        img.save(img_path, format="PNG", optimize=True)
        cmd = [
            tesseract,
            str(img_path),
            "stdout",
            "-l",
            str(lang),
            "--psm",
            str(int(psm)),
        ]
        # NOTE: stderr can be noisy; suppress unless failure.
        proc = subprocess.run(cmd, text=True, stdout=subprocess.PIPE, stderr=subprocess.PIPE)
        if proc.returncode != 0:
            raise RuntimeError(f"tesseract failed: {proc.stderr.strip()[:200]}")
        return proc.stdout or ""


def _xlsx_upsert_ssot_override(
    *,
    case_xlsx: Path,
    sample_page: int,
    override_file_path: str,
    override_page: int,
    width_mm: float,
    dpi: int,
    crop: str | None,
    src_id: str,
    note: str,
    dry_run: bool,
) -> dict[str, Any]:
    wb = openpyxl.load_workbook(case_xlsx)
    if "SSOT_PAGE_OVERRIDES" not in wb.sheetnames:
        raise SystemExit(f"Missing sheet SSOT_PAGE_OVERRIDES in: {case_xlsx}")
    ws = wb["SSOT_PAGE_OVERRIDES"]

    headers = [c.value for c in ws[1]]
    header_map = {str(h).strip(): idx for idx, h in enumerate(headers, start=1) if h}
    required = ["sample_page", "override_file_path", "override_page", "width_mm", "dpi", "crop", "src_id", "note"]
    missing = [c for c in required if c not in header_map]
    if missing:
        raise SystemExit(f"SSOT_PAGE_OVERRIDES headers missing columns: {missing}")

    # find existing row by sample_page
    target_row = None
    for r in range(2, ws.max_row + 1):
        v = ws.cell(row=r, column=header_map["sample_page"]).value
        try:
            v_int = int(v) if v is not None else None
        except Exception:
            v_int = None
        if v_int == int(sample_page):
            target_row = r
            break

    if target_row is None:
        target_row = ws.max_row + 1

    new_row = {
        "sample_page": int(sample_page),
        "override_file_path": str(override_file_path),
        "override_page": int(override_page),
        "width_mm": float(width_mm),
        "dpi": int(dpi),
        "crop": str(crop or ""),
        "src_id": str(src_id),
        "note": str(note),
    }

    if dry_run:
        return {"action": "dry_run", "row": target_row, "values": new_row}

    for k, v in new_row.items():
        ws.cell(row=target_row, column=header_map[k]).value = v

    wb.save(case_xlsx)
    return {"action": "upsert", "row": target_row, "values": new_row}


def search_and_preview(
    *,
    pdf: Path,
    keywords: list[str],
    page_start: int,
    page_end: int,
    dpi: int,
    crop_top_ratio: float,
    contrast: float,
    threshold: int,
    lang: str,
    psm: int,
    preview_dir: Path,
    max_matches: int,
) -> list[Match]:
    matches: list[Match] = []

    pdf = pdf.resolve()
    preview_dir.mkdir(parents=True, exist_ok=True)
    pdf_tag = re.sub(r"[^0-9A-Za-z가-힣]+", "_", pdf.stem)[:60] or "pdf"

    for p in range(int(page_start), int(page_end) + 1):
        with tempfile.TemporaryDirectory() as td:
            td_path = Path(td)
            full_png = td_path / f"p{p}.png"
            _render_pdf_page_png(pdf, p, dpi=int(dpi), out_png=full_png)

            img = Image.open(full_png).convert("RGB")
            ocr_img = _preprocess_for_ocr(
                img, crop_top_ratio=float(crop_top_ratio), contrast=float(contrast), threshold=int(threshold)
            )
            txt = _tesseract_ocr(ocr_img, lang=lang, psm=int(psm))
            one = _WS_RE.sub(" ", txt).strip()
            comp = _compact(one)

            score = 1
            if keywords:
                score = sum(comp.count(kw) for kw in keywords)
                if score <= 0:
                    continue

            # Save previews only for matches
            out_full = preview_dir / f"{pdf_tag}__p{p:04d}__full.png"
            out_top = preview_dir / f"{pdf_tag}__p{p:04d}__top.png"
            img.save(out_full, format="PNG", optimize=True)
            ocr_img.save(out_top, format="PNG", optimize=True)

            matches.append(
                Match(
                    page=p,
                    score=int(score),
                    text=one,
                    compact=comp,
                    preview_full_png=str(out_full),
                    preview_top_png=str(out_top),
                )
            )

            if len(matches) > int(max_matches):
                matches.sort(key=lambda m: (-m.score, m.page))
                matches = matches[: int(max_matches)]

    matches.sort(key=lambda m: (-m.score, m.page))
    return matches


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "SSOT_PAGE_OVERRIDES helper: (1) search PDF pages by OCR keywords, "
            "(2) generate preview PNGs, (3) optionally upsert SSOT_PAGE_OVERRIDES row in case.xlsx."
        )
    )
    ap.add_argument("--pdf", required=True, type=Path, help="Override PDF path (e.g., attachments/normalized/ATT-0001__*.pdf)")
    ap.add_argument("--case-xlsx", type=Path, default=None, help="case.xlsx path (required for --apply).")
    ap.add_argument("--sample-page", type=int, default=None, help="Sample page number to override (required for --apply).")
    ap.add_argument("--pick-page", type=int, default=None, help="Override PDF page to apply (1-based). If omitted, only search/preview.")
    ap.add_argument("--apply", action="store_true", help="Write to case.xlsx:SSOT_PAGE_OVERRIDES (UPSERT by sample_page).")
    ap.add_argument("--dry-run", action="store_true", help="Do not write xlsx; just show intended row values.")

    ap.add_argument("--keywords", default="", help="comma/semicolon separated. Example: '위치도;계획평면도;배수계획'")
    ap.add_argument("--page-start", type=int, default=1)
    ap.add_argument("--page-end", type=int, default=None)
    ap.add_argument("--max-pages", type=int, default=30, help="If --page-end omitted, scan at most this many pages.")
    ap.add_argument("--max-matches", type=int, default=30, help="Keep at most this many matches in memory (ranked).")
    ap.add_argument("--top-n", type=int, default=10, help="Show only top N matches (ranked by keyword hit count). Use 0 to show all.")

    ap.add_argument("--dpi", type=int, default=160, help="Rasterize DPI for OCR + preview.")
    ap.add_argument("--crop-top-ratio", type=float, default=0.35)
    ap.add_argument("--contrast", type=float, default=2.0)
    ap.add_argument("--threshold", type=int, default=180)
    ap.add_argument("--lang", default="kor+eng")
    ap.add_argument("--psm", type=int, default=6)

    ap.add_argument("--width-mm", type=float, default=170.0)
    ap.add_argument("--crop", default="", help="Optional crop string to store in SSOT_PAGE_OVERRIDES (e.g., 'auto').")
    ap.add_argument("--src-id", default="S-TBD", help="Source id(s) for this override (semicolon separated).")
    ap.add_argument("--note", default="", help="Optional note for SSOT_PAGE_OVERRIDES.")
    ap.add_argument("--preview-dir", type=Path, default=None, help="Directory to write preview PNGs.")
    ap.add_argument("--out-json", type=Path, default=None, help="Optional JSON output path.")

    args = ap.parse_args()

    pdf = args.pdf.expanduser().resolve()
    if not pdf.exists():
        raise SystemExit(f"PDF not found: {pdf}")

    # Choose preview dir: case_dir/attachments/derived/ssot_previews by default (if case-xlsx provided)
    if args.preview_dir:
        preview_dir = args.preview_dir.expanduser().resolve()
    else:
        if args.case_xlsx:
            preview_dir = args.case_xlsx.expanduser().resolve().parent / "attachments" / "derived" / "ssot_previews"
        else:
            preview_dir = Path("output/ssot_previews").resolve()

    raw_keywords = re.split(r"[;,]\s*", str(args.keywords or ""))
    keywords = _normalize_keywords(raw_keywords)

    page_count = _pdf_page_count(pdf)
    page_start = max(1, int(args.page_start or 1))
    if args.page_end is not None:
        page_end = int(args.page_end)
    else:
        if page_count:
            page_end = min(page_count, page_start + max(0, int(args.max_pages) - 1))
        else:
            page_end = page_start + max(0, int(args.max_pages) - 1)

    if page_count and page_end > page_count:
        page_end = page_count
    if page_end < page_start:
        raise SystemExit(f"Invalid page range: {page_start}..{page_end}")

    # Search + preview
    matches = search_and_preview(
        pdf=pdf,
        keywords=keywords,
        page_start=page_start,
        page_end=page_end,
        dpi=int(args.dpi),
        crop_top_ratio=float(args.crop_top_ratio),
        contrast=float(args.contrast),
        threshold=int(args.threshold),
        lang=str(args.lang),
        psm=int(args.psm),
        preview_dir=preview_dir,
        max_matches=int(args.max_matches),
    )

    print(f"PDF: {pdf}")
    if page_count:
        print(f"Pages: {page_count}")
    print(f"Scanned: {page_start}..{page_end}")
    print(f"Keywords: {keywords}")
    top_n = int(getattr(args, 'top_n', 10) or 0)
    shown = matches if (top_n <= 0) else matches[:top_n]

    print(f"Matches: {len(matches)} (showing {len(shown)})")
    for m in shown:
        snip = (m.text or "")[:120]
        print(f"- p{m.page:>4} score={m.score}: {snip}")
        print(f"  preview_full: {m.preview_full_png}")
        print(f"  preview_top : {m.preview_top_png}")

    out_obj: dict[str, Any] = {
        "pdf_path": str(pdf),
        "page_count": page_count,
        "scanned_pages": {"start": page_start, "end": page_end},
        "settings": {
            "dpi": int(args.dpi),
            "crop_top_ratio": float(args.crop_top_ratio),
            "contrast": float(args.contrast),
            "threshold": int(args.threshold),
            "lang": str(args.lang),
            "psm": int(args.psm),
            "keywords": keywords,
        },
        "preview_dir": str(preview_dir),
        "matches": [m.__dict__ for m in matches],
    }

    # Apply override
    if args.apply or args.dry_run:
        if not args.case_xlsx or not args.sample_page or not args.pick_page:
            raise SystemExit("--apply requires --case-xlsx, --sample-page, --pick-page")

        case_xlsx = args.case_xlsx.expanduser().resolve()
        if not case_xlsx.exists():
            raise SystemExit(f"case.xlsx not found: {case_xlsx}")
        case_dir = case_xlsx.parent.resolve()

        pick_page = int(args.pick_page)
        if page_count and (pick_page < 1 or pick_page > page_count):
            raise SystemExit(f"pick_page out of range: {pick_page} (1..{page_count})")

        # store relative path where possible (portability)
        try:
            rel = pdf.relative_to(case_dir)
            file_path = str(rel).replace("\\", "/")
        except Exception:
            file_path = str(pdf)

        note = str(args.note or "").strip()
        if keywords and not note:
            note = f"keywords={';'.join(keywords)}"

        row_update = _xlsx_upsert_ssot_override(
            case_xlsx=case_xlsx,
            sample_page=int(args.sample_page),
            override_file_path=file_path,
            override_page=pick_page,
            width_mm=float(args.width_mm),
            dpi=int(args.dpi),
            crop=str(args.crop or "").strip() or None,
            src_id=str(args.src_id),
            note=note,
            dry_run=bool(args.dry_run),
        )
        out_obj["applied_override"] = row_update
        print(f"SSOT_PAGE_OVERRIDES: {row_update['action']} row={row_update['row']}")

    if args.out_json:
        out_path = args.out_json.expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(json.dumps(out_obj, ensure_ascii=False, indent=2), encoding="utf-8")
        print(f"OK wrote JSON: {out_path}")


if __name__ == "__main__":
    main()

