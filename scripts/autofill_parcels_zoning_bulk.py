#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
import re
from pathlib import Path

from openpyxl import load_workbook


def _parse_parcel_tokens(parcel_no: str) -> list[str]:
    s = str(parcel_no or "").strip()
    if not s:
        return []
    tail = s.split()[-1]
    if "," not in tail:
        return [tail]

    parts = [p.strip() for p in tail.split(",") if p.strip()]
    if not parts:
        return []

    first = parts[0]
    base = first.split("-", 1)[0].strip()
    tokens: list[str] = []
    for p in parts:
        if "-" in p:
            tokens.append(p)
            continue
        if base:
            tokens.append(f"{base}-{p}")
    return tokens


def _norm_key(v: str) -> str:
    s = str(v or "").strip()
    if not s:
        return ""
    # Keep only the tail token to tolerate "OO리 123-4" formats.
    return s.split()[-1]


def _load_mapping_csv(path: Path) -> dict[str, str]:
    mapping: dict[str, str] = {}
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        r = csv.DictReader(f)
        for row in r:
            if not row:
                continue
            p = _norm_key(str(row.get("parcel_no") or row.get("jibun") or ""))
            z = str(row.get("zoning") or "").strip()
            if not p or not z:
                continue
            mapping[p] = z
    return mapping


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Auto-fill PARCELS.zoning in case.xlsx(v2) by applying a constant value or a parcel_no→zoning CSV mapping.\n"
            "This script does not guess zoning; it only applies user-provided values (to avoid false entries)."
        )
    )
    ap.add_argument("--xlsx", required=True, type=Path, help="Path to case.xlsx(v2)")
    ap.add_argument("--out", type=Path, default=None, help="Write to this path (default: in-place)")
    ap.add_argument("--zoning", type=str, default="", help="Fill all blank PARCELS.zoning with this value")
    ap.add_argument(
        "--mapping-csv",
        type=Path,
        default=None,
        help="CSV with headers: parcel_no,zoning (optional: jibun as alias for parcel_no).",
    )
    ap.add_argument(
        "--append-src-id",
        type=str,
        default="",
        help="When filling a row, append this src_id token to PARCELS.src_id (e.g., S-02 or EVD-LURIS-2026-01-04).",
    )
    ap.add_argument("--dry-run", action="store_true", help="Print planned changes without writing")
    args = ap.parse_args()

    xlsx = args.xlsx.expanduser().resolve()
    out_path = args.out.expanduser().resolve() if args.out else xlsx

    mapping: dict[str, str] = {}
    if args.mapping_csv:
        mapping = _load_mapping_csv(args.mapping_csv.expanduser().resolve())
        if not mapping:
            raise SystemExit(f"Mapping CSV has no usable rows: {args.mapping_csv}")

    fill_value = str(args.zoning or "").strip()
    if not mapping and not fill_value:
        raise SystemExit("Provide either --zoning or --mapping-csv")

    wb = load_workbook(xlsx)
    if "PARCELS" not in wb.sheetnames:
        raise SystemExit("Missing PARCELS sheet in case.xlsx(v2)")
    ws = wb["PARCELS"]

    header = [c.value for c in ws[1]]
    hm: dict[str, int] = {}
    for idx, h in enumerate(header, start=1):
        if not h:
            continue
        hm[str(h).strip()] = idx

    def _col(name: str) -> int:
        c = hm.get(name)
        if not c:
            raise SystemExit(f"PARCELS missing required column: {name}")
        return c

    col_parcel_no = _col("parcel_no")
    col_zoning = _col("zoning")
    col_src = hm.get("src_id") or 0

    append_src = str(args.append_src_id or "").strip()

    changes: list[tuple[int, str, str, str]] = []
    for row_idx in range(2, ws.max_row + 1):
        parcel_no = ws.cell(row=row_idx, column=col_parcel_no).value
        if parcel_no is None or (isinstance(parcel_no, str) and not parcel_no.strip()):
            continue

        cur = ws.cell(row=row_idx, column=col_zoning).value
        if cur is not None and str(cur).strip():
            continue  # don't overwrite

        chosen: str | None = None
        if mapping:
            tokens = _parse_parcel_tokens(str(parcel_no))
            if not tokens:
                continue
            zs = []
            for t in tokens:
                z = mapping.get(_norm_key(t))
                if not z:
                    z = mapping.get(_norm_key(str(parcel_no)))
                if not z:
                    zs = []
                    break
                zs.append(z)
            if zs and len(set(zs)) == 1:
                chosen = zs[0]
        else:
            chosen = fill_value

        if not chosen:
            continue

        src_before = ""
        src_after = ""
        if col_src and append_src:
            src_val = ws.cell(row=row_idx, column=col_src).value
            src_before = "" if src_val is None else str(src_val).strip()
            parts = [s.strip() for s in re.split(r"[;,]", src_before) if s.strip()] if src_before else []
            if append_src not in parts:
                parts.append(append_src)
            src_after = ";".join(parts)

        changes.append((row_idx, str(parcel_no).strip(), chosen, src_after))

        if not args.dry_run:
            ws.cell(row=row_idx, column=col_zoning).value = chosen
            if col_src and append_src and src_after:
                ws.cell(row=row_idx, column=col_src).value = src_after

    print(f"PARCELS rows updated: {len(changes)}")
    for row_idx, parcel_no, zoning, src_after in changes[:40]:
        extra = f", src_id+={append_src}" if append_src and src_after else ""
        print(f"- row {row_idx}: {parcel_no} -> zoning='{zoning}'{extra}")

    if args.dry_run:
        return

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"OK wrote {out_path}")


if __name__ == "__main__":
    main()

