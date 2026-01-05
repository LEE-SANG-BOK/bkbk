#!/usr/bin/env python3
from __future__ import annotations

import argparse
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import openpyxl


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _as_int(v: Any) -> int | None:
    s = _as_str(v)
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _load_sheet(path: Path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [_as_str(c.value) for c in ws[1]]
    header_map = {h: i for i, h in enumerate(headers, start=1) if h}
    return ws, header_map


def _col(header_map: dict[str, int], name: str) -> int:
    if name not in header_map:
        raise SystemExit(f"Missing required column '{name}' in coverage xlsx")
    return header_map[name]


def _collect_suggestions(row: dict[str, Any]) -> list[dict[str, str]]:
    """
    Reads suggest_* columns if present.
    Output items: {id, score, scope, caption}
    """
    out: list[dict[str, str]] = []
    i = 1
    while True:
        sid = _as_str(row.get(f"suggest_{i}_id"))
        if not sid:
            break
        out.append(
            {
                "id": sid,
                "score": _as_str(row.get(f"suggest_{i}_score")),
                "scope": _as_str(row.get(f"suggest_{i}_scope")),
                "caption": _as_str(row.get(f"suggest_{i}_caption")),
            }
        )
        i += 1
    return out


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Generate a markdown backlog of NEED/UNCLASSIFIED items from coverage_matrix_seed(.suggested).xlsx. "
            "This helps turn the coverage matrix into actionable tickets without manual copy-paste."
        )
    )
    ap.add_argument("--in-xlsx", type=Path, required=True, help="Input coverage_matrix_seed(.suggested).xlsx path")
    ap.add_argument(
        "--out-md",
        type=Path,
        default=None,
        help="Output markdown path (default: <in>.need_tickets.md)",
    )
    ap.add_argument(
        "--include-unclassified",
        action="store_true",
        help="Include rows with empty status(EXIST/NEED/IGNORE) as UNCLASSIFIED tickets",
    )
    ap.add_argument(
        "--include-chapters",
        action="store_true",
        help="Include kind=CHAPTER rows (default: skip CHAPTER to keep tickets focused on FIGURE/TABLE)",
    )
    ap.add_argument(
        "--stable",
        action="store_true",
        help=(
            "Make output more reproducible for SSOT storage: omit timestamps and prefer repo-relative paths "
            "(when possible)."
        ),
    )
    args = ap.parse_args()

    in_xlsx = args.in_xlsx.expanduser().resolve()
    if not in_xlsx.exists():
        raise SystemExit(f"Input xlsx not found: {in_xlsx}")

    repo_root = Path(__file__).resolve().parents[1]  # eia-gen/

    def _display_path(p: Path) -> str:
        try:
            return str(p.relative_to(repo_root))
        except Exception:
            return str(p)

    out_md = args.out_md
    if out_md is None:
        out_md = in_xlsx.with_suffix("").with_name(in_xlsx.stem + ".need_tickets.md")
    out_md = out_md.expanduser().resolve()

    ws, header_map = _load_sheet(in_xlsx)

    col_sample = _col(header_map, "sample_page")
    col_kind = _col(header_map, "kind")
    col_label = _col(header_map, "label")
    col_caption = _col(header_map, "caption")
    col_our = _col(header_map, "our_spec_id")
    col_status = _col(header_map, "status(EXIST/NEED/IGNORE)")
    col_note = _col(header_map, "note")

    # optional suggest headers: just read by name later
    # Build row dicts using header names
    rows: list[dict[str, Any]] = []
    for r in range(2, ws.max_row + 1):
        sample_page = _as_int(ws.cell(row=r, column=col_sample).value)
        kind = _as_str(ws.cell(row=r, column=col_kind).value).upper()
        label = _as_str(ws.cell(row=r, column=col_label).value)
        caption = _as_str(ws.cell(row=r, column=col_caption).value)
        our = _as_str(ws.cell(row=r, column=col_our).value)
        status = _as_str(ws.cell(row=r, column=col_status).value).upper()
        note = _as_str(ws.cell(row=r, column=col_note).value)

        if sample_page is None and not any([kind, label, caption, our, status, note]):
            continue
        if kind == "CHAPTER" and not args.include_chapters:
            continue

        row: dict[str, Any] = {
            "row": r,
            "sample_page": sample_page,
            "kind": kind,
            "label": label,
            "caption": caption,
            "our_spec_id": our,
            "status": status,
            "note": note,
        }

        # suggestions if present
        for h in header_map.keys():
            if h.startswith("suggest_"):
                row[h] = ws.cell(row=r, column=header_map[h]).value

        rows.append(row)

    counts = {"EXIST": 0, "NEED": 0, "IGNORE": 0, "UNCLASSIFIED": 0}
    tickets: list[dict[str, Any]] = []
    for row in rows:
        st = row.get("status") or ""
        if st in ("EXIST", "NEED", "IGNORE"):
            counts[st] += 1
        else:
            counts["UNCLASSIFIED"] += 1

        if st == "NEED" or (args.include_unclassified and st not in ("EXIST", "IGNORE")):
            tickets.append(row)

    # Render markdown
    lines: list[str] = []
    lines.append("# Coverage Matrix — NEED 티켓(자동 생성)")
    lines.append("")
    if not args.stable:
        lines.append(f"- generated_at: `{_utc_now_iso()}`")
    lines.append(f"- source_xlsx: `{_display_path(in_xlsx)}`")
    lines.append(f"- total_rows: `{len(rows)}`")
    lines.append(
        f"- counts: `EXIST={counts['EXIST']}, NEED={counts['NEED']}, IGNORE={counts['IGNORE']}, "
        f"UNCLASSIFIED={counts['UNCLASSIFIED']}`"
    )
    lines.append("")
    lines.append("## 읽는 법")
    lines.append("- `NEED`: 샘플에 있으나 우리 스펙/입력/렌더가 부족한 항목(= 작업 티켓).")
    lines.append("- `UNCLASSIFIED`: 아직 분류되지 않은 항목(선택적으로 티켓화).")
    lines.append("")
    lines.append("## 티켓 목록")
    if not tickets:
        lines.append("- (없음)")
    else:
        for t in tickets:
            sp = t.get("sample_page")
            kind = t.get("kind") or ""
            label = t.get("label") or ""
            caption = t.get("caption") or ""
            our = t.get("our_spec_id") or ""
            status = t.get("status") or "UNCLASSIFIED"
            note = t.get("note") or ""
            rowno = t.get("row")

            title = f"p{sp} {kind} {label} — {caption}".strip()
            lines.append(f"### {title}")
            lines.append(f"- status: `{status}`")
            lines.append(f"- our_spec_id: `{our or '(empty)'}`")
            lines.append(f"- source_row: `{in_xlsx.name}:row{rowno}`")
            if note:
                lines.append(f"- note: {note}")

            sugg = _collect_suggestions(t)
            if sugg:
                lines.append("- suggested_spec_ids:")
                for s in sugg:
                    lines.append(
                        f"  - `{s['id']}` (score={s['score']}, scope={s['scope']}) — {s['caption']}"
                    )

            lines.append("- next_actions:")
            if status == "UNCLASSIFIED":
                lines.append("  - status(EXIST/NEED/IGNORE) 먼저 결정")
            if not our:
                lines.append("  - our_spec_id 결정(또는 새 id 초안 작성)")
            lines.append("  - 필요한 입력 시트/컬럼 및 근거(src/evidence) 기록")
            lines.append("  - 완료정의(DoD): 샘플과 ‘보이는 결과물’이 동등한지로 정의")
            lines.append("")

    out_md.parent.mkdir(parents=True, exist_ok=True)
    out_md.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    print(f"OK wrote: {out_md}")


if __name__ == "__main__":
    main()
