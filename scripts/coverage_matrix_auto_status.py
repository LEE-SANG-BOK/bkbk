#!/usr/bin/env python3
from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any

import openpyxl
import yaml


def _as_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _load_spec_ids(repo_root: Path) -> set[str]:
    """
    Collect known spec ids from spec/spec_dia YAMLs.
    """
    ids: set[str] = set()
    spec_files = [
        repo_root / "spec" / "figure_specs.yaml",
        repo_root / "spec" / "table_specs.yaml",
        repo_root / "spec_dia" / "figure_specs.yaml",
        repo_root / "spec_dia" / "table_specs.yaml",
    ]
    for p in spec_files:
        if not p.exists():
            continue
        obj = yaml.safe_load(p.read_text(encoding="utf-8")) or {}
        for k in ("figures", "tables"):
            items = obj.get(k) or []
            for it in items:
                sid = _as_str((it or {}).get("id"))
                if sid:
                    ids.add(sid)
    return ids


def _header_map(ws) -> dict[str, int]:
    headers = [_as_str(c.value) for c in ws[1]]
    return {h: i for i, h in enumerate(headers, start=1) if h}


def _col(hmap: dict[str, int], name: str) -> int:
    if name not in hmap:
        raise SystemExit(f"Missing required column '{name}' in coverage xlsx")
    return hmap[name]


def _append_note(existing: str, msg: str) -> str:
    existing = (existing or "").strip()
    if not existing:
        return msg
    if msg in existing:
        return existing
    return f"{existing} | {msg}"


def main() -> None:
    ap = argparse.ArgumentParser(
        description=(
            "Auto-fill status(EXIST/NEED/IGNORE) for coverage_matrix_seed(.suggested).xlsx.\n"
            "- If our_spec_id exists and matches known spec ids -> status=EXIST\n"
            "- If our_spec_id is set but unknown -> status=NEED + note appended\n"
            "- (optional) If our_spec_id is empty -> status=NEED + note appended\n"
            "This script only updates rows where status is empty by default (idempotent)."
        )
    )
    ap.add_argument("--in-xlsx", type=Path, required=True, help="Input coverage matrix xlsx path")
    ap.add_argument(
        "--out-xlsx",
        type=Path,
        default=None,
        help="Output xlsx path (default: <in>.status.xlsx)",
    )
    ap.add_argument(
        "--in-place",
        action="store_true",
        help="Write back to --in-xlsx (overwrites file).",
    )
    ap.add_argument(
        "--force",
        action="store_true",
        help="Overwrite non-empty status values (default: only fill empty status).",
    )
    ap.add_argument(
        "--auto-ignore-chapters",
        action="store_true",
        help="If kind=CHAPTER and status is empty, set status=IGNORE (default: leave as-is).",
    )
    ap.add_argument(
        "--default-missing-to-need",
        action="store_true",
        help=(
            "If kind is FIGURE/TABLE and our_spec_id is empty (and status is empty), "
            "set status=NEED and append a note. This helps complete 3-way classification "
            "(EXIST/NEED/IGNORE) without leaving many UNCLASSIFIED rows."
        ),
    )
    ap.add_argument(
        "--dry-run",
        action="store_true",
        help="Do not write any file; print summary only.",
    )
    args = ap.parse_args()

    in_xlsx = args.in_xlsx.expanduser().resolve()
    if not in_xlsx.exists():
        raise SystemExit(f"Input xlsx not found: {in_xlsx}")

    repo_root = Path(__file__).resolve().parents[1]  # eia-gen/
    known_ids = _load_spec_ids(repo_root)
    if not known_ids:
        raise SystemExit("No spec ids found. Check spec/spec_dia files.")

    wb = openpyxl.load_workbook(in_xlsx)
    ws = wb.active
    hmap = _header_map(ws)

    col_kind = _col(hmap, "kind")
    col_our = _col(hmap, "our_spec_id")
    col_status = _col(hmap, "status(EXIST/NEED/IGNORE)")
    col_note = _col(hmap, "note")

    updated_exist = 0
    updated_need = 0
    updated_ignore = 0
    skipped = 0
    unchanged = 0

    for r in range(2, ws.max_row + 1):
        kind = _as_str(ws.cell(row=r, column=col_kind).value).upper()
        our = _as_str(ws.cell(row=r, column=col_our).value)
        status = _as_str(ws.cell(row=r, column=col_status).value).upper()
        note = _as_str(ws.cell(row=r, column=col_note).value)

        if not args.force and status:
            skipped += 1
            continue

        # Optional: CHAPTER auto-ignore (only when status empty or force)
        if kind == "CHAPTER" and args.auto_ignore_chapters and (args.force or not status):
            if status != "IGNORE":
                ws.cell(row=r, column=col_status).value = "IGNORE"
                updated_ignore += 1
            else:
                unchanged += 1
            continue

        # Optional: complete classification for missing our_spec_id (FIGURE/TABLE only).
        if args.default_missing_to_need and kind in {"FIGURE", "TABLE"} and not our:
            if status != "NEED":
                ws.cell(row=r, column=col_status).value = "NEED"
                updated_need += 1
            else:
                unchanged += 1
            msg = "missing our_spec_id: map to spec id, or set status=IGNORE if out-of-scope"
            ws.cell(row=r, column=col_note).value = _append_note(note, msg)
            continue

        if not our:
            unchanged += 1
            continue

        if our in known_ids:
            if status != "EXIST":
                ws.cell(row=r, column=col_status).value = "EXIST"
                updated_exist += 1
            else:
                unchanged += 1
            continue

        # our_spec_id set but unknown -> NEED + note
        if status != "NEED":
            ws.cell(row=r, column=col_status).value = "NEED"
            updated_need += 1
        else:
            unchanged += 1

        msg = f"our_spec_id not found in spec/spec_dia: {our}"
        ws.cell(row=r, column=col_note).value = _append_note(note, msg)

    total_updates = updated_exist + updated_need + updated_ignore
    print(
        "SUMMARY:",
        f"updated_exist={updated_exist}",
        f"updated_need={updated_need}",
        f"updated_ignore={updated_ignore}",
        f"unchanged={unchanged}",
        f"skipped={skipped}",
    )

    if args.dry_run:
        print("DRY-RUN: not writing output")
        return

    if args.in_place and args.out_xlsx is not None:
        raise SystemExit("Use either --in-place or --out-xlsx (not both).")

    out_xlsx = args.out_xlsx
    if args.in_place:
        out_xlsx = in_xlsx
    elif out_xlsx is None:
        out_xlsx = in_xlsx.with_suffix("").with_name(in_xlsx.stem + ".status.xlsx")

    out_xlsx = out_xlsx.expanduser().resolve()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_xlsx)
    print(f"WROTE: {out_xlsx}")


if __name__ == "__main__":
    main()
