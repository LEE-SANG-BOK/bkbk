from __future__ import annotations

from typing import Any

from openpyxl.workbook.workbook import Workbook

from eia_gen.models.case import Case
from eia_gen.services.canonicalize import canonicalize_case


def _split_ids(value: Any) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    parts: list[str] = []
    for chunk in s.replace(",", ";").split(";"):
        t = chunk.strip()
        if t:
            parts.append(t)
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def _is_empty_row(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _sheet_header_map(ws) -> dict[str, int]:
    header = [c.value for c in ws[1]]
    mapping: dict[str, int] = {}
    for idx, h in enumerate(header):
        if not h:
            continue
        mapping[str(h).strip()] = idx
    return mapping


def _get(header_map: dict[str, int], row_values: list[Any], key: str) -> Any:
    idx = header_map.get(key)
    if idx is None or idx >= len(row_values):
        return None
    return row_values[idx]


def _tf(value: Any, src_ids: list[str] | None = None) -> dict[str, Any]:
    d: dict[str, Any] = {"t": "" if value is None else str(value).strip()}
    if src_ids:
        d["src"] = src_ids
    return d


def _qf(value: Any, unit: str, src_ids: list[str] | None = None) -> dict[str, Any]:
    v = None
    if value is None or (isinstance(value, str) and not value.strip()):
        v = None
    else:
        try:
            v = float(value)
        except Exception:
            v = None
    d: dict[str, Any] = {"v": v, "u": unit}
    if src_ids:
        d["src"] = src_ids
    return d


def _to_ym(value: Any) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if len(s) >= 7 and s[4] == "-":
        return s[:7]
    return s


def _map_figure_type_to_asset_type(raw: str) -> str:
    v = (raw or "").strip()
    if not v:
        return "other"
    # Already internal?
    if v in {
        "location_map",
        "layout_plan",
        "influence_area_map",
        "landuse_plan",
        "aerial_photo",
        "drainage_map",
        "stormwater_plan_map",
        "dia_target_area_map",
        "photo",
        "photo_sheet",
        "simulation",
    }:
        return v

    up = v.upper()
    mapping = {
        "LOCATION_MAP": "location_map",
        "INFLUENCE_AREA_MAP": "influence_area_map",
        "LANDUSE_PLAN": "landuse_plan",
        "AERIAL_PHOTO": "aerial_photo",
        "LAYOUT_PLAN": "layout_plan",
        "SITE_PLAN": "layout_plan",
        "DRAINAGE_MAP": "drainage_map",
        "STORMWATER_PLAN_MAP": "stormwater_plan_map",
        "CATCHMENT_MAP": "dia_target_area_map",
        "ECO_ROUTE_MAP": "eco_route_map",
        "PHOTO_SHEET": "photo_sheet",
        "PHOTO": "photo",
        "SIMULATION": "simulation",
        "CHART": "simulation",
    }
    return mapping.get(up, "other")


def _flag_to_bool(value: Any) -> bool:
    if value is None:
        return False
    s = str(value).strip().upper()
    return s in {"Y", "YES", "TRUE", "T", "1"}


def _parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _parse_int(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    s = str(value).strip()
    if not s:
        return None
    try:
        return int(float(s))
    except Exception:
        return None


def _parse_epsg(value: Any, default: int = 4326) -> int:
    s = str(value or "").strip().upper()
    if not s:
        return default
    if s.startswith("EPSG:"):
        s = s.split("EPSG:", 1)[1].strip()
    try:
        return int(s)
    except Exception:
        return default


def _map_env_status_to_korean(status: str) -> str:
    v = (status or "").strip().upper()
    if v in {"FOCUS", "중점", "중점평가항목"}:
        return "중점"
    if v in {"BASELINE", "현황", "현황조사항목"}:
        return "현황"
    if v in {"EXCLUDE", "제외", "평가제외항목"}:
        return "제외"
    return status


def _map_include_to_yes_no_unknown(value: Any) -> str:
    v = str(value or "").strip().upper()
    if v in {"Y", "YES", "TRUE", "O", "〇", "●"}:
        return "YES"
    if v in {"N", "NO", "FALSE", "X", "✕", "×"}:
        return "NO"
    if v in {"NA", "UNKNOWN", "UNK", ""}:
        return "UNKNOWN"
    return v


def _map_water_param_to_key(param: str) -> str:
    p = (param or "").strip().upper()
    mapping = {
        "BOD": "bod_mgL",
        "COD": "cod_mgL",
        "SS": "ss_mgL",
        "TN": "tn_mgL",
        "TP": "tp_mgL",
        "PH": "ph",
        "DO": "do_mgL",
        "탁도": "turbidity",
    }
    return mapping.get(p, p.lower() if p else "unknown")


def load_case_from_workbook_v2(wb: Workbook) -> Case:
    """Load v2 case.xlsx (snake_case + LOOKUPS) into existing Case model (best-effort)."""
    data: dict[str, Any] = {}
    utilities_drainage: list[dict[str, Any]] = []
    evidence_src_ids_by_id: dict[str, list[str]] = {}

    def _filter_tbd_src_ids(src_ids: list[str]) -> list[str]:
        return [s for s in src_ids if s and s not in {"S-TBD", "SRC-TBD"}]

    def _src_ids_from_row_or_evidence(hm: dict[str, int], row: list[Any]) -> list[str]:
        src_ids = _filter_tbd_src_ids(_split_ids(_get(hm, row, "src_id")))
        if src_ids:
            return src_ids
        ev_id = str(_get(hm, row, "evidence_id") or "").strip()
        if not ev_id:
            return []
        return list(evidence_src_ids_by_id.get(ev_id, []))

    # META
    case_id = ""
    if "META" in wb.sheetnames:
        ws = wb["META"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        case_id = str(_get(hm, row, "case_id") or "").strip()
        template_version = str(_get(hm, row, "template_version") or "").strip()
        locale = str(_get(hm, row, "locale") or "").strip()
        output_targets = str(_get(hm, row, "output_targets") or "").strip()

        meta: dict[str, Any] = {}
        if case_id:
            meta["case_id"] = case_id
        if template_version:
            meta["template_version"] = template_version
        if locale:
            meta["language"] = locale
        if output_targets:
            meta["report_type"] = output_targets
        if meta:
            data["meta"] = meta

    # LOCATION (used for address/src)
    addr_text = ""
    address_src_ids: list[str] = []
    coord_src_ids: list[str] = []
    admin_sido = ""
    admin_sigungu = ""
    admin_eupmyeon = ""
    center_lat = None
    center_lon = None
    epsg = 4326
    boundary_file = ""
    if "LOCATION" in wb.sheetnames:
        ws = wb["LOCATION"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        address_src_ids = _split_ids(_get(hm, row, "src_id"))
        coord_src_ids = address_src_ids[:]

        addr_jibeon = _get(hm, row, "address_jibeon")
        addr_road = _get(hm, row, "address_road")
        addr_text = str(addr_jibeon or addr_road or "").strip()

        admin_si = str(_get(hm, row, "admin_si") or "").strip()
        admin_eup = str(_get(hm, row, "admin_eupmyeon") or "").strip()
        # v2 splits differently; store best-effort into sido/sigungu/eupmyeon
        admin_sido = admin_si
        admin_sigungu = admin_si
        admin_eupmyeon = admin_eup

        center_lat = _get(hm, row, "center_lat")
        center_lon = _get(hm, row, "center_lon")
        epsg = _parse_epsg(_get(hm, row, "crs"), default=4326)
        boundary_file = str(_get(hm, row, "boundary_file") or "").strip()

    # PARTIES (optional; used to populate cover fields when PROJECT columns are missing)
    party_by_role: dict[str, tuple[str, list[str]]] = {}
    if "PARTIES" in wb.sheetnames:
        ws = wb["PARTIES"]
        hm = _sheet_header_map(ws)
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            role = str(_get(hm, row, "role") or "").strip()
            name = str(_get(hm, row, "name") or "").strip()
            if not role or not name:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            party_by_role.setdefault(role, (name, src_ids))

    # PROJECT
    if "PROJECT" in wb.sheetnames:
        ws = wb["PROJECT"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(hm, row, "src_id")) or address_src_ids or ["S-TBD"]

        project_name = _get(hm, row, "project_name")
        project_type = _get(hm, row, "project_type")
        doc_env_required = str(_get(hm, row, "doc_env_required") or "").strip()
        doc_drr_required = str(_get(hm, row, "doc_drr_required") or "").strip()
        total_area_m2 = _get(hm, row, "total_area_m2")
        purpose_hint = _get(hm, row, "main_facilities_summary") or _get(hm, row, "stormwater_plan")

        client_name = _get(hm, row, "client_name")
        proponent_name = _get(hm, row, "proponent_name")
        author_org = _get(hm, row, "author_org")
        submit_date = _get(hm, row, "submit_date")
        approving_authority = _get(hm, row, "approving_authority")
        consultation_agency = _get(hm, row, "consultation_agency")

        def _cover_text(
            *,
            explicit_value: Any,
            fallback_roles: list[str],
            default_text: str = "",
        ) -> dict[str, Any]:
            text = str(explicit_value or "").strip()
            if text:
                return _tf(text, src_ids)
            for role in fallback_roles:
                p = party_by_role.get(role)
                if p:
                    name, p_src_ids = p
                    return _tf(name, p_src_ids)
            return _tf(default_text, src_ids if default_text else [])

        # cover + project_overview core
        data["cover"] = {
            "project_name": _tf(project_name, src_ids),
            "client_name": _cover_text(explicit_value=client_name, fallback_roles=["사업자"]),
            "proponent_name": _cover_text(explicit_value=proponent_name, fallback_roles=["사업자"]),
            "author_org": _cover_text(explicit_value=author_org, fallback_roles=["대행자", "작성자"]),
            "submit_date": _cover_text(explicit_value=submit_date, fallback_roles=[]),
            "approving_authority": _cover_text(explicit_value=approving_authority, fallback_roles=[]),
            "consultation_agency": _cover_text(explicit_value=consultation_agency, fallback_roles=[]),
        }
        data["project_overview"] = {
            "purpose_need": _tf(purpose_hint, src_ids),
            "location": {
                "address": _tf(addr_text, address_src_ids or src_ids),
                "admin": {
                    "sido": _tf(admin_sido, address_src_ids or src_ids),
                    "sigungu": _tf(admin_sigungu, address_src_ids or src_ids),
                    "eupmyeon": _tf(admin_eupmyeon, address_src_ids or src_ids),
                },
                "center_coord": {
                    "epsg": epsg,
                    "lat": _qf(center_lat, "deg", coord_src_ids or src_ids),
                    "lon": _qf(center_lon, "deg", coord_src_ids or src_ids),
                },
                # v2 optional: allow figure generator to locate a boundary file explicitly
                "boundary_file": boundary_file,
            },
            "area": {"total_area_m2": _qf(total_area_m2, "m2", src_ids)},
        }

        # meta.project_type best-effort
        if project_type is not None and str(project_type).strip():
            data.setdefault("meta", {})
            data["meta"]["project_type"] = str(project_type).strip()
        if doc_env_required:
            data.setdefault("meta", {})
            data["meta"]["doc_env_required"] = doc_env_required
        if doc_drr_required:
            data.setdefault("meta", {})
            data["meta"]["doc_drr_required"] = doc_drr_required

    # PARCELS
    if "PARCELS" in wb.sheetnames:
        ws = wb["PARCELS"]
        hm = _sheet_header_map(ws)
        parcels: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            parcels.append(
                {
                    "jibun": _tf(_get(hm, row, "parcel_no"), src_ids),
                    "pnu": _tf("", src_ids),
                    "land_category": _tf(_get(hm, row, "jimok"), src_ids),
                    "zoning": _tf(_get(hm, row, "zoning"), src_ids),
                    "area_m2": _qf(_get(hm, row, "area_m2"), "m2", src_ids),
                    "note": _tf(_get(hm, row, "note"), src_ids),
                }
            )
        if parcels:
            data.setdefault("project_overview", {}).setdefault("area", {})
            data["project_overview"]["area"]["parcels"] = parcels

    # ZONING_BREAKDOWN (optional; explicit land-use zoning area breakdown)
    if "ZONING_BREAKDOWN" in wb.sheetnames:
        ws = wb["ZONING_BREAKDOWN"]
        hm = _sheet_header_map(ws)
        zoning_area: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            zoning = str(_get(hm, row, "zoning") or "").strip()
            if not zoning:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            zoning_area[zoning] = _qf(_get(hm, row, "area_m2"), "m2", src_ids)
        if zoning_area:
            data.setdefault("project_overview", {}).setdefault("area", {})
            data["project_overview"]["area"]["zoning_area_m2"] = zoning_area

    # FACILITIES (map to project_overview.contents_scale.facilities)
    if "FACILITIES" in wb.sheetnames:
        ws = wb["FACILITIES"]
        hm = _sheet_header_map(ws)
        facilities: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            qty_unit = _get(hm, row, "qty_unit")
            facilities.append(
                {
                    "category": _tf(_get(hm, row, "type"), src_ids),
                    "name": _tf(_get(hm, row, "name"), src_ids),
                    "qty": _qf(_get(hm, row, "qty"), "" if qty_unit is None else str(qty_unit).strip(), src_ids),
                    "area_m2": _qf(_get(hm, row, "area_m2"), "m2", src_ids),
                    "capacity_person": _qf(_get(hm, row, "capacity_person"), "명", src_ids),
                    "note": _tf(_get(hm, row, "note"), src_ids),
                }
            )
        if facilities:
            data.setdefault("project_overview", {}).setdefault("contents_scale", {})
            data["project_overview"]["contents_scale"]["facilities"] = facilities

    # PLAN_LANDUSE (after-stage summary)
    if "PLAN_LANDUSE" in wb.sheetnames:
        ws = wb["PLAN_LANDUSE"]
        hm = _sheet_header_map(ws)
        summary: dict[str, dict[str, Any]] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            stage = str(_get(hm, row, "stage") or "").strip().upper()
            # Prefer "AFTER"(계획 반영)만 요약
            if stage and stage not in {"AFTER", "후"}:
                continue
            category = str(_get(hm, row, "category") or "").strip()
            if not category:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            area = _get(hm, row, "area_m2")

            if category not in summary:
                summary[category] = _qf(area, "m2", src_ids)
                continue

            # accumulate numeric
            try:
                prev = summary[category].get("v")
                summary[category]["v"] = (float(prev) if prev is not None else 0.0) + float(area or 0.0)
            except Exception:
                pass
            # merge src ids
            merged = _split_ids(";".join([*(summary[category].get("src") or []), *src_ids]))
            if merged:
                summary[category]["src"] = merged

        if summary:
            data.setdefault("project_overview", {}).setdefault("contents_scale", {})
            data["project_overview"]["contents_scale"]["land_use_plan_summary"] = summary

    # UTILITIES (우수/배수 중심 → DIA drainage facilities)
    if "UTILITIES" in wb.sheetnames:
        ws = wb["UTILITIES"]
        hm = _sheet_header_map(ws)
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            util_type = str(_get(hm, row, "util_type") or "").strip()
            # 우수/배수만 DIA 배수시설 표로 연결(상수/오수는 제외)
            if util_type and ("우수" not in util_type and "배수" not in util_type):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            utilities_drainage.append(
                {
                    "facility_id": _tf(_get(hm, row, "util_id"), src_ids),
                    "type": _tf(util_type or "우수", src_ids),
                    "size_desc": _tf(_get(hm, row, "drawing_ref"), src_ids),
                    "capacity": _tf(_get(hm, row, "capacity"), src_ids),
                    "discharge_to": _tf(_get(hm, row, "discharge_point"), src_ids),
                    "maintenance_class": _tf("", src_ids),
                }
            )

    # SCHEDULE (map to YYYY-MM milestones)
    if "SCHEDULE" in wb.sheetnames:
        ws = wb["SCHEDULE"]
        hm = _sheet_header_map(ws)
        milestones: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            milestones.append(
                {
                    "phase": _tf(_get(hm, row, "phase"), src_ids),
                    "start": _tf(_to_ym(_get(hm, row, "start_date")), src_ids),
                    "end": _tf(_to_ym(_get(hm, row, "end_date")), src_ids),
                }
            )
        if milestones:
            data.setdefault("project_overview", {}).setdefault("schedule", {})
            data["project_overview"]["schedule"]["milestones"] = milestones

    # FIGURES (map to Case.assets)
    if "FIGURES" in wb.sheetnames:
        ws = wb["FIGURES"]
        hm = _sheet_header_map(ws)
        assets: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            fig_id = str(_get(hm, row, "fig_id") or "").strip()
            if not fig_id:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            ftype = _map_figure_type_to_asset_type(str(_get(hm, row, "figure_type") or "").strip())
            file_path = str(_get(hm, row, "file_path") or "").strip()
            caption = str(_get(hm, row, "caption") or "").strip()
            title = str(_get(hm, row, "title") or "").strip()
            gen_method = str(_get(hm, row, "gen_method") or "").strip()
            geom_ref = str(_get(hm, row, "geom_ref") or "").strip()
            crop = str(_get(hm, row, "crop") or "").strip()
            width_mm = _parse_float(_get(hm, row, "width_mm"))
            source_origin = str(_get(hm, row, "source_origin") or "").strip()
            asset_role = str(_get(hm, row, "asset_role") or "").strip()
            source_class = str(_get(hm, row, "source_class") or "").strip()
            authenticity = str(_get(hm, row, "authenticity") or "").strip()
            usage_scope = str(_get(hm, row, "usage_scope") or "").strip()
            fallback_mode = str(_get(hm, row, "fallback_mode") or "").strip()
            doc_scope = str(_get(hm, row, "doc_scope") or "").strip()

            viewpoint = ""
            if fig_id.startswith("FIG-VP-"):
                viewpoint = fig_id.replace("FIG-", "", 1)  # VP-01

            # Backward compatibility: allow deriving authenticity from legacy source_origin.
            if not authenticity:
                so_u = source_origin.strip().upper()
                if so_u in {"REFERENCE", "REF"}:
                    authenticity = "REFERENCE"
                elif so_u == "OFFICIAL":
                    authenticity = "OFFICIAL"

            assets.append(
                {
                    "asset_id": fig_id,
                    "type": ftype,
                    "file_path": file_path,
                    "caption": _tf(caption, src_ids),
                    "source_ids": src_ids,
                    "sensitive": _flag_to_bool(_get(hm, row, "sensitive")),
                    # Not a claim; keep as plain string to avoid noisy exports.
                    "insert_anchor": str(_get(hm, row, "insert_anchor") or "").strip(),
                    # Optional figure controls (v2)
                    "title": title,
                    "doc_scope": doc_scope,
                    "source_origin": source_origin,
                    "asset_role": asset_role,
                    "source_class": source_class,
                    "authenticity": authenticity,
                    "usage_scope": usage_scope,
                    "fallback_mode": fallback_mode,
                    "gen_method": gen_method,
                    "geom_ref": geom_ref,
                    "width_mm": width_mm,
                    "crop": crop,
                    # Used for FIG-VP-* selection (spec_figures.py)
                    "viewpoint": _tf(viewpoint, src_ids) if viewpoint else _tf("", src_ids),
                }
            )
        if assets:
            data["assets"] = assets

    # ZONING_OVERLAY (optional; for regulatory/protected-area O/X tables)
    if "ZONING_OVERLAY" in wb.sheetnames:
        ws = wb["ZONING_OVERLAY"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "overlay_id": _tf(_get(hm, row, "overlay_id"), src_ids),
                    "category": _tf(_get(hm, row, "category"), src_ids),
                    "designation_name": _tf(_get(hm, row, "designation_name"), src_ids),
                    "is_applicable": _tf(_get(hm, row, "is_applicable"), src_ids),
                    "distance_m": _qf(_get(hm, row, "distance_m"), "m", src_ids),
                    "direction": _tf(_get(hm, row, "direction"), src_ids),
                    "basis": _tf(_get(hm, row, "basis"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            data["zoning_overlay"] = rows

    # ATTACHMENTS / FIELD_SURVEY_LOG as extras (compliance gates / appendix support)
    #
    # Note: older v2 templates accidentally created duplicate sheets like "ATTACHMENTS1".
    # We merge them best-effort to avoid losing user-entered evidence rows.
    def _parse_attachments_sheet(ws) -> list[dict[str, Any]]:
        hm = _sheet_header_map(ws)
        parsed: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            parsed.append(
                {
                    "evidence_id": _tf(_get(hm, row, "evidence_id"), src_ids),
                    "evidence_type": _tf(_get(hm, row, "evidence_type"), src_ids),
                    "title": _tf(_get(hm, row, "title"), src_ids),
                    "file_path": _tf(_get(hm, row, "file_path"), src_ids),
                    "related_fig_id": _tf(_get(hm, row, "related_fig_id"), src_ids),
                    "used_in": _tf(_get(hm, row, "used_in"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                    "sensitive": _tf(_get(hm, row, "sensitive"), src_ids),
                    "note": _tf(_get(hm, row, "note"), src_ids),
                }
            )
        return parsed

    attachment_rows: list[dict[str, Any]] = []
    for sheet_name in ("ATTACHMENTS", "ATTACHMENTS1"):
        if sheet_name not in wb.sheetnames:
            continue
        attachment_rows.extend(_parse_attachments_sheet(wb[sheet_name]))

    if attachment_rows:
        data["attachments_manifest"] = attachment_rows
        for it in attachment_rows:
            ev_id = str((it.get("evidence_id") or {}).get("t") or "").strip()
            src_ids = _filter_tbd_src_ids(list((it.get("evidence_id") or {}).get("src") or []))
            if not ev_id or not src_ids:
                continue
            existing = evidence_src_ids_by_id.get(ev_id, [])
            for s in src_ids:
                if s not in existing:
                    existing.append(s)
            evidence_src_ids_by_id[ev_id] = existing

    # DATA_REQUESTS (optional): store as extras for auditability / future enrich pipeline
    if "DATA_REQUESTS" in wb.sheetnames:
        ws = wb["DATA_REQUESTS"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            rows.append(
                {
                    "req_id": str(_get(hm, row, "req_id") or "").strip(),
                    "enabled": _get(hm, row, "enabled"),
                    "priority": _get(hm, row, "priority"),
                    "connector": str(_get(hm, row, "connector") or "").strip(),
                    "purpose": str(_get(hm, row, "purpose") or "").strip(),
                    "src_id": str(_get(hm, row, "src_id") or "").strip(),
                    "params_json": str(_get(hm, row, "params_json") or "").strip(),
                    "output_sheet": str(_get(hm, row, "output_sheet") or "").strip(),
                    "merge_strategy": str(_get(hm, row, "merge_strategy") or "").strip(),
                    "upsert_keys": str(_get(hm, row, "upsert_keys") or "").strip(),
                    "run_mode": str(_get(hm, row, "run_mode") or "").strip(),
                    "last_run_at": str(_get(hm, row, "last_run_at") or "").strip(),
                    "last_evidence_ids": str(_get(hm, row, "last_evidence_ids") or "").strip(),
                    "note": str(_get(hm, row, "note") or "").strip(),
                }
            )
        if rows:
            data["data_requests_manifest"] = rows

    if "FIELD_SURVEY_LOG" in wb.sheetnames:
        ws = wb["FIELD_SURVEY_LOG"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "survey_id": _tf(_get(hm, row, "survey_id"), src_ids),
                    "survey_date": _tf(_get(hm, row, "survey_date"), src_ids),
                    "survey_time_range": _tf(_get(hm, row, "survey_time_range"), src_ids),
                    "surveyors": _tf(_get(hm, row, "surveyors"), src_ids),
                    "weather": _tf(_get(hm, row, "weather"), src_ids),
                    "scope": _tf(_get(hm, row, "scope"), src_ids),
                    "route_desc": _tf(_get(hm, row, "route_desc"), src_ids),
                    "photo_folder": _tf(_get(hm, row, "photo_folder"), src_ids),
                    "notes": _tf(_get(hm, row, "notes"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            data["field_survey_log"] = rows

    # ENV_SCOPING -> Case.scoping_matrix
    if "ENV_SCOPING" in wb.sheetnames:
        ws = wb["ENV_SCOPING"]
        hm = _sheet_header_map(ws)
        scoping: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            item_id = str(_get(hm, row, "item_id") or "").strip()
            if not item_id:
                continue
            status = _map_env_status_to_korean(str(_get(hm, row, "status") or "").strip())
            item_name = str(_get(hm, row, "item_name") or "").strip()
            exclude_reason = str(_get(hm, row, "if_excluded_reason") or "").strip()
            baseline_method = str(_get(hm, row, "method") or "").strip()
            src_ids = _split_ids(_get(hm, row, "src_id"))
            scoping.append(
                {
                    "item_id": item_id,
                    "item_name": item_name or item_id,
                    "category": _tf(status, src_ids),
                    "exclude_reason": _tf(exclude_reason, src_ids),
                    "baseline_method": _tf(baseline_method, src_ids),
                    # If the user didn't provide a separate prediction/evaluation method,
                    # reuse the baseline method to avoid placeholder-only output.
                    "prediction_method": _tf(baseline_method, src_ids),
                    "src_expected": src_ids,
                }
            )
        if scoping:
            data["scoping_matrix"] = scoping

    # ENV_BASE_AIR -> baseline.air_quality (pm10/pm25/o3 only)
    if "ENV_BASE_AIR" in wb.sheetnames:
        ws = wb["ENV_BASE_AIR"]
        hm = _sheet_header_map(ws)
        station_name = ""
        pm10 = None
        pm25 = None
        o3 = None
        src_ids_any: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            src_ids_any = src_ids_any or src_ids
            if not station_name:
                station_name = str(_get(hm, row, "station_name") or "").strip()
            pol = str(_get(hm, row, "pollutant") or "").strip().upper()
            val = _get(hm, row, "value_avg")
            unit = str(_get(hm, row, "unit") or "").strip() or ("µg/m3" if pol in {"PM10", "PM2.5", "PM2_5"} else "")
            if pol == "PM10":
                pm10 = _qf(val, unit or "µg/m3", src_ids)
            if pol in {"PM2.5", "PM2_5"}:
                pm25 = _qf(val, unit or "µg/m3", src_ids)
            if pol in {"O3", "OZONE"}:
                o3 = _qf(val, unit or "ppm", src_ids)
        if station_name or pm10 or pm25 or o3:
            data.setdefault("baseline", {})
            data["baseline"]["air_quality"] = {
                "station_name": _tf(station_name, src_ids_any),
                "pm10_ugm3": pm10 or _qf(None, "µg/m3", src_ids_any),
                "pm25_ugm3": pm25 or _qf(None, "µg/m3", src_ids_any),
                "ozone_ppm": o3 or _qf(None, "ppm", src_ids_any),
            }

    # ENV_BASE_SOCIO -> baseline.population_traffic (best-effort; extra fields)
    if "ENV_BASE_SOCIO" in wb.sheetnames:
        ws = wb["ENV_BASE_SOCIO"]
        hm = _sheet_header_map(ws)
        admin_code = ""
        admin_name = ""
        stats: list[dict[str, Any]] = []
        src_ids_any: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue

            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            src_ids_any = src_ids_any or src_ids

            if not admin_code:
                admin_code = str(_get(hm, row, "admin_code") or "").strip()
            if not admin_name:
                admin_name = str(_get(hm, row, "admin_name") or "").strip()

            year = str(_get(hm, row, "year") or "").strip()
            if not year:
                continue

            stats.append(
                {
                    "year": _tf(year, src_ids),
                    "population_total": _qf(_get(hm, row, "population_total"), "명", src_ids),
                    "households": _qf(_get(hm, row, "households"), "세대", src_ids),
                    "housing_total": _qf(_get(hm, row, "housing_total"), "호", src_ids),
                }
            )

        if stats or admin_code or admin_name:
            data.setdefault("baseline", {}).setdefault("population_traffic", {})
            if admin_code:
                data["baseline"]["population_traffic"]["admin_code"] = _tf(admin_code, src_ids_any)
            if admin_name:
                data["baseline"]["population_traffic"]["admin_name"] = _tf(admin_name, src_ids_any)
            if stats:
                data["baseline"]["population_traffic"]["socio_stats"] = stats

    # ENV_BASE_WATER -> streams + water_quality
    if "ENV_BASE_WATER" in wb.sheetnames:
        ws = wb["ENV_BASE_WATER"]
        hm = _sheet_header_map(ws)
        streams: list[dict[str, Any]] = []
        wq: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            name = str(_get(hm, row, "waterbody_name") or "").strip()
            if name:
                streams.append(
                    {
                        "name": _tf(name, src_ids),
                        "distance_m": _qf(_get(hm, row, "distance_m"), "m", src_ids),
                        "flow_direction": _tf("", src_ids),
                        "note": _tf(str(_get(hm, row, "relation") or "").strip(), src_ids),
                    }
                )
            param = str(_get(hm, row, "parameter") or "").strip()
            if param:
                key = _map_water_param_to_key(param)
                unit = str(_get(hm, row, "unit") or "").strip() or "mg/L"
                wq[key] = _qf(_get(hm, row, "value"), unit, src_ids)
        if streams or wq:
            data.setdefault("baseline", {}).setdefault("water_environment", {})
            if streams:
                data["baseline"]["water_environment"]["streams"] = streams
            if wq:
                data["baseline"]["water_environment"]["water_quality"] = wq

    # ENV_BASE_NOISE -> one receptor
    if "ENV_BASE_NOISE" in wb.sheetnames:
        ws = wb["ENV_BASE_NOISE"]
        hm = _sheet_header_map(ws)
        receptors: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            receptors.append(
                {
                    "name": _tf(_get(hm, row, "point_name"), src_ids),
                    "distance_m": _qf(None, "m", src_ids),
                    "baseline_day_db": _qf(_get(hm, row, "day_leq"), "dB(A)", src_ids),
                    "baseline_night_db": _qf(_get(hm, row, "night_leq"), "dB(A)", src_ids),
                    "measured": _tf("false", src_ids),
                }
            )
        if receptors:
            data.setdefault("baseline", {}).setdefault("noise_vibration", {})
            data["baseline"]["noise_vibration"]["receptors"] = receptors

    # ENV_BASE_GEO -> baseline.topography_geology (best-effort)
    if "ENV_BASE_GEO" in wb.sheetnames:
        ws = wb["ENV_BASE_GEO"]
        hm = _sheet_header_map(ws)
        topo: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            topic = str(_get(hm, row, "topic") or "").strip()
            summary = str(_get(hm, row, "summary") or "").strip()
            if not topic or not summary:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]

            if "표고" in topic or "elevation" in topic.lower():
                topo["elevation_range_m"] = _tf(summary, src_ids)
                continue

            if "지질" in topic or "geology" in topic.lower():
                topo["geology_summary"] = _tf(summary, src_ids)
                continue

            if "토양" in topic or "soil" in topic.lower():
                topo["soil_summary"] = _tf(summary, src_ids)
                continue

            # fallback: append to geology_summary
            prev = topo.get("geology_summary")
            prev_txt = prev.get("t") if isinstance(prev, dict) else ""
            topo["geology_summary"] = _tf((prev_txt + "\n" if prev_txt else "") + f"{topic}: {summary}", src_ids)

        if topo:
            data.setdefault("baseline", {}).setdefault("topography_geology", {})
            data["baseline"]["topography_geology"].update(topo)

    # ENV_ECO_EVENTS / ENV_ECO_OBS -> baseline.ecology (best-effort)
    if "ENV_ECO_EVENTS" in wb.sheetnames:
        ws = wb["ENV_ECO_EVENTS"]
        hm = _sheet_header_map(ws)
        dates: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            d = _get(hm, row, "date")
            if d is None or not str(d).strip():
                continue
            src_ids = _src_ids_from_row_or_evidence(hm, row)
            dates.append(_tf(d, src_ids or None))
        if dates:
            data.setdefault("baseline", {}).setdefault("ecology", {})
            data["baseline"]["ecology"]["survey_dates"] = dates

    if "ENV_ECO_OBS" in wb.sheetnames:
        ws = wb["ENV_ECO_OBS"]
        hm = _sheet_header_map(ws)
        flora: list[dict[str, Any]] = []
        fauna: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            taxon = str(_get(hm, row, "taxon_group") or "").strip()
            ko = str(_get(hm, row, "korean_name") or "").strip()
            if not taxon and not ko:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            entry = {
                "species_ko": _tf(ko, src_ids),
                "scientific": _tf(_get(hm, row, "scientific_name"), src_ids),
                "protected": _tf(_get(hm, row, "protected_status"), src_ids),
                "note": _tf(_get(hm, row, "note"), src_ids),
                "evidence": _tf(_get(hm, row, "evidence_type"), src_ids),
            }
            if "식물" in taxon:
                flora.append(entry)
            else:
                fauna.append(entry)
        if flora or fauna:
            data.setdefault("baseline", {}).setdefault("ecology", {})
            if flora:
                data["baseline"]["ecology"]["flora_list"] = flora
            if fauna:
                data["baseline"]["ecology"]["fauna_list"] = fauna

    # ENV_LANDSCAPE -> baseline.landuse_landscape.key_viewpoints
    if "ENV_LANDSCAPE" in wb.sheetnames:
        ws = wb["ENV_LANDSCAPE"]
        hm = _sheet_header_map(ws)
        vps: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            vid = str(_get(hm, row, "view_id") or "").strip()
            if not vid:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            vps.append(
                {
                    "vp_id": _tf(vid, src_ids),
                    "location_desc": _tf(_get(hm, row, "description") or _get(hm, row, "viewpoint_name"), src_ids),
                    "photo_asset_id": _tf(_get(hm, row, "photo_fig_id"), src_ids),
                    "note": _tf("", src_ids),
                }
            )
        if vps:
            data.setdefault("baseline", {}).setdefault("landuse_landscape", {})
            data["baseline"]["landuse_landscape"]["key_viewpoints"] = vps

    # ENV_MITIGATION -> Case.mitigation.measures
    if "ENV_MITIGATION" in wb.sheetnames:
        ws = wb["ENV_MITIGATION"]
        hm = _sheet_header_map(ws)
        measures: list[dict[str, Any]] = []

        def _stage_to_phase(stage: str) -> str:
            v = (stage or "").strip().upper()
            if v in {"CONSTRUCTION", "공사"}:
                return "공사"
            if v in {"OPERATION", "운영"}:
                return "운영"
            if v in {"BOTH"}:
                return "공사/운영"
            return stage

        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            mid = str(_get(hm, row, "mit_id") or "").strip()
            if not mid:
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            related = _split_ids(_get(hm, row, "target_item"))
            measures.append(
                {
                    "measure_id": mid,
                    "phase": _tf(_stage_to_phase(str(_get(hm, row, "stage") or "")), src_ids),
                    "title": _tf(_get(hm, row, "measure"), src_ids),
                    "description": _tf(_get(hm, row, "measure"), src_ids),
                    "location_ref": _tf(_get(hm, row, "location") or _get(hm, row, "evidence_id"), src_ids),
                    "monitoring": _tf(_get(hm, row, "responsible"), src_ids),
                    "related_impacts": related,
                }
            )
        if measures:
            data["mitigation"] = {"measures": measures}
            # Best-effort: if summary_inputs is empty, derive it from mitigation rows.
            # (Keeps CH0_SUMMARY from being blank when user already filled ENV_MITIGATION.)
            if not data.get("summary_inputs"):
                key_issues: list[dict[str, Any]] = []
                key_measures: list[dict[str, Any]] = []
                seen_issues: set[str] = set()
                seen_measures: set[str] = set()
                for m in measures:
                    m_title = str((m.get("title") or {}).get("t") or "").strip()
                    m_src_ids = list((m.get("title") or {}).get("src") or [])
                    if m_title and m_title not in seen_measures:
                        key_measures.append(_tf(m_title, m_src_ids or ["S-TBD"]))
                        seen_measures.add(m_title)
                    for imp in (m.get("related_impacts") or []):
                        imp_t = str(imp or "").strip()
                        if not imp_t or imp_t in seen_issues:
                            continue
                        key_issues.append(_tf(imp_t, m_src_ids or ["S-TBD"]))
                        seen_issues.add(imp_t)
                if key_issues or key_measures:
                    data["summary_inputs"] = {"key_issues": key_issues, "key_measures": key_measures}

    # ENV_MANAGEMENT -> Case.management_plan.implementation_register (best-effort)
    if "ENV_MANAGEMENT" in wb.sheetnames:
        ws = wb["ENV_MANAGEMENT"]
        hm = _sheet_header_map(ws)
        items: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _src_ids_from_row_or_evidence(hm, row) or None
            items.append(
                {
                    "item": _tf(_get(hm, row, "condition_text"), src_ids),
                    "measure_id": _tf(_get(hm, row, "cond_id"), src_ids),
                    "when": _tf(_get(hm, row, "compliance_plan") or _get(hm, row, "status"), src_ids),
                    "evidence": _tf(_get(hm, row, "evidence_id"), src_ids),
                    "responsible": _tf("", src_ids),
                }
            )
        if items:
            data["management_plan"] = {"implementation_register": items}

    # DRR_* -> case.disaster (spec_dia compatibility; best-effort)
    disaster: dict[str, Any] = {}
    zoning_disaster_overlays: list[dict[str, Any]] = []
    if "ZONING_OVERLAY" in wb.sheetnames:
        ws = wb["ZONING_OVERLAY"]
        hm = _sheet_header_map(ws)
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            category = str(_get(hm, row, "category") or "").strip().upper()
            if category != "DISASTER":
                continue
            zoning_disaster_overlays.append(
                {
                    "overlay_id": str(_get(hm, row, "overlay_id") or "").strip(),
                    "designation_name": str(_get(hm, row, "designation_name") or "").strip(),
                    "is_applicable": str(_get(hm, row, "is_applicable") or "").strip(),
                    "basis": str(_get(hm, row, "basis") or "").strip(),
                    "data_origin": str(_get(hm, row, "data_origin") or "").strip(),
                    "src_ids": _split_ids(_get(hm, row, "src_id")) or ["S-TBD"],
                }
            )

    if "DRR_SCOPING" in wb.sheetnames:
        ws = wb["DRR_SCOPING"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            include = _get(hm, row, "include")
            applicable = _map_include_to_yes_no_unknown(include)
            reason = _get(hm, row, "reason")
            rows.append(
                {
                    "hazard_item": _tf(_get(hm, row, "hazard_type"), src_ids),
                    "applicable": _tf(applicable, src_ids),
                    "analysis_level": _tf(_get(hm, row, "method"), src_ids),
                    "exclude_reason": _tf(reason if applicable == "NO" else "", src_ids),
                }
            )
        if not rows and zoning_disaster_overlays:
            for it in zoning_disaster_overlays:
                src_ids = it.get("src_ids") or ["S-TBD"]
                applicable = _map_include_to_yes_no_unknown(it.get("is_applicable"))
                basis = str(it.get("basis") or it.get("data_origin") or "").strip()
                analysis_level = f"WMS overlay ({basis})".strip() if basis else "WMS overlay"
                exclude_reason = "WMS overlay 중첩 없음" if applicable == "NO" else ""
                hazard_item = str(it.get("designation_name") or it.get("overlay_id") or "").strip()
                if not hazard_item:
                    continue
                rows.append(
                    {
                        "hazard_item": _tf(hazard_item, src_ids),
                        "applicable": _tf(applicable, src_ids),
                        "analysis_level": _tf(analysis_level, src_ids),
                        "exclude_reason": _tf(exclude_reason, src_ids),
                    }
                )
        if rows:
            disaster["scoping_matrix"] = rows

    default_basin_area_km2: float | None = None
    if "DRR_TARGET_AREA" in wb.sheetnames:
        ws = wb["DRR_TARGET_AREA"]
        hm = _sheet_header_map(ws)
        targets: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            if default_basin_area_km2 is None:
                try:
                    v = _get(hm, row, "upstream_area_km2")
                    default_basin_area_km2 = float(v) if v is not None and str(v).strip() != "" else None
                except Exception:
                    default_basin_area_km2 = None
            targets.append(
                {
                    "concept": _tf(_get(hm, row, "concept"), src_ids),
                    "upstream_area_km2": _qf(_get(hm, row, "upstream_area_km2"), "km2", src_ids),
                    "downstream_to": _tf(_get(hm, row, "downstream_to"), src_ids),
                    "affected_neighborhood": _tf(_get(hm, row, "affected_neighborhood"), src_ids),
                    "map_fig_id": _tf(_get(hm, row, "map_fig_id"), src_ids),
                }
            )
        if targets:
            disaster["target_area"] = targets

    if "DRR_TARGET_AREA_PARTS" in wb.sheetnames:
        ws = wb["DRR_TARGET_AREA_PARTS"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "part": _tf(_get(hm, row, "part"), src_ids),
                    "included": _tf(_map_include_to_yes_no_unknown(_get(hm, row, "included")), src_ids),
                    "reason": _tf(_get(hm, row, "reason"), src_ids),
                    "exclude_reason": _tf(_get(hm, row, "exclude_reason"), src_ids),
                    "geom_ref": _tf(_get(hm, row, "geom_ref"), src_ids),
                    "figure_id": _tf(_get(hm, row, "figure_id"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            disaster["target_area_parts"] = rows

    if "DRR_BASE_HAZARD" in wb.sheetnames:
        ws = wb["DRR_BASE_HAZARD"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "hazard_id": _tf(_get(hm, row, "hazard_id"), src_ids),
                    "hazard_type": _tf(_get(hm, row, "hazard_type"), src_ids),
                    "occurred": _tf(_get(hm, row, "occurred"), src_ids),
                    "interview_done": _tf(_get(hm, row, "interview_done"), src_ids),
                    "interview_summary": _tf(_get(hm, row, "interview_summary"), src_ids),
                    "photo_fig_id": _tf(_get(hm, row, "photo_fig_id"), src_ids),
                    "evidence_id": _tf(_get(hm, row, "evidence_id"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            disaster["hazard_history"] = rows

    if "DRR_INTERVIEWS" in wb.sheetnames:
        ws = wb["DRR_INTERVIEWS"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "interview_id": _tf(_get(hm, row, "interview_id"), src_ids),
                    "respondent_id": _tf(_get(hm, row, "respondent_id"), src_ids),
                    "residence_years": _qf(_get(hm, row, "residence_years"), "yr", src_ids),
                    "location_desc": _tf(_get(hm, row, "location_desc"), src_ids),
                    "summary": _tf(_get(hm, row, "summary"), src_ids),
                    "photo_fig_id": _tf(_get(hm, row, "photo_fig_id"), src_ids),
                    "evidence_id": _tf(_get(hm, row, "evidence_id"), src_ids),
                    "data_origin": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            disaster["interviews"] = rows

    if "DRR_HYDRO_RAIN" in wb.sheetnames:
        ws = wb["DRR_HYDRO_RAIN"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            station_label = _get(hm, row, "station_name")
            if not station_label:
                station_label = _get(hm, row, "source_basis")
            dur_hr = _get(hm, row, "duration_hr")
            dur_min = None
            try:
                dur_min = float(dur_hr) * 60.0 if dur_hr is not None else None
            except Exception:
                dur_min = None
            rows.append(
                {
                    "station_name": _tf(station_label, src_ids),
                    "duration_min": _qf(dur_min, "min", src_ids),
                    "frequency_year": _qf(_get(hm, row, "return_period_yr"), "yr", src_ids),
                    "rainfall_mm": _qf(_get(hm, row, "rainfall_mm"), "mm", src_ids),
                    "source_type": _tf(_get(hm, row, "data_origin"), src_ids),
                }
            )
        if rows:
            disaster["rainfall"] = rows

    if "DRR_HYDRO_MODEL" in wb.sheetnames:
        ws = wb["DRR_HYDRO_MODEL"]
        hm = _sheet_header_map(ws)
        basins: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            pre = _get(hm, row, "peak_cms_before")
            post = _get(hm, row, "peak_cms_after")
            delta = None
            try:
                if pre is not None and post is not None:
                    delta = float(post) - float(pre)
            except Exception:
                delta = None
            basins.append(
                {
                    "basin_id": _tf(_get(hm, row, "hydro_id"), src_ids),
                    "basin_area_km2": _qf(default_basin_area_km2, "km2", src_ids),
                    "tc_min": _qf(_get(hm, row, "tc_min"), "min", src_ids),
                    "cn_value": _qf(_get(hm, row, "cn_or_c"), "", src_ids),
                    "pre_peak_cms": _qf(pre, "m3/s", src_ids),
                    "post_peak_cms": _qf(post, "m3/s", src_ids),
                    "delta_peak_cms": _qf(delta, "m3/s", src_ids),
                    "model": _tf(_get(hm, row, "model"), src_ids),
                }
            )
        if basins:
            disaster["runoff_basins"] = basins

    if utilities_drainage and "drainage_facilities" not in disaster:
        disaster["drainage_facilities"] = utilities_drainage

    if "DRR_MITIGATION" in wb.sheetnames:
        ws = wb["DRR_MITIGATION"]
        hm = _sheet_header_map(ws)
        measures: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            measures.append(
                {
                    "measure_id": _tf(_get(hm, row, "drr_mit_id"), src_ids),
                    "target_hazard": _tf(_get(hm, row, "hazard_type"), src_ids),
                    "stage": _tf("BOTH", src_ids),
                    "description": _tf(_get(hm, row, "description"), src_ids),
                    "linked_facility_id": _tf("", src_ids),
                }
            )
        if measures:
            disaster["measures"] = measures

    if "DRR_MAINTENANCE" in wb.sheetnames:
        ws = wb["DRR_MAINTENANCE"]
        hm = _sheet_header_map(ws)
        ledger: list[dict[str, Any]] = []
        maintenance_summary: list[str] = []
        maintenance_src_ids: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            maintenance_src_ids = _split_ids(";".join([*maintenance_src_ids, *src_ids])) or maintenance_src_ids
            facility = str(_get(hm, row, "facility_name") or "").strip()
            cycle = str(_get(hm, row, "inspection_cycle") or "").strip()
            method = str(_get(hm, row, "maintenance_method") or "").strip()
            summary = f"{facility}:{cycle} 점검".strip(":")
            if method:
                summary = f"{summary}({method})"
            if summary:
                maintenance_summary.append(summary)
            ledger.append(
                {
                    "asset_id": _tf(_get(hm, row, "facility_name"), src_ids),
                    "inspection_cycle": _tf(_get(hm, row, "inspection_cycle"), src_ids),
                    "inspection_item": _tf(_get(hm, row, "maintenance_method"), src_ids),
                    "responsible_role": _tf(_get(hm, row, "responsible"), src_ids),
                    "record_format": _tf(_get(hm, row, "ledger_template"), src_ids),
                    "evidence_id_template": _tf(_get(hm, row, "evidence_id"), src_ids),
                }
            )
        if ledger:
            disaster["maintenance_ledger"] = ledger
            summary_text = " / ".join([s for s in maintenance_summary if s])
            if summary_text and utilities_drainage:
                for fac in utilities_drainage:
                    fac_src_ids = fac.get("facility_id", {}).get("src") or []
                    merged = _split_ids(";".join([*fac_src_ids, *maintenance_src_ids])) or (fac_src_ids or maintenance_src_ids or ["S-TBD"])
                    fac["maintenance_class"] = _tf(summary_text, merged)

            if "measures" not in disaster and utilities_drainage:
                measures: list[dict[str, Any]] = []
                for fac in utilities_drainage:
                    fac_id = str((fac.get("facility_id") or {}).get("t") or "").strip()
                    if not fac_id:
                        continue
                    fac_src_ids = fac.get("facility_id", {}).get("src") or []
                    merged = _split_ids(";".join([*fac_src_ids, *maintenance_src_ids])) or (fac_src_ids or maintenance_src_ids or ["S-TBD"])

                    cap = str((fac.get("capacity") or {}).get("t") or "").strip()
                    discharge = str((fac.get("discharge_to") or {}).get("t") or "").strip()
                    mnt = str((fac.get("maintenance_class") or {}).get("t") or "").strip()
                    desc_parts = []
                    if cap:
                        desc_parts.append(f"규모: {cap}")
                    if discharge:
                        desc_parts.append(f"방류처: {discharge}")
                    if mnt:
                        desc_parts.append(f"유지관리: {mnt}")
                    desc = " / ".join(desc_parts) or "우수/배수 계획 및 유지관리 반영"

                    measures.append(
                        {
                            "measure_id": _tf(f"AUTO-{fac_id}", merged),
                            "target_hazard": _tf("INLAND", merged),
                            "stage": _tf("BOTH", merged),
                            "description": _tf(desc, merged),
                            "linked_facility_id": _tf(fac_id, merged),
                        }
                    )
                if measures:
                    disaster["measures"] = measures

    if "DRR_SEDIMENT" in wb.sheetnames:
        ws = wb["DRR_SEDIMENT"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "sed_id": _tf(_get(hm, row, "sed_id"), src_ids),
                    "method": _tf(_get(hm, row, "method"), src_ids),
                    "r_factor": _qf(_get(hm, row, "r_factor"), "", src_ids),
                    "k_factor": _qf(_get(hm, row, "k_factor"), "", src_ids),
                    "ls_factor": _qf(_get(hm, row, "ls_factor"), "", src_ids),
                    "c_factor": _qf(_get(hm, row, "c_factor"), "", src_ids),
                    "p_factor": _qf(_get(hm, row, "p_factor"), "", src_ids),
                    "soil_loss_before": _qf(_get(hm, row, "soil_loss_t_ha_yr_before"), "t/ha/yr", src_ids),
                    "soil_loss_after": _qf(_get(hm, row, "soil_loss_t_ha_yr_after"), "t/ha/yr", src_ids),
                }
            )
        if not rows:
            # Self-use fallback: keep the table non-placeholder without inventing numbers.
            # Use S-AUTHOR-INPUT (declared in sources.yaml by default).
            src_ids = ["S-AUTHOR-INPUT"]
            rows.append(
                {
                    "sed_id": _tf("AUTO-SEDIMENT-001", src_ids),
                    "method": _tf("자료 미확보(미산정, 추후 보완)", src_ids),
                    "r_factor": _qf(None, "", src_ids),
                    "k_factor": _qf(None, "", src_ids),
                    "ls_factor": _qf(None, "", src_ids),
                    "c_factor": _qf(None, "", src_ids),
                    "p_factor": _qf(None, "", src_ids),
                    "soil_loss_before": _qf(None, "t/ha/yr", src_ids),
                    "soil_loss_after": _qf(None, "t/ha/yr", src_ids),
                }
            )
        if rows:
            disaster["sediment_erosion"] = rows

    if "DRR_SLOPE" in wb.sheetnames:
        ws = wb["DRR_SLOPE"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            rows.append(
                {
                    "slope_id": _tf(_get(hm, row, "slope_id"), src_ids),
                    "exists": _tf(_get(hm, row, "has_slope_work"), src_ids),
                    "height_m": _qf(_get(hm, row, "height_m"), "m", src_ids),
                    "risk_grade": _qf(_get(hm, row, "risk_grade"), "", src_ids),
                    "stabilization": _tf(_get(hm, row, "mitigation_ref"), src_ids),
                    "hazard_map_layer_used": _tf("", src_ids),
                }
            )
        if not rows and zoning_disaster_overlays:
            for it in zoning_disaster_overlays:
                oid = str(it.get("overlay_id") or "").strip().upper()
                name = str(it.get("designation_name") or "").strip()
                if "LANDSLIDE" not in oid and "산사태" not in name:
                    continue
                src_ids = it.get("src_ids") or ["S-TBD"]
                applicable = _map_include_to_yes_no_unknown(it.get("is_applicable"))
                rows.append(
                    {
                        "slope_id": _tf("AUTO-SLOPE-01", src_ids),
                        "exists": _tf(applicable, src_ids),
                        "height_m": _qf(None, "m", src_ids),
                        "risk_grade": _qf(None, "", src_ids),
                        "stabilization": _tf("", src_ids),
                        "hazard_map_layer_used": _tf(name or oid, src_ids),
                    }
                )
                break
        if rows:
            disaster["slope_landslide"] = rows

    if "SSOT_PAGE_OVERRIDES" in wb.sheetnames:
        ws = wb["SSOT_PAGE_OVERRIDES"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue

            sample_page = _parse_int(_get(hm, row, "sample_page"))
            override_file_path = str(_get(hm, row, "override_file_path") or "").strip()
            override_page = _parse_int(_get(hm, row, "override_page"))
            if not sample_page or not override_file_path or not override_page:
                continue

            width_mm = _parse_float(_get(hm, row, "width_mm"))
            dpi = _parse_int(_get(hm, row, "dpi"))
            crop = str(_get(hm, row, "crop") or "").strip() or None
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            note = str(_get(hm, row, "note") or "").strip()

            rows.append(
                {
                    "sample_page": sample_page,
                    "file_path": override_file_path,
                    "page": override_page,
                    "width_mm": width_mm,
                    "dpi": dpi,
                    "crop": crop,
                    "src_ids": src_ids,
                    "note": note,
                }
            )
        if rows:
            data["ssot_page_overrides"] = rows

    if "APPENDIX_INSERTS" in wb.sheetnames:
        ws = wb["APPENDIX_INSERTS"]
        hm = _sheet_header_map(ws)
        rows: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue

            ins_id = str(_get(hm, row, "ins_id") or "").strip()
            file_path = str(_get(hm, row, "file_path") or "").strip()
            pdf_page = _parse_int(_get(hm, row, "pdf_page"))
            if not ins_id or not file_path or not pdf_page:
                continue

            order = _parse_int(_get(hm, row, "order")) or 0
            caption = str(_get(hm, row, "caption") or "").strip()
            width_mm = _parse_float(_get(hm, row, "width_mm"))
            dpi = _parse_int(_get(hm, row, "dpi"))
            crop = str(_get(hm, row, "crop") or "").strip() or None
            src_ids = _split_ids(_get(hm, row, "src_id")) or ["S-TBD"]
            note = str(_get(hm, row, "note") or "").strip()

            rows.append(
                {
                    "ins_id": ins_id,
                    "order": order,
                    "file_path": file_path,
                    "page": pdf_page,
                    "caption": caption,
                    "width_mm": width_mm,
                    "dpi": dpi,
                    "crop": crop,
                    "src_ids": src_ids,
                    "note": note,
                }
            )

        if rows:
            data["appendix_inserts"] = rows

    if disaster:
        data["disaster"] = disaster

    case = Case.model_validate(data or {})
    return canonicalize_case(case)
