from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class _SheetSpec:
    name: str
    headers: list[str]


# v2 (snake_case) 통합 입력 템플릿: ENV(환경) + DRR(재해) 병렬 입력
_SHEETS_V2: list[_SheetSpec] = [
    _SheetSpec("META", ["case_id", "template_version", "created_at", "author", "locale", "output_targets"]),
    _SheetSpec("PARTIES", ["party_id", "role", "name", "biz_reg_no", "address", "contact", "note", "src_id"]),
    _SheetSpec(
        "PROJECT",
        [
            "project_name",
            "project_type",
            "doc_env_required",
            "doc_drr_required",
            "total_area_m2",
            "zonings_summary",
            "main_facilities_summary",
            "max_occupancy",
            "expected_visitors_day",
            "water_supply",
            "sewage_treatment",
            "stormwater_plan",
            "client_name",
            "proponent_name",
            "author_org",
            "submit_date",
            "approving_authority",
            "consultation_agency",
            "sensitive_fields_note",
            "src_id",
        ],
    ),
    _SheetSpec(
        "LOCATION",
        [
            "address_road",
            "address_jibeon",
            "admin_si",
            "admin_eupmyeon",
            "center_lat",
            "center_lon",
            "crs",
            "boundary_file",
            "bbox_wkt",
            "src_id",
        ],
    ),
    _SheetSpec(
        "PARCELS",
        [
            "parcel_id",
            "parcel_no",
            "jimok",
            "area_m2",
            "official_value",
            "zoning",
            "ownership",
            "note",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ZONING_BREAKDOWN",
        [
            "zoning",
            "area_m2",
            "src_id",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "ZONING_OVERLAY",
        [
            "overlay_id",
            "category",
            "designation_name",
            "is_applicable",
            "distance_m",
            "direction",
            "basis",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "PLAN_LANDUSE",
        ["landuse_id", "stage", "category", "area_m2", "impervious", "runoff_coeff", "note", "src_id"],
    ),
    _SheetSpec(
        "BUILDINGS",
        [
            "bldg_id",
            "name",
            "floors",
            "main_use",
            "sub_uses",
            "building_area_m2",
            "gross_floor_area_m2",
            "structure",
            "occupancy",
            "src_id",
        ],
    ),
    _SheetSpec(
        "FACILITIES",
        [
            "fac_id",
            "type",
            "name",
            "qty",
            "qty_unit",
            "area_m2",
            "capacity_person",
            "length_m",
            "max_depth_m",
            "material",
            "note",
            "src_id",
        ],
    ),
    _SheetSpec(
        "UTILITIES",
        ["util_id", "util_type", "description", "capacity", "discharge_point", "drawing_ref", "src_id"],
    ),
    _SheetSpec("SCHEDULE", ["sched_id", "phase", "start_date", "end_date", "note", "src_id"]),
    _SheetSpec(
        "FIGURES",
        [
            "fig_id",
            "doc_scope",
            "figure_type",
            "title",
            "caption",
            "source_origin",
            "asset_role",
            "source_class",
            "authenticity",
            "usage_scope",
            "fallback_mode",
            "file_path",
            "gen_method",
            "geom_ref",
            "width_mm",
            "crop",
            "src_id",
            "sensitive",
            "insert_anchor",
        ],
    ),
    _SheetSpec(
        "ATTACHMENTS",
        [
            "evidence_id",
            "evidence_type",
            "title",
            "file_path",
            "related_fig_id",
            "used_in",
            "data_origin",
            "src_id",
            "sensitive",
            "note",
        ],
    ),
    _SheetSpec(
        "SSOT_PAGE_OVERRIDES",
        [
            "sample_page",
            "override_file_path",
            "override_page",
            "width_mm",
            "dpi",
            "crop",
            "src_id",
            "note",
        ],
    ),
    _SheetSpec(
        "APPENDIX_INSERTS",
        [
            "ins_id",
            "order",
            "file_path",
            "pdf_page",
            "caption",
            "width_mm",
            "dpi",
            "crop",
            "src_id",
            "note",
        ],
    ),
    _SheetSpec(
        "DATA_REQUESTS",
        [
            "req_id",
            "enabled",
            "priority",
            "connector",
            "purpose",
            "src_id",
            "params_json",
            "output_sheet",
            "merge_strategy",
            "upsert_keys",
            "run_mode",
            "last_run_at",
            "last_evidence_ids",
            "note",
        ],
    ),
    _SheetSpec(
        "FIELD_SURVEY_LOG",
        [
            "survey_id",
            "survey_date",
            "survey_time_range",
            "surveyors",
            "weather",
            "scope",
            "route_desc",
            "photo_folder",
            "notes",
            "data_origin",
            "src_id",
        ],
    ),
    # ENV (EIA)
    _SheetSpec(
        "ENV_SCOPING",
        [
            "item_id",
            "category",
            "item_name",
            "status",
            "reason",
            "data_origin",
            "method",
            "survey_needed",
            "if_excluded_reason",
            "prior_assessment_exists",
            "omission_basis",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ENV_BASE_AIR",
        [
            "air_id",
            "station_name",
            "station_distance_km",
            "period_start",
            "period_end",
            "pollutant",
            "value_avg",
            "unit",
            "data_origin",
            "src_id",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "ENV_BASE_SOCIO",
        [
            "socio_id",
            "admin_code",
            "admin_name",
            "year",
            "population_total",
            "households",
            "housing_total",
            "data_origin",
            "src_id",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "ENV_BASE_WATER",
        [
            "water_id",
            "waterbody_name",
            "relation",
            "distance_m",
            "parameter",
            "value",
            "unit",
            "sampling_date",
            "data_origin",
            "src_id",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "ENV_BASE_NOISE",
        ["noise_id", "point_name", "receptor_type", "day_leq", "night_leq", "unit", "data_origin", "note", "src_id"],
    ),
    _SheetSpec(
        "ENV_BASE_GEO",
        ["geo_id", "topic", "summary", "data_origin", "source_map", "src_id", "evidence_id"],
    ),
    _SheetSpec("ENV_ECO_EVENTS", ["eco_event_id", "date", "season", "survey_team", "weather", "area_desc", "data_origin", "evidence_id"]),
    _SheetSpec(
        "ENV_ECO_OBS",
        [
            "obs_id",
            "eco_event_id",
            "taxon_group",
            "scientific_name",
            "korean_name",
            "count",
            "evidence_type",
            "protected_status",
            "data_origin",
            "src_id",
            "note",
        ],
    ),
    _SheetSpec(
        "ENV_LANDSCAPE",
        [
            "view_id",
            "viewpoint_name",
            "lat",
            "lon",
            "description",
            "photo_fig_id",
            "simulation_fig_id",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ENV_IMPACT_AIR",
        [
            "air_imp_id",
            "pollutant",
            "calc_method",
            "emission_gps",
            "model_name",
            "background_value",
            "increment_value",
            "predicted_value",
            "standard_value",
            "exceed",
            "data_origin",
            "evidence_id",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ENV_IMPACT_WATER",
        [
            "water_imp_id",
            "impact_type",
            "earthwork_m3",
            "slope_area_m2",
            "mitigation_ref",
            "data_origin",
            "evidence_id",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ENV_MITIGATION",
        [
            "mit_id",
            "stage",
            "target_item",
            "measure",
            "location",
            "responsible",
            "evidence_id",
            "src_id",
        ],
    ),
    _SheetSpec(
        "ENV_MANAGEMENT",
        ["cond_id", "condition_text", "compliance_plan", "evidence_id", "status", "note"],
    ),
    # DRR (DIA)
    _SheetSpec(
        "DRR_SCOPING",
        ["drr_item_id", "hazard_type", "include", "reason", "method", "data_origin", "src_id"],
    ),
    _SheetSpec(
        "DRR_TARGET_AREA",
        [
            "area_id",
            "concept",
            "upstream_area_km2",
            "downstream_to",
            "affected_neighborhood",
            "map_fig_id",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_TARGET_AREA_PARTS",
        [
            "part",
            "included",
            "reason",
            "exclude_reason",
            "geom_ref",
            "figure_id",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_BASE_HAZARD",
        [
            "hazard_id",
            "hazard_type",
            "occurred",
            "interview_done",
            "interview_summary",
            "photo_fig_id",
            "evidence_id",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_INTERVIEWS",
        [
            "interview_id",
            "respondent_id",
            "residence_years",
            "location_desc",
            "summary",
            "photo_fig_id",
            "evidence_id",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_HYDRO_RAIN",
        [
            "rain_id",
            "source_basis",
            "return_period_yr",
            "duration_hr",
            "rainfall_mm",
            "intensity_formula",
            "temporal_dist",
            "data_origin",
            "src_id",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "DRR_HYDRO_MODEL",
        [
            "hydro_id",
            "model",
            "cn_or_c",
            "tc_min",
            "k_storage",
            "peak_cms_before",
            "peak_cms_after",
            "vol_m3_before",
            "vol_m3_after",
            "critical_duration_hr",
            "data_origin",
            "evidence_id",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_SEDIMENT",
        [
            "sed_id",
            "method",
            "r_factor",
            "k_factor",
            "ls_factor",
            "c_factor",
            "p_factor",
            "soil_loss_t_ha_yr_before",
            "soil_loss_t_ha_yr_after",
            "data_origin",
            "evidence_id",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_SLOPE",
        [
            "slope_id",
            "has_slope_work",
            "slope_type",
            "height_m",
            "length_m",
            "risk_grade",
            "mitigation_ref",
            "exclude",
            "exclude_reason",
            "data_origin",
            "evidence_id",
        ],
    ),
    _SheetSpec(
        "DRR_MITIGATION",
        [
            "drr_mit_id",
            "hazard_type",
            "measure_type",
            "description",
            "location",
            "design_ref",
            "maintenance_ref",
            "data_origin",
            "src_id",
        ],
    ),
    _SheetSpec(
        "DRR_MAINTENANCE",
        [
            "mnt_id",
            "facility_name",
            "inspection_cycle",
            "maintenance_method",
            "responsible",
            "ledger_template",
            "evidence_id",
            "src_id",
        ],
    ),
]


def write_case_template_v2_xlsx(path: str | Path) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # LOOKUPS (data validation sources; column-based lists)
    ws_lists = wb.active
    ws_lists.title = "LOOKUPS"

    lists: dict[str, list[str]] = {
        "YES_NO": ["Y", "N"],
        "YN_NA": ["Y", "N", "NA"],
        "REQUIRE_DECISION": ["Y", "N", "판정필요"],
        "ENV_ITEM_STATUS": ["FOCUS", "BASELINE", "EXCLUDE"],
        "DOC_SCOPE": ["ENV", "DRR", "BOTH"],
        "DATA_ORIGIN": ["FIELD_SURVEY", "OFFICIAL_DB", "LITERATURE", "MODEL_OUTPUT", "CLIENT_PROVIDED"],
        "STAGE": ["CONSTRUCTION", "OPERATION", "BOTH"],
        "FIGURE_SOURCE_ORIGIN": ["OFFICIAL", "REFERENCE", "REF", "UNKNOWN"],
        "FIGURE_ASSET_ROLE": ["MAP", "DRAWING", "PHOTO", "PHOTO_SHEET", "CHART", "TABLE_CAPTURE", "OTHER"],
        "FIGURE_SOURCE_CLASS": ["USER_PROVIDED", "WMS", "WMTS", "API", "DERIVED"],
        "FIGURE_AUTHENTICITY": ["OFFICIAL", "REFERENCE"],
        "FIGURE_USAGE_SCOPE": ["DISPLAY_ONLY", "SUPPORTING", "DECISION_PROHIBITED"],
        "FIGURE_FALLBACK_MODE": ["PLACEHOLDER", "SKIP", "ERROR"],
        "FIGURE_TYPE": [
            "LOCATION_MAP",
            "INFLUENCE_AREA_MAP",
            "LANDUSE_PLAN",
            "AERIAL_PHOTO",
            "SITE_PLAN",
            "DRAINAGE_MAP",
            "STORMWATER_PLAN_MAP",
            "CATCHMENT_MAP",
            "ECO_ROUTE_MAP",
            "PHOTO_SHEET",
            "PHOTO",
            "SIMULATION",
            "CHART",
            "OTHER",
        ],
        "PROJECT_TYPE": ["관광농원", "관광지", "체육시설", "기타"],
        "PARTY_ROLE": ["사업자", "대행자", "작성자", "설계자", "시공사"],
        "DRR_HAZARD_TYPE": ["FLOOD", "INLAND", "SEDIMENT", "SLOPE", "ETC"],
        "REVIEW_LEVEL": ["적정", "보완", "미흡"],
        "SEASON": ["봄", "여름", "가을", "겨울"],
        "ECO_TAXON_GROUP": ["식물", "포유류", "조류", "양서·파충류", "어류", "저서성대형무척추", "곤충", "기타"],
        "ECO_EVIDENCE_TYPE": ["목격", "청음", "흔적", "채집", "문헌"],
        "WATER_SUPPLY": ["상수도", "지하수", "기타"],
        "OWNERSHIP": ["국유", "공유", "사유", "기타"],
        "SLOPE_TYPE": ["절토", "성토", "기존", "옹벽", "축대"],
        "SLOPE_EXCLUDE": ["Y", "N"],
        "HAZARD_OCCURRED": ["Y", "N", "UNKNOWN"],
        "INTERVIEW_DONE": ["Y", "N"],
        "EVIDENCE_TYPE": ["도면", "사진", "계산서", "측정원시자료", "회의록", "기타"],
        "SENSITIVE_FLAG": ["Y", "N"],
        "SURVEY_SCOPE": ["생태", "경관", "배수", "소음", "수질", "기타"],
        "DRR_CONCEPT": ["면적", "선", "점"],
        "DRR_AREA_PART": ["PROJECT", "UPSTREAM", "DOWNSTREAM", "SURROUNDING"],
        "LANDUSE_STAGE": ["BEFORE", "AFTER"],
        "IMPERVIOUS": ["Y", "N", "부분"],
        "INSPECTION_CYCLE": ["수시", "월1", "분기1", "반기1", "연1"],
        "ZONING_CATEGORY": ["NATURE", "AIR", "WATER", "SOCIAL", "DISASTER", "OTHER"],
        "OX_UNKNOWN": ["O", "X", "UNKNOWN"],
        "DATA_REQ_CONNECTOR": ["WMS", "KMA_ASOS", "AIRKOREA", "KOSIS", "AUTO_GIS", "GEOCODE", "NIER_WATER", "PDF_PAGE"],
        "DATA_REQ_PURPOSE": ["OVERLAY", "AIR_BASELINE", "DRR_RAINFALL", "SOCIO_STATS", "WATER_BASELINE", "EVIDENCE"],
        "DATA_REQ_MERGE": ["REPLACE_SHEET", "UPSERT_KEYS", "APPEND"],
        "DATA_REQ_RUN_MODE": ["AUTO", "ONCE", "NEVER"],
    }

    list_ranges: dict[str, str] = {}
    for idx, (name, values) in enumerate(lists.items(), start=1):
        col = get_column_letter(idx)
        ws_lists[f"{col}1"].value = name
        for r, v in enumerate(values, start=2):
            ws_lists[f"{col}{r}"].value = v
        list_ranges[name] = f"=LOOKUPS!${col}$2:${col}${len(values)+1}"
        ws_lists.column_dimensions[col].width = 24
    ws_lists.freeze_panes = "A2"

    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(vertical="top", wrap_text=True)

    def init_sheet(spec: _SheetSpec) -> None:
        ws = wb.create_sheet(spec.name)
        ws.append(spec.headers)
        for i, h in enumerate(spec.headers, start=1):
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(44, len(h) + 6))
        ws.freeze_panes = "A2"

    for s in _SHEETS_V2:
        init_sheet(s)

    # Drop-down validations
    def add_list_validation(sheet: str, header: str, list_name: str, max_rows: int = 500) -> None:
        ws = wb[sheet]
        header_row = [c.value for c in ws[1]]
        if header not in header_row:
            return
        col_idx = header_row.index(header) + 1
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=list_ranges[list_name], allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")

    add_list_validation("PARTIES", "role", "PARTY_ROLE")
    add_list_validation("PROJECT", "project_type", "PROJECT_TYPE")
    add_list_validation("PROJECT", "doc_env_required", "REQUIRE_DECISION")
    add_list_validation("PROJECT", "doc_drr_required", "REQUIRE_DECISION")
    add_list_validation("PROJECT", "water_supply", "WATER_SUPPLY")
    add_list_validation("PARCELS", "ownership", "OWNERSHIP")
    add_list_validation("ZONING_OVERLAY", "category", "ZONING_CATEGORY")
    add_list_validation("ZONING_OVERLAY", "is_applicable", "OX_UNKNOWN")
    add_list_validation("ZONING_OVERLAY", "data_origin", "DATA_ORIGIN")
    add_list_validation("PLAN_LANDUSE", "stage", "LANDUSE_STAGE")
    add_list_validation("PLAN_LANDUSE", "impervious", "IMPERVIOUS")
    add_list_validation("FIGURES", "doc_scope", "DOC_SCOPE")
    add_list_validation("FIGURES", "figure_type", "FIGURE_TYPE")
    add_list_validation("FIGURES", "source_origin", "FIGURE_SOURCE_ORIGIN")
    add_list_validation("FIGURES", "asset_role", "FIGURE_ASSET_ROLE")
    add_list_validation("FIGURES", "source_class", "FIGURE_SOURCE_CLASS")
    add_list_validation("FIGURES", "authenticity", "FIGURE_AUTHENTICITY")
    add_list_validation("FIGURES", "usage_scope", "FIGURE_USAGE_SCOPE")
    add_list_validation("FIGURES", "fallback_mode", "FIGURE_FALLBACK_MODE")
    add_list_validation("FIGURES", "sensitive", "SENSITIVE_FLAG")
    add_list_validation("ATTACHMENTS", "evidence_type", "EVIDENCE_TYPE")
    add_list_validation("ATTACHMENTS", "data_origin", "DATA_ORIGIN")
    add_list_validation("ATTACHMENTS", "sensitive", "SENSITIVE_FLAG")
    add_list_validation("DATA_REQUESTS", "connector", "DATA_REQ_CONNECTOR")
    add_list_validation("DATA_REQUESTS", "purpose", "DATA_REQ_PURPOSE")
    add_list_validation("DATA_REQUESTS", "merge_strategy", "DATA_REQ_MERGE")
    add_list_validation("DATA_REQUESTS", "run_mode", "DATA_REQ_RUN_MODE")
    add_list_validation("FIELD_SURVEY_LOG", "scope", "SURVEY_SCOPE")
    add_list_validation("FIELD_SURVEY_LOG", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_SCOPING", "status", "ENV_ITEM_STATUS")
    add_list_validation("ENV_SCOPING", "survey_needed", "YES_NO")
    add_list_validation("ENV_SCOPING", "prior_assessment_exists", "YES_NO")
    add_list_validation("ENV_SCOPING", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_BASE_AIR", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_BASE_SOCIO", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_BASE_WATER", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_BASE_NOISE", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_BASE_GEO", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_ECO_EVENTS", "season", "SEASON")
    add_list_validation("ENV_ECO_EVENTS", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_ECO_OBS", "taxon_group", "ECO_TAXON_GROUP")
    add_list_validation("ENV_ECO_OBS", "evidence_type", "ECO_EVIDENCE_TYPE")
    add_list_validation("ENV_ECO_OBS", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_LANDSCAPE", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_IMPACT_AIR", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_IMPACT_WATER", "data_origin", "DATA_ORIGIN")
    add_list_validation("ENV_MITIGATION", "stage", "STAGE")
    add_list_validation("DRR_SCOPING", "hazard_type", "DRR_HAZARD_TYPE")
    add_list_validation("DRR_SCOPING", "include", "YN_NA")
    add_list_validation("DRR_SCOPING", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_TARGET_AREA", "concept", "DRR_CONCEPT")
    add_list_validation("DRR_TARGET_AREA", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_TARGET_AREA_PARTS", "part", "DRR_AREA_PART")
    add_list_validation("DRR_TARGET_AREA_PARTS", "included", "YN_NA")
    add_list_validation("DRR_TARGET_AREA_PARTS", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_BASE_HAZARD", "hazard_type", "DRR_HAZARD_TYPE")
    add_list_validation("DRR_BASE_HAZARD", "occurred", "HAZARD_OCCURRED")
    add_list_validation("DRR_BASE_HAZARD", "interview_done", "INTERVIEW_DONE")
    add_list_validation("DRR_BASE_HAZARD", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_INTERVIEWS", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_HYDRO_RAIN", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_HYDRO_MODEL", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_SEDIMENT", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_SLOPE", "slope_type", "SLOPE_TYPE")
    add_list_validation("DRR_SLOPE", "exclude", "SLOPE_EXCLUDE")
    add_list_validation("DRR_SLOPE", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_MITIGATION", "hazard_type", "DRR_HAZARD_TYPE")
    add_list_validation("DRR_MITIGATION", "data_origin", "DATA_ORIGIN")
    add_list_validation("DRR_MAINTENANCE", "inspection_cycle", "INSPECTION_CYCLE")

    wb.save(out)
    return out
