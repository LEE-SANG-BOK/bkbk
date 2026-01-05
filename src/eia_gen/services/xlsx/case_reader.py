from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from eia_gen.models.case import Case
from eia_gen.services.canonicalize import canonicalize_case


def _split_ids(value: Any) -> list[str]:
    if value is None:
        return []
    s = str(value).strip()
    if not s:
        return []
    # allow ; or , as separators
    parts: list[str] = []
    for chunk in s.replace(",", ";").split(";"):
        t = chunk.strip()
        if t:
            parts.append(t)
    # de-dupe preserving order
    seen: set[str] = set()
    out: list[str] = []
    for p in parts:
        if p in seen:
            continue
        seen.add(p)
        out.append(p)
    return out


def _is_empty_row(values: list[Any]) -> bool:
    for v in values:
        if v is None:
            continue
        if isinstance(v, str) and not v.strip():
            continue
        return False
    return True


def _sheet_header_map(ws) -> dict[str, int]:
    header = [c.value for c in ws[1]]
    mapping: dict[str, int] = {}
    for idx, h in enumerate(header):
        if not h:
            continue
        mapping[str(h).strip()] = idx
    return mapping


def _get(ws, header_map: dict[str, int], row_values: list[Any], key: str) -> Any:
    idx = header_map.get(key)
    if idx is None or idx >= len(row_values):
        return None
    return row_values[idx]


def _tf(value: Any, src_ids: list[str] | None = None) -> dict[str, Any]:
    d: dict[str, Any] = {"t": "" if value is None else str(value).strip()}
    if src_ids:
        d["src"] = src_ids
    return d


def _qf(value: Any, unit: str, src_ids: list[str] | None = None) -> dict[str, Any]:
    v = None
    if value is None or (isinstance(value, str) and not value.strip()):
        v = None
    else:
        try:
            v = float(value)
        except Exception:
            v = None
    d: dict[str, Any] = {"v": v, "u": unit}
    if src_ids:
        d["src"] = src_ids
    return d


def _nested_set(obj: dict[str, Any], path: str, value: Any) -> None:
    """Set value into nested dict using dot-path (e.g., 'dust.method')."""
    cur = obj
    parts = [p for p in path.split(".") if p]
    for p in parts[:-1]:
        nxt = cur.get(p)
        if not isinstance(nxt, dict):
            nxt = {}
            cur[p] = nxt
        cur = nxt
    if parts:
        cur[parts[-1]] = value


def load_case_from_xlsx(path: str | Path) -> Case:
    """Load case.xlsx (defined in docs/03_case_xlsx_spec.md) into Case model."""
    xlsx = Path(path)
    wb = load_workbook(xlsx, data_only=True)

    # v2 (snake_case + LOOKUPS sheet) auto-detection
    if "LOOKUPS" in wb.sheetnames:
        from eia_gen.services.xlsx.case_reader_v2 import load_case_from_workbook_v2

        return load_case_from_workbook_v2(wb)

    data: dict[str, Any] = {}

    # META (optional)
    if "META" in wb.sheetnames:
        ws = wb["META"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        if not _is_empty_row(row):
            meta: dict[str, Any] = {}
            for key in ["template_version", "project_type", "report_type", "language"]:
                v = _get(ws, hm, row, key)
                if v is not None and str(v).strip():
                    meta[key] = str(v).strip()
            if meta:
                data["meta"] = meta

    # COVER (single record)
    if "COVER" in wb.sheetnames:
        ws = wb["COVER"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
        cover: dict[str, Any] = {
            "project_name": _tf(_get(ws, hm, row, "project_name"), src_ids),
            "submit_date": _tf(_get(ws, hm, row, "submit_date"), src_ids),
            "approving_authority": _tf(_get(ws, hm, row, "approving_authority"), src_ids),
            "consultation_agency": _tf(_get(ws, hm, row, "consultation_agency"), src_ids),
            "author_org": _tf(_get(ws, hm, row, "author_org"), src_ids),
            "client_name": _tf(_get(ws, hm, row, "client_name")),
            "proponent_name": _tf(_get(ws, hm, row, "proponent_name")),
        }
        data["cover"] = cover

    # SUMMARY_ISSUES / SUMMARY_MEASURES
    summary_inputs: dict[str, Any] = {}
    if "SUMMARY_ISSUES" in wb.sheetnames:
        ws = wb["SUMMARY_ISSUES"]
        hm = _sheet_header_map(ws)
        issues: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            issues.append(_tf(_get(ws, hm, row, "issue"), src_ids))
        if issues:
            summary_inputs["key_issues"] = issues

    if "SUMMARY_MEASURES" in wb.sheetnames:
        ws = wb["SUMMARY_MEASURES"]
        hm = _sheet_header_map(ws)
        measures: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            measures.append(_tf(_get(ws, hm, row, "measure"), src_ids))
        if measures:
            summary_inputs["key_measures"] = measures

    if summary_inputs:
        data["summary_inputs"] = summary_inputs

    # PROJECT (single record)
    project_overview: dict[str, Any] = {}
    if "PROJECT" in wb.sheetnames:
        ws = wb["PROJECT"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]

        address_src = _split_ids(_get(ws, hm, row, "address_src_ids"))
        coord_src = _split_ids(_get(ws, hm, row, "coord_src_ids"))

        project_overview["purpose_need"] = _tf(_get(ws, hm, row, "purpose_need"), address_src)
        project_overview["location"] = {
            "address": _tf(_get(ws, hm, row, "address"), address_src),
            "admin": {
                "sido": _tf(_get(ws, hm, row, "sido"), address_src),
                "sigungu": _tf(_get(ws, hm, row, "sigungu"), address_src),
                "eupmyeon": _tf(_get(ws, hm, row, "eupmyeon"), address_src),
            },
            "center_coord": {
                "epsg": 4326,
                "lat": _qf(_get(ws, hm, row, "lat"), "deg", coord_src),
                "lon": _qf(_get(ws, hm, row, "lon"), "deg", coord_src),
            },
        }

    # AREA (single record)
    if "AREA" in wb.sheetnames:
        ws = wb["AREA"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
        project_overview.setdefault("area", {})
        project_overview["area"]["total_area_m2"] = _qf(_get(ws, hm, row, "total_area_m2"), "m2", src_ids)

    # PARCELS
    if "PARCELS" in wb.sheetnames:
        ws = wb["PARCELS"]
        hm = _sheet_header_map(ws)
        parcels: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            parcels.append(
                {
                    "jibun": _tf(_get(ws, hm, row, "jibun"), src_ids),
                    "pnu": _tf(_get(ws, hm, row, "pnu"), src_ids),
                    "land_category": _tf(_get(ws, hm, row, "land_category"), src_ids),
                    "zoning": _tf(_get(ws, hm, row, "zoning"), src_ids),
                    "area_m2": _qf(_get(ws, hm, row, "area_m2"), "m2", src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        if parcels:
            project_overview.setdefault("area", {})
            project_overview["area"]["parcels"] = parcels

    # ZONING_BREAKDOWN (dict)
    if "ZONING_BREAKDOWN" in wb.sheetnames:
        ws = wb["ZONING_BREAKDOWN"]
        hm = _sheet_header_map(ws)
        zoning_area: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            zoning = _get(ws, hm, row, "zoning")
            if zoning is None or not str(zoning).strip():
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            zoning_area[str(zoning).strip()] = _qf(_get(ws, hm, row, "area_m2"), "m2", src_ids)
        if zoning_area:
            project_overview.setdefault("area", {})
            project_overview["area"]["zoning_area_m2"] = zoning_area

    # FACILITIES
    if "FACILITIES" in wb.sheetnames:
        ws = wb["FACILITIES"]
        hm = _sheet_header_map(ws)
        facilities: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            facilities.append(
                {
                    "category": _tf(_get(ws, hm, row, "category"), src_ids),
                    "name": _tf(_get(ws, hm, row, "name"), src_ids),
                    "qty": _qf(_get(ws, hm, row, "qty"), str(_get(ws, hm, row, "qty_unit") or "").strip(), src_ids),
                    "area_m2": _qf(_get(ws, hm, row, "area_m2"), "m2", src_ids),
                    "capacity_person": _qf(_get(ws, hm, row, "capacity_person"), "명", src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        if facilities:
            project_overview.setdefault("contents_scale", {})
            project_overview["contents_scale"]["facilities"] = facilities

    # SCHEDULE
    if "SCHEDULE" in wb.sheetnames:
        ws = wb["SCHEDULE"]
        hm = _sheet_header_map(ws)
        milestones: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            milestones.append(
                {
                    "phase": _tf(_get(ws, hm, row, "phase"), src_ids),
                    "start": _tf(_get(ws, hm, row, "start_ym"), src_ids),
                    "end": _tf(_get(ws, hm, row, "end_ym"), src_ids),
                }
            )
        if milestones:
            project_overview.setdefault("schedule", {})
            project_overview["schedule"]["milestones"] = milestones

    # PERMITS
    if "PERMITS" in wb.sheetnames:
        ws = wb["PERMITS"]
        hm = _sheet_header_map(ws)
        permit_list: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            permit_list.append(
                {
                    "name": _tf(_get(ws, hm, row, "name"), src_ids),
                    "status": _tf(_get(ws, hm, row, "status"), src_ids),
                    "authority": _tf(_get(ws, hm, row, "authority"), src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        if permit_list:
            project_overview.setdefault("legal_permits", {})
            project_overview["legal_permits"]["permit_list"] = permit_list

    if project_overview:
        data["project_overview"] = project_overview

    # SURVEY_PLAN
    if "SURVEY_PLAN" in wb.sheetnames:
        ws = wb["SURVEY_PLAN"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        radius_src = _split_ids(_get(ws, hm, row, "radius_src_ids"))
        just_src = _split_ids(_get(ws, hm, row, "justification_src_ids"))
        survey_plan = {
            "influence_area": {
                "radius_m": _qf(_get(ws, hm, row, "influence_radius_m"), "m", radius_src),
                "justification": _tf(_get(ws, hm, row, "justification"), just_src),
            },
            "methods": {
                "literature_db": _tf(_get(ws, hm, row, "method_literature_db")),
                "field_survey": _tf(_get(ws, hm, row, "method_field_survey")),
            },
        }
        data["survey_plan"] = survey_plan

    # SCOPING
    if "SCOPING" in wb.sheetnames:
        ws = wb["SCOPING"]
        hm = _sheet_header_map(ws)
        scoping: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_expected = _split_ids(_get(ws, hm, row, "src_expected"))
            scoping.append(
                {
                    "item_id": str(_get(ws, hm, row, "item_id") or "").strip(),
                    "item_name": str(_get(ws, hm, row, "item_name") or "").strip(),
                    "category": _tf(_get(ws, hm, row, "category")),
                    "exclude_reason": _tf(_get(ws, hm, row, "exclude_reason")),
                    "baseline_method": _tf(_get(ws, hm, row, "baseline_method")),
                    "prediction_method": _tf(_get(ws, hm, row, "prediction_method")),
                    "src_expected": src_expected,
                }
            )
        if scoping:
            data["scoping_matrix"] = scoping

    # BASELINE*
    baseline: dict[str, Any] = {}

    if "BASELINE_TOPO" in wb.sheetnames:
        ws = wb["BASELINE_TOPO"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
        baseline["topography_geology"] = {
            "elevation_range_m": _tf(_get(ws, hm, row, "elevation_range_m"), src_ids),
            "mean_slope_deg": _qf(_get(ws, hm, row, "mean_slope_deg"), "deg", src_ids),
            "geology_summary": _tf(_get(ws, hm, row, "geology_summary"), src_ids),
            "soil_summary": _tf(_get(ws, hm, row, "soil_summary"), src_ids),
        }

    # ECO dates / flora / fauna
    if "ECO_DATES" in wb.sheetnames:
        ws = wb["ECO_DATES"]
        hm = _sheet_header_map(ws)
        dates: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            dates.append(_tf(_get(ws, hm, row, "survey_date"), src_ids))
        baseline.setdefault("ecology", {})
        if dates:
            baseline["ecology"]["survey_dates"] = dates

    if "ECO_FLORA" in wb.sheetnames:
        ws = wb["ECO_FLORA"]
        hm = _sheet_header_map(ws)
        flora: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            flora.append(
                {
                    "species_ko": _tf(_get(ws, hm, row, "species_ko"), src_ids),
                    "scientific": _tf(_get(ws, hm, row, "scientific"), src_ids),
                    "protected": _tf(_get(ws, hm, row, "protected"), src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        baseline.setdefault("ecology", {})
        if flora:
            baseline["ecology"]["flora_list"] = flora

    if "ECO_FAUNA" in wb.sheetnames:
        ws = wb["ECO_FAUNA"]
        hm = _sheet_header_map(ws)
        fauna: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            fauna.append(
                {
                    "species_ko": _tf(_get(ws, hm, row, "species_ko"), src_ids),
                    "scientific": _tf(_get(ws, hm, row, "scientific"), src_ids),
                    "protected": _tf(_get(ws, hm, row, "protected"), src_ids),
                    "evidence": _tf(_get(ws, hm, row, "evidence"), src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        baseline.setdefault("ecology", {})
        if fauna:
            baseline["ecology"]["fauna_list"] = fauna

    # WATER_STREAMS + WATER_QUALITY
    if "WATER_STREAMS" in wb.sheetnames:
        ws = wb["WATER_STREAMS"]
        hm = _sheet_header_map(ws)
        streams: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            streams.append(
                {
                    "name": _tf(_get(ws, hm, row, "name"), src_ids),
                    "distance_m": _qf(_get(ws, hm, row, "distance_m"), "m", src_ids),
                    "flow_direction": _tf(_get(ws, hm, row, "flow_direction"), src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        baseline.setdefault("water_environment", {})
        if streams:
            baseline["water_environment"]["streams"] = streams

    if "WATER_QUALITY" in wb.sheetnames:
        ws = wb["WATER_QUALITY"]
        hm = _sheet_header_map(ws)
        wq: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            key = _get(ws, hm, row, "key")
            if key is None or not str(key).strip():
                continue
            k = str(key).strip()
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            unit = str(_get(ws, hm, row, "unit") or "").strip()
            val = _get(ws, hm, row, "value")
            if isinstance(val, (int, float)) and unit:
                wq[k] = _qf(val, unit, src_ids)
            elif isinstance(val, (int, float)):
                wq[k] = {"v": float(val), "src": src_ids}
                if unit:
                    wq[k]["u"] = unit
            else:
                wq[k] = _tf(val, src_ids)
        baseline.setdefault("water_environment", {})
        if wq:
            baseline["water_environment"]["water_quality"] = wq

    # AIR
    if "AIR" in wb.sheetnames:
        ws = wb["AIR"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
        baseline["air_quality"] = {
            "station_name": _tf(_get(ws, hm, row, "station_name"), src_ids),
            "pm10_ugm3": _qf(_get(ws, hm, row, "pm10_ugm3"), "µg/m3", src_ids),
            "pm25_ugm3": _qf(_get(ws, hm, row, "pm25_ugm3"), "µg/m3", src_ids),
            "ozone_ppm": _qf(_get(ws, hm, row, "ozone_ppm"), "ppm", src_ids),
        }

    # NOISE_RECEPTORS
    if "NOISE_RECEPTORS" in wb.sheetnames:
        ws = wb["NOISE_RECEPTORS"]
        hm = _sheet_header_map(ws)
        receptors: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            receptors.append(
                {
                    "name": _tf(_get(ws, hm, row, "name"), src_ids),
                    "distance_m": _qf(_get(ws, hm, row, "distance_m"), "m", src_ids),
                    "baseline_day_db": _qf(_get(ws, hm, row, "baseline_day_db"), "dB(A)", src_ids),
                    "baseline_night_db": _qf(_get(ws, hm, row, "baseline_night_db"), "dB(A)", src_ids),
                    "measured": _tf(_get(ws, hm, row, "measured"), src_ids),
                }
            )
        baseline.setdefault("noise_vibration", {})
        if receptors:
            baseline["noise_vibration"]["receptors"] = receptors

    # LANDUSE (single row)
    if "LANDUSE" in wb.sheetnames:
        ws = wb["LANDUSE"]
        hm = _sheet_header_map(ws)
        row = [c.value for c in ws[2]]
        src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
        baseline.setdefault("landuse_landscape", {})
        baseline["landuse_landscape"].update(
            {
                "current_landcover_summary": _tf(_get(ws, hm, row, "current_landcover_summary"), src_ids),
                "protected_areas_overlap": _tf(_get(ws, hm, row, "protected_areas_overlap"), src_ids),
            }
        )

    # VIEWPOINTS
    if "VIEWPOINTS" in wb.sheetnames:
        ws = wb["VIEWPOINTS"]
        hm = _sheet_header_map(ws)
        vps: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            vps.append(
                {
                    "vp_id": _tf(_get(ws, hm, row, "vp_id"), src_ids),
                    "location_desc": _tf(_get(ws, hm, row, "location_desc"), src_ids),
                    "photo_asset_id": _tf(_get(ws, hm, row, "photo_asset_id"), src_ids),
                    "note": _tf(_get(ws, hm, row, "note"), src_ids),
                }
            )
        baseline.setdefault("landuse_landscape", {})
        if vps:
            baseline["landuse_landscape"]["key_viewpoints"] = vps

    # POP_TRAFFIC
    if "POP_TRAFFIC" in wb.sheetnames:
        ws = wb["POP_TRAFFIC"]
        hm = _sheet_header_map(ws)
        nearest = None
        distance = None
        vehicles = None
        access_roads: list[dict[str, Any]] = []
        src_ids_last: list[str] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            src_ids_last = src_ids_last or src_ids
            if nearest is None:
                nearest = _get(ws, hm, row, "nearest_village")
            if distance is None:
                distance = _get(ws, hm, row, "distance_to_village_m")
            if vehicles is None:
                vehicles = _get(ws, hm, row, "expected_vehicles_per_day")
            road = _get(ws, hm, row, "access_road")
            if road is not None and str(road).strip():
                access_roads.append(_tf(road, src_ids))
        if nearest is not None or distance is not None or vehicles is not None or access_roads:
            baseline["population_traffic"] = {
                "nearest_village": _tf(nearest, src_ids_last),
                "distance_to_village_m": _qf(distance, "m", src_ids_last),
                "access_roads": access_roads,
                "expected_vehicles_per_day": _qf(vehicles, "대/일", src_ids_last),
            }

    if baseline:
        data["baseline"] = baseline

    # IMPACT_* (dict)
    impact: dict[str, Any] = {}
    if "IMPACT_CONS" in wb.sheetnames:
        ws = wb["IMPACT_CONS"]
        hm = _sheet_header_map(ws)
        cons: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            key = _get(ws, hm, row, "item_key")
            if key is None or not str(key).strip():
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            _nested_set(cons, str(key).strip(), _tf(_get(ws, hm, row, "text"), src_ids))
        if cons:
            impact["construction"] = cons

    if "IMPACT_OPER" in wb.sheetnames:
        ws = wb["IMPACT_OPER"]
        hm = _sheet_header_map(ws)
        oper: dict[str, Any] = {}
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            key = _get(ws, hm, row, "item_key")
            if key is None or not str(key).strip():
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            _nested_set(oper, str(key).strip(), _tf(_get(ws, hm, row, "text"), src_ids))
        if oper:
            impact["operation"] = oper

    if impact:
        data["impact_prediction"] = impact

    # MITIGATION
    if "MITIGATION" in wb.sheetnames:
        ws = wb["MITIGATION"]
        hm = _sheet_header_map(ws)
        measures: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            related = _split_ids(_get(ws, hm, row, "related_impacts"))
            measures.append(
                {
                    "measure_id": str(_get(ws, hm, row, "measure_id") or "").strip(),
                    "phase": _tf(_get(ws, hm, row, "phase"), src_ids),
                    "title": _tf(_get(ws, hm, row, "title"), src_ids),
                    "description": _tf(_get(ws, hm, row, "description"), src_ids),
                    "location_ref": _tf(_get(ws, hm, row, "location_ref"), src_ids),
                    "monitoring": _tf(_get(ws, hm, row, "monitoring"), src_ids),
                    "related_impacts": related,
                }
            )
        if measures:
            data["mitigation"] = {"measures": measures}

    # CONDITION_TRACKER
    if "CONDITION_TRACKER" in wb.sheetnames:
        ws = wb["CONDITION_TRACKER"]
        hm = _sheet_header_map(ws)
        items: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            items.append(
                {
                    "item": _tf(_get(ws, hm, row, "item"), src_ids),
                    "measure_id": _tf(_get(ws, hm, row, "measure_id"), src_ids),
                    "when": _tf(_get(ws, hm, row, "when"), src_ids),
                    "evidence": _tf(_get(ws, hm, row, "evidence"), src_ids),
                    "responsible": _tf(_get(ws, hm, row, "responsible")),
                }
            )
        if items:
            data["management_plan"] = {"implementation_register": items}

    # RESIDENT_OPINION
    if "RESIDENT_OPINION" in wb.sheetnames:
        ws = wb["RESIDENT_OPINION"]
        hm = _sheet_header_map(ws)
        applicable = None
        summary = None
        src_ids_last: list[str] = []
        responses: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            src_ids_last = src_ids_last or src_ids
            if applicable is None:
                applicable = _get(ws, hm, row, "applicable")
            if summary is None:
                summary = _get(ws, hm, row, "summary")
            resp = _get(ws, hm, row, "response")
            if resp is not None and str(resp).strip():
                responses.append(_tf(resp, src_ids))
        if applicable is not None or summary is not None or responses:
            data["resident_opinion"] = {
                "applicable": _tf(applicable, src_ids_last),
                "summary": _tf(summary, src_ids_last),
                "responses": responses,
            }

    # ASSETS
    if "ASSETS" in wb.sheetnames:
        ws = wb["ASSETS"]
        hm = _sheet_header_map(ws)
        assets: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            asset_id = str(_get(ws, hm, row, "asset_id") or "").strip()
            if not asset_id:
                continue
            asset_type = str(_get(ws, hm, row, "type") or "").strip()
            file_path = str(_get(ws, hm, row, "file_path") or "").strip()
            source_ids = _split_ids(_get(ws, hm, row, "source_ids"))
            asset: dict[str, Any] = {
                "asset_id": asset_id,
                "type": asset_type,
                "file_path": file_path,
                "caption": _tf(_get(ws, hm, row, "caption"), source_ids),
                "source_ids": source_ids,
                "viewpoint": _tf(_get(ws, hm, row, "viewpoint")),
                "shooting_date": _tf(_get(ws, hm, row, "shooting_date")),
            }
            scale = _get(ws, hm, row, "scale")
            if scale is not None and str(scale).strip():
                asset["scale"] = str(scale).strip()
            assets.append(asset)
        if assets:
            data["assets"] = assets

    # DIA_* (소규모 재해영향평가/재해영향성검토)
    disaster: dict[str, Any] = {}

    if "DIA_SCOPE" in wb.sheetnames:
        ws = wb["DIA_SCOPE"]
        hm = _sheet_header_map(ws)
        items: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            items.append(
                {
                    "hazard_item": _tf(_get(ws, hm, row, "hazard_item"), src_ids),
                    "applicable": _tf(_get(ws, hm, row, "applicable"), src_ids),
                    "analysis_level": _tf(_get(ws, hm, row, "analysis_level"), src_ids),
                    "exclude_reason": _tf(_get(ws, hm, row, "exclude_reason"), src_ids),
                }
            )
        if items:
            disaster["scoping_matrix"] = items

    if "DIA_RAINFALL" in wb.sheetnames:
        ws = wb["DIA_RAINFALL"]
        hm = _sheet_header_map(ws)
        rainfall: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            rainfall.append(
                {
                    "station_name": _tf(_get(ws, hm, row, "station_name"), src_ids),
                    "duration_min": _qf(_get(ws, hm, row, "duration_min"), "min", src_ids),
                    "frequency_year": _qf(_get(ws, hm, row, "frequency_year"), "yr", src_ids),
                    "rainfall_mm": _qf(_get(ws, hm, row, "rainfall_mm"), "mm", src_ids),
                    "source_type": _tf(_get(ws, hm, row, "source_type"), src_ids),
                }
            )
        if rainfall:
            disaster["rainfall"] = rainfall

    if "DIA_RUNOFF" in wb.sheetnames:
        ws = wb["DIA_RUNOFF"]
        hm = _sheet_header_map(ws)
        basins: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            basins.append(
                {
                    "basin_id": _tf(_get(ws, hm, row, "basin_id"), src_ids),
                    "basin_area_km2": _qf(_get(ws, hm, row, "basin_area_km2"), "km2", src_ids),
                    "tc_min": _qf(_get(ws, hm, row, "tc_min"), "min", src_ids),
                    "cn_value": _qf(_get(ws, hm, row, "cn_value"), "", src_ids),
                    "pre_peak_cms": _qf(_get(ws, hm, row, "pre_peak_cms"), "m3/s", src_ids),
                    "post_peak_cms": _qf(_get(ws, hm, row, "post_peak_cms"), "m3/s", src_ids),
                    "delta_peak_cms": _qf(_get(ws, hm, row, "delta_peak_cms"), "m3/s", src_ids),
                    "model": _tf(_get(ws, hm, row, "model"), src_ids),
                }
            )
        if basins:
            disaster["runoff_basins"] = basins

    if "DIA_DRAINAGE" in wb.sheetnames:
        ws = wb["DIA_DRAINAGE"]
        hm = _sheet_header_map(ws)
        drainage: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            drainage.append(
                {
                    "facility_id": _tf(_get(ws, hm, row, "facility_id"), src_ids),
                    "type": _tf(_get(ws, hm, row, "type"), src_ids),
                    "size_desc": _tf(_get(ws, hm, row, "size_desc"), src_ids),
                    "capacity": _tf(_get(ws, hm, row, "capacity"), src_ids),
                    "discharge_to": _tf(_get(ws, hm, row, "discharge_to"), src_ids),
                    "maintenance_class": _tf(_get(ws, hm, row, "maintenance_class"), src_ids),
                }
            )
        if drainage:
            disaster["drainage_facilities"] = drainage

    if "DIA_MEASURES" in wb.sheetnames:
        ws = wb["DIA_MEASURES"]
        hm = _sheet_header_map(ws)
        measures: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            measures.append(
                {
                    "measure_id": _tf(_get(ws, hm, row, "measure_id"), src_ids),
                    "target_hazard": _tf(_get(ws, hm, row, "target_hazard"), src_ids),
                    "stage": _tf(_get(ws, hm, row, "stage"), src_ids),
                    "description": _tf(_get(ws, hm, row, "description"), src_ids),
                    "linked_facility_id": _tf(_get(ws, hm, row, "linked_facility_id"), src_ids),
                }
            )
        if measures:
            disaster["measures"] = measures

    if "DIA_MAINTENANCE" in wb.sheetnames:
        ws = wb["DIA_MAINTENANCE"]
        hm = _sheet_header_map(ws)
        ledger: list[dict[str, Any]] = []
        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if _is_empty_row(row):
                continue
            src_ids = _split_ids(_get(ws, hm, row, "src_ids"))
            ledger.append(
                {
                    "asset_id": _tf(_get(ws, hm, row, "asset_id"), src_ids),
                    "inspection_cycle": _tf(_get(ws, hm, row, "inspection_cycle"), src_ids),
                    "inspection_item": _tf(_get(ws, hm, row, "inspection_item"), src_ids),
                    "responsible_role": _tf(_get(ws, hm, row, "responsible_role"), src_ids),
                    "record_format": _tf(_get(ws, hm, row, "record_format"), src_ids),
                    "evidence_id_template": _tf(_get(ws, hm, row, "evidence_id_template"), src_ids),
                }
            )
        if ledger:
            disaster["maintenance_ledger"] = ledger

    if disaster:
        data["disaster"] = disaster

    case = Case.model_validate(data)
    case = canonicalize_case(case)
    return case
