from __future__ import annotations

import tempfile
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl

from eia_gen.services.xlsx.case_template_v2 import write_case_template_v2_xlsx


@dataclass(frozen=True)
class XlsxUpgradeReport:
    input_path: Path
    output_path: Path
    added_sheets: list[str]
    preserved_extra_sheets: list[str]
    added_columns_by_sheet: dict[str, list[str]]


def _norm_header(v: object) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    return s


def _is_empty_cell(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _row_is_empty(values: tuple[Any, ...]) -> bool:
    return all(_is_empty_cell(v) for v in values)


def upgrade_case_xlsx_v2(
    *,
    xlsx_in: Path,
    xlsx_out: Path,
    preserve_unknown_sheets: bool = True,
    preserve_unknown_columns: bool = True,
) -> XlsxUpgradeReport:
    """Upgrade an existing case.xlsx(v2) to the latest template schema.

    Strategy:
    - Create a fresh template workbook from `case_template_v2.py` (keeps validations/styles).
    - Copy cell values from the input workbook into the template workbook by header name.
    - Optionally preserve unknown sheets/columns to avoid data loss.
    """

    xlsx_in = xlsx_in.expanduser().resolve()
    xlsx_out = xlsx_out.expanduser().resolve()

    if not xlsx_in.exists():
        raise FileNotFoundError(xlsx_in)

    with tempfile.TemporaryDirectory(prefix="eia_gen_xlsx_upgrade_") as td:
        template_path = (Path(td) / "case_template.v2.xlsx").resolve()
        write_case_template_v2_xlsx(template_path)

        src_wb = openpyxl.load_workbook(xlsx_in, data_only=False)
        tpl_wb = openpyxl.load_workbook(template_path, data_only=False)

        # Prefer template LOOKUPS for stable validations.
        skip_copy_sheets = {"LOOKUPS"}

        added_columns_by_sheet: dict[str, list[str]] = {}

        tpl_sheetnames = set(tpl_wb.sheetnames)
        src_sheetnames = list(src_wb.sheetnames)

        for sheet_name in src_sheetnames:
            if sheet_name in skip_copy_sheets:
                continue
            if sheet_name not in tpl_sheetnames:
                continue

            src_ws = src_wb[sheet_name]
            tpl_ws = tpl_wb[sheet_name]

            src_headers_raw = [c.value for c in src_ws[1]]
            tpl_headers_raw = [c.value for c in tpl_ws[1]]

            src_headers = [_norm_header(h) for h in src_headers_raw]
            tpl_headers = [_norm_header(h) for h in tpl_headers_raw]

            src_hm = {h: i for i, h in enumerate(src_headers, start=1) if h}
            tpl_hm = {h: i for i, h in enumerate(tpl_headers, start=1) if h}

            added_cols: list[str] = []
            if preserve_unknown_columns:
                for h in src_headers:
                    if not h or h in tpl_hm:
                        continue
                    added_cols.append(h)

                if added_cols:
                    base_style_cell = tpl_ws.cell(row=1, column=1)
                    for h in added_cols:
                        tpl_headers.append(h)
                        col_idx = len(tpl_headers)
                        c = tpl_ws.cell(row=1, column=col_idx)
                        c.value = h
                        # Keep header styling consistent (best-effort).
                        c.font = base_style_cell.font
                        c.fill = base_style_cell.fill
                        c.alignment = base_style_cell.alignment
                        tpl_hm[h] = col_idx
                    added_columns_by_sheet[sheet_name] = list(added_cols)

            # Copy values row-by-row.
            max_src_col = max(1, len(src_headers))
            max_dst_col = max(1, len(tpl_headers))

            # Find last non-empty row in source (avoid bloating the file).
            last_nonempty_row = 1
            for ridx, row in enumerate(
                src_ws.iter_rows(min_row=2, max_col=max_src_col, values_only=True), start=2
            ):
                if row and (not _row_is_empty(tuple(row))):
                    last_nonempty_row = ridx

            if last_nonempty_row < 2:
                continue

            for ridx, row in enumerate(
                src_ws.iter_rows(min_row=2, max_row=last_nonempty_row, values_only=True),
                start=2,
            ):
                row = tuple(row or ())
                if not row:
                    continue

                for dst_col in range(1, max_dst_col + 1):
                    header = tpl_headers[dst_col - 1] if (dst_col - 1) < len(tpl_headers) else ""
                    if not header:
                        continue
                    src_col = src_hm.get(header)
                    if not src_col:
                        continue
                    v = row[src_col - 1] if (src_col - 1) < len(row) else None
                    if v is None:
                        continue
                    tpl_ws.cell(row=ridx, column=dst_col).value = v

        preserved_extra_sheets: list[str] = []
        if preserve_unknown_sheets:
            for sheet_name in src_sheetnames:
                if sheet_name in tpl_sheetnames:
                    continue
                preserved_extra_sheets.append(sheet_name)
                src_ws = src_wb[sheet_name]
                new_ws = tpl_wb.create_sheet(sheet_name)
                for row in src_ws.iter_rows(values_only=True):
                    new_ws.append(list(row or []))

        # Identify sheets added by the template (missing in input).
        added_sheets = [s for s in tpl_wb.sheetnames if s not in src_wb.sheetnames]

        xlsx_out.parent.mkdir(parents=True, exist_ok=True)
        tpl_wb.save(xlsx_out)

    return XlsxUpgradeReport(
        input_path=xlsx_in,
        output_path=xlsx_out,
        added_sheets=added_sheets,
        preserved_extra_sheets=preserved_extra_sheets,
        added_columns_by_sheet=added_columns_by_sheet,
    )
