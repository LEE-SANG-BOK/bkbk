from __future__ import annotations

import json
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _norm_header(v: Any) -> str:
    s = str(v or "").strip()
    return s


_META_COLS = {
    "src_id",
    "src_ids",
    "note",
    "notes",
    "sensitive",
    "review_flag",
    "qa_flag",
    "att_id",
    "att_ids",
}


def _is_id_col(name: str) -> bool:
    if not name:
        return False
    if name in {"row_id", "case_id", "req_id", "evidence_id", "fig_id", "fac_id", "bldg_id"}:
        return True
    if name.endswith("_id"):
        return True
    return False


_PATH_EXTS = {
    ".pdf",
    ".png",
    ".jpg",
    ".jpeg",
    ".webp",
    ".tif",
    ".tiff",
    ".geojson",
    ".json",
    ".shp",
    ".gpkg",
    ".xlsx",
    ".docx",
    ".txt",
}


def _looks_like_path_value(v: Any) -> bool:
    if _is_empty(v):
        return False
    s = str(v).strip()
    if not s:
        return False
    if "/" in s or "\\" in s:
        return True
    sl = s.lower()
    return any(sl.endswith(ext) for ext in _PATH_EXTS)


def _looks_like_path_col(name: str) -> bool:
    n = name.lower()
    if n in {"file_path", "boundary_file", "geometry_file"}:
        return True
    if n.endswith("_file") or n.endswith("_path"):
        return True
    return False


@dataclass(frozen=True)
class SheetFillStats:
    sheet: str
    row_count: int

    data_columns: int
    data_total_cells: int
    data_filled_cells: int
    data_fill_ratio: float

    id_columns: int
    id_total_cells: int
    id_filled_cells: int
    id_fill_ratio: float

    src_id_empty_rows: int
    src_id_tbd_rows: int


@dataclass(frozen=True)
class XlsxStatus:
    xlsx_path: str
    version: str
    sheet_stats: list[SheetFillStats]

    total_rows: int
    total_data_fill_ratio: float
    total_id_fill_ratio: float

    src_id_empty_rows: int
    src_id_tbd_rows: int

    referenced_files: int
    missing_files: list[str]

    validation_eia: dict[str, Any] | None = None
    validation_dia: dict[str, Any] | None = None

    def to_dict(self) -> dict[str, Any]:
        return {
            **asdict(self),
            "sheet_stats": [asdict(s) for s in self.sheet_stats],
        }


def compute_xlsx_status(xlsx: Path) -> XlsxStatus:
    xlsx = Path(xlsx)
    wb = load_workbook(xlsx, data_only=False)

    version = "v2" if "LOOKUPS" in wb.sheetnames else "v1"
    base_dir = xlsx.parent.resolve()

    sheet_stats: list[SheetFillStats] = []

    total_rows = 0
    total_data_cells = 0
    total_data_filled = 0
    total_id_cells = 0
    total_id_filled = 0
    src_id_empty_rows = 0
    src_id_tbd_rows = 0

    referenced_files = 0
    missing_files: list[str] = []

    for sheet_name in wb.sheetnames:
        if sheet_name == "LOOKUPS":
            continue

        ws = wb[sheet_name]
        if ws.max_row < 2:
            # header-only sheet
            sheet_stats.append(
                SheetFillStats(
                    sheet=sheet_name,
                    row_count=0,
                    data_columns=0,
                    data_total_cells=0,
                    data_filled_cells=0,
                    data_fill_ratio=0.0,
                    id_columns=0,
                    id_total_cells=0,
                    id_filled_cells=0,
                    id_fill_ratio=0.0,
                    src_id_empty_rows=0,
                    src_id_tbd_rows=0,
                )
            )
            continue

        headers = [_norm_header(c.value) for c in ws[1]]
        col_idx = {h: i for i, h in enumerate(headers) if h}

        id_cols = [i for i, h in enumerate(headers) if h and _is_id_col(h) and h not in _META_COLS]
        meta_cols = [i for i, h in enumerate(headers) if h and h in _META_COLS]
        data_cols = [
            i
            for i, h in enumerate(headers)
            if h and (i not in id_cols) and (i not in meta_cols) and h not in _META_COLS
        ]

        src_col = col_idx.get("src_id") or col_idx.get("src_ids")

        row_count = 0
        data_cells = 0
        data_filled = 0
        id_cells = 0
        id_filled = 0
        src_empty = 0
        src_tbd = 0

        path_cols = [i for i, h in enumerate(headers) if h and _looks_like_path_col(h)]

        for r in ws.iter_rows(min_row=2, values_only=True):
            row = list(r)
            if all(_is_empty(v) for v in row):
                continue

            row_count += 1

            # data
            data_cells += len(data_cols)
            for i in data_cols:
                if i < len(row) and not _is_empty(row[i]):
                    data_filled += 1

            # ids
            id_cells += len(id_cols)
            for i in id_cols:
                if i < len(row) and not _is_empty(row[i]):
                    id_filled += 1

            # src
            if src_col is not None and src_col < len(row):
                sv = row[src_col]
                if _is_empty(sv):
                    src_empty += 1
                else:
                    s = str(sv).upper()
                    if "TBD" in s or "미정" in s:
                        src_tbd += 1

            # file refs
            for i in path_cols:
                if i >= len(row):
                    continue
                pv = row[i]
                if _is_empty(pv):
                    continue
                if not _looks_like_path_value(pv):
                    continue
                p = Path(str(pv).strip())
                if not p.is_absolute():
                    p = (base_dir / p).resolve()
                referenced_files += 1
                if not p.exists():
                    missing_files.append(f"{sheet_name}.{headers[i]} -> {p}")

        total_rows += row_count
        total_data_cells += data_cells
        total_data_filled += data_filled
        total_id_cells += id_cells
        total_id_filled += id_filled
        src_id_empty_rows += src_empty
        src_id_tbd_rows += src_tbd

        data_ratio = float(data_filled) / float(data_cells) if data_cells else 0.0
        id_ratio = float(id_filled) / float(id_cells) if id_cells else 0.0
        sheet_stats.append(
            SheetFillStats(
                sheet=sheet_name,
                row_count=row_count,
                data_columns=len(data_cols),
                data_total_cells=data_cells,
                data_filled_cells=data_filled,
                data_fill_ratio=round(data_ratio, 4),
                id_columns=len(id_cols),
                id_total_cells=id_cells,
                id_filled_cells=id_filled,
                id_fill_ratio=round(id_ratio, 4),
                src_id_empty_rows=src_empty,
                src_id_tbd_rows=src_tbd,
            )
        )

    total_data_ratio = float(total_data_filled) / float(total_data_cells) if total_data_cells else 0.0
    total_id_ratio = float(total_id_filled) / float(total_id_cells) if total_id_cells else 0.0

    # Optional: attach validation stats when present (same folder).
    validation_eia = None
    validation_dia = None
    eia_path = base_dir / "validation_report_eia.json"
    dia_path = base_dir / "validation_report_dia.json"
    try:
        if eia_path.exists():
            validation_eia = json.loads(eia_path.read_text(encoding="utf-8"))
    except Exception:
        validation_eia = None
    try:
        if dia_path.exists():
            validation_dia = json.loads(dia_path.read_text(encoding="utf-8"))
    except Exception:
        validation_dia = None

    return XlsxStatus(
        xlsx_path=str(xlsx),
        version=version,
        sheet_stats=sheet_stats,
        total_rows=total_rows,
        total_data_fill_ratio=round(total_data_ratio, 4),
        total_id_fill_ratio=round(total_id_ratio, 4),
        src_id_empty_rows=src_id_empty_rows,
        src_id_tbd_rows=src_id_tbd_rows,
        referenced_files=referenced_files,
        missing_files=sorted(set(missing_files)),
        validation_eia=validation_eia,
        validation_dia=validation_dia,
    )
