from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation


@dataclass(frozen=True)
class _SheetSpec:
    name: str
    headers: list[str]


_SHEETS: list[_SheetSpec] = [
    _SheetSpec("META", ["template_version", "project_type", "report_type", "language"]),
    _SheetSpec(
        "COVER",
        [
            "project_name",
            "submit_date",
            "approving_authority",
            "consultation_agency",
            "author_org",
            "client_name",
            "proponent_name",
            "src_ids",
        ],
    ),
    _SheetSpec("SUMMARY_ISSUES", ["issue", "src_ids"]),
    _SheetSpec("SUMMARY_MEASURES", ["measure", "src_ids"]),
    _SheetSpec(
        "PROJECT",
        [
            "purpose_need",
            "address",
            "sido",
            "sigungu",
            "eupmyeon",
            "lat",
            "lon",
            "coord_src_ids",
            "address_src_ids",
        ],
    ),
    _SheetSpec("AREA", ["total_area_m2", "src_ids"]),
    _SheetSpec("PARCELS", ["jibun", "pnu", "land_category", "zoning", "area_m2", "note", "src_ids"]),
    _SheetSpec("ZONING_BREAKDOWN", ["zoning", "area_m2", "src_ids"]),
    _SheetSpec(
        "FACILITIES",
        ["category", "name", "qty", "qty_unit", "area_m2", "capacity_person", "note", "src_ids"],
    ),
    _SheetSpec("SCHEDULE", ["phase", "start_ym", "end_ym", "src_ids"]),
    _SheetSpec("PERMITS", ["name", "status", "authority", "note", "src_ids"]),
    _SheetSpec(
        "SURVEY_PLAN",
        [
            "influence_radius_m",
            "radius_src_ids",
            "justification",
            "justification_src_ids",
            "method_literature_db",
            "method_field_survey",
        ],
    ),
    _SheetSpec(
        "SCOPING",
        [
            "item_id",
            "item_name",
            "category",
            "exclude_reason",
            "baseline_method",
            "prediction_method",
            "src_expected",
        ],
    ),
    _SheetSpec("BASELINE_TOPO", ["elevation_range_m", "mean_slope_deg", "geology_summary", "soil_summary", "src_ids"]),
    _SheetSpec("ECO_DATES", ["survey_date", "src_ids"]),
    _SheetSpec("ECO_FLORA", ["species_ko", "scientific", "protected", "note", "src_ids"]),
    _SheetSpec("ECO_FAUNA", ["species_ko", "scientific", "protected", "evidence", "note", "src_ids"]),
    _SheetSpec("WATER_STREAMS", ["name", "distance_m", "flow_direction", "note", "src_ids"]),
    _SheetSpec("WATER_QUALITY", ["key", "value", "unit", "src_ids"]),
    _SheetSpec("AIR", ["station_name", "pm10_ugm3", "pm25_ugm3", "ozone_ppm", "src_ids"]),
    _SheetSpec(
        "NOISE_RECEPTORS",
        ["name", "distance_m", "baseline_day_db", "baseline_night_db", "measured", "src_ids"],
    ),
    _SheetSpec("LANDUSE", ["current_landcover_summary", "protected_areas_overlap", "src_ids"]),
    _SheetSpec("VIEWPOINTS", ["vp_id", "location_desc", "photo_asset_id", "note", "src_ids"]),
    _SheetSpec(
        "POP_TRAFFIC",
        ["nearest_village", "distance_to_village_m", "access_road", "expected_vehicles_per_day", "src_ids"],
    ),
    _SheetSpec("IMPACT_CONS", ["item_key", "text", "src_ids"]),
    _SheetSpec("IMPACT_OPER", ["item_key", "text", "src_ids"]),
    _SheetSpec(
        "MITIGATION",
        ["measure_id", "phase", "title", "description", "location_ref", "monitoring", "related_impacts", "src_ids"],
    ),
    _SheetSpec(
        "CONDITION_TRACKER", ["item", "measure_id", "when", "evidence", "responsible", "src_ids"]
    ),
    _SheetSpec("RESIDENT_OPINION", ["applicable", "summary", "response", "src_ids"]),
    _SheetSpec(
        "ASSETS",
        ["asset_id", "type", "file_path", "caption", "source_ids", "viewpoint", "shooting_date", "scale"],
    ),
    # DIA (소규모 재해영향평가/재해영향성검토) 입력 시트(옵션)
    _SheetSpec("DIA_SCOPE", ["hazard_item", "applicable", "analysis_level", "exclude_reason", "src_ids"]),
    _SheetSpec(
        "DIA_RAINFALL",
        ["station_name", "duration_min", "frequency_year", "rainfall_mm", "source_type", "src_ids"],
    ),
    _SheetSpec(
        "DIA_RUNOFF",
        [
            "basin_id",
            "basin_area_km2",
            "tc_min",
            "cn_value",
            "pre_peak_cms",
            "post_peak_cms",
            "delta_peak_cms",
            "model",
            "src_ids",
        ],
    ),
    _SheetSpec(
        "DIA_DRAINAGE",
        [
            "facility_id",
            "type",
            "size_desc",
            "capacity",
            "discharge_to",
            "maintenance_class",
            "src_ids",
        ],
    ),
    _SheetSpec(
        "DIA_MEASURES",
        ["measure_id", "target_hazard", "stage", "description", "linked_facility_id", "src_ids"],
    ),
    _SheetSpec(
        "DIA_MAINTENANCE",
        [
            "asset_id",
            "inspection_cycle",
            "inspection_item",
            "responsible_role",
            "record_format",
            "evidence_id_template",
            "src_ids",
        ],
    ),
]


def write_case_template_xlsx(path: str | Path) -> Path:
    out = Path(path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # LISTS (data validation sources)
    ws_lists = wb.active
    ws_lists.title = "LISTS"

    lists: dict[str, list[str]] = {
        "YES_NO": ["true", "false"],
        "SCOPING_CATEGORY": ["중점", "현황", "제외"],
        "PHASE": ["공사", "운영"],
        "ASSET_TYPE": [
            "location_map",
            "influence_area_map",
            "layout_plan",
            "landuse_plan",
            "aerial_photo",
            "drainage_map",
            "stormwater_plan_map",
            "eco_route_map",
            "photo_sheet",
            "photo",
            "simulation",
            # DIA
            "dia_target_area_map",
        ],
        "DIA_APPLICABLE": ["YES", "NO", "UNKNOWN"],
        "DIA_ANALYSIS_LEVEL": ["SIMPLIFIED", "STANDARD"],
        "DIA_STAGE": ["CONSTRUCTION", "OPERATION", "BOTH"],
        "DIA_INSPECTION_CYCLE": ["WEEKLY", "MONTHLY", "QUARTERLY", "YEARLY", "AFTER_RAIN"],
    }

    list_ranges: dict[str, str] = {}
    for idx, (name, values) in enumerate(lists.items(), start=1):
        col = get_column_letter(idx)
        ws_lists[f"{col}1"].value = name
        for r, v in enumerate(values, start=2):
            ws_lists[f"{col}{r}"].value = v
        list_ranges[name] = f"=LISTS!${col}$2:${col}${len(values)+1}"
        ws_lists.column_dimensions[col].width = 22

    ws_lists.freeze_panes = "A2"

    # Common header style
    header_font = Font(bold=True)
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    header_align = Alignment(vertical="top", wrap_text=True)

    def init_sheet(spec: _SheetSpec) -> None:
        ws = wb.create_sheet(spec.name)
        ws.append(spec.headers)
        for i, h in enumerate(spec.headers, start=1):
            cell = ws.cell(row=1, column=i)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_align
            ws.column_dimensions[get_column_letter(i)].width = max(14, min(42, len(h) + 6))
        ws.freeze_panes = "A2"

    for s in _SHEETS:
        init_sheet(s)

    # Drop-down validations (apply to a reasonable range; users can copy/paste down)
    def add_list_validation(sheet: str, header: str, list_name: str, max_rows: int = 200) -> None:
        ws = wb[sheet]
        try:
            col_idx = ws[1].index(next(c for c in ws[1] if c.value == header)) + 1
        except StopIteration:
            return
        col_letter = get_column_letter(col_idx)
        dv = DataValidation(type="list", formula1=list_ranges[list_name], allow_blank=True)
        ws.add_data_validation(dv)
        dv.add(f"{col_letter}2:{col_letter}{max_rows}")

    add_list_validation("SCOPING", "category", "SCOPING_CATEGORY")
    add_list_validation("MITIGATION", "phase", "PHASE")
    add_list_validation("ASSETS", "type", "ASSET_TYPE")
    add_list_validation("RESIDENT_OPINION", "applicable", "YES_NO")
    add_list_validation("NOISE_RECEPTORS", "measured", "YES_NO")
    add_list_validation("DIA_SCOPE", "applicable", "DIA_APPLICABLE")
    add_list_validation("DIA_SCOPE", "analysis_level", "DIA_ANALYSIS_LEVEL")
    add_list_validation("DIA_MEASURES", "stage", "DIA_STAGE")
    add_list_validation("DIA_MAINTENANCE", "inspection_cycle", "DIA_INSPECTION_CYCLE")

    wb.save(out)
    return out
