from __future__ import annotations

import json
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook as openpyxl_load_workbook
from openpyxl.workbook.workbook import Workbook

from eia_gen.services.data_requests.models import DataRequest, Evidence, now_iso


DATA_REQUESTS_SHEET = "DATA_REQUESTS"
ATTACHMENTS_SHEET = "ATTACHMENTS"
LOCATION_SHEET = "LOCATION"


def _is_empty(v: Any) -> bool:
    if v is None:
        return True
    if isinstance(v, str) and not v.strip():
        return True
    return False


def _is_empty_row(row: list[Any]) -> bool:
    return all(_is_empty(v) for v in row)


def _header_map(ws) -> dict[str, int]:
    header = [c.value for c in ws[1]]
    mapping: dict[str, int] = {}
    for idx, h in enumerate(header):
        if h is None:
            continue
        s = str(h).strip()
        if s:
            mapping[s] = idx
    return mapping


def _get(hm: dict[str, int], row: list[Any], key: str) -> Any:
    idx = hm.get(key)
    if idx is None or idx >= len(row):
        return None
    return row[idx]


def _parse_bool(v: Any, default: bool = False) -> bool:
    if isinstance(v, bool):
        return v
    s = str(v or "").strip().upper()
    if s in {"TRUE", "T", "Y", "YES", "1"}:
        return True
    if s in {"FALSE", "F", "N", "NO", "0"}:
        return False
    return default


def _parse_int(v: Any, default: int) -> int:
    if v is None or (isinstance(v, str) and not v.strip()):
        return default
    try:
        return int(v)
    except Exception:
        try:
            return int(float(v))
        except Exception:
            return default


def _parse_list(v: Any) -> list[str]:
    s = str(v or "").strip()
    if not s:
        return []
    return [t.strip() for t in s.split(";") if t.strip()]


def read_sheet_dicts(wb: Workbook, sheet_name: str) -> list[dict[str, Any]]:
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    hm = _header_map(ws)
    if not hm:
        return []
    out: list[dict[str, Any]] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue
        d: dict[str, Any] = {}
        for k, idx in hm.items():
            if idx < len(row):
                d[k] = row[idx]
        out.append(d)
    return out


def apply_rows_to_sheet(
    wb: Workbook,
    *,
    sheet_name: str,
    rows: list[dict[str, Any]],
    merge_strategy: str,
    upsert_keys: list[str] | None = None,
) -> list[str]:
    """Apply row dicts into an existing sheet without destroying validations/styles."""
    warnings: list[str] = []
    if sheet_name not in wb.sheetnames:
        wb.create_sheet(sheet_name)
    ws = wb[sheet_name]

    hm = _header_map(ws)
    if not hm:
        # Create header from keys (best-effort).
        headers: list[str] = []
        for r in rows:
            for k in r.keys():
                if k not in headers:
                    headers.append(k)
        if not headers:
            return warnings
        # `ws.append()` appends after the current max_row, which can leave an empty row 1
        # (and thus an empty header map) when the sheet exists but its header row was cleared.
        # Prefer writing the header into row 1 in that case.
        row1 = [c.value for c in ws[1]]
        if _is_empty_row(list(row1)):
            for col, h in enumerate(headers, start=1):
                ws.cell(row=1, column=col).value = h
        else:
            ws.append(headers)
        hm = _header_map(ws)

    def _make_row_values(d: dict[str, Any]) -> list[Any]:
        out = [None] * max(len(hm), 1)
        for k, idx in hm.items():
            if k in d and idx < len(out):
                out[idx] = d[k]
        return out

    strategy = (merge_strategy or "REPLACE_SHEET").strip().upper()
    if strategy == "REPLACE_SHEET":
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        for d in rows:
            ws.append(_make_row_values(d))
        return warnings

    if strategy == "APPEND":
        for d in rows:
            ws.append(_make_row_values(d))
        return warnings

    if strategy == "UPSERT_KEYS":
        keys = [k.strip() for k in (upsert_keys or []) if k.strip()]
        if not keys:
            return [f"UPSERT_KEYS requires upsert_keys for sheet {sheet_name}"]
        missing = [k for k in keys if k not in hm]
        if missing:
            return [f"UPSERT_KEYS keys not found in sheet {sheet_name}: {', '.join(missing)}"]

        key_cols = [hm[k] for k in keys]
        index: dict[tuple[str, ...], int] = {}
        for ridx, r in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
            vals = []
            for cidx in key_cols:
                if cidx >= len(r):
                    vals.append("")
                else:
                    vals.append(str(r[cidx].value or "").strip())
            if any(not v for v in vals):
                continue
            index[tuple(vals)] = ridx

        for d in rows:
            kvals = [str(d.get(k) or "").strip() for k in keys]
            if any(not v for v in kvals):
                warnings.append(f"UPSERT skipped row with missing keys in {sheet_name}: {kvals}")
                continue
            ktuple = tuple(kvals)
            ridx = index.get(ktuple)
            if ridx is None:
                ws.append(_make_row_values(d))
                index[ktuple] = ws.max_row
                continue
            # Update existing row in place
            for k, idx in hm.items():
                if k not in d:
                    continue
                ws.cell(row=ridx, column=idx + 1).value = d[k]

        return warnings

    return [f"Unknown merge_strategy: {merge_strategy}"]


def load_workbook(path: str | Path) -> Workbook:
    return openpyxl_load_workbook(Path(path), data_only=False)


def read_data_requests(wb: Workbook) -> list[DataRequest]:
    if DATA_REQUESTS_SHEET not in wb.sheetnames:
        return []
    ws = wb[DATA_REQUESTS_SHEET]
    hm = _header_map(ws)
    out: list[DataRequest] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        row = list(r)
        if _is_empty_row(row):
            continue

        req_id = str(_get(hm, row, "req_id") or "").strip()
        if not req_id:
            continue

        params_json = str(_get(hm, row, "params_json") or "").strip()
        try:
            params = json.loads(params_json) if params_json else {}
        except Exception:
            params = {}

        out.append(
            DataRequest(
                req_id=req_id,
                enabled=_parse_bool(_get(hm, row, "enabled"), default=True),
                priority=_parse_int(_get(hm, row, "priority"), default=100),
                connector=str(_get(hm, row, "connector") or "").strip(),
                purpose=str(_get(hm, row, "purpose") or "").strip(),
                src_id=str(_get(hm, row, "src_id") or "").strip(),
                params_json=params_json,
                params=params,
                output_sheet=str(_get(hm, row, "output_sheet") or "").strip(),
                merge_strategy=str(_get(hm, row, "merge_strategy") or "").strip(),
                upsert_keys=_parse_list(_get(hm, row, "upsert_keys")),
                run_mode=str(_get(hm, row, "run_mode") or "AUTO").strip() or "AUTO",
                last_run_at=str(_get(hm, row, "last_run_at") or "").strip(),
                last_evidence_ids=_parse_list(_get(hm, row, "last_evidence_ids")),
                note=str(_get(hm, row, "note") or "").strip(),
            )
        )

    out.sort(key=lambda x: (x.priority, x.req_id))
    return out


def write_data_requests(wb: Workbook, requests: list[DataRequest]) -> None:
    if DATA_REQUESTS_SHEET not in wb.sheetnames:
        wb.create_sheet(DATA_REQUESTS_SHEET)
    ws = wb[DATA_REQUESTS_SHEET]

    # Ensure headers exist; if not, create from spec ordering.
    headers = [
        "req_id",
        "enabled",
        "priority",
        "connector",
        "purpose",
        "src_id",
        "params_json",
        "output_sheet",
        "merge_strategy",
        "upsert_keys",
        "run_mode",
        "last_run_at",
        "last_evidence_ids",
        "note",
    ]
    existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    if [str(x).strip() for x in existing if x is not None] != headers:
        ws.delete_rows(1, ws.max_row)
        ws.append(headers)
    else:
        # Clear body
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)

    for req in requests:
        ws.append(
            [
                req.req_id,
                bool(req.enabled),
                int(req.priority),
                req.connector,
                req.purpose,
                req.src_id,
                req.params_json,
                req.output_sheet,
                req.merge_strategy,
                ";".join(req.upsert_keys),
                req.run_mode,
                req.last_run_at,
                ";".join(req.last_evidence_ids),
                req.note,
            ]
        )


def update_request_run(wb: Workbook, *, req_id: str, evidence_ids: list[str]) -> None:
    if DATA_REQUESTS_SHEET not in wb.sheetnames:
        return
    ws = wb[DATA_REQUESTS_SHEET]
    hm = _header_map(ws)
    if "req_id" not in hm:
        return
    # Find row by req_id
    for ridx, r in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
        cell = r[hm["req_id"]]
        if str(cell.value or "").strip() != req_id:
            continue
        if "last_run_at" in hm:
            r[hm["last_run_at"]].value = now_iso()
        if "last_evidence_ids" in hm:
            r[hm["last_evidence_ids"]].value = ";".join(evidence_ids)
        return


def append_attachment(wb: Workbook, ev: Evidence) -> None:
    ws, hm = _ensure_attachments_sheet_and_headers(wb)

    def _col(name: str) -> int:
        return hm.get(name, -1)

    row = [None] * max(len(hm), 10)
    mapping = {
        "evidence_id": ev.evidence_id,
        "evidence_type": ev.evidence_type,
        "title": ev.title,
        "file_path": ev.file_path,
        "related_fig_id": "",
        "used_in": ev.used_in,
        "data_origin": ev.data_origin,
        "src_id": ev.src_id,
        "sensitive": "N",
        "note": ev.note,
    }
    for k, v in mapping.items():
        idx = _col(k)
        if idx >= 0:
            if idx >= len(row):
                row.extend([None] * (idx - len(row) + 1))
            row[idx] = v

    ws.append(row)


def upsert_attachment_by_used_in(wb: Workbook, ev: Evidence) -> None:
    """Upsert an ATTACHMENTS row by `used_in` value.

    This is useful for ONCE-style DATA_REQUESTS where re-runs should update/replace
    the previous evidence row instead of appending duplicates.
    """
    ws, hm = _ensure_attachments_sheet_and_headers(wb)

    used_in = str(ev.used_in or "").strip()
    if not used_in:
        append_attachment(wb, ev)
        return

    target_row_idx: int | None = None
    col_used_in = hm.get("used_in")
    if col_used_in is not None:
        for ridx in range(2, ws.max_row + 1):
            cell = ws.cell(row=ridx, column=col_used_in + 1)
            if str(cell.value or "").strip() == used_in:
                target_row_idx = ridx
                break

    mapping = {
        "evidence_id": ev.evidence_id,
        "evidence_type": ev.evidence_type,
        "title": ev.title,
        "file_path": ev.file_path,
        "related_fig_id": "",
        "used_in": ev.used_in,
        "data_origin": ev.data_origin,
        "src_id": ev.src_id,
        "sensitive": "N",
        "note": ev.note,
    }

    if target_row_idx is None:
        append_attachment(wb, ev)
        return

    for k, v in mapping.items():
        idx = hm.get(k)
        if idx is None:
            continue
        ws.cell(row=target_row_idx, column=idx + 1).value = v


def _ensure_attachments_sheet_and_headers(wb: Workbook):
    if ATTACHMENTS_SHEET not in wb.sheetnames:
        wb.create_sheet(ATTACHMENTS_SHEET)
    ws = wb[ATTACHMENTS_SHEET]
    hm = _header_map(ws)

    # Ensure headers exist (minimal)
    if not hm:
        ws.append([])

    # Ensure expected headers exist; older v2 templates may have partial columns.
    expected = [
        "evidence_id",
        "evidence_type",
        "title",
        "file_path",
        "related_fig_id",
        "used_in",
        "data_origin",
        "src_id",
        "sensitive",
        "note",
    ]

    existing = [c.value for c in ws[1]] if ws.max_row >= 1 else []
    existing_norm = [str(v).strip() for v in existing if v is not None and str(v).strip()]
    if not existing_norm:
        for i, h in enumerate(expected, start=1):
            ws.cell(row=1, column=i).value = h
    else:
        last_col = ws.max_column
        hm = _header_map(ws)
        for h in expected:
            if h in hm:
                continue
            last_col += 1
            ws.cell(row=1, column=last_col).value = h

    return ws, _header_map(ws)


def save_workbook(wb: Workbook, path: str | Path) -> None:
    wb.save(Path(path))


def read_location_hint(wb: Workbook) -> dict[str, Any]:
    """Read minimal bbox/coord hints from LOCATION sheet (v2)."""
    if LOCATION_SHEET not in wb.sheetnames:
        return {}
    ws = wb[LOCATION_SHEET]
    hm = _header_map(ws)
    row = [c.value for c in ws[2]]
    if _is_empty_row(row):
        return {}

    def _float(x: Any) -> float | None:
        if x is None or (isinstance(x, str) and not x.strip()):
            return None
        try:
            return float(x)
        except Exception:
            return None

    crs_raw = str(_get(hm, row, "crs") or "").strip()
    epsg = 4326
    if crs_raw:
        s = crs_raw.upper().strip()
        if s.startswith("EPSG:"):
            s = s.split("EPSG:", 1)[1].strip()
        try:
            epsg = int(s)
        except Exception:
            epsg = 4326

    return {
        "center_lat": _float(_get(hm, row, "center_lat")),
        "center_lon": _float(_get(hm, row, "center_lon")),
        "epsg": epsg,
        "boundary_file": str(_get(hm, row, "boundary_file") or "").strip(),
        "bbox_wkt": str(_get(hm, row, "bbox_wkt") or "").strip(),
    }
