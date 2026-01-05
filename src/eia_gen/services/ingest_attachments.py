from __future__ import annotations

import hashlib
import json
import shutil
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.workbook.workbook import Workbook

from eia_gen.services.data_requests.xlsx_io import apply_rows_to_sheet, load_workbook, save_workbook


def _now_iso() -> str:
    return datetime.now().isoformat(timespec="seconds")


def _sha256_file(p: Path) -> str:
    h = hashlib.sha256()
    with p.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _is_file_candidate(p: Path) -> bool:
    if not p.is_file():
        return False
    if p.name.startswith("."):
        return False
    return True


def _infer_evidence_type(path: Path) -> str:
    ext = path.suffix.lower()
    if ext in {".jpg", ".jpeg", ".png", ".webp", ".tif", ".tiff"}:
        return "사진"
    if ext in {".xlsx", ".csv"}:
        return "계산서"
    if ext in {".pdf"}:
        # could be drawings or reports; default to 기타
        return "기타"
    return "기타"


def _infer_sensitive(path: Path) -> str:
    # Conservative default: not sensitive.
    return "N"


@dataclass(frozen=True)
class Ingested:
    evidence_id: str
    src_path: str
    dst_path: str
    sha256: str


def _next_attachment_id(existing: list[str]) -> str:
    # Generate ATT-0001 style ids.
    mx = 0
    for eid in existing:
        s = str(eid or "")
        if not s.startswith("ATT-"):
            continue
        tail = s.replace("ATT-", "", 1)
        try:
            n = int(tail)
        except Exception:
            continue
        mx = max(mx, n)
    return f"ATT-{mx + 1:04d}"


def _read_existing_evidence_ids(wb: Workbook) -> list[str]:
    if "ATTACHMENTS" not in wb.sheetnames:
        return []
    ws = wb["ATTACHMENTS"]
    headers = [c.value for c in ws[1]]
    if "evidence_id" not in headers:
        return []
    idx = headers.index("evidence_id")
    out: list[str] = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or idx >= len(r):
            continue
        v = r[idx]
        if v is None:
            continue
        s = str(v).strip()
        if s:
            out.append(s)
    return out


def ingest_inbox(
    *,
    xlsx: Path,
    inbox_dir: Path,
    normalized_dir: Path,
    src_id: str,
    data_origin: str = "CLIENT_PROVIDED",
    evidence_type: str | None = None,
    move: bool = True,
) -> dict[str, Any]:
    """Ingest files from inbox -> normalized and register them in ATTACHMENTS sheet."""
    xlsx = xlsx.resolve()
    case_dir = xlsx.parent.resolve()

    inbox_dir = inbox_dir.resolve()
    normalized_dir = normalized_dir.resolve()
    normalized_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx)
    existing_ids = _read_existing_evidence_ids(wb)

    ingested: list[Ingested] = []
    manifest_entries: list[dict[str, Any]] = []

    for p in sorted(inbox_dir.glob("*")):
        if not _is_file_candidate(p):
            continue

        sha = _sha256_file(p)
        # de-dup by sha256 in manifest (best-effort) would be better, but keep simple.
        evidence_id = _next_attachment_id(existing_ids + [x.evidence_id for x in ingested])

        dst_name = f"{evidence_id}__{p.name}"
        dst_path = normalized_dir / dst_name

        if move:
            shutil.move(str(p), str(dst_path))
        else:
            shutil.copy2(str(p), str(dst_path))

        rel_dst = dst_path.relative_to(case_dir)

        ingested.append(Ingested(evidence_id=evidence_id, src_path=str(p), dst_path=str(rel_dst), sha256=sha))
        manifest_entries.append(
            {
                "evidence_id": evidence_id,
                "src_path": str(p),
                "dst_path": str(rel_dst).replace("\\", "/"),
                "sha256": sha,
                "ingested_at": _now_iso(),
            }
        )

    if not ingested:
        return {"ingested": 0, "items": []}

    rows: list[dict[str, Any]] = []
    for it in ingested:
        p = Path(it.dst_path)
        et = evidence_type or _infer_evidence_type(p)
        rows.append(
            {
                "evidence_id": it.evidence_id,
                "evidence_type": et,
                "title": p.name,
                "file_path": it.dst_path.replace("\\", "/"),
                "related_fig_id": "",
                "used_in": "INGEST",
                "data_origin": data_origin,
                "src_id": src_id,
                "sensitive": _infer_sensitive(p),
                "note": f"sha256={it.sha256}",
            }
        )

    apply_rows_to_sheet(
        wb,
        sheet_name="ATTACHMENTS",
        rows=rows,
        merge_strategy="UPSERT_KEYS",
        upsert_keys=["evidence_id"],
    )

    save_workbook(wb, xlsx)

    manifest_path = case_dir / "attachments" / "attachments_manifest.json"
    try:
        existing = json.loads(manifest_path.read_text(encoding="utf-8")) if manifest_path.exists() else {}
    except Exception:
        existing = {}
    if not isinstance(existing, dict):
        existing = {}
    existing.setdefault("created_at", _now_iso())
    existing["updated_at"] = _now_iso()
    existing.setdefault("items", [])
    if isinstance(existing.get("items"), list):
        existing["items"].extend(manifest_entries)
    else:
        existing["items"] = manifest_entries

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(existing, ensure_ascii=False, indent=2), encoding="utf-8")

    return {"ingested": len(ingested), "items": [it.__dict__ for it in ingested], "manifest": str(manifest_path)}
