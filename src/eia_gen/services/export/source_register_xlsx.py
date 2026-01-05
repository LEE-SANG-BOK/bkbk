from __future__ import annotations

import io
import json
import re
from collections import Counter, defaultdict
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel

from eia_gen.models.case import Case
from eia_gen.models.fields import QuantityField, TextField
from eia_gen.models.sources import SourceRegistry
from eia_gen.services.draft import ReportDraft


_CITATION_BLOCK_RE = re.compile(r"〔([^〕]+)〕")
_CLAIM_STRIP_CITATIONS_RE = re.compile(r"\s*〔[^〕]+〕\s*$")
_SHA1_RE = re.compile(r"\bsha1=([0-9a-fA-F]{40})\b")
_SHA256_RE = re.compile(r"\bsha256=([0-9a-fA-F]{64})\b")


def _normalize_ids(ids: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for s in ids:
        s2 = (s or "").strip()
        if not s2:
            continue
        if s2 in {"SRC-TBD", "S-TBD"}:
            continue
        if s2 in seen:
            continue
        seen.add(s2)
        out.append(s2)
    return out


def _extract_citation_ids(text: str) -> list[str]:
    ids: list[str] = []
    for m in _CITATION_BLOCK_RE.finditer(text or ""):
        inside = m.group(1)
        for raw in inside.split(","):
            token = raw.strip()
            if not token:
                continue
            if token.upper().startswith("SRC:"):
                token = token.split(":", 1)[1].strip()
            ids.append(token)
    return _normalize_ids(ids)


def _strip_trailing_citations(text: str) -> str:
    return _CLAIM_STRIP_CITATIONS_RE.sub("", text or "").strip()


def _split_joined_ids(value: str) -> list[str]:
    # supports "A, B" and "A;B" styles in XLSX-exported columns.
    if not value:
        return []
    raw = [p.strip() for p in re.split(r"[;,]", value) if p.strip()]
    return _normalize_ids(raw)


def _parse_doc_target(section_id: str) -> tuple[str, str]:
    # When generating EIA+DIA together we prefix SectionDraft.section_id with "EIA:"/"DIA:".
    if ":" in section_id:
        head, rest = section_id.split(":", 1)
        if head in {"EIA", "DIA"}:
            return head, rest
    return "", section_id


def _walk_source_ids(obj: Any):
    # Prefer concrete field types first.
    if isinstance(obj, TextField):
        for s in obj.src or []:
            yield s
        return
    if isinstance(obj, QuantityField):
        for s in obj.src or []:
            yield s
        return

    # Pydantic models
    if isinstance(obj, BaseModel):
        for k in obj.__class__.model_fields:
            yield from _walk_source_ids(getattr(obj, k))
        extra = getattr(obj, "model_extra", None) or {}
        if isinstance(extra, dict):
            for v in extra.values():
                yield from _walk_source_ids(v)
        return

    if isinstance(obj, dict):
        # dict-style field: {"t": "...", "src":[...]} or {"v":..., "src":[...]}
        src = obj.get("src")
        if isinstance(src, list):
            for s in src:
                yield s
        elif isinstance(src, str):
            yield src
        for v in obj.values():
            yield from _walk_source_ids(v)
        return

    if isinstance(obj, list):
        for v in obj:
            yield from _walk_source_ids(v)
        return


def _extract_hash_from_note(note: Any) -> str:
    s = str(note or "").strip()
    if not s:
        return ""
    if s.startswith("{") and s.endswith("}"):
        try:
            obj = json.loads(s)
        except Exception:
            obj = None
        if isinstance(obj, dict):
            for k in ("hash_sha1", "sha1", "out_sha1", "out_sha256", "sha256", "file_sha256", "file_sha1"):
                v = obj.get(k)
                if isinstance(v, str) and v.strip():
                    return v.strip()
    m = _SHA1_RE.search(s)
    if m:
        return m.group(1)
    m = _SHA256_RE.search(s)
    if m:
        return m.group(1)
    return ""


def _apply_header_style(ws) -> None:
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")
        cell.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def _set_column_widths(ws, widths: dict[int, int]) -> None:
    for idx, w in widths.items():
        ws.column_dimensions[get_column_letter(idx)].width = w


def build_source_register_xlsx_bytes(
    case: Case,
    sources: SourceRegistry,
    draft: ReportDraft | None = None,
    *,
    validation_reports: list[tuple[str, Any]] | None = None,
    report_tag: str | None = None,
    draft_section_id_allowlist: set[str] | None = None,
) -> bytes:
    """Build `source_register.xlsx` bytes.

    - Always includes all entries from `sources.yaml`.
    - Adds usage info from case/draft (best-effort).
    - Also emits canonical 4-sheet registry (P4-3) while keeping legacy sheets.
    """

    # Usage from case fields
    case_ids = _normalize_ids([str(x) for x in _walk_source_ids(case)])

    # Usage from draft citations (by section)
    #
    # NOTE: In some workflows, a spec may include extra sections (e.g., SSOT reuse blocks)
    # that are not present in the actual template used for rendering. In that case, those
    # sections should not contribute "used source IDs" to the source register.
    # `draft_section_id_allowlist` is an optional escape hatch computed by the caller
    # (typically based on template anchors present in the DOCX).
    used_in_sections: dict[str, set[str]] = defaultdict(set)
    citation_counts: Counter[str] = Counter()
    if draft is not None:
        for sec in draft.sections:
            if draft_section_id_allowlist is not None and sec.section_id not in draft_section_id_allowlist:
                continue
            for p in sec.paragraphs:
                for cid in _extract_citation_ids(p):
                    used_in_sections[cid].add(sec.section_id)
                    citation_counts[cid] += 1
            # Best-effort include tables/figures if present in draft
            for t in getattr(sec, "tables", []) or []:
                for cid in _normalize_ids(getattr(t, "source_ids", []) or []):
                    used_in_sections[cid].add(sec.section_id)
                    citation_counts[cid] += 1
            for f in getattr(sec, "figures", []) or []:
                for cid in _normalize_ids(getattr(f, "source_ids", []) or []):
                    used_in_sections[cid].add(sec.section_id)
                    citation_counts[cid] += 1

    used_ids = set(case_ids) | set(used_in_sections.keys())
    registered_ids = {s.id for s in sources.sources}
    missing_ids = sorted(used_ids - registered_ids)

    wb = Workbook()
    ws = wb.active
    ws.title = "Source Register"

    header = [
        "Source ID",
        "유형",
        "자료명",
        "제공기관",
        "기준일",
        "기간",
        "측정소/지점",
        "이격거리(km)",
        "파일/링크",
        "비고",
        "사용여부",
        "인용횟수",
        "사용처(섹션)",
    ]
    ws.append(header)
    _apply_header_style(ws)

    def add_row(
        source_id: str,
        type_: str,
        title: str,
        publisher: str,
        date: str,
        period: str,
        station: str,
        dist_km: str,
        file_or_url: str,
        note: str,
    ) -> None:
        used = "Y" if source_id in used_ids else ""
        count = citation_counts.get(source_id, 0)
        sections = ", ".join(sorted(used_in_sections.get(source_id, set())))
        ws.append(
            [
                source_id,
                type_,
                title,
                publisher,
                date,
                period,
                station,
                dist_km,
                file_or_url,
                note,
                used,
                count,
                sections,
            ]
        )

    for s in sources.sources:
        add_row(
            source_id=s.id,
            type_=s.type or "",
            title=s.title or "",
            publisher=s.publisher or "",
            date=s.date or "",
            period=s.period or "",
            station=s.station_name or "",
            dist_km=str(s.station_distance_km) if s.station_distance_km is not None else "",
            file_or_url=s.file_path or s.file_or_url or "",
            note=s.note or s.notes or "",
        )

    # If case/draft referenced unknown IDs, append them for quick fix.
    for mid in missing_ids:
        add_row(
            source_id=mid,
            type_="",
            title="【미등록】sources.yaml에 Source ID 추가 필요",
            publisher="",
            date="",
            period="",
            station="",
            dist_km="",
            file_or_url="",
            note="",
        )

    # Basic column widths (best-effort)
    widths = {
        1: 14,
        2: 10,
        3: 34,
        4: 18,
        5: 12,
        6: 16,
        7: 18,
        8: 14,
        9: 30,
        10: 28,
        11: 8,
        12: 8,
        13: 30,
    }
    _set_column_widths(ws, widths)

    # Wrap text for long columns
    wrap_cols = {3, 9, 10, 13}
    for row in ws.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            if i in wrap_cols:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top")

    # Evidence Register (best-effort): attachments/figures manifest for auditability.
    ev = wb.create_sheet("Evidence Register")
    ev_header = [
        "evidence_or_fig_id",
        "kind",
        "title_or_caption",
        "file_path",
        "related_fig_id",
        "used_in",
        "data_origin",
        "sensitive",
        "source_ids",
        "note",
    ]
    ev.append(ev_header)
    _apply_header_style(ev)

    def _tfv(x: Any) -> str:
        if isinstance(x, dict):
            if "t" in x:
                return str(x.get("t") or "").strip()
            if "v" in x:
                return str(x.get("v") or "").strip()
        t = getattr(x, "t", None)
        if isinstance(t, str):
            return t.strip()
        v = getattr(x, "v", None)
        if v is not None:
            return str(v).strip()
        return str(x or "").strip()

    def _row_src_ids(row: Any) -> str:
        ids: list[str] = []

        # v2 ATTACHMENTS rows are plain dicts with `src_id` (not field-style `src`).
        if isinstance(row, dict):
            sid = str(row.get("src_id") or "").strip()
            if sid:
                ids.append(sid)
            # Support semi-normalized joined style too.
            ids.extend(_split_joined_ids(str(row.get("src_ids") or "")))

        ids.extend([str(s) for s in _walk_source_ids(row)])
        return ";".join(_normalize_ids(ids))

    extra = case.model_extra or {}
    attachments_manifest = extra.get("attachments_manifest")
    derived_manifest = extra.get("derived_evidence_manifest")
    if isinstance(attachments_manifest, list):
        for r in attachments_manifest:
            if not isinstance(r, dict):
                continue
            ev.append(
                [
                    _tfv(r.get("evidence_id")),
                    _tfv(r.get("evidence_type")),
                    _tfv(r.get("title")),
                    _tfv(r.get("file_path")),
                    _tfv(r.get("related_fig_id")),
                    _tfv(r.get("used_in")),
                    _tfv(r.get("data_origin")),
                    _tfv(r.get("sensitive")),
                    _row_src_ids(r),
                    _tfv(r.get("note")),
                ]
            )

    # Engine-produced artifacts (e.g., materialized figure images) recorded during this run.
    if isinstance(derived_manifest, list):
        for r in derived_manifest:
            if not isinstance(r, dict):
                continue
            ev.append(
                [
                    _tfv(r.get("evidence_id")),
                    _tfv(r.get("evidence_type")),
                    _tfv(r.get("title")),
                    _tfv(r.get("file_path")),
                    _tfv(r.get("related_fig_id")),
                    _tfv(r.get("used_in")),
                    _tfv(r.get("data_origin")),
                    _tfv(r.get("sensitive")),
                    _row_src_ids(r),
                    _tfv(r.get("note")),
                ]
            )

    # Figures from `case.assets` (FIGURES sheet in v2)
    if getattr(case, "assets", None):
        for a in case.assets:
            try:
                ev.append(
                    [
                        getattr(a, "asset_id", ""),
                        getattr(a, "type", ""),
                        _tfv(getattr(a, "caption", None)),
                        getattr(a, "file_path", ""),
                        "",
                        _tfv(getattr(a, "insert_anchor", None)),
                        "",
                        "Y" if getattr(a, "sensitive", False) else "N",
                        ";".join(_normalize_ids([str(s) for s in getattr(a, "source_ids", []) or []])),
                        "",
                    ]
                )
            except Exception:
                continue

    widths_ev = {1: 18, 2: 14, 3: 34, 4: 44, 5: 16, 6: 18, 7: 14, 8: 10, 9: 28, 10: 28}
    _set_column_widths(ev, widths_ev)
    wrap_cols_ev = {3, 4, 6, 9, 10}
    for row in ev.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols_ev))

    # Figure→evidence mapping for claim-level traceability (best-effort).
    fig_to_evidence_ids: dict[str, list[str]] = defaultdict(list)

    def _add_fig_evidence(related_fig_id: str, evidence_id: str) -> None:
        rid = (related_fig_id or "").strip()
        eid = (evidence_id or "").strip()
        if not rid or not eid:
            return
        if eid in fig_to_evidence_ids[rid]:
            return
        fig_to_evidence_ids[rid].append(eid)

    for manifest in (attachments_manifest, derived_manifest):
        if not isinstance(manifest, list):
            continue
        for r in manifest:
            if not isinstance(r, dict):
                continue
            _add_fig_evidence(_tfv(r.get("related_fig_id")), _tfv(r.get("evidence_id")))

    # Claim-level view (best-effort): paragraph/table/figure → source_id 연결
    if draft is not None:
        ws2 = wb.create_sheet("Claims")
        header2 = [
            "claim_id",
            "doc_target",
            "section_id",
            "section_title",
            "report_anchor",
            "object_type",
            "field_path",
            "value",
            "unit",
            "computed",
            "computation_note",
            "method",
            "source_id",
            "evidence_id",
            "source_locator",
            "retrieved_at",
            "confidence",
            "qa_status",
            "qa_note",
            "reviewer_note",
        ]
        ws2.append(header2)
        _apply_header_style(ws2)

        def add_claim_row(
            claim_id: str,
            doc_target: str,
            section_id: str,
            section_title: str,
            report_anchor: str,
            object_type: str,
            field_path: str,
            value: str,
            src_ids: list[str],
            qa_status: str,
            qa_note: str = "",
            evidence_id: str = "",
        ) -> None:
            ws2.append(
                [
                    claim_id,
                    doc_target,
                    section_id,
                    section_title,
                    report_anchor,
                    object_type,
                    field_path,
                    value,
                    "",  # unit
                    "",  # computed
                    "",  # computation_note
                    "",  # method
                    ";".join(src_ids),
                    evidence_id,
                    "",  # source_locator
                    "",  # retrieved_at
                    "",  # confidence
                    qa_status,
                    qa_note,
                    "",  # reviewer_note
                ]
            )

        for sec in draft.sections:
            if draft_section_id_allowlist is not None and sec.section_id not in draft_section_id_allowlist:
                continue
            doc_target, raw_section_id = _parse_doc_target(sec.section_id)

            # Paragraph-level claims
            for i, p in enumerate(sec.paragraphs, start=1):
                src_ids = _extract_citation_ids(p)
                qa_status = "pass" if src_ids else "fail"
                qa_note = "" if src_ids else "missing source_id (or only SRC-TBD)"
                add_claim_row(
                    claim_id=f"{sec.section_id}:P{i:04d}",
                    doc_target=doc_target,
                    section_id=raw_section_id,
                    section_title=sec.title,
                    report_anchor=raw_section_id,
                    object_type="sentence",
                    field_path=f"{raw_section_id}.paragraphs[{i}]",
                    value=_strip_trailing_citations(p),
                    src_ids=src_ids,
                    evidence_id="",
                    qa_status=qa_status,
                    qa_note=qa_note,
                )

            # Table-level claims (row granularity when SRC column exists)
            for t in getattr(sec, "tables", []) or []:
                table_id = getattr(t, "table_id", "") or ""
                src_col = None
                if getattr(t, "headers", None):
                    try:
                        if str(t.headers[-1]).strip().upper() == "SRC":
                            src_col = len(t.headers) - 1
                    except Exception:
                        src_col = None

                rows = getattr(t, "rows", []) or []
                for r_idx, row in enumerate(rows, start=1):
                    row_src_ids: list[str] = []
                    if src_col is not None and src_col < len(row):
                        row_src_ids = _split_joined_ids(str(row[src_col]))
                    if not row_src_ids:
                        row_src_ids = _normalize_ids(getattr(t, "source_ids", []) or [])
                    qa_status = "pass" if row_src_ids else "fail"
                    qa_note = "" if row_src_ids else "missing source_id (or only S-TBD)"

                    row_value_cells = [str(v) for j, v in enumerate(row) if j != src_col]
                    add_claim_row(
                        claim_id=f"{sec.section_id}:T:{table_id}:R{r_idx:04d}",
                        doc_target=doc_target,
                        section_id=raw_section_id,
                        section_title=sec.title,
                        report_anchor=table_id,
                        object_type="table_cell",
                        field_path=f"{table_id}.rows[{r_idx}]",
                        value=" | ".join(row_value_cells),
                        src_ids=row_src_ids,
                        evidence_id="",
                        qa_status=qa_status,
                        qa_note=qa_note,
                    )

            # Figure-level claims
            for f in getattr(sec, "figures", []) or []:
                fig_id = getattr(f, "figure_id", "") or ""
                src_ids = _normalize_ids(getattr(f, "source_ids", []) or [])
                evidence_id = ";".join(fig_to_evidence_ids.get(fig_id, []))
                qa_status = "pass" if src_ids else "fail"
                qa_note = "" if src_ids else "missing source_id (or only S-TBD)"
                add_claim_row(
                    claim_id=f"{sec.section_id}:F:{fig_id}",
                    doc_target=doc_target,
                    section_id=raw_section_id,
                    section_title=sec.title,
                    report_anchor=fig_id,
                    object_type="figure",
                    field_path=f"{fig_id}",
                    value=str(getattr(f, "caption", "") or ""),
                    src_ids=src_ids,
                    evidence_id=evidence_id,
                    qa_status=qa_status,
                    qa_note=qa_note,
                )

        widths2 = {
            1: 34,  # claim_id
            2: 8,  # doc_target
            3: 18,  # section_id
            4: 26,  # section_title
            5: 18,  # report_anchor
            6: 12,  # object_type
            7: 26,  # field_path
            8: 60,  # value
            9: 10,
            10: 10,
            11: 18,
            12: 10,
            13: 28,  # source_id
            14: 16,
            15: 22,
            16: 18,
            17: 12,
            18: 10,
            19: 26,
            20: 26,
        }
        _set_column_widths(ws2, widths2)

        wrap_cols2 = {4, 8, 13, 19, 20}
        for row in ws2.iter_rows(min_row=2):
            for i, cell in enumerate(row, start=1):
                cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols2))

    # Canonical 4-sheet registry (P4-3): SOURCE_CATALOG/EVIDENCE_INDEX/USAGE_REGISTER/VALIDATION_SUMMARY.

    # SOURCE_CATALOG
    sc = wb.create_sheet("SOURCE_CATALOG")
    sc_header = [
        "src_id",
        "type",
        "title",
        "publisher",
        "ref",
        "published",
        "updated",
        "accessed",
        "version",
        "reliability",
        "confidential",
        "license",
        "citation_ko",
        "notes",
        "used",
        "used_count",
        "used_in_docs",
    ]
    sc.append(sc_header)
    _apply_header_style(sc)

    def _extra_dict(obj: Any) -> dict[str, Any]:
        d = getattr(obj, "model_extra", None) or {}
        return d if isinstance(d, dict) else {}

    def _bool_to_cell(v: Any) -> str:
        if v is True:
            return "TRUE"
        if v is False:
            return "FALSE"
        return ""

    for s in sources.sources:
        ex = _extra_dict(s)
        citation = ex.get("citation")
        citation_ko = ""
        if isinstance(citation, dict):
            citation_ko = str(citation.get("short") or citation.get("full") or "").strip()
        elif isinstance(ex.get("citation_ko"), str):
            citation_ko = str(ex.get("citation_ko") or "").strip()

        src_id = s.id
        used = "Y" if src_id in used_ids else ""
        used_count = int(citation_counts.get(src_id, 0))
        used_in_docs = ", ".join(sorted(used_in_sections.get(src_id, set())))

        sc.append(
            [
                src_id,
                str(s.type or "").strip(),
                str(s.title or "").strip(),
                str(s.publisher or "").strip(),
                str(s.file_path or s.file_or_url or "").strip(),
                str(s.date or "").strip(),
                str(ex.get("updated") or ex.get("updated_at") or "").strip(),
                str(s.access_date or "").strip(),
                str(ex.get("version") or "").strip(),
                str(ex.get("reliability") or "").strip(),
                _bool_to_cell(s.confidential if s.confidential is not None else ex.get("confidential")),
                str(ex.get("license") or "").strip(),
                citation_ko,
                str(s.note or s.notes or "").strip(),
                used,
                used_count,
                used_in_docs,
            ]
        )

    for mid in missing_ids:
        sc.append(
            [
                mid,
                "",
                "【미등록】sources.yaml에 Source ID 추가 필요",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "Y",
                int(citation_counts.get(mid, 0)),
                ", ".join(sorted(used_in_sections.get(mid, set()))),
            ]
        )

    widths_sc = {
        1: 16,
        2: 12,
        3: 38,
        4: 18,
        5: 34,
        6: 14,
        7: 14,
        8: 14,
        9: 12,
        10: 14,
        11: 12,
        12: 14,
        13: 26,
        14: 28,
        15: 8,
        16: 10,
        17: 30,
    }
    _set_column_widths(sc, widths_sc)
    wrap_cols_sc = {3, 5, 13, 14, 17}
    for row in sc.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols_sc))

    # EVIDENCE_INDEX
    ei = wb.create_sheet("EVIDENCE_INDEX")
    ei_header = [
        "evidence_id",
        "src_id",
        "evidence_type",
        "retrieved_at",
        "request_url",
        "request_params",
        "hash_sha1",
        "path",
    ]
    ei.append(ei_header)
    _apply_header_style(ei)

    def _append_evidence_index_row(r: dict[str, Any]) -> None:
        evidence_id = _tfv(r.get("evidence_id"))
        if not evidence_id:
            return
        path = _tfv(r.get("file_path"))
        src_id = _tfv(r.get("src_id")) or _tfv(r.get("src_ids"))
        if not src_id:
            src_id = _row_src_ids(r)

        def _stable_json(v: Any) -> str:
            try:
                return json.dumps(v, ensure_ascii=False, separators=(",", ":"), sort_keys=True)
            except Exception:
                return str(v)

        retrieved_at = _tfv(r.get("retrieved_at")) or _tfv(r.get("ingested_at"))
        request_url = _tfv(r.get("request_url"))
        request_params_raw = r.get("request_params")
        if isinstance(request_params_raw, (dict, list)):
            request_params = _stable_json(request_params_raw)
        else:
            request_params = _tfv(request_params_raw)

        note_text = _tfv(r.get("note"))
        if note_text.startswith("{") and note_text.endswith("}"):
            try:
                note_obj = json.loads(note_text)
            except Exception:
                note_obj = None
            if isinstance(note_obj, dict):
                if not retrieved_at:
                    retrieved_at = str(note_obj.get("retrieved_at") or note_obj.get("ingested_at") or "").strip()
                if not request_url:
                    request_url = str(note_obj.get("request_url") or note_obj.get("url") or "").strip()
                if not request_params:
                    rp = note_obj.get("request_params") or note_obj.get("params")
                    if isinstance(rp, (dict, list)):
                        request_params = _stable_json(rp)
                    elif rp is not None:
                        request_params = str(rp).strip()

        hash_ = _extract_hash_from_note(note_text)
        ei.append(
            [
                evidence_id,
                src_id,
                _tfv(r.get("evidence_type")),
                retrieved_at,
                request_url,
                request_params,
                hash_,
                path,
            ]
        )

    if isinstance(attachments_manifest, list):
        for r in attachments_manifest:
            if isinstance(r, dict):
                _append_evidence_index_row(r)
    if isinstance(derived_manifest, list):
        for r in derived_manifest:
            if isinstance(r, dict):
                _append_evidence_index_row(r)

    if getattr(case, "assets", None):
        for a in case.assets:
            try:
                evidence_id = str(getattr(a, "asset_id", "") or "").strip()
                if not evidence_id:
                    continue
                src_ids = ";".join(_normalize_ids([str(s) for s in getattr(a, "source_ids", []) or []]))
                ei.append(
                    [
                        evidence_id,
                        src_ids,
                        str(getattr(a, "type", "") or "").strip(),
                        "",
                        "",
                        "",
                        "",
                        str(getattr(a, "file_path", "") or "").strip(),
                    ]
                )
            except Exception:
                continue

    widths_ei = {1: 20, 2: 18, 3: 16, 4: 20, 5: 30, 6: 34, 7: 66, 8: 44}
    _set_column_widths(ei, widths_ei)
    wrap_cols_ei = {5, 6, 7, 8}
    for row in ei.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols_ei))

    # USAGE_REGISTER
    ur = wb.create_sheet("USAGE_REGISTER")
    ur_header = [
        "use_id",
        "report",
        "anchor",
        "section_path",
        "content_kind",
        "content_id",
        "claim_text",
        "value",
        "unit",
        "src_ids",
        "evidence_ids",
        "rendered_citation",
        "qa_flag",
        "notes",
    ]
    ur.append(ur_header)
    _apply_header_style(ur)

    def _append_usage_row(
        *,
        use_id: str,
        report: str,
        anchor: str,
        section_path: str,
        content_kind: str,
        content_id: str,
        claim_text: str,
        src_ids: list[str],
        evidence_ids: str = "",
        qa_flag: str = "OK",
        notes: str = "",
    ) -> None:
        ur.append(
            [
                use_id,
                report,
                anchor,
                section_path,
                content_kind,
                content_id,
                claim_text,
                "",
                "",
                ";".join(src_ids),
                evidence_ids,
                "",
                qa_flag,
                notes,
            ]
        )

    if draft is not None:
        for sec in draft.sections:
            if draft_section_id_allowlist is not None and sec.section_id not in draft_section_id_allowlist:
                continue
            doc_target, raw_section_id = _parse_doc_target(sec.section_id)
            report_cell = doc_target or (report_tag or "")
            section_path = raw_section_id

            for i, p in enumerate(sec.paragraphs, start=1):
                src_ids = _extract_citation_ids(p)
                qa_flag = "OK" if src_ids else "TODO"
                notes = "" if src_ids else "missing source_id (or only SRC-TBD)"
                _append_usage_row(
                    use_id=f"{sec.section_id}:P{i:04d}",
                    report=report_cell,
                    anchor=raw_section_id,
                    section_path=section_path,
                    content_kind="SENTENCE",
                    content_id=f"{raw_section_id}:P{i:04d}",
                    claim_text=_strip_trailing_citations(p),
                    src_ids=src_ids,
                    qa_flag=qa_flag,
                    notes=notes,
                )

            for t in getattr(sec, "tables", []) or []:
                table_id = getattr(t, "table_id", "") or ""
                src_col = None
                if getattr(t, "headers", None):
                    try:
                        if str(t.headers[-1]).strip().upper() == "SRC":
                            src_col = len(t.headers) - 1
                    except Exception:
                        src_col = None
                rows = getattr(t, "rows", []) or []
                for r_idx, row in enumerate(rows, start=1):
                    row_src_ids: list[str] = []
                    if src_col is not None and src_col < len(row):
                        row_src_ids = _split_joined_ids(str(row[src_col]))
                    if not row_src_ids:
                        row_src_ids = _normalize_ids(getattr(t, "source_ids", []) or [])
                    qa_flag = "OK" if row_src_ids else "TODO"
                    notes = "" if row_src_ids else "missing source_id (or only S-TBD)"
                    row_value_cells = [str(v) for j, v in enumerate(row) if j != src_col]
                    _append_usage_row(
                        use_id=f"{sec.section_id}:T:{table_id}:R{r_idx:04d}",
                        report=report_cell,
                        anchor=table_id,
                        section_path=section_path,
                        content_kind="TABLE",
                        content_id=f"{table_id}:R{r_idx:04d}",
                        claim_text=" | ".join(row_value_cells),
                        src_ids=row_src_ids,
                        qa_flag=qa_flag,
                        notes=notes,
                    )

            for f in getattr(sec, "figures", []) or []:
                fig_id = getattr(f, "figure_id", "") or ""
                if not fig_id:
                    continue
                src_ids = _normalize_ids(getattr(f, "source_ids", []) or [])
                evidence_id = ";".join(fig_to_evidence_ids.get(fig_id, []))
                qa_flag = "OK" if src_ids else "TODO"
                notes = "" if src_ids else "missing source_id (or only S-TBD)"
                if src_ids and not evidence_id:
                    qa_flag = "WARN"
                    notes = "missing evidence_id for this figure"
                _append_usage_row(
                    use_id=f"{sec.section_id}:F:{fig_id}",
                    report=report_cell,
                    anchor=fig_id,
                    section_path=section_path,
                    content_kind="FIGURE",
                    content_id=fig_id,
                    claim_text=str(getattr(f, "caption", "") or ""),
                    src_ids=src_ids,
                    evidence_ids=evidence_id,
                    qa_flag=qa_flag,
                    notes=notes,
                )

    widths_ur = {
        1: 34,
        2: 8,
        3: 18,
        4: 20,
        5: 12,
        6: 22,
        7: 60,
        8: 10,
        9: 10,
        10: 28,
        11: 22,
        12: 18,
        13: 12,
        14: 30,
    }
    _set_column_widths(ur, widths_ur)
    wrap_cols_ur = {7, 10, 11, 14}
    for row in ur.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols_ur))

    # VALIDATION_SUMMARY
    vs = wb.create_sheet("VALIDATION_SUMMARY")
    vs_header = ["report", "severity", "code", "message", "related_anchor", "related_sheet", "related_row_id"]
    vs.append(vs_header)
    _apply_header_style(vs)

    if validation_reports:
        for tag, rep in validation_reports:
            if rep is None:
                continue
            results = getattr(rep, "results", None)
            if not isinstance(results, list):
                continue
            for r in results:
                try:
                    severity = str(getattr(r, "severity", "") or "").strip()
                    code = str(getattr(r, "rule_id", "") or "").strip()
                    message = str(getattr(r, "message", "") or "").strip()
                    related_anchor = str(getattr(r, "related_anchor", "") or getattr(r, "path", "") or "").strip()
                    related_sheet = str(getattr(r, "related_sheet", "") or "").strip()
                    related_row_id = str(getattr(r, "related_row_id", "") or "").strip()
                except Exception:
                    continue
                if not (severity or code or message):
                    continue
                vs.append(
                    [
                        str(tag or report_tag or "").strip(),
                        severity,
                        code,
                        message,
                        related_anchor,
                        related_sheet,
                        related_row_id,
                    ]
                )

    widths_vs = {1: 10, 2: 10, 3: 16, 4: 72, 5: 22, 6: 16, 7: 14}
    _set_column_widths(vs, widths_vs)
    wrap_cols_vs = {4}
    for row in vs.iter_rows(min_row=2):
        for i, cell in enumerate(row, start=1):
            cell.alignment = Alignment(vertical="top", wrap_text=(i in wrap_cols_vs))

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
